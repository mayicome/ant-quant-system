import importlib.util
import os
import sys
import time
import argparse
from datetime import datetime
import re
from typing import Any, List, Tuple

_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
_TRIM_PNG_PATH = os.path.join(_ROOT_DIR, "tools", "trim_export_png.py")
if not os.path.isfile(_TRIM_PNG_PATH):
    raise ImportError(
        f"未找到 PNG 裁剪模块: {_TRIM_PNG_PATH}\n"
        "请将工程中的 tools/trim_export_png.py 与本脚本保持在同一目录结构下。"
    )
_spec_trim = importlib.util.spec_from_file_location("trim_export_png", _TRIM_PNG_PATH)
if _spec_trim is None or _spec_trim.loader is None:
    raise ImportError(f"无法加载: {_TRIM_PNG_PATH}")
_trim_png_mod = importlib.util.module_from_spec(_spec_trim)
_spec_trim.loader.exec_module(_trim_png_mod)
_trim_export_png_margins = _trim_png_mod.trim_export_png_margins

import akshare as ak
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from PyQt5.QtCore import QDate, QThread, QTimer, pyqtSignal
from PyQt5.QtWidgets import (
    QApplication,
    QDateEdit,
    QDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
)

MAX_SHOW_ROWS = 500


def _wps_range_value_to_matrix(val: Any) -> List[List[Any]]:
    """Excel COM Range.Value → 二维列表（行优先）。单格为标量；单行多为扁平元组。"""
    if val is None:
        return []
    if not isinstance(val, (tuple, list)):
        return [[val]]
    seq = list(val)
    if not seq:
        return []
    first = seq[0]
    if isinstance(first, (tuple, list)):
        return [list(row) for row in seq]
    return [list(seq)]


def _wps_cell_nonempty(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return bool(v.strip())
    if isinstance(v, float):
        try:
            import math

            if math.isnan(v):
                return False
        except Exception:
            pass
        return True
    if isinstance(v, bool):
        return True
    if isinstance(v, int):
        return True
    return True


def _wps_trim_used_matrix_bounds(mat: List[List[Any]]) -> Tuple[int, int, int, int]:
    """
    在矩阵内裁掉四周全空行列，返回相对矩阵左上角的 (r0, c0, height, width) 0-based。
    若全空则退回整张表范围。
    """
    if not mat:
        return 0, 0, 1, 1
    nrows = len(mat)
    ncols = max((len(r) for r in mat), default=0)
    if ncols == 0:
        return 0, 0, 1, 1
    grid = [list(r) + [None] * (ncols - len(r)) for r in mat]
    min_r, max_r = nrows, -1
    min_c, max_c = ncols, -1
    for i in range(nrows):
        for j in range(ncols):
            if _wps_cell_nonempty(grid[i][j]):
                min_r = min(min_r, i)
                max_r = max(max_r, i)
                min_c = min(min_c, j)
                max_c = max(max_c, j)
    if max_r < 0:
        return 0, 0, nrows, ncols
    return min_r, min_c, max_r - min_r + 1, max_c - min_c + 1


def apply_excel_borders(xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
    wb.save(xlsx_path)


def _header_two_lines(name: str) -> str:
    s = str(name or "").strip()
    if s == "机构态度":
        return "机构\n态度"
    if s == "北向态度":
        return "北向\n态度"
    if s == "机构/北向态度":
        return "机构/北向\n态度"
    if s == "机构家数":
        return "机构\n家数"
    if s == "机构买/净(万)":
        return "机构买/净\n(万)"
    if s == "北向买/净(万)":
        return "北向买/净\n(万)"
    m = re.match(r"^(.+?)([（(].+[）)])$", s)
    if not m:
        return s
    return f"{m.group(1)}\n{m.group(2)}"


def _fmt_wan_int(v: object) -> str:
    """金额（万）显示为整数，直接去掉小数点及之后部分。"""
    try:
        return str(int(float(v or 0)))
    except (TypeError, ValueError):
        return "0"


def _zfill_stock_code(code: object) -> str:
    """股票代码统一为 6 位数字字符串，避免 Excel/导图丢掉前导零。"""
    s = re.sub(r"\D", "", str(code or ""))
    if not s:
        return ""
    return s.zfill(6) if len(s) <= 6 else s[-6:]


def _parse_slash_pair(val: object) -> Tuple[float, float]:
    s = str(val or "").strip().replace(",", "").replace("，", "")
    if "/" not in s:
        return 0.0, 0.0
    a, b = s.split("/", 1)
    try:
        left = float(a.strip() or 0)
    except ValueError:
        left = 0.0
    try:
        right = float(b.strip() or 0)
    except ValueError:
        right = 0.0
    return left, right


def extract_lhb_seat_amounts(row: object) -> Tuple[float, float, float, float, float, float]:
    """从龙虎榜「主力情报」行提取席位金额（兼容旧分列与新「买/净」合并列）。

    Returns:
        (机构家数, 机构买万, 机构净万, 北向买万, 北向净万, 游资净万)
    """

    def _get(*names: str):
        for n in names:
            try:
                if hasattr(row, "index") and n in row.index:
                    return row.get(n)
            except Exception:
                pass
            if isinstance(row, dict) and n in row:
                return row.get(n)
        return None

    jg_cnt = float(pd.to_numeric(_get("机构家数"), errors="coerce") or 0)
    jg_pair = _get("机构买/净(万)", "机构买/净（万）")
    if jg_pair is not None and "/" in str(jg_pair):
        jg_buy, jg_net = _parse_slash_pair(jg_pair)
    else:
        jg_buy = float(pd.to_numeric(_get("机构买(万)"), errors="coerce") or 0)
        jg_net = float(pd.to_numeric(_get("机构净(万)"), errors="coerce") or 0)

    bx_pair = _get("北向买/净(万)", "北向买/净（万）")
    if bx_pair is not None and "/" in str(bx_pair):
        bx_buy, bx_net = _parse_slash_pair(bx_pair)
    else:
        bx_buy = float(pd.to_numeric(_get("北向买(万)"), errors="coerce") or 0)
        bx_net = float(pd.to_numeric(_get("北向净(万)"), errors="coerce") or 0)

    yz_net = float(pd.to_numeric(_get("游资净(万)", "游资净（万）"), errors="coerce") or 0)
    return jg_cnt, jg_buy, jg_net, bx_buy, bx_net, yz_net


def compute_lhb_strength_and_premium(
    jg_cnt: float,
    jg_buy: float,
    jg_net: float,
    bx_buy: float,
    bx_net: float,
    yz_net: float,
) -> Tuple[int, str]:
    """主力强度(0-100) + 次日溢价预测文案（强/中/弱三档）。

    规则基于 history_data 龙虎榜解析 × 次日 K 线回测（约 59 日 / 4k+ 样本）：
    - 强：机构净买 + 北向净买 + 机构≥2家（次日高开最稳）
    - 中：双买但机构不足 2 家 / 单边较强买 / 普通净买
    - 弱：机构多家未净买、双净卖、游资大卖等
    """
    jg_cnt_i = int(jg_cnt or 0)
    jg_buy_f = float(jg_buy or 0.0)
    jg_net_f = float(jg_net or 0.0)
    bx_net_f = float(bx_net or 0.0)
    yz_net_f = float(yz_net or 0.0)

    # —— 分档（决定文案）——
    if jg_net_f > 0 and bx_net_f > 0 and jg_cnt_i >= 2:
        tier = "强"
        premium = "强 -> 次日大概率高开"
    elif jg_cnt_i >= 2 and jg_net_f <= 0:
        tier = "弱"
        premium = "弱 -> 主力出逃"
    elif jg_net_f < 0 and bx_net_f < 0:
        tier = "弱"
        premium = "弱 -> 主力出逃"
    elif jg_net_f > 0 and bx_net_f > 0:
        # 双买但机构席位不足：历史开盘偏弱，勿标强
        tier = "中"
        premium = "中 -> 震荡"
    elif (jg_net_f > 2000 and bx_net_f >= 0) or (bx_net_f > 2000 and jg_net_f >= 0):
        tier = "中"
        premium = "中 -> 可冲高"
    elif jg_net_f > 0 or bx_net_f > 0:
        tier = "中"
        premium = "中 -> 震荡"
    elif yz_net_f < -20000 or (jg_buy_f > 3000 and jg_net_f <= 0):
        tier = "弱"
        premium = "弱 -> 小心低开"
    else:
        tier = "弱"
        premium = "弱 -> 小心低开"

    # —— 连续分（表格排序用；与分档大致对齐）——
    score = 40
    if jg_net_f > 5000:
        score += 22
    elif jg_net_f > 2000:
        score += 16
    elif jg_net_f > 0:
        score += 10
    elif jg_net_f < -5000:
        score -= 20
    elif jg_net_f < 0:
        score -= 10

    if jg_cnt_i >= 3:
        score += 10 if jg_net_f > 0 else -14
    elif jg_cnt_i >= 2:
        score += 7 if jg_net_f > 0 else -10

    if bx_net_f > 3000:
        score += 14
    elif bx_net_f > 0:
        score += 7
    elif bx_net_f < -3000:
        score -= 12
    elif bx_net_f < 0:
        score -= 6

    if jg_net_f > 0 and bx_net_f > 0:
        score += 12
        if jg_cnt_i >= 2:
            score += 12
    if jg_net_f < 0 and bx_net_f < 0:
        score -= 12

    if yz_net_f < -20000:
        score -= 10
    elif yz_net_f < -5000:
        score -= 5

    if jg_buy_f > 3000 and jg_net_f <= 0:
        score -= 8

    # 档位锚点，避免连续分与文案档严重错位
    anchor = {"强": 88, "中": 50, "弱": 18}.get(tier, 40)
    score = int(round(0.65 * score + 0.35 * anchor))
    score = max(0, min(100, score))
    return score, premium


def build_lhb_advanced_tables(date_str: str) -> dict:
    df_lhb_all = ak.stock_lhb_detail_em(start_date=date_str, end_date=date_str)
    stock_list = df_lhb_all[["代码", "名称"]].drop_duplicates().reset_index(drop=True)

    total_detail = []
    for _, row in stock_list.iterrows():
        code = _zfill_stock_code(row["代码"])
        name = row["名称"]
        if not code:
            continue
        try:
            df_one = ak.stock_lhb_stock_detail_em(symbol=code, date=date_str)
            df_one["股票代码"] = code
            df_one["股票名称"] = name
            df_one["是否机构"] = df_one["交易营业部名称"].astype(str).str.contains("机构专用", na=False)
            df_one["是否北向"] = df_one["交易营业部名称"].astype(str).str.contains("股通专用", na=False)
            total_detail.append(df_one)
        except Exception:
            pass
        time.sleep(0.2)

    if total_detail:
        df = pd.concat(total_detail, ignore_index=True)
    else:
        df = pd.DataFrame(
            columns=["股票代码", "股票名称", "交易营业部名称", "买入金额", "卖出金额", "净额", "是否机构", "是否北向"]
        )

    df["买(万)"] = (pd.to_numeric(df.get("买入金额"), errors="coerce").fillna(0.0) / 10000.0).round(2)
    df["卖(万)"] = (pd.to_numeric(df.get("卖出金额"), errors="coerce").fillna(0.0) / 10000.0).round(2)
    df["净(万)"] = (pd.to_numeric(df.get("净额"), errors="coerce").fillna(0.0) / 10000.0).round(2)

    stock_stat = []
    for (code, name), g in df.groupby(["股票代码", "股票名称"]):
        code = _zfill_stock_code(code)
        jg = g[g["是否机构"]]
        jg_buy = jg["买(万)"].sum()
        jg_sell = jg["卖(万)"].sum()
        jg_net = jg["净(万)"].sum()
        jg_cnt = int(jg.shape[0])

        bx = g[g["是否北向"]]
        bx_buy = bx["买(万)"].sum()
        bx_sell = bx["卖(万)"].sum()
        bx_net = bx["净(万)"].sum()

        yz = g[~g["是否机构"] & ~g["是否北向"]]
        yz_net = yz["净(万)"].sum()

        jg_type = "净买" if jg_net > 0 else ("净卖" if jg_net < 0 else "观望")
        bx_type = "净买" if bx_net > 0 else ("净卖" if bx_net < 0 else "观望")

        score, premium = compute_lhb_strength_and_premium(
            jg_cnt=jg_cnt,
            jg_buy=jg_buy,
            jg_net=jg_net,
            bx_buy=bx_buy,
            bx_net=bx_net,
            yz_net=yz_net,
        )

        stock_stat.append(
            [
                code,
                name,
                jg_cnt,
                f"{_fmt_wan_int(jg_buy)}/{_fmt_wan_int(jg_net)}",
                f"{_fmt_wan_int(bx_buy)}/{_fmt_wan_int(bx_net)}",
                _fmt_wan_int(yz_net),
                f"{jg_type}/{bx_type}",
                score,
                premium,
            ]
        )

    cols = [
        "代码",
        "名称",
        "机构家数",
        "机构买/净(万)",
        "北向买/净(万)",
        "游资净(万)",
        "机构/北向态度",
        "主力强度(0-100)",
        "次日溢价预测",
    ]
    df_result = pd.DataFrame(stock_stat, columns=cols).sort_values("主力强度(0-100)", ascending=False).reset_index(drop=True)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(script_dir, "history_data")
    os.makedirs(out_dir, exist_ok=True)
    out_fp = os.path.join(out_dir, f"龙虎榜解析_{date_str}.xlsx")
    df_result_export = df_result.copy()
    df_result_export.columns = [_header_two_lines(c) for c in df_result_export.columns]

    with pd.ExcelWriter(out_fp, engine="openpyxl") as writer:
        df_result_export.to_excel(writer, index=False, sheet_name="主力情报")
        df_lhb_all.to_excel(writer, index=False, sheet_name="原始数据_总榜")
        df.to_excel(writer, index=False, sheet_name="原始数据_席位明细")
    apply_excel_borders(out_fp)
    wb = load_workbook(out_fp)
    ws = wb["主力情报"] if "主力情报" in wb.sheetnames else None
    if ws is not None:
        # 代码列 + 含「/」的合并列强制文本，避免 WPS/Excel 把 003013 显示成 3013、把 净买/净买 显示异常
        text_cols = {1, 4, 5, 6, 7, 9}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c in text_cols:
                cell.number_format = "@"
            for r in range(2, (ws.max_row or 1) + 1):
                data_cell = ws.cell(row=r, column=c)
                if c in text_cols:
                    if data_cell.value is not None:
                        data_cell.value = str(data_cell.value)
                    data_cell.number_format = "@"
                    try:
                        data_cell.quote_prefix = True
                    except Exception:
                        pass
        ws.row_dimensions[1].height = 32
    wb.save(out_fp)

    jsonl_info = None
    try:
        from tools.export_lhb_to_jsonl import write_daily_lhb_jsonl_shards

        jsonl_info = write_daily_lhb_jsonl_shards(out_fp, str(date_str))
        print(
            "[龙虎榜] JSONL 分片已写:",
            jsonl_info.get("counts"),
            "→",
            jsonl_info.get("out_dir"),
        )
    except Exception as e:
        print("[龙虎榜] JSONL 分片写出失败:", e)

    return {
        "date": date_str,
        "summary_df": df_result,
        "raw_all_df": df_lhb_all,
        "raw_detail_df": df,
        "export_path": out_fp,
        "jsonl_info": jsonl_info,
    }


class Worker(QThread):
    success = pyqtSignal(dict)
    failed = pyqtSignal(str)

    def __init__(self, date_str: str):
        super().__init__()
        self.date_str = date_str

    def run(self):
        try:
            self.success.emit(build_lhb_advanced_tables(self.date_str))
        except Exception as e:
            self.failed.emit(str(e))


class Dialog(QDialog):
    def __init__(self, auto_run: bool = False):
        super().__init__()
        self._auto_run = bool(auto_run)
        self._auto_finished = False
        self.worker = None
        self.setWindowTitle("蚂蚁量化 - 龙虎榜挖掘分析（机构+北向+游资）")
        self.resize(1100, 780)

        layout = QVBoxLayout(self)
        top = QHBoxLayout()
        layout.addLayout(top)

        top.addWidget(QLabel("分析日期:"))
        self.date_edit = QDateEdit(self)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        top.addWidget(self.date_edit)

        self.run_btn = QPushButton("开始分析并导出Excel", self)
        self.run_btn.clicked.connect(self._run)
        top.addWidget(self.run_btn)
        top.addStretch(1)

        self.status = QLabel("待执行", self)
        layout.addWidget(self.status)

        self.result_text = QTextEdit(self)
        self.result_text.setReadOnly(True)
        self.result_text.setPlaceholderText("生成后的主力情报会显示在这里")
        layout.addWidget(self.result_text, 1)

        self.raw_text = QTextEdit(self)
        self.raw_text.setReadOnly(True)
        self.raw_text.setPlaceholderText("获取到的原始数据会显示在这里")
        layout.addWidget(self.raw_text, 1)

    def _auto_adjust_excel_column_widths(self, xlsx_path: str) -> bool:
        try:
            from openpyxl.utils import get_column_letter
        except Exception:
            return False
        if not os.path.exists(xlsx_path):
            return False
        try:
            wb = load_workbook(xlsx_path)
            ws = wb["主力情报"] if "主力情报" in wb.sheetnames else wb.active
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, max_row + 1):
                    val = ws.cell(row=r, column=c).value
                    text = "" if val is None else str(val).replace("\n", " ")
                    if len(text) > max_len:
                        max_len = len(text)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)
            wb.save(xlsx_path)
            return True
        except Exception:
            return False

    def _excel_to_png_via_wps(self, xlsx_path: str, png_path: str) -> bool:
        if not os.path.exists(xlsx_path):
            return False
        try:
            import pythoncom
            import win32com.client as win32
            from PIL import Image, ImageGrab
        except Exception:
            return False
        app = None
        wb = None
        inited = False
        try:
            pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
            inited = True
            for progid in ("ket.Application", "et.Application", "KET.Application", "ET.Application"):
                try:
                    app = win32.DispatchEx(progid)
                    if app is not None:
                        break
                except Exception:
                    app = None
            if app is None:
                return False
            app.Visible = False
            app.DisplayAlerts = False
            wb = app.Workbooks.Open(os.path.abspath(xlsx_path))
            try:
                ws = wb.Worksheets("主力情报")
            except Exception:
                ws = wb.Worksheets(1)
            ws.Activate()
            used = ws.UsedRange
            if used is None:
                return False

            base_r, base_c = int(used.Row), int(used.Column)
            mat = _wps_range_value_to_matrix(used.Value)
            dr, dc, nh, nw = _wps_trim_used_matrix_bounds(mat)
            sr, sc = base_r + dr, base_c + dc
            er, ec = sr + nh - 1, sc + nw - 1
            pic_rng = ws.Range(ws.Cells(sr, sc), ws.Cells(er, ec))

            try:
                app.ScreenUpdating = False
            except Exception:
                pass

            try:
                pic_rng.Columns.AutoFit()
            except Exception:
                try:
                    pic_rng.EntireColumn.AutoFit()
                except Exception:
                    used.EntireColumn.AutoFit()

            copied = False
            try:
                pic_rng.CopyPicture(1, 2)
                copied = True
            except Exception:
                copied = False
            if not copied:
                try:
                    pic_rng.CopyPicture()
                    copied = True
                except Exception:
                    copied = False
            if not copied:
                used.EntireColumn.AutoFit()
                try:
                    used.CopyPicture(1, 2)
                    copied = True
                except Exception:
                    used.CopyPicture()

            img = None
            for _ in range(10):
                time.sleep(0.2)
                try:
                    img = ImageGrab.grabclipboard()
                except Exception:
                    img = None
                if img is not None:
                    break
            if img is None or isinstance(img, list):
                return False
            if getattr(img, "mode", "") in ("RGBA", "LA"):
                white_bg = Image.new("RGB", img.size, "white")
                alpha = img.split()[-1]
                white_bg.paste(img, mask=alpha)
                to_save = white_bg
            elif getattr(img, "mode", "") != "RGB":
                to_save = img.convert("RGB")
            else:
                to_save = img.copy()
            try:
                to_save = _trim_export_png_margins(to_save)
            except Exception:
                pass
            to_save.save(png_path, "PNG")
            return os.path.exists(png_path)
        except Exception:
            return False
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            if inited:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _summary_df_to_png(self, df: pd.DataFrame, png_path: str) -> bool:
        try:
            import matplotlib.pyplot as plt
        except Exception:
            return False
        try:
            if df is None or df.empty:
                return False
            # 保证中文可显示；代码列补零
            plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS", "DejaVu Sans"]
            plt.rcParams["axes.unicode_minus"] = False
            show_df = df.head(80).copy()
            if "代码" in show_df.columns:
                show_df["代码"] = show_df["代码"].map(_zfill_stock_code)
            fig_h = max(4.5, min(22, 0.35 * (len(show_df) + 2)))
            fig_w = max(12, min(26, 1.25 * len(show_df.columns)))
            fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=180)
            fig.patch.set_facecolor("white")
            ax.set_facecolor("white")
            ax.axis("off")
            table = ax.table(
                cellText=show_df.astype(str).values.tolist(),
                colLabels=[str(c) for c in show_df.columns],
                loc="center",
                cellLoc="center",
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.2)
            for _, cell in table.get_celld().items():
                cell.set_edgecolor("#000000")
                cell.set_linewidth(0.6)
                cell.set_facecolor("white")
            fig.tight_layout()
            fig.savefig(png_path, bbox_inches="tight", facecolor="white")
            plt.close(fig)
            return os.path.exists(png_path)
        except Exception:
            return False

    def _run(self):
        if self.worker and self.worker.isRunning():
            return
        date_str = self.date_edit.date().toString("yyyyMMdd")
        self.run_btn.setEnabled(False)
        self.status.setText(f"正在分析 {date_str}，请稍候...")
        self.result_text.setPlainText("")
        self.raw_text.setPlainText("")
        self.worker = Worker(date_str)
        self.worker.success.connect(self._on_success)
        self.worker.failed.connect(self._on_failed)
        self.worker.finished.connect(lambda: self.run_btn.setEnabled(True))
        self.worker.start()

    def _on_success(self, data: dict):
        df_result = data["summary_df"]
        df_raw_all = data["raw_all_df"]
        df_raw_detail = data["raw_detail_df"]
        out_fp = data["export_path"]
        date_str = data["date"]

        self.status.setText(f"分析完成：{date_str}")
        result_show = df_result.head(MAX_SHOW_ROWS)
        raw_all_show = df_raw_all.head(MAX_SHOW_ROWS)
        raw_detail_show = df_raw_detail.head(MAX_SHOW_ROWS)

        self.result_text.setPlainText(
            f"主力情报（前{len(result_show)}行，共{len(df_result)}行）\n"
            f"{result_show.to_string(index=False)}\n\n已导出：{out_fp}"
        )
        self.raw_text.setPlainText(
            f"原始数据1 stock_lhb_detail_em（前{len(raw_all_show)}行，共{len(df_raw_all)}行）\n"
            f"{raw_all_show.to_string(index=False)}\n\n"
            f"原始数据2 stock_lhb_stock_detail_em 汇总（前{len(raw_detail_show)}行，共{len(df_raw_detail)}行）\n"
            f"{raw_detail_show.to_string(index=False)}"
        )
        png_path = os.path.splitext(out_fp)[0] + ".png"
        self._auto_adjust_excel_column_widths(out_fp)
        ok_png = self._excel_to_png_via_wps(out_fp, png_path)
        if not ok_png:
            ok_png = self._summary_df_to_png(df_result, png_path)
        exit_tip = "；10秒后自动退出" if self._auto_run else ""
        jsonl_info = data.get("jsonl_info") or {}
        jsonl_tip = ""
        if jsonl_info.get("paths"):
            jsonl_tip = "；JSONL：" + "、".join(
                os.path.basename(p) for p in (jsonl_info.get("paths") or {}).values()
            )
        if ok_png:
            self.status.setText(
                f"分析完成：{date_str}；已导出：{out_fp}、{png_path}{jsonl_tip}{exit_tip}。"
            )
        else:
            self.status.setText(
                f"分析完成：{date_str}；已导出Excel但图片导出失败{jsonl_tip}{exit_tip}。"
            )
        if self._auto_run and (not self._auto_finished):
            self._auto_finished = True
            QTimer.singleShot(10000, self.accept)

    def _on_failed(self, err: str):
        self.status.setText("分析失败")
        if self._auto_run:
            self._auto_finished = True
            self.status.setText(f"分析失败：{err}；10秒后自动退出。")
            QTimer.singleShot(10000, self.accept)
            return
        QMessageBox.warning(self, "龙虎榜挖掘分析失败", str(err))


def main():
    parser = argparse.ArgumentParser(description="龙虎榜挖掘分析（机构+北向+游资）")
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="启动后自动执行分析并导出 Excel，10秒后自动退出",
    )
    args, _unknown = parser.parse_known_args()

    app = QApplication(sys.argv)
    dlg = Dialog(auto_run=bool(args.auto_run))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg._run)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

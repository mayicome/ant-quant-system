import importlib.util
import os
import re
import sys
import warnings
import argparse
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import time

import akshare as ak
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QFontMetrics
from PyQt5.QtWidgets import (
    QApplication,
    QDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
)

MAX_RAW_ROWS_TO_SHOW = 600

_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
_TRIM_PNG_PATH = os.path.join(_ROOT_DIR, "tools", "trim_export_png.py")
if not os.path.isfile(_TRIM_PNG_PATH):
    raise ImportError(
        f"未找到 PNG 裁剪模块: {_TRIM_PNG_PATH}\n"
        "请将工程中的 tools/trim_export_png.py 与本脚本保持在同一目录结构下。"
    )
_spec_trim = importlib.util.spec_from_file_location("trim_export_png", _TRIM_PNG_PATH)
if _spec_trim is None or _spec_trim.loader is None:
    raise ImportError(f"无法加载: {_TRIM_PNG_PATH}")
_trim_png_mod = importlib.util.module_from_spec(_spec_trim)
_spec_trim.loader.exec_module(_trim_png_mod)
_trim_image_outer_white_bands = _trim_png_mod.trim_export_png_margins


def _wps_range_value_to_matrix(val: Any) -> List[List[Any]]:
    """Excel COM Range.Value → 二维列表（行优先）。单格为标量；单行多为扁平元组。"""
    if val is None:
        return []
    if not isinstance(val, (tuple, list)):
        return [[val]]
    seq = list(val)
    if not seq:
        return []
    first = seq[0]
    if isinstance(first, (tuple, list)):
        return [list(row) for row in seq]
    return [list(seq)]


def _wps_cell_nonempty(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return bool(v.strip())
    if isinstance(v, float):
        try:
            import math

            if math.isnan(v):
                return False
        except Exception:
            pass
        return True
    if isinstance(v, bool):
        return True
    if isinstance(v, int):
        return True
    return True


def _wps_trim_used_matrix_bounds(mat: List[List[Any]]) -> Tuple[int, int, int, int]:
    """
    在矩阵内裁掉四周全空行列，返回相对矩阵左上角的 (r0, c0, height, width) 0-based。
    若全空则退回整张表范围。
    """
    if not mat:
        return 0, 0, 1, 1
    nrows = len(mat)
    ncols = max((len(r) for r in mat), default=0)
    if ncols == 0:
        return 0, 0, 1, 1
    grid = [list(r) + [None] * (ncols - len(r)) for r in mat]
    min_r, max_r = nrows, -1
    min_c, max_c = ncols, -1
    for i in range(nrows):
        for j in range(ncols):
            if _wps_cell_nonempty(grid[i][j]):
                min_r = min(min_r, i)
                max_r = max(max_r, i)
                min_c = min(min_c, j)
                max_c = max(max_c, j)
    if max_r < 0:
        return 0, 0, nrows, ncols
    return min_r, min_c, max_r - min_r + 1, max_c - min_c + 1


def _excel_stock_cell_str_for_export(v: str) -> str:
    """6位纯数字代码前加零宽空格，减轻 Excel/WPS 智能格式提醒。"""
    s = str(v or "").strip()
    return ("\u200b" + s) if (len(s) == 6 and s.isdigit()) else s


def _normalize_yyyymmdd(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 8:
        return digits[:8]
    return ""


def _detect_data_date(df: pd.DataFrame) -> str:
    candidates = ["交易日期", "上榜日", "上榜日期", "trade_date"]
    for col in candidates:
        if col in df.columns:
            vals = [_normalize_yyyymmdd(x) for x in df[col].tolist()]
            vals = [x for x in vals if len(x) == 8]
            if vals:
                d = max(vals)
                try:
                    run_dt = datetime.now().date()
                    data_dt = datetime.strptime(d, "%Y%m%d").date()
                    if abs((run_dt - data_dt).days) > 30:
                        continue
                except Exception:
                    continue
                return d
    return ""


warnings.filterwarnings("ignore", category=DeprecationWarning, message=r".*sipPyTypeDict.*")


def _pick_recent_dates(df: pd.DataFrame, date_col: str, n: int) -> List[str]:
    if date_col not in df.columns:
        return []
    vals = [_normalize_yyyymmdd(x) for x in df[date_col].tolist()]
    vals = sorted({x for x in vals if len(x) == 8})
    return vals[-n:] if vals else []


def _is_a_share_code(code: object) -> bool:
    s = str(code or "").strip()
    if len(s) != 6 or (not s.isdigit()):
        return False
    return s.startswith(("000", "001", "002", "003", "300", "301", "600", "601", "603", "605", "688"))


def apply_excel_styles(xlsx_path: str, code_col_names: list[str]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    code_cols_1based = []
    header_map = {}
    for c in range(1, ws.max_column + 1):
        header_map[str(ws.cell(row=1, column=c).value or "").strip()] = c
    for n in code_col_names:
        cidx = header_map.get(str(n).strip())
        if cidx:
            code_cols_1based.append(cidx)

    for cidx in code_cols_1based:
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=cidx)
            cell.number_format = "@"
            qp = getattr(cell, "quote_prefix", None)
            if qp is not None:
                try:
                    cell.quote_prefix = True
                except Exception:
                    pass

    if code_cols_1based and ws.max_row >= 1:
        try:
            from openpyxl.worksheet.cell_range import CellRange
            from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

            ign = []
            for cidx in code_cols_1based:
                letter = get_column_letter(cidx)
                rng = f"{letter}1:{letter}{ws.max_row}"
                ign.append(IgnoredError(sqref=CellRange(rng), numberStoredAsText=True))
            ws.ignored_errors = IgnoredErrors(ignoredError=tuple(ign))
        except Exception:
            pass
    wb.save(xlsx_path)


def export_inst_tables_one_sheet(tables: Dict[str, pd.DataFrame], out_fp: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    data_font = Font(size=12)

    # 表头过长的列（表内第 4/5/7/8 列）拆成两行，便于缩窄列宽、放大数字
    header_two_line = {
        "当日净买占流通%": "当日净买\n占流通%",
        "三日净买占流通%": "三日净买\n占流通%",
        "机构净买当日万": "机构净买\n当日万",
        "机构净买三日万": "机构净买\n三日万",
        "当日净卖占流通%": "当日净卖\n占流通%",
        "三日净卖占流通%": "三日净卖\n占流通%",
        "机构净卖当日万": "机构净卖\n当日万",
        "机构净卖三日万": "机构净卖\n三日万",
        "买方次数": "买方\n次数",
        "卖方次数": "卖方\n次数",
    }
    wrap_header_names = set(header_two_line.keys())

    start_row = 1
    start_col = 1
    gap_rows = 2
    for title, df in tables.items():
        exp = df.reset_index().rename(columns={"index": "排名"})
        if "股票代码" in exp.columns:
            exp["股票代码"] = exp["股票代码"].map(_excel_stock_cell_str_for_export)

        headers = list(exp.columns)
        width = max(1, len(headers))
        end_col = start_col + width - 1
        header_row = start_row + 1

        ws.cell(row=start_row, column=start_col, value=title)
        ws.cell(row=start_row, column=start_col).font = title_font
        ws.cell(row=start_row, column=start_col).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        if end_col > start_col:
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=start_row,
                end_column=end_col,
            )

        for i, h in enumerate(headers):
            cell = ws.cell(
                row=header_row,
                column=start_col + i,
                value=header_two_line.get(str(h), h),
            )
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=str(h) in wrap_header_names,
            )
        ws.row_dimensions[header_row].height = 32

        for r_idx, row in enumerate(exp.itertuples(index=False), start=start_row + 2):
            for c_idx, val in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = data_font

        block_end_row = max(header_row, start_row + 1 + len(exp))
        for r in range(header_row, block_end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = border
                if r != header_row:
                    ws.cell(row=r, column=c).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

        for i, h in enumerate(headers):
            c = start_col + i
            letter = get_column_letter(c)
            name = str(h)
            if name in wrap_header_names:
                ws.column_dimensions[letter].width = 10
            elif name in ("股票代码", "股票名称"):
                ws.column_dimensions[letter].width = 12
            elif name == "流通市值":
                ws.column_dimensions[letter].width = 11
            elif name in ("排名",):
                ws.column_dimensions[letter].width = 9
            else:
                ws.column_dimensions[letter].width = 12

        start_row = block_end_row + gap_rows + 1

    wb.save(out_fp)


def build_longhubang_reports() -> dict:
    run_date_str = datetime.now().strftime("%Y%m%d")
    date_disp = datetime.now().strftime("%m月%d日")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(script_dir, "history_data")
    os.makedirs(out_dir, exist_ok=True)

    df_jg_raw = ak.stock_lhb_jgmx_sina()
    df_jg = df_jg_raw.copy()
    date_col = "日期" if "日期" in df_jg.columns else ("交易日期" if "交易日期" in df_jg.columns else "")
    code_col = "代码" if "代码" in df_jg.columns else ("股票代码" if "股票代码" in df_jg.columns else "")
    name_col = "名称" if "名称" in df_jg.columns else ("股票名称" if "股票名称" in df_jg.columns else "")
    buy_col = "机构买入额" if "机构买入额" in df_jg.columns else ("机构席位买入额" if "机构席位买入额" in df_jg.columns else "")
    sell_col = "机构卖出额" if "机构卖出额" in df_jg.columns else ("机构席位卖出额" if "机构席位卖出额" in df_jg.columns else "")

    if not all([date_col, code_col, name_col, buy_col, sell_col]):
        raise ValueError(f"机构数据缺少关键列，现有列: {list(df_jg.columns)}")

    df_jg["trade_date"] = df_jg[date_col].map(_normalize_yyyymmdd)
    df_jg["code"] = df_jg[code_col].astype(str).str.zfill(6)
    df_jg["name"] = df_jg[name_col].astype(str).str.strip()
    df_jg["buy_raw"] = pd.to_numeric(df_jg.get(buy_col), errors="coerce").fillna(0.0)
    df_jg["sell_raw"] = pd.to_numeric(df_jg.get(sell_col), errors="coerce").fillna(0.0)
    df_jg = df_jg[df_jg["code"].map(_is_a_share_code)].copy()
    dedup_cols = ["code", "name", "trade_date", "buy_raw", "sell_raw"]
    if "类型" in df_jg.columns:
        dedup_cols.append("类型")
    df_jg = df_jg.drop_duplicates(subset=[c for c in dedup_cols if c in df_jg.columns]).reset_index(drop=True)

    jg_dates = _pick_recent_dates(df_jg, "trade_date", 3)
    data_date_jg = jg_dates[-1] if jg_dates else _detect_data_date(df_jg)
    file_date_jg = data_date_jg or run_date_str

    jg_daily = (
        df_jg.groupby(["code", "name", "trade_date"], as_index=False)
        .agg(买入万=("buy_raw", "sum"), 卖出万=("sell_raw", "sum"))
    )
    jg_daily["净额万"] = (jg_daily["买入万"] - jg_daily["卖出万"]).round(2)

    jg_3d_daily = jg_daily[jg_daily["trade_date"].isin(jg_dates)].copy() if jg_dates else jg_daily.iloc[0:0].copy()
    jg_3d_stat = (
        jg_3d_daily.groupby(["code", "name"], as_index=False)
        .agg(
            买入三日万=("买入万", "sum"),
            卖出三日万=("卖出万", "sum"),
            净额三日万=("净额万", "sum"),
            买方次数=("净额万", lambda s: int((pd.to_numeric(s, errors="coerce").fillna(0.0) > 0).sum())),
            卖方次数=("净额万", lambda s: int((pd.to_numeric(s, errors="coerce").fillna(0.0) < 0).sum())),
        )
    )

    jg_day_daily = jg_daily[jg_daily["trade_date"] == data_date_jg].copy() if data_date_jg else jg_daily.iloc[0:0].copy()
    jg_day_stat = (
        jg_day_daily.groupby(["code", "name"], as_index=False)
        .agg(买入当日万=("买入万", "sum"), 卖出当日万=("卖出万", "sum"), 净额当日万=("净额万", "sum"))
    )

    jg_mix = jg_day_stat.merge(jg_3d_stat, on=["code", "name"], how="left")
    for c in ["买入三日万", "卖出三日万", "净额三日万", "买方次数", "卖方次数"]:
        jg_mix[c] = pd.to_numeric(jg_mix.get(c), errors="coerce").fillna(0.0)
    jg_mix["买方次数"] = jg_mix["买方次数"].astype(int)
    jg_mix["卖方次数"] = jg_mix["卖方次数"].astype(int)
    jg_3d = jg_3d_stat.copy()

    # 流通市值归一：净额(万)*1e4 / 流通市值(元) * 100 → 占流通%
    # 只读主力净流入 CSV，不回退 xtdata / miniQMT。
    from utils.main_force_inflow_rank import load_float_market_cap_yuan_map, yuan_to_display

    cap_map = load_float_market_cap_yuan_map(data_date_jg or run_date_str)

    def _attach_cap_ratios(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        caps: List[str] = []
        day_ratios: List[float] = []
        d3_ratios: List[float] = []
        for _, r in out.iterrows():
            code = str(r.get("code") or "").zfill(6)
            day_wan = float(pd.to_numeric(r.get("净额当日万"), errors="coerce") or 0.0)
            d3_wan = float(pd.to_numeric(r.get("净额三日万"), errors="coerce") or 0.0)
            cap = cap_map.get(code)
            if cap is not None and cap > 0:
                day_ratios.append(round(day_wan * 1e4 / cap * 100.0, 2))
                d3_ratios.append(round(d3_wan * 1e4 / cap * 100.0, 2))
                caps.append(yuan_to_display(cap))
            else:
                day_ratios.append(float("nan"))
                d3_ratios.append(float("nan"))
                caps.append("")
        out["流通市值"] = caps
        out["当日占流通%"] = day_ratios
        out["三日占流通%"] = d3_ratios
        return out

    jg_mix = _attach_cap_ratios(jg_mix)

    def _top_buy(df_mix: pd.DataFrame, topn: int) -> pd.DataFrame:
        base = df_mix.dropna(subset=["当日占流通%"]).copy()
        tmp = (
            base[base["净额当日万"] > 0]
            .sort_values("当日占流通%", ascending=False)
            .head(topn)
            .copy()
        )
        tmp = tmp.rename(
            columns={
                "当日占流通%": "当日净买占流通%",
                "三日占流通%": "三日净买占流通%",
                "净额当日万": "机构净买当日万",
                "净额三日万": "机构净买三日万",
            }
        )
        out_cols = [
            "code",
            "name",
            "当日净买占流通%",
            "三日净买占流通%",
            "流通市值",
            "机构净买当日万",
            "机构净买三日万",
            "买方次数",
            "卖方次数",
        ]
        out = tmp[out_cols].reset_index(drop=True)
        out = out.rename(columns={"code": "股票代码", "name": "股票名称"})
        out.index += 1
        return out

    def _top_sell(df_mix: pd.DataFrame, topn: int) -> pd.DataFrame:
        base = df_mix.dropna(subset=["当日占流通%"]).copy()
        tmp = (
            base[base["净额当日万"] < 0]
            .sort_values("当日占流通%", ascending=True)
            .head(topn)
            .copy()
        )
        tmp["机构净卖当日万"] = tmp["净额当日万"].abs().round(2)
        tmp["机构净卖三日万"] = tmp["净额三日万"].abs().round(2)
        tmp["当日净卖占流通%"] = tmp["当日占流通%"].abs().round(2)
        tmp["三日净卖占流通%"] = tmp["三日占流通%"].abs().round(2)
        out_cols = [
            "code",
            "name",
            "当日净卖占流通%",
            "三日净卖占流通%",
            "流通市值",
            "机构净卖当日万",
            "机构净卖三日万",
            "买方次数",
            "卖方次数",
        ]
        out = tmp[out_cols].reset_index(drop=True)
        out = out.rename(columns={"code": "股票代码", "name": "股票名称"})
        out.index += 1
        return out

    jg_tables = {
        "机构净买占流通比 TOP10（按当日）": _top_buy(jg_mix, 10),
        "机构净卖占流通比 TOP10（按当日）": _top_sell(jg_mix, 10),
    }
    merged_title = "机构四榜连排"
    merged_fp = os.path.join(out_dir, f"{merged_title}_{file_date_jg}.xlsx")
    export_inst_tables_one_sheet(jg_tables, merged_fp)
    jg_paths: Dict[str, str] = {title: merged_fp for title in jg_tables.keys()}
    jg_paths[merged_title] = merged_fp

    return {
        "run_date": run_date_str,
        "date_disp": date_disp,
        "jg_data_date": data_date_jg,
        "jg_field_mapping": {
            "date_col": date_col,
            "code_col": code_col,
            "name_col": name_col,
            "buy_col": buy_col,
            "sell_col": sell_col,
        },
        "jg_raw_rows": len(df_jg_raw),
        "jg_dedup_rows": len(df_jg),
        "jg_dates_used": jg_dates,
        "jg_tables": jg_tables,
        "jg_paths": jg_paths,
        "jg_raw_df": df_jg_raw,
    }


class LonghubangWorker(QThread):
    success = pyqtSignal(dict)
    failed = pyqtSignal(str)

    def run(self):
        try:
            result = build_longhubang_reports()
            self.success.emit(result)
        except Exception as e:
            self.failed.emit(str(e))


class LonghubangDialog(QDialog):
    def __init__(self, parent=None, auto_run: bool = False, wps_strict: bool = True):
        super().__init__(parent)
        self.setWindowTitle("蚂蚁量化 - 机构净买净卖排行（独立版）")
        self.resize(980, 760)
        self.worker = None
        self._auto_run = bool(auto_run)
        self._wps_strict = bool(wps_strict)

        layout = QVBoxLayout(self)
        top = QHBoxLayout()
        layout.addLayout(top)

        self.status = QLabel("待执行：点击“开始分析并导出机构榜单”", self)
        self.status.setFont(QFont("", 10))
        top.addWidget(self.status, 1)

        self.run_btn = QPushButton("开始分析并导出机构榜单", self)
        self.run_btn.clicked.connect(self._run_analysis)
        top.addWidget(self.run_btn)

        self.text = QTextEdit(self)
        self.text.setReadOnly(True)
        layout.addWidget(self.text, 1)

        self.raw_text = QTextEdit(self)
        self.raw_text.setReadOnly(True)
        self.raw_text.setPlaceholderText("原始数据会显示在这里")
        layout.addWidget(self.raw_text, 1)

    def _run_analysis(self):
        if self.worker and self.worker.isRunning():
            return
        self.run_btn.setEnabled(False)
        self.status.setText("正在抓取机构净买净卖数据并导出，请稍候...")
        self.text.setPlainText("")
        self.raw_text.setPlainText("")

        self.worker = LonghubangWorker()
        self.worker.success.connect(self._on_success)
        self.worker.failed.connect(self._on_failed)
        self.worker.finished.connect(lambda: self.run_btn.setEnabled(True))
        self.worker.start()

    def _on_success(self, result: dict):
        lines = []
        lines.append(f"运行日期: {result.get('run_date', '')}")
        lines.append(
            f"机构数据基准日: {result.get('jg_data_date', '')}；三日窗口: {', '.join(result.get('jg_dates_used', [])) or '无'}"
        )
        fmap = result.get("jg_field_mapping") or {}
        lines.append(
            "机构字段映射: "
            f"日期={fmap.get('date_col','')}, 代码={fmap.get('code_col','')}, 名称={fmap.get('name_col','')}, "
            f"买入={fmap.get('buy_col','')}, 卖出={fmap.get('sell_col','')}"
        )
        lines.append(f"机构原始行数: {result.get('jg_raw_rows', 0)}；去重后行数: {result.get('jg_dedup_rows', 0)}")
        lines.append("")

        for title, df in (result.get("jg_tables") or {}).items():
            lines.append("==================================================")
            lines.append(title)
            lines.append("==================================================")
            lines.append(df.to_string(index=True) if df is not None and not df.empty else "(无数据)")
            p = (result.get("jg_paths") or {}).get(title, "")
            if p:
                lines.append(f"已导出：{p}")
            lines.append("")

        self.text.setPlainText("\n".join(lines))

        def _fmt_raw(df: pd.DataFrame, title: str) -> str:
            if df is None:
                return f"{title}\n(无数据)"
            total = len(df)
            show_df = df.head(MAX_RAW_ROWS_TO_SHOW)
            tip = ""
            if total > MAX_RAW_ROWS_TO_SHOW:
                tip = f"\n(仅显示前 {MAX_RAW_ROWS_TO_SHOW} 行，共 {total} 行)\n"
            return (
                f"{title}\n"
                f"列: {list(df.columns)}\n"
                f"行数: {total}{tip}\n"
                f"{show_df.to_string(index=False)}"
            )

        raw_blocks = [
            _fmt_raw(result.get("jg_raw_df"), "原始数据1: ak.stock_lhb_jgmx_sina()"),
        ]
        self.raw_text.setPlainText("\n".join(raw_blocks))
        self.status.setText("分析完成，文件已导出到 history_data。")
        if self._auto_run:
            self._export_form_image_and_exit(result)

    def _on_failed(self, err: str):
        self.status.setText("分析失败")
        self.text.setPlainText(f"运行失败：{err}")
        self.raw_text.setPlainText("")
        if self._auto_run:
            # 自动模式下不弹框阻塞，直接退出
            QTimer.singleShot(10000, self.accept)
        else:
            QMessageBox.warning(self, "机构净买净卖排行失败", str(err))

    def _excel_to_png_via_wps(self, xlsx_fp: str, png_fp: str) -> None:
        """
        通过 WPS 表格（COM）把 Excel 原样导出为图片：
        - 按 UsedRange.Value 裁出「有内容」的最小矩形（去掉右侧/下方被格式撑大的空列空行）
        - 仅对该矩形列宽自适应后 CopyPicture，避免导图右侧大片空白
        - 从剪贴板取图保存 PNG
        """
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
            from PIL import ImageGrab  # type: ignore
        except Exception as e:
            raise RuntimeError(f"缺少依赖（pywin32/Pillow），无法调用 WPS 导图: {e}")

        last_error = None
        # WPS COM 偶发 RPC/服务器异常，做多次全流程重试
        for _attempt in range(1, 4):
            app = None
            wb = None
            try:
                # 明确用 STA 模式初始化 COM，减少 WPS 自动化不稳定
                try:
                    pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
                except Exception:
                    pythoncom.CoInitialize()

                last_err = None
                for progid in ("ket.Application", "et.Application", "KET.Application", "ET.Application"):
                    try:
                        app = win32com.client.DispatchEx(progid)
                        break
                    except Exception as _e:
                        last_err = _e
                        app = None
                if app is None:
                    raise RuntimeError(f"未找到可用的 WPS COM ProgID（ket/et.Application）: {last_err}")

                app.Visible = False
                app.DisplayAlerts = False
                wb = app.Workbooks.Open(os.path.abspath(xlsx_fp))
                ws = wb.Worksheets(1)
                ws.Activate()
                used = ws.UsedRange
                if used is None:
                    raise RuntimeError("工作表为空，无法导图")

                base_r, base_c = int(used.Row), int(used.Column)
                mat = _wps_range_value_to_matrix(used.Value)
                dr, dc, nh, nw = _wps_trim_used_matrix_bounds(mat)
                sr, sc = base_r + dr, base_c + dc
                er, ec = sr + nh - 1, sc + nw - 1
                pic_rng = ws.Range(ws.Cells(sr, sc), ws.Cells(er, ec))

                try:
                    app.ScreenUpdating = False
                except Exception:
                    pass

                # 只调有数据列的列宽，避免 UsedRange 右侧幽灵列把图片拉宽
                try:
                    pic_rng.Columns.AutoFit()
                except Exception:
                    try:
                        pic_rng.EntireColumn.AutoFit()
                    except Exception:
                        used.EntireColumn.AutoFit()

                # 两行表头列：压窄列宽并加高表头行，避免 AutoFit 又把列拉宽
                try:
                    for c in range(sc, ec + 1):
                        for r in range(sr, min(er, sr + 40) + 1):
                            cell = ws.Cells(r, c)
                            raw = cell.Value
                            if not isinstance(raw, str):
                                continue
                            if "\n" not in raw and "\r" not in raw:
                                continue
                            try:
                                cell.WrapText = True
                                cell.HorizontalAlignment = -4108  # xlCenter
                                cell.VerticalAlignment = -4108
                            except Exception:
                                pass
                            try:
                                ws.Rows(r).RowHeight = max(float(ws.Rows(r).RowHeight or 0), 32)
                            except Exception:
                                pass
                            try:
                                # WPS/Excel 列宽单位约等于字符数
                                if float(ws.Columns(c).ColumnWidth or 0) > 11:
                                    ws.Columns(c).ColumnWidth = 10
                            except Exception:
                                pass
                            break
                except Exception:
                    pass

                # CopyPicture 到剪贴板：1=xlScreen, 2=xlPicture
                copied = False
                try:
                    pic_rng.CopyPicture(1, 2)
                    copied = True
                except Exception:
                    copied = False

                if not copied:
                    try:
                        pic_rng.CopyPicture()
                        copied = True
                    except Exception:
                        copied = False

                if not copied:
                    used.EntireColumn.AutoFit()
                    try:
                        used.CopyPicture(1, 2)
                        copied = True
                    except Exception:
                        used.CopyPicture()
                        copied = True

                if not copied:
                    raise RuntimeError("CopyPicture 未成功")

                # 等待剪贴板稳定并取图
                img = None
                for _ in range(10):
                    time.sleep(0.2)
                    try:
                        img = ImageGrab.grabclipboard()
                    except Exception:
                        img = None
                    if img is not None:
                        break
                if img is None:
                    raise RuntimeError("剪贴板未拿到图片（ImageGrab.grabclipboard() 返回空）")
                from PIL import Image as _PILImage

                if getattr(img, "mode", "") in ("RGBA", "LA"):
                    _bg = _PILImage.new("RGB", img.size, "white")
                    _alpha = img.split()[-1]
                    _bg.paste(img, mask=_alpha)
                    _to_save = _bg
                elif getattr(img, "mode", "") != "RGB":
                    _to_save = img.convert("RGB")
                else:
                    _to_save = img.copy()
                try:
                    _to_save = _trim_image_outer_white_bands(_to_save)
                except Exception:
                    pass
                _to_save.save(os.path.abspath(png_fp), "PNG")
                if not os.path.exists(png_fp):
                    raise RuntimeError("PNG 保存失败")

                # 保存一次，确保 AutoFit 已写回
                wb.Save()
                return
            except Exception as e:
                last_error = e
                time.sleep(0.4)
            finally:
                try:
                    if wb is not None:
                        wb.Close(SaveChanges=False)
                except Exception:
                    pass
                try:
                    if app is not None:
                        app.Quit()
                except Exception:
                    pass
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        raise RuntimeError(f"WPS 导图失败: {last_error}")

    def _excel_to_png(self, xlsx_fp: str, png_fp: str, max_rows: int = 220, max_cols: int = 16) -> None:
        """将导出的 Excel 首个 sheet 渲染为 PNG 图片。"""
        wb = load_workbook(xlsx_fp, data_only=True)
        ws = wb.active
        rows: List[List[str]] = []
        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if ridx > max_rows:
                break
            vals = [("" if v is None else str(v)) for v in row[:max_cols]]
            # 去掉末尾空列，减小图片宽度
            while vals and vals[-1] == "":
                vals.pop()
            rows.append(vals)
        wb.close()

        if not rows:
            rows = [["(空表)"]]
        max_c = max(len(r) for r in rows) if rows else 1
        norm_rows = [r + [""] * (max_c - len(r)) for r in rows]

        fig_w = min(36, max(8, max_c * 1.6))
        fig_h = min(120, max(6, len(norm_rows) * 0.33))
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))
        ax.axis("off")
        table = ax.table(
            cellText=norm_rows,
            loc="center",
            cellLoc="center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1.0, 1.2)
        fig.tight_layout()
        fig.savefig(png_fp, dpi=180, bbox_inches="tight")
        plt.close(fig)

    def _export_form_image_and_exit(self, result: dict):
        """自动模式：把导出的 Excel 转成图片，并自动退出。"""
        try:
            out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history_data")
            os.makedirs(out_dir, exist_ok=True)
            d8 = str(result.get("jg_data_date") or result.get("run_date") or datetime.now().strftime("%Y%m%d"))
            log_fp = os.path.join(out_dir, f"机构净买净卖排行_导图日志_{d8}.txt")
            png_fp = os.path.join(out_dir, f"机构净买净卖排行_{d8}.png")
            xlsx_fp = (result.get("jg_paths") or {}).get("机构四榜连排")
            if not xlsx_fp:
                raise RuntimeError("未找到机构四榜连排导出文件路径")
            # 默认严格走 WPS；仅在非严格模式下才允许回退
            try:
                self._excel_to_png_via_wps(xlsx_fp, png_fp)
                with open(log_fp, "w", encoding="utf-8") as f:
                    f.write("method=WPS_COM\n")
                    f.write(f"xlsx={os.path.abspath(xlsx_fp)}\n")
                    f.write(f"png={os.path.abspath(png_fp)}\n")
                self.status.setText(f"自动运行完成（WPS导图），已导出图片：{os.path.basename(png_fp)}，即将退出")
            except Exception as e_wps:
                if self._wps_strict:
                    with open(log_fp, "w", encoding="utf-8") as f:
                        f.write("method=WPS_COM_FAILED_STRICT\n")
                        f.write(f"xlsx={os.path.abspath(xlsx_fp)}\n")
                        f.write(f"error={e_wps}\n")
                    raise
                self._excel_to_png(xlsx_fp, png_fp)
                with open(log_fp, "w", encoding="utf-8") as f:
                    f.write("method=MATPLOTLIB_FALLBACK\n")
                    f.write(f"xlsx={os.path.abspath(xlsx_fp)}\n")
                    f.write(f"png={os.path.abspath(png_fp)}\n")
                    f.write(f"wps_error={e_wps}\n")
                self.status.setText(
                    f"自动运行完成（WPS失败已回退），已导出图片：{os.path.basename(png_fp)}，即将退出"
                )
        except Exception as e:
            self.status.setText(f"自动模式截图失败：{e}，即将退出")
        finally:
            QTimer.singleShot(10000, self.accept)


def main():
    parser = argparse.ArgumentParser(description="机构净买净卖排行（独立版）")
    parser.add_argument("--auto-run", action="store_true", help="启动后自动运行，导出后截图并退出")
    parser.add_argument(
        "--allow-fallback",
        action="store_true",
        help="允许 WPS 导图失败时回退到 matplotlib（默认严格不回退）",
    )
    args, _unknown = parser.parse_known_args()

    app = QApplication(sys.argv)
    dlg = LonghubangDialog(auto_run=bool(args.auto_run), wps_strict=(not bool(args.allow_fallback)))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg._run_analysis)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

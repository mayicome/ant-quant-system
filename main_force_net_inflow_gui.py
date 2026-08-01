#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主力净流入统计（独立版）

读取 history_data/个股主力净流入/ 下「个股主力净流入_YYYYMMDD.csv」，
对今日 / 近2～3日：用「累计净流入 ÷ 截止日流通市值」排名，各取前 N 名。
"""

from __future__ import annotations

import csv
import argparse
import importlib.util
import io
import json
import os
import re
import sys
import warnings
from datetime import date, datetime
from typing import Dict, List, Optional, Set, Tuple

from utils.trading_day import is_tradeday, last_tradeday_on_or_before

_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
_TRIM_PNG_PATH = os.path.join(_ROOT_DIR, "tools", "trim_export_png.py")
if not os.path.isfile(_TRIM_PNG_PATH):
    raise ImportError(
        f"未找到 PNG 裁剪模块: {_TRIM_PNG_PATH}\n"
        "请将工程中的 tools/trim_export_png.py 与本脚本保持在同一目录结构下。"
    )
_spec_trim = importlib.util.spec_from_file_location("trim_export_png", _TRIM_PNG_PATH)
if _spec_trim is None or _spec_trim.loader is None:
    raise ImportError(f"无法加载: {_TRIM_PNG_PATH}")
_trim_png_mod = importlib.util.module_from_spec(_spec_trim)
_spec_trim.loader.exec_module(_trim_png_mod)
_trim_export_png_margins = _trim_png_mod.trim_export_png_margins

from PyQt5.QtCore import QBuffer, QByteArray, QDate, QIODevice, Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtGui import QColor, QBrush, QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QDateEdit,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

warnings.filterwarnings(
    "ignore",
    category=DeprecationWarning,
    message=r".*sipPyTypeDict.*",
)

from utils.main_force_inflow_path import (  # noqa: E402
    FLOW_FILE_PREFIX,
    FLOW_FILE_SUFFIX,
    flow_data_dir,
    list_flow_file_dates,
    resolve_flow_csv_path,
)

CONFIG_FILENAME = "main_force_net_inflow_types.json"
# 与 _window_dates_upto_anchor(..., 10) 一致：最多 10 个交易日；N 日统计不得超过该窗口
MAX_N_DAYS_ALLOWED = 10
# 固定窗口：今日 / 近2～3日累计净流入 ÷ 最近一日流通市值
DEFAULT_RANK_WINDOWS = (1, 2, 3)
DEFAULT_TOP_K = 5  # 近2～3日累计占比取前 K
DEFAULT_TODAY_TOP_K = 20  # 今日占比取前 K（与累计档分开）

# 主结果表按「统计类型」分组交替行背景（便于区分相邻类型块）
_ROW_BG_TYPE_A = QColor(210, 220, 235)  # 略深
_ROW_BG_TYPE_B = QColor(245, 247, 250)  # 略浅


def _qt_color_to_excel_pattern_fill(qc: QColor):
    """与界面 QColor 一致的实心填充，供导出 xlsx。"""
    from openpyxl.styles import PatternFill

    hx = f"{qc.red():02X}{qc.green():02X}{qc.blue():02X}"
    return PatternFill(fill_type="solid", fgColor=hx)

# 统一一行：前 4 列固定 (统计类型, 代码, 名称, 起始日期)，之后 k 列依次为 0日…-(k-1)日净流入（k = 配置中 N 的最大值，且 ≤ 窗口长度）
UnifiedRow = Tuple[str, ...]


def _zfill_code(raw: str) -> str:
    s = "".join(ch for ch in str(raw or "").strip() if ch.isdigit())
    if not s:
        return ""
    return s[:6].zfill(6) if len(s) >= 6 else s.zfill(6)


# 与 limit_up_gene_analysis_gui 导出一致：6 位代码前加零宽空格，减轻 Excel/WPS 绿色智能标记
_EXCEL_STOCK_TEXT_PREFIX = "\u200b"


def _excel_stock_cell_str_for_export(s: str) -> str:
    t = (s or "").strip()
    if len(t) == 6 and t.isdigit():
        return _EXCEL_STOCK_TEXT_PREFIX + t
    return t


def _parse_main_flow_to_yuan(text: object) -> float:
    """将「21.52亿」「2432.77万」等转为元（元）；负数支持。"""
    if text is None:
        return 0.0
    if isinstance(text, (int, float)):
        return float(text)
    s = str(text).strip().replace(",", "").replace(" ", "")
    if not s or s in ("--", "-", "nan", "None"):
        return 0.0
    neg = 1.0
    if s.startswith("-"):
        neg = -1.0
        s = s[1:].strip()
    # 亿
    m = re.match(r"^(\d+\.?\d*)\s*亿", s)
    if m:
        return neg * float(m.group(1)) * 1e8
    m = re.match(r"^(\d+\.?\d*)\s*万", s)
    if m:
        return neg * float(m.group(1)) * 1e4
    m = re.match(r"^(\d+\.?\d*)$", s)
    if m:
        return neg * float(m.group(1))
    return 0.0


def _yuan_to_display(yuan: float) -> str:
    a = abs(yuan)
    sign = "-" if yuan < 0 else ""
    if a >= 1e8:
        return f"{sign}{a / 1e8:.2f}亿"
    if a >= 1e4:
        # 「万」单位只保留整数，不显示小数
        return f"{sign}{int(a / 1e4)}万"
    if a > 0:
        return f"{sign}{a:.2f}元"
    return "0"


def _rank_type_label(n_days: int, top_k: int) -> str:
    if int(n_days) <= 1:
        return f"今日占比前{int(top_k)}"
    return f"近{int(n_days)}日累计占比前{int(top_k)}"


def _list_flow_file_dates(history_dir: str) -> List[str]:
    """列出可用交易日 YYYYMMDD；过滤非交易日。"""
    out: List[str] = []
    for part in list_flow_file_dates(history_dir):
        try:
            d = datetime.strptime(part, "%Y%m%d").date()
        except ValueError:
            continue
        if not is_tradeday(d):
            continue
        out.append(part)
    return out


def _flow_csv_for_date(history_dir: str, ymd: str) -> Optional[str]:
    return resolve_flow_csv_path(ymd, history_dir)


def _find_anchor_date(user_yyyymmdd: str, available: List[str]) -> Optional[str]:
    if not available:
        return None
    try:
        user_dt = datetime.strptime(user_yyyymmdd, "%Y%m%d").date()
    except ValueError:
        return None
    target = last_tradeday_on_or_before(user_dt)
    if target is None:
        return None
    target_str = target.strftime("%Y%m%d")
    earlier = [d for d in available if d <= target_str]
    if not earlier:
        return None
    return max(earlier)


def _window_dates_upto_anchor(available: List[str], anchor: str, n: int = 10) -> List[str]:
    """anchor 及之前、在 available 中的连续序列里，取最多 n 个交易日（含 anchor）。"""
    idx = available.index(anchor)
    start = max(0, idx - (n - 1))
    return available[start : idx + 1]


def _read_flow_csv(filepath: str) -> Dict[str, Tuple[float, str, str, Optional[float], Optional[float]]]:
    """
    返回 code6 -> (净流入元, 显示串, 证券名称, 净流入占流通%, 流通市值元)
    """
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None

    def _norm_key(k: object) -> str:
        return re.sub(r"\s+", "", str(k or "").replace("\n", "").replace("\xa0", ""))

    for enc in encodings:
        try:
            out: Dict[str, Tuple[float, str, str, Optional[float], Optional[float]]] = {}
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue
                key_map = {_norm_key(k): k for k in reader.fieldnames}

                def col(*needles: str) -> Optional[str]:
                    for nk, orig in key_map.items():
                        if all(n in nk for n in needles):
                            return orig
                    return None

                c_code = col("代码")
                c_name = col("名称")
                c_flow = col("今日主力净流入") or col("主力净流入")
                c_pct = col("净流入占流通")
                c_cap = col("流通市值")
                if not c_code or not c_flow:
                    continue
                for row in reader:
                    code = _zfill_code(row.get(c_code) or "")
                    if not code:
                        continue
                    raw = row.get(c_flow)
                    yuan = _parse_main_flow_to_yuan(raw)
                    name = (row.get(c_name) or "").strip() if c_name else ""
                    pct: Optional[float] = None
                    cap_yuan: Optional[float] = None
                    if c_cap:
                        cap_yuan = _parse_main_flow_to_yuan(row.get(c_cap))
                        if cap_yuan is not None and cap_yuan <= 0:
                            cap_yuan = None
                    if c_pct:
                        try:
                            v = row.get(c_pct)
                            if v is not None and str(v).strip() not in ("", "--", "nan", "None"):
                                pct = float(str(v).strip())
                        except (TypeError, ValueError):
                            pct = None
                    if pct is None and cap_yuan and yuan != 0:
                        pct = yuan / cap_yuan * 100.0
                    if (cap_yuan is None or cap_yuan <= 0) and pct and pct != 0 and yuan != 0:
                        cap_yuan = yuan / (pct / 100.0)
                    if code not in out:
                        disp = _yuan_to_display(yuan)
                        if pct is not None:
                            disp = f"{disp}({pct:.2f}%)"
                        out[code] = (yuan, disp, name, pct, cap_yuan)
            return out
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"读取失败: {filepath}; {last_err}")


def _date_fmt(ds: str) -> str:
    if len(ds) == 8:
        return f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
    return ds


def _date_mmdd(ds: str) -> str:
    if len(ds) == 8:
        return f"{ds[4:6]}-{ds[6:8]}"
    if len(ds) == 10 and "-" in ds:
        return ds[5:10]
    return ds


def analyze_window(
    history_dir: str,
    window_dates: List[str],
    type_configs: Optional[List[Dict[str, object]]] = None,
    *,
    top_k: int = DEFAULT_TOP_K,
    today_top_k: int = DEFAULT_TODAY_TOP_K,
    rank_windows: Optional[Tuple[int, ...]] = None,
) -> List[UnifiedRow]:
    """
    按「N 日累计净流入 / 最近一日流通市值」排名。
    今日窗口取前 today_top_k；近2～3日窗口取前 top_k。

    多日：把窗口内各日净流入金额相加，再除以截止日流通市值。
    全量日 CSV 下：窗口内任一日缺记录则跳过（停牌/未上市等）；
    旧版「≥3000万截断」日榜与全量日混用时，缺日仍不能当 0。

    每行: (统计类型, 代码, 名称, 累计占流通%, 0日…流入显示)
    """
    _ = type_configs  # 兼容旧签名
    windows = tuple(rank_windows) if rank_windows else DEFAULT_RANK_WINDOWS
    top_k = max(1, int(top_k or DEFAULT_TOP_K))
    today_top_k = max(1, int(today_top_k or DEFAULT_TODAY_TOP_K))

    by_code_date: Dict[str, Dict[str, Tuple[float, str, Optional[float], Optional[float]]]] = {}
    code_to_name: Dict[str, str] = {}
    all_codes: Set[str] = set()

    for d in window_dates:
        path = _flow_csv_for_date(history_dir, d)
        if not path:
            continue
        day_map = _read_flow_csv(path)
        for code, (yuan, disp, name, pct, cap) in day_map.items():
            all_codes.add(code)
            by_code_date.setdefault(code, {})[d] = (yuan, disp, pct, cap)
            if name:
                code_to_name[code] = name

    def flow_on(
        code: str, d: str
    ) -> Optional[Tuple[float, str, Optional[float], Optional[float]]]:
        return by_code_date.get(code, {}).get(d)

    out: List[UnifiedRow] = []
    W = window_dates
    if not W:
        return out

    latest_day = W[-1]
    max_n = max(windows) if windows else 1
    flow_cols = min(max_n, len(W), MAX_N_DAYS_ALLOWED)

    for n_days in windows:
        n_days = int(n_days)
        if n_days <= 0 or n_days > len(W):
            continue
        take_k = today_top_k if n_days <= 1 else top_k
        label = _rank_type_label(n_days, take_k)
        sub = W[-n_days:]
        scored: List[Tuple[float, str, str, str, Tuple[str, ...]]] = []

        for code in all_codes:
            latest = flow_on(code, latest_day)
            if not latest:
                continue
            cap = latest[3]
            if cap is None or cap <= 0:
                continue

            # 缺日不按 0：窗口内任一日不在日榜则不参与该 N 日排名
            day_flows = []
            incomplete = False
            for d in sub:
                fx = flow_on(code, d)
                if fx is None:
                    incomplete = True
                    break
                day_flows.append(fx)
            if incomplete:
                continue

            sum_yuan = sum(float(fx[0] or 0.0) for fx in day_flows)
            ratio = sum_yuan / cap * 100.0

            def _disp(offset: int) -> str:
                idx = len(W) - 1 - offset
                if idx < 0:
                    return ""
                d = W[idx]
                if d not in sub:
                    fx = flow_on(code, d)
                    return fx[1] if fx else ""
                fx = flow_on(code, d)
                return fx[1] if fx else "缺"

            flows = tuple(_disp(i) for i in range(flow_cols))
            scored.append(
                (
                    ratio,
                    code,
                    code_to_name.get(code, ""),
                    f"{ratio:.2f}",
                    flows,
                )
            )

        scored.sort(key=lambda x: (-x[0], x[1]))
        for ratio, code, nm, ratio_txt, flows in scored[:take_k]:
            out.append((label, code, nm, ratio_txt) + flows)

    return out


class MainForceAnalyzeThread(QThread):
    finished_ok = pyqtSignal(list, str, str)
    # rows, anchor_yyyymmdd, debug_msg
    failed = pyqtSignal(str)

    def __init__(
        self,
        history_dir: str,
        user_date: str,
        top_k: int = DEFAULT_TOP_K,
        today_top_k: int = DEFAULT_TODAY_TOP_K,
        parent=None,
    ):
        super().__init__(parent)
        self.history_dir = history_dir
        self.user_date = user_date
        self.top_k = int(top_k or DEFAULT_TOP_K)
        self.today_top_k = int(today_top_k or DEFAULT_TODAY_TOP_K)

    def run(self) -> None:
        try:
            available = _list_flow_file_dates(self.history_dir)
            if not available:
                self.failed.emit(
                    f"{flow_data_dir('history_data')} 下未找到「个股主力净流入_*.csv」文件。"
                )
                return
            anchor = _find_anchor_date(self.user_date, available)
            if not anchor:
                self.failed.emit(
                    f"在 {self.user_date} 及之前没有可用的主力净流入文件（请检查日期与文件命名）。"
                )
                return
            need = max(DEFAULT_RANK_WINDOWS) + 2
            window = _window_dates_upto_anchor(available, anchor, max(need, 10))
            rows = analyze_window(
                self.history_dir,
                window,
                top_k=self.top_k,
                today_top_k=self.today_top_k,
                rank_windows=DEFAULT_RANK_WINDOWS,
            )
            win_str = ",".join(_date_fmt(d) for d in window)
            dbg = (
                f"截止输入日={self.user_date}；锚定交易日={_date_fmt(anchor)}；"
                f"共载入 {len(window)} 个交易日：{win_str}；"
                f"排名=累计净流入/截止日流通市值；"
                f"今日前{self.today_top_k}、近2～3日前{self.top_k}；"
                f"近N日须窗口内每日均有记录（缺日不按0；全量日后多日才可比）"
            )
            self.finished_ok.emit(rows, anchor, dbg)
        except Exception as e:
            self.failed.emit(f"{type(e).__name__}: {e}")


class MainForceNetInflowDialog(QDialog):
    def __init__(self, parent=None, auto_run: bool = False):
        super().__init__(parent)
        self._auto_run = bool(auto_run)
        self._auto_export_done = False
        self._auto_last_export_xlsx = ""
        self._auto_last_export_png = ""
        self.setWindowTitle("蚂蚁量化 - 主力净流入统计（独立版）")
        self.resize(1040, 620)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        self._rows: List[UnifiedRow] = []
        self._anchor = ""
        self._thread: Optional[MainForceAnalyzeThread] = None
        self._top_k: int = DEFAULT_TOP_K
        self._today_top_k: int = DEFAULT_TODAY_TOP_K
        self._n_flow_cols: int = max(DEFAULT_RANK_WINDOWS)
        self.headers: List[str] = []
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self._config_path = os.path.join(script_dir, "data", CONFIG_FILENAME)

        root = QVBoxLayout(self)

        row = QHBoxLayout()
        root.addLayout(row)
        row.addWidget(QLabel("截止日期："))
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        anchor = last_tradeday_on_or_before(date.today())
        if anchor is not None:
            self.date_edit.setDate(QDate(anchor.year, anchor.month, anchor.day))
        else:
            self.date_edit.setDate(QDate.currentDate())
        row.addWidget(self.date_edit)

        row.addWidget(QLabel("今日取前："))
        self.today_top_k_spin = QSpinBox()
        self.today_top_k_spin.setRange(1, 100)
        self.today_top_k_spin.setValue(DEFAULT_TODAY_TOP_K)
        row.addWidget(self.today_top_k_spin)

        row.addWidget(QLabel("近2～3日取前："))
        self.top_k_spin = QSpinBox()
        self.top_k_spin.setRange(1, 50)
        self.top_k_spin.setValue(DEFAULT_TOP_K)
        row.addWidget(self.top_k_spin)

        self.btn_run = QPushButton("统计")
        row.addWidget(self.btn_run)
        self.btn_export = QPushButton("导出 Excel…")
        self.btn_export.setEnabled(False)
        row.addWidget(self.btn_export)
        row.addStretch()

        self.status = QLabel(
            f"今日占比前{DEFAULT_TODAY_TOP_K}；近2～3日累计占比前{DEFAULT_TOP_K}。"
            "口径：累计净流入÷截止日流通市值。流入列仅展示 0/-1/-2 日。"
            "近N日须窗口内每日均有记录（全量接口落盘后才有真实多日累计）。"
        )
        self.status.setWordWrap(True)
        root.addWidget(self.status)

        self.table = QTableWidget()
        self._apply_flow_headers(max(DEFAULT_RANK_WINDOWS))
        root.addWidget(self.table, 1)

        self.btn_run.clicked.connect(self._on_run)
        self.btn_export.clicked.connect(self._on_export)

        self._load_rank_config()

    def _apply_flow_headers(self, n_flow: int) -> None:
        n_flow = max(1, min(int(n_flow), MAX_N_DAYS_ALLOWED))
        self._n_flow_cols = n_flow
        flow_names = ["0日流入(占比)" if i == 0 else f"-{i}日流入(占比)" for i in range(n_flow)]
        self.headers = ["统计类型", "股票代码", "股票名称", "累计占流通%"] + flow_names
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)

    def _load_rank_config(self) -> None:
        top_k = DEFAULT_TOP_K
        today_top_k = DEFAULT_TODAY_TOP_K
        if os.path.isfile(self._config_path):
            try:
                with open(self._config_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    if raw.get("top_k") is not None:
                        top_k = int(raw.get("top_k") or DEFAULT_TOP_K)
                    if raw.get("today_top_k") is not None:
                        today_top_k = int(raw.get("today_top_k") or DEFAULT_TODAY_TOP_K)
                elif isinstance(raw, list):
                    top_k = DEFAULT_TOP_K
                    today_top_k = DEFAULT_TODAY_TOP_K
            except Exception:
                top_k = DEFAULT_TOP_K
                today_top_k = DEFAULT_TODAY_TOP_K
        self._top_k = max(1, min(50, top_k))
        self._today_top_k = max(1, min(100, today_top_k))
        self.top_k_spin.setValue(self._top_k)
        self.today_top_k_spin.setValue(self._today_top_k)

    def _save_rank_config(self) -> None:
        self._top_k = int(self.top_k_spin.value())
        self._today_top_k = int(self.today_top_k_spin.value())
        try:
            os.makedirs(os.path.dirname(self._config_path), exist_ok=True)
            with open(self._config_path, "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "today_top_k": self._today_top_k,
                        "top_k": self._top_k,
                        "windows": list(DEFAULT_RANK_WINDOWS),
                        "metric": "sum_inflow / latest_float_mktcap",
                    },
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
        except Exception:
            pass

    def _on_run(self) -> None:
        qd = self.date_edit.date()
        user_date = f"{qd.year():04d}{qd.month():02d}{qd.day():02d}"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        history_dir = os.path.join(script_dir, "history_data")
        self._save_rank_config()

        self.btn_run.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.status.setText("正在读取 CSV 并按累计占比排名，请稍候…")
        self._auto_export_done = False

        self._thread = MainForceAnalyzeThread(
            history_dir,
            user_date,
            top_k=self._top_k,
            today_top_k=self._today_top_k,
            parent=self,
        )
        self._thread.finished_ok.connect(self._on_done)
        self._thread.failed.connect(self._on_fail)
        self._thread.start()

    def _on_done(self, rows: list, anchor: str, dbg: str) -> None:
        self._rows = rows
        self._anchor = anchor
        if rows:
            self._fill_unified_table(rows)
        else:
            self._apply_flow_headers(max(DEFAULT_RANK_WINDOWS))
            self.table.setRowCount(0)
        parts: List[str] = []
        for n in DEFAULT_RANK_WINDOWS:
            k = self._today_top_k if int(n) <= 1 else self._top_k
            label = _rank_type_label(n, k)
            cnt = sum(1 for row in rows if row and row[0] == label)
            parts.append(f"「{label}」{cnt} 条")
        detail = "；".join(parts) if parts else ""
        self.status.setText(
            f"完成。{dbg}\n"
            f"共 {len(rows)} 条。{detail}"
        )
        self.btn_run.setEnabled(True)
        self.btn_export.setEnabled(True)
        if self._auto_run and (not self._auto_export_done):
            self._auto_export_done = True
            self._auto_export_excel_and_png()
            QTimer.singleShot(10000, self.accept)

    def _fill_unified_table(self, rows: List[UnifiedRow]) -> None:
        if rows:
            self._apply_flow_headers(len(rows[0]) - 4)
        self.table.setRowCount(len(rows))
        self.table.setAlternatingRowColors(False)
        prev_label: Optional[str] = None
        use_alt_palette = False  # False -> A 色，True -> B 色；类型切换时交替
        for r, tup in enumerate(rows):
            label = str(tup[0]) if tup else ""
            if r > 0 and label != prev_label:
                use_alt_palette = not use_alt_palette
            prev_label = label
            row_bg = QBrush(_ROW_BG_TYPE_B if use_alt_palette else _ROW_BG_TYPE_A)
            for c, val in enumerate(tup):
                text = val if val is not None else ""
                item = QTableWidgetItem(text)
                item.setBackground(row_bg)
                if c == 0:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                elif c == 1:
                    item.setTextAlignment(Qt.AlignCenter)
                elif c == 2:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                elif c == 3:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                elif c >= 4:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(r, c, item)
        self.table.resizeRowsToContents()

    def _on_fail(self, msg: str) -> None:
        self.status.setText(f"出错：{msg}")
        self.btn_run.setEnabled(True)
        self.btn_export.setEnabled(False)
        if self._auto_run:
            QTimer.singleShot(10000, self.accept)

    def _export_table_to_png(self, png_path: str) -> bool:
        try:
            nrows, ncols = self.table.rowCount(), self.table.columnCount()
            if nrows <= 0 or ncols <= 0:
                return False
            tmp = QTableWidget(nrows, ncols)
            headers = []
            for c in range(ncols):
                hi = self.table.horizontalHeaderItem(c)
                headers.append(hi.text() if hi else "")
            tmp.setHorizontalHeaderLabels(headers)
            for r in range(nrows):
                for c in range(ncols):
                    it = self.table.item(r, c)
                    new_it = QTableWidgetItem(it.text() if it else "")
                    if it is not None:
                        new_it.setTextAlignment(it.textAlignment())
                        b = it.background()
                        if b is not None:
                            new_it.setBackground(b)
                    tmp.setItem(r, c, new_it)
            tmp.resizeColumnsToContents()
            tmp.resizeRowsToContents()
            total_w = tmp.verticalHeader().width() + 2 * tmp.frameWidth()
            total_h = tmp.horizontalHeader().height() + 2 * tmp.frameWidth()
            for c in range(ncols):
                total_w += tmp.columnWidth(c)
            for r in range(nrows):
                total_h += tmp.rowHeight(r)
            # 原先用 780 下限会把窄表强行拉宽，导出 PNG 右侧一大片空白
            total_w = max(total_w, 80)
            total_h = max(total_h, 60)
            tmp.resize(total_w, total_h)
            tmp.setStyleSheet("QTableWidget { background: #ffffff; }")
            tmp.viewport().setStyleSheet("background: #ffffff;")
            QApplication.processEvents()

            pix = QPixmap(tmp.size())
            pix.fill(Qt.white)
            tmp.render(pix)
            try:
                from PIL import Image

                ba = QByteArray()
                qbuf = QBuffer(ba)
                qbuf.open(QIODevice.WriteOnly)
                if not pix.save(qbuf, "PNG"):
                    qbuf.close()
                    return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
                qbuf.close()
                im = Image.open(io.BytesIO(bytes(ba))).convert("RGB")
                im = _trim_export_png_margins(im)
                im.save(png_path, "PNG")
                return os.path.exists(png_path)
            except Exception:
                return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
        except Exception:
            return False

    def _export_excel_file(self, *, show_dialogs: bool = True, path: str = "") -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if show_dialogs:
                QMessageBox.warning(
                    self,
                    "缺少依赖",
                    "导出 Excel 需要：pip install openpyxl",
                )
            return ""

        fill_type_a = _qt_color_to_excel_pattern_fill(_ROW_BG_TYPE_A)
        fill_type_b = _qt_color_to_excel_pattern_fill(_ROW_BG_TYPE_B)

        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        default_name = f"主力净流入统计_{self._anchor or 'export'}.xlsx"
        if show_dialogs:
            path, _ = QFileDialog.getSaveFileName(
                self,
                "导出 Excel",
                os.path.join(out_dir, default_name),
                "Excel 工作簿 (*.xlsx);;所有文件 (*.*)",
            )
            if not path:
                return ""
        else:
            if not path:
                path = os.path.join(out_dir, default_name)
            os.makedirs(os.path.dirname(path), exist_ok=True)
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "主力净流入统计"[:31]

        header_font = Font(bold=True)
        header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        # 与界面列一致：类型左、代码中、名称左、起始日期中、各日净流入右
        ncols = len(self.headers)
        col_h_align = ["left", "center", "left", "center"] + ["right"] * (ncols - 4)

        # 第 2 列为「股票代码」，与涨停基因分析导出相同：文本格式 + 零宽空格 + quote_prefix + ignoredErrors
        code_col_1based = 2

        for c, h in enumerate(self.headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = cell_border
            if c == code_col_1based:
                cell.number_format = "@"

        prev_label: Optional[str] = None
        use_alt_palette = False
        for data_idx, tup in enumerate(self._rows):
            r = data_idx + 2
            label = str(tup[0]) if tup else ""
            if data_idx > 0 and label != prev_label:
                use_alt_palette = not use_alt_palette
            prev_label = label
            row_fill = fill_type_b if use_alt_palette else fill_type_a
            for c, val in enumerate(tup, start=1):
                v = val if val is not None else ""
                if c == code_col_1based:
                    v = _excel_stock_cell_str_for_export(str(v))
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = cell_border
                cell.fill = row_fill
                if c == code_col_1based:
                    cell.number_format = "@"
                    qp = getattr(cell, "quote_prefix", None)
                    if qp is not None:
                        try:
                            cell.quote_prefix = True
                        except Exception:
                            pass
                ha = col_h_align[c - 1]
                if "\n" in str(v):
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="top", horizontal=ha
                    )
                else:
                    cell.alignment = Alignment(vertical="center", horizontal=ha)

        n_data = len(self._rows)
        if n_data > 0:
            try:
                from openpyxl.worksheet.cell_range import CellRange
                from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

                letter = get_column_letter(code_col_1based)
                last_row = n_data + 1
                rng = f"{letter}1:{letter}{last_row}"
                ws.ignored_errors = IgnoredErrors(
                    ignoredError=(
                        IgnoredError(sqref=CellRange(rng), numberStoredAsText=True),
                    )
                )
            except Exception:
                pass

        # 第 1 列「统计类型」列宽（openpyxl 列宽约等于字符宽度）
        ws.column_dimensions[get_column_letter(1)].width = 20.0

        try:
            wb.save(path)
        except Exception as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", str(e))
            return ""
        if show_dialogs:
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        return path

    def _on_export(self) -> None:
        self._export_excel_file(show_dialogs=True)

    def _auto_export_excel_and_png(self) -> bool:
        if not self._rows:
            self.status.setText("自动运行完成，但无统计结果，未导出。")
            return False
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        base = f"主力净流入统计_{self._anchor or 'export'}"
        xlsx_path = os.path.join(out_dir, f"{base}.xlsx")
        png_path = os.path.join(out_dir, f"{base}.png")
        out_xlsx = self._export_excel_file(show_dialogs=False, path=xlsx_path)
        if not out_xlsx:
            self.status.setText("自动导出失败：Excel 导出失败。")
            return False
        ok_png = self._export_table_to_png(png_path)
        if not ok_png:
            self.status.setText("自动导出失败：图片导出失败。")
            return False
        self._auto_last_export_xlsx = out_xlsx
        self._auto_last_export_png = png_path
        self.status.setText(
            f"自动运行完成，已导出：{os.path.basename(out_xlsx)}、{os.path.basename(png_path)}；10秒后退出。"
        )
        return True


def main() -> None:
    parser = argparse.ArgumentParser(description="主力净流入统计（独立版）")
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="启动后自动统计并导出 Excel 和图片，10秒后自动退出",
    )
    args, _unknown = parser.parse_known_args()
    app = QApplication(sys.argv)
    dlg = MainForceNetInflowDialog(auto_run=bool(args.auto_run))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg._on_run)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

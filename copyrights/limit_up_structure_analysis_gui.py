#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
涨停基因分析：按所选股票池统计历史涨停基因（与当日是否涨停无必然关系）。

输入来源：
1) 全部今日涨停：读取 history_data 下最新「涨停板数据_*.csv」中的全部涨停股（剔除 ST）
2) 指定股票列表文件：txt/csv 每行或首列为代码，分析列表内全部股票（不要求当日涨停）
3) 直接输入代码：逗号分隔，同上

若指定股票当日未涨停，仍参与分析，在「首板统计」表中展示，今日连板数为 0。
"""

import csv
import html
import os
import re
import sys
import argparse
import warnings
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from PyQt5.QtCore import QDate, QThread, QTimer, pyqtSignal
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QDateEdit,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

# PyQt5 与新版 sip 在子类化 QObject 时会触发 sipPyTypeDict 弃用提示，属绑定层内部，应用代码无法消除根因
warnings.filterwarnings(
    "ignore",
    category=DeprecationWarning,
    message=r".*sipPyTypeDict.*",
)

try:
    import xtquant.xtdata as xtdata
except Exception:
    xtdata = None


def _to_full_stock_code(stock_code_6: str) -> str:
    """把 6 位股票代码映射为 xtdata 常用的全代码（.SH/.SZ/.BJ）。"""
    s = str(stock_code_6).strip()
    if "." in s:
        return s
    s = _zfill_6(s)
    if not s:
        return s
    if s.startswith(("0", "1", "3")):
        return f"{s}.SZ"
    if s.startswith("6"):
        return f"{s}.SH"
    if s.startswith(("8", "4")):
        return f"{s}.BJ"
    if s.startswith("920"):
        return f"{s}.BJ"
    return s


def _extract_open_close_map(stock_df) -> Dict[str, Tuple[float, float, float, float]]:
    """
    stock_df: xtdata.get_market_data_ex 返回的单只股票 DataFrame（列至少 open/high/low/close，index为时间戳或字符串）。
    返回：date_yyyymmdd -> (open, high, low, close)
    """
    out: Dict[str, Tuple[float, float, float, float]] = {}
    if stock_df is None or getattr(stock_df, "empty", True):
        return out
    if (
        "open" not in stock_df.columns
        or "high" not in stock_df.columns
        or "low" not in stock_df.columns
        or "close" not in stock_df.columns
    ):
        return out

    for idx in stock_df.index:
        try:
            if isinstance(idx, (int, float)):
                dt = datetime.fromtimestamp(float(idx) / 1000.0)
                ds = dt.strftime("%Y%m%d")
            elif isinstance(idx, str):
                ds = idx.replace("-", "")[:8]
            else:
                ds = str(idx).replace("-", "")[:8]
            if not ds or len(ds) != 8:
                continue

            o = float(stock_df.loc[idx, "open"])
            h = float(stock_df.loc[idx, "high"])
            l = float(stock_df.loc[idx, "low"])
            c = float(stock_df.loc[idx, "close"])
            out[ds] = (o, h, l, c)
        except Exception:
            continue
    return out


def _load_or_fetch_tick_df(stock_code_6: str, trade_date_yyyymmdd: str):
    """每次直接调用 QMT 拉取当日 tick 数据。"""
    if xtdata is None:
        return None

    full_code = _to_full_stock_code(stock_code_6)
    tick_start = f"{trade_date_yyyymmdd}000000"
    tick_end = f"{trade_date_yyyymmdd}235959"
    try:
        xtdata.download_history_data(full_code, "tick", tick_start, tick_end)
        tick_df_map = xtdata.get_market_data_ex(
            [],
            [full_code],
            period="tick",
            start_time=tick_start,
            end_time=tick_end,
            count=-1,
        )
        tick_df = tick_df_map.get(full_code) if isinstance(tick_df_map, dict) else None
    except Exception:
        tick_df = None

    return tick_df


def _zfill_6(code: str) -> str:
    s = "".join(ch for ch in str(code or "").strip() if ch.isdigit())
    if not s:
        return ""
    return s[:6].zfill(6) if len(s) >= 6 else s.zfill(6)


def _is_valid_stock_code_6(code: str) -> bool:
    """六位数字代码；排除占位/异常产生的 000000（_zfill_6('0') 会得到该值）。"""
    if not code or len(code) != 6 or not code.isdigit():
        return False
    return code != "000000"


# 导出 xlsx 时加在 6 位代码前，打断 Excel/WPS「纯数字当文本」的绿色角标；单元格显示仍为 6 位数字
_EXCEL_STOCK_TEXT_PREFIX = "\u200b"


def _excel_stock_cell_str_for_export(s: str) -> str:
    """仅用于写入 Excel：6 位全数字则前加零宽空格，避免智能标记；其它字符串原样。"""
    t = (s or "").strip()
    if len(t) == 6 and t.isdigit():
        return _EXCEL_STOCK_TEXT_PREFIX + t
    return t


def _read_codes_from_txt_or_csv(filepath: str) -> List[str]:
    """
    读取股票列表文件：
    - txt：每行一个代码
    - csv：默认取首列作为代码（允许存在表头）
    """
    codes: List[str] = []
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                for line in f:
                    line = (line or "").strip()
                    if not line:
                        continue
                    first = line.split(",")[0].strip()
                    if not first:
                        continue
                    # 跳过可能的表头
                    if not any(ch.isdigit() for ch in first):
                        continue
                    code = _zfill_6(first)
                    if code and _is_valid_stock_code_6(code):
                        codes.append(code)
            break
        except Exception as e:
            last_err = e
            continue

    if not codes and last_err:
        raise RuntimeError(f"读取股票列表失败: {filepath}; {type(last_err).__name__}: {last_err}")

    # 去重保持顺序
    seen: Set[str] = set()
    out: List[str] = []
    for c in codes:
        if c in seen:
            continue
        seen.add(c)
        out.append(c)
    return out


def _parse_codes_from_comma_text(text: str) -> List[str]:
    """从逗号/中文逗号/分号/空白分隔的字符串解析 6 位股票代码，去重保序。"""
    raw = (text or "").replace("，", ",").replace(";", ",").replace("；", ",")
    parts = [p.strip() for p in raw.split(",")]
    seen: Set[str] = set()
    out: List[str] = []
    for p in parts:
        if not p:
            continue
        code = _zfill_6(p)
        if not code or not _is_valid_stock_code_6(code):
            continue
        if code in seen:
            continue
        seen.add(code)
        out.append(code)
    return out


def _ratio_to_chinese_desc(ratio: float) -> str:
    """将 0~1 的晋级比例转为中文：优先常见分数（如三分之一），否则「百分之整数」。"""
    try:
        r = float(ratio)
    except (TypeError, ValueError):
        return str(ratio)
    if r <= 1e-9:
        return "百分之零"
    if r >= 1.0 - 1e-9:
        return "百分之百"
    candidates = [
        (1.0 / 2, "百分之五十"),
        (1.0 / 3, "三分之一"),
        (2.0 / 3, "三分之二"),
        (1.0 / 4, "四分之一"),
        (3.0 / 4, "四分之三"),
        (1.0 / 5, "五分之一"),
        (2.0 / 5, "五分之二"),
        (3.0 / 5, "五分之三"),
        (4.0 / 5, "五分之四"),
        (1.0 / 6, "六分之一"),
        (5.0 / 6, "六分之五"),
        (1.0 / 8, "八分之一"),
    ]
    tol = 0.008
    for val, label in candidates:
        if abs(r - val) < tol:
            return label
    pct = int(round(r * 100))
    pct = max(0, min(100, pct))
    if pct == 100:
        return "百分之百"
    if pct == 0:
        return "百分之零"
    return f"百分之{pct}"


def _fmt_px(x) -> str:
    """表格中展示价格，空/无效则空串。"""
    if x is None:
        return ""
    try:
        return f"{float(x):.2f}"
    except (TypeError, ValueError):
        return str(x)


def _html_esc(s: str) -> str:
    return html.escape(s or "", quote=False)


def _qtable_headers(table: QTableWidget) -> List[str]:
    n = table.columnCount()
    out: List[str] = []
    for c in range(n):
        hi = table.horizontalHeaderItem(c)
        out.append(hi.text() if hi else "")
    return out


def _qtable_row_texts(table: QTableWidget, row: int) -> List[str]:
    n = table.columnCount()
    out: List[str] = []
    for c in range(n):
        it = table.item(row, c)
        t = it.text() if it else ""
        out.append(t.replace("\r\n", "\n").replace("\r", "\n"))
    return out


def _detail_subtable_html(details: List[dict]) -> str:
    """单只股票走势明细：与弹窗表一致（日期、开盘、最高、最低、收盘、备注）。"""
    if not details:
        return '<p style="color:#888;margin:6px 0;font-size:13px;">（暂无走势明细）</p>'

    # 每次从“首板涨停”开始，到“后次日/跌停”结束，作为一个颜色段；两种颜色交替
    color_a = "#2f5597"
    color_b = "#7a4f01"
    seg_idx = 0
    in_cycle = False
    rows_html = []
    for rec in details:
        remark = str(rec.get("remark") or "")
        if "首板涨停" in remark:
            in_cycle = True
        row_color = color_a if (seg_idx % 2 == 0) else color_b
        rows_html.append(
            "<tr>"
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;color:{row_color};">{_html_esc(str(rec.get("date") or ""))}</td>'
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;text-align:right;color:{row_color};">{_html_esc(_fmt_px(rec.get("open")))}</td>'
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;text-align:right;color:{row_color};">{_html_esc(_fmt_px(rec.get("high")))}</td>'
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;text-align:right;color:{row_color};">{_html_esc(_fmt_px(rec.get("low")))}</td>'
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;text-align:right;color:{row_color};">{_html_esc(_fmt_px(rec.get("close")))}</td>'
            f'<td style="padding:4px 8px;border:1px solid #e0e0e0;text-align:left;color:{row_color};">{_html_esc(remark)}</td>'
            "</tr>"
        )
        if in_cycle and (("后次日" in remark) or ("跌停" in remark)):
            in_cycle = False
            seg_idx += 1
    return (
        '<table style="width:100%;max-width:640px;border-collapse:collapse;font-size:13px;margin-top:6px;">'
        "<thead><tr>"
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;">日期</th>'
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;">开盘</th>'
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;">最高</th>'
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;">最低</th>'
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;">收盘</th>'
        '<th style="padding:4px 8px;border:1px solid #ccc;background:#f0f0f0;text-align:left;">备注</th>'
        "</tr></thead><tbody>"
        + "".join(rows_html)
        + "</tbody></table>"
    )


def _wechat_stats_block_html(
    section_title: str,
    table: QTableWidget,
    detail_by_code: Dict[str, List[dict]],
    section_key: str = "sec",
    include_details: bool = True,
) -> str:
    """一块统计表：每只股票一行，并在行下方直接展示走势明细（公众号端无需折叠/展开）。"""
    if table.rowCount() == 0 or table.columnCount() == 0:
        return f"<h2>{_html_esc(section_title)}</h2><p>（无数据）</p>"
    # 分区配色：首板/连板使用不同背景色，长文里更容易定位
    if section_key == "fb":
        section_title_bg = "#eaf4ff"
        section_title_border = "#bcdcff"
        inner_head_bg = "#eaf4ff"
        row_bg = "#f7fbff"
    elif section_key == "st":
        section_title_bg = "#fff4e8"
        section_title_border = "#ffd6a8"
        inner_head_bg = "#fff4e8"
        row_bg = "#fffaf5"
    else:
        section_title_bg = "#f5f5f5"
        section_title_border = "#dddddd"
        inner_head_bg = "#f3f3f3"
        row_bg = "#fafafa"

    headers = _qtable_headers(table)

    # 导出公众号专用列处理：
    # - “股票”列：把「代码+名称」合并为 1 列（两行显示：上代码，下名称）
    # - 首板统计：去掉“今日连板数”（首板恒为1，信息冗余）
    drop_idx = set()
    if section_key == "fb":
        for i, h in enumerate(headers):
            if (h or "").strip() == "今日连板数":
                drop_idx.add(i)
                break

    # 期望原表前两列为：股票代码、股票名称
    has_code_name = len(headers) >= 2

    # “次日涨停次数 + 次日收涨次数 + 次日收跌次数” 合并为一列显示
    lu_idx = None
    up_idx = None
    down_idx = None
    for i, h in enumerate(headers):
        hh = (h or "").strip()
        if hh == "次日涨停次数":
            lu_idx = i
        elif hh == "次日收涨次数":
            up_idx = i
        elif hh == "次日收跌次数":
            down_idx = i

    display_headers: List[str] = []
    display_cols: List[Any] = []
    if has_code_name:
        display_headers.append("股票")
        # 从第3列开始继续输出（跳过 code/name 两列），并应用 drop_idx
        for i in range(2, len(headers)):
            if i in drop_idx:
                continue
            # 合并“次日涨停次数 / 次日收涨次数 / 次日收跌次数”
            if lu_idx is not None and up_idx is not None and down_idx is not None:
                if i == lu_idx:
                    display_headers.append("次日涨停/收涨/收跌次数")
                    display_cols.append(("LUDOWN", lu_idx, up_idx, down_idx))
                    continue
                if i == up_idx:
                    continue
                if i == down_idx:
                    continue
            display_headers.append(headers[i])
            display_cols.append(i)
    else:
        # 兼容：若表结构不是 code+name 开头，则不做合并，仅按 drop_idx 输出
        for i, h in enumerate(headers):
            if i in drop_idx:
                continue
            if lu_idx is not None and up_idx is not None and down_idx is not None:
                if i == lu_idx:
                    display_headers.append("次日涨停/收涨/收跌次数")
                    display_cols.append(("LUDOWN", lu_idx, up_idx, down_idx))
                    continue
                if i == up_idx:
                    continue
                if i == down_idx:
                    continue
            display_headers.append(h)
            display_cols.append(i)

    ncol = len(display_headers)
    body_parts: List[str] = []
    # 固定列宽，保证表头和数据列对齐
    if ncol <= 1:
        col_widths = [100.0]
    else:
        first_col = 24.0
        rest = (100.0 - first_col) / float(ncol - 1)
        col_widths = [first_col] + [rest] * (ncol - 1)
    colgroup_html = (
        "<colgroup>"
        + "".join(f'<col style="width:{w:.2f}%;">' for w in col_widths)
        + "</colgroup>"
    )
    thead_html = (
        "<thead><tr>"
        + "".join(
            f'<th style="padding:6px 8px;border:1px solid #ccc;background:{inner_head_bg};font-size:13px;">{_html_esc(h)}</th>'
            for h in display_headers
        )
        + "</tr></thead>"
    )

    # 总体情况（不展开）模式：单表头 + 多行，不再每只股票重复表头
    if not include_details:
        row_parts: List[str] = []
        for r in range(table.rowCount()):
            cells = _qtable_row_texts(table, r)
            inner_cells: List[str] = []
            if has_code_name:
                code_text = _html_esc((cells[0] if len(cells) > 0 else "").strip())
                name_text = _html_esc((cells[1] if len(cells) > 1 else "").strip())
                stock_cell = (
                    '<div style="line-height:1.25;">'
                    f'<div style="font-weight:600;">{code_text}</div>'
                    f'<div style="color:#666;font-size:12px;margin-top:2px;">{name_text}</div>'
                    "</div>"
                )
                inner_cells.append(
                    '<td style="padding:8px 10px;border:1px solid #eee;vertical-align:middle;">'
                    + stock_cell
                    + "</td>"
                )
                for src_i in display_cols:
                    if isinstance(src_i, tuple) and len(src_i) == 4 and src_i[0] == "LUDOWN":
                        lu_i = int(src_i[1]); up_i = int(src_i[2]); dn_i = int(src_i[3])
                        lu_v = (cells[lu_i] if lu_i < len(cells) else "").strip()
                        up_v = (cells[up_i] if up_i < len(cells) else "").strip()
                        dn_v = (cells[dn_i] if dn_i < len(cells) else "").strip()
                        cell_val = _html_esc(f"{lu_v} / {up_v} / {dn_v}")
                    else:
                        cell_val = _html_esc(cells[src_i] if src_i < len(cells) else "")
                    inner_cells.append(
                        f'<td style="padding:4px 6px;border:1px solid #eee;vertical-align:middle;">{cell_val}</td>'
                    )
            else:
                for idx, src_i in enumerate(display_cols):
                    if isinstance(src_i, tuple) and len(src_i) == 4 and src_i[0] == "LUDOWN":
                        lu_i = int(src_i[1]); up_i = int(src_i[2]); dn_i = int(src_i[3])
                        lu_v = (cells[lu_i] if lu_i < len(cells) else "").strip()
                        up_v = (cells[up_i] if up_i < len(cells) else "").strip()
                        dn_v = (cells[dn_i] if dn_i < len(cells) else "").strip()
                        cell_val = _html_esc(f"{lu_v} / {up_v} / {dn_v}")
                    else:
                        cell_val = _html_esc(cells[src_i] if src_i < len(cells) else "")
                    cell_style = "padding:8px 10px;border:1px solid #eee;vertical-align:middle;" if idx == 0 else "padding:4px 6px;border:1px solid #eee;vertical-align:middle;"
                    inner_cells.append(f'<td style="{cell_style}">{cell_val}</td>')
            row_parts.append(f'<tr style="background:{row_bg};">{"".join(inner_cells)}</tr>')
        return (
            f"<h2 style=\"margin:20px 0 10px 0;font-size:18px;padding:8px 10px;background:{section_title_bg};border:1px solid {section_title_border};\">{_html_esc(section_title)}</h2>"
            + '<table style="width:100%;border-collapse:collapse;max-width:100%;box-sizing:border-box;table-layout:fixed;">'
            + colgroup_html
            + thead_html
            + "<tbody>"
            + "".join(row_parts)
            + "</tbody></table>"
        )

    # 展开版：每只股票块内重复显示表头
    inner_thead_html = thead_html
    for r in range(table.rowCount()):
        cells = _qtable_row_texts(table, r)
        code = (cells[0] if cells else "").strip()
        details = list(detail_by_code.get(code) or [])
        inner_cells: List[str] = []

        if has_code_name:
            # “股票”列：两行展示（上代码，下名称）
            code_text = _html_esc((cells[0] if len(cells) > 0 else "").strip())
            name_text = _html_esc((cells[1] if len(cells) > 1 else "").strip())
            stock_cell = (
                '<div style="line-height:1.25;">'
                f'<div style="font-weight:600;">{code_text}</div>'
                f'<div style="color:#666;font-size:12px;margin-top:2px;">{name_text}</div>'
                "</div>"
            )
            inner_cells.append(
                '<td style="padding:8px 10px;border:1px solid #eee;vertical-align:middle;">'
                + stock_cell
                + "</td>"
            )
            # 其余列
            for src_i in display_cols:
                if isinstance(src_i, tuple) and len(src_i) == 4 and src_i[0] == "LUDOWN":
                    lu_i = int(src_i[1])
                    up_i = int(src_i[2])
                    dn_i = int(src_i[3])
                    lu_v = (cells[lu_i] if lu_i < len(cells) else "").strip()
                    up_v = (cells[up_i] if up_i < len(cells) else "").strip()
                    dn_v = (cells[dn_i] if dn_i < len(cells) else "").strip()
                    cell_val = _html_esc(f"{lu_v} / {up_v} / {dn_v}")
                else:
                    cell_val = _html_esc(cells[src_i] if src_i < len(cells) else "")
                inner_cells.append(
                    f'<td style="padding:4px 6px;border:1px solid #eee;vertical-align:middle;">{cell_val}</td>'
                )
        else:
            # 兼容：不合并时按 display_cols 输出
            for idx, src_i in enumerate(display_cols):
                if isinstance(src_i, tuple) and len(src_i) == 4 and src_i[0] == "LUDOWN":
                    lu_i = int(src_i[1])
                    up_i = int(src_i[2])
                    dn_i = int(src_i[3])
                    lu_v = (cells[lu_i] if lu_i < len(cells) else "").strip()
                    up_v = (cells[up_i] if up_i < len(cells) else "").strip()
                    dn_v = (cells[dn_i] if dn_i < len(cells) else "").strip()
                    cell_val = _html_esc(f"{lu_v} / {up_v} / {dn_v}")
                else:
                    cell_val = _html_esc(cells[src_i] if src_i < len(cells) else "")
                cell_style = "padding:8px 10px;border:1px solid #eee;vertical-align:middle;" if idx == 0 else "padding:4px 6px;border:1px solid #eee;vertical-align:middle;"
                inner_cells.append(f'<td style="{cell_style}">{cell_val}</td>')

        inner_cells = "".join(inner_cells)
        inner_row = f'<tr style="background:{row_bg};">{inner_cells}</tr>'
        inner_table = (
            '<table style="width:100%;border-collapse:collapse;font-size:14px;table-layout:fixed;">'
            + colgroup_html
            + inner_thead_html
            + "<tbody>"
            + inner_row
            + "</tbody></table>"
        )
        body_parts.append(
            "<tr><td colspan=\""
            + str(ncol)
            + '" style="padding:0;border:1px solid #ddd;">'
            '<div style="padding:8px 6px;">'
            + inner_table
            + "</div>"
            + (
                '<div style="padding:8px 12px 12px;max-height:9999px;overflow:hidden;border-top:1px solid #eee;background:#fff;">'
                + _detail_subtable_html(details)
                + "</div>"
                if include_details
                else ""
            )
            + "</td></tr>"
        )
    return (
        f"<h2 style=\"margin:20px 0 10px 0;font-size:18px;padding:8px 10px;background:{section_title_bg};border:1px solid {section_title_border};\">{_html_esc(section_title)}</h2>"
        '<table style="width:100%;border-collapse:collapse;max-width:100%;box-sizing:border-box;">'
        + "<tbody>"
        + "".join(body_parts)
        + "</tbody></table>"
    )


def _build_wechat_limitup_article_html(
    limitup_date: str,
    first_board_table: QTableWidget,
    streak_table: QTableWidget,
    detail_by_code: Dict[str, List[dict]],
    *,
    include_details: bool = True,
) -> str:
    date_disp = limitup_date or ""
    if len(date_disp) == 8 and date_disp.isdigit():
        date_disp = f"{date_disp[:4]}-{date_disp[4:6]}-{date_disp[6:8]}"
    intro = (
        "<p><strong>数据日期：</strong>"
        + _html_esc(date_disp or limitup_date or "—")
        + "</p>"
        '<p style="color:#666;font-size:13px;line-height:1.6;">'
        "以下为今天涨停的股票近一年来的统计情况，今天首板的在「首板统计」，今天连板的在「连板统计」。"
        "首板统计按股票代码升序；连板统计按今日连板数从大到小，同连板数再按代码升序。"
        "</p>"
    )
    fb = _wechat_stats_block_html("首板统计", first_board_table, detail_by_code, section_key="fb", include_details=include_details)
    st = _wechat_stats_block_html("连板统计", streak_table, detail_by_code, section_key="st", include_details=include_details)
    style = """
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, "PingFang SC", "Microsoft YaHei", sans-serif; font-size: 15px; color: #333; padding: 12px; max-width: 720px; margin: 0 auto; }
"""
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_html_esc("涨停基因统计 " + (limitup_date or ""))}</title>
<style>{style}</style>
</head>
<body>
{intro}
{fb}
{st}
<p style="margin-top:24px;color:#aaa;font-size:12px;">由 蚂蚁量化 · 涨停基因分析 导出</p>
</body>
</html>
"""


def _limit_ratio_for_board(code_6: str) -> float:
    """与统计线程内日K涨停判定一致的板块涨跌停比例（按代码前缀，不含 ST 名称）。"""
    s = _zfill_6(code_6)[:6]
    if s.startswith(("300", "301", "688", "689")):
        return 0.20
    if s.startswith(("8", "4", "920")):
        return 0.30
    return 0.10


def _load_all_a_stocks(csv_path: str) -> Dict[str, str]:
    """
    加载 data/all_a_stocks.csv：
    返回 code -> name
    只保留主A股 code 前缀：0/3/4/6/8/9
    """
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None

    for enc in encodings:
        try:
            with open(csv_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                code_col = "证券代码" if "证券代码" in (reader.fieldnames or []) else (reader.fieldnames or [None])[0]
                name_col = "证券简称" if "证券简称" in (reader.fieldnames or []) else None
                out: Dict[str, str] = {}
                for row in reader:
                    raw_code = row.get(code_col) if code_col else ""
                    code = _zfill_6(raw_code)
                    if not code:
                        continue
                    if not code.startswith(("0", "3", "4", "6", "8", "9")):
                        continue
                    name = ""
                    if name_col:
                        name = (row.get(name_col) or "").strip()
                    out[code] = name
                return out
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"加载 all_a_stocks.csv 失败: {csv_path}; {type(last_err).__name__}: {last_err}")


def _find_latest_limitup_file(history_dir: str, today_yyyymmdd: str) -> Tuple[str, str]:
    """
    找到 history_dir 下 <= today_yyyymmdd 的最新文件：
    返回 (filepath, yyyymmdd)
    """
    prefix = "涨停板数据_"
    suffix = ".csv"
    latest_date = ""
    latest_path = ""

    if not os.path.exists(history_dir):
        return "", ""

    for fname in os.listdir(history_dir):
        if not fname.startswith(prefix) or not fname.endswith(suffix):
            continue
        date_part = fname[len(prefix) : -len(suffix)].strip()  # 2026-03-20
        date_norm = date_part.replace("-", "")
        if len(date_norm) != 8 or not date_norm.isdigit():
            continue
        if date_norm > today_yyyymmdd:
            continue
        if date_norm >= latest_date:
            latest_date = date_norm
            latest_path = os.path.join(history_dir, fname)

    return latest_path, latest_date


def _load_limitup_codes(filepath: str) -> Dict[str, str]:
    """
    读取单日涨停板数据 csv：
    返回 code -> name
    """
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                code_col = "代码" if "代码" in (reader.fieldnames or []) else (reader.fieldnames or [None])[0]
                name_col = "名称" if "名称" in (reader.fieldnames or []) else None
                out: Dict[str, str] = {}
                for row in reader:
                    code = _zfill_6(row.get(code_col) or "")
                    if not code or not _is_valid_stock_code_6(code):
                        continue
                    name = (row.get(name_col) or "").strip() if name_col else ""
                    out[code] = name
                return out
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"读取涨停板数据失败: {filepath}; {type(last_err).__name__}: {last_err}")


def _load_limitup_codes_set(filepath: str) -> Set[str]:
    """只返回某日涨停股票 set（用于快速判断是否涨停）"""
    d = _load_limitup_codes(filepath)
    return set(d.keys())


def _extract_tick_preview_rows(raw_df, limit_n: Optional[int] = 200) -> Tuple[List[str], List[List[str]]]:
    """从 xtdata tick DataFrame 提取全部可用字段（含索引列）。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return [], []

    def _format_time_like(val) -> Optional[str]:
        if val is None:
            return None
        try:
            if isinstance(val, (int, float)):
                iv = int(val)
                # 毫秒时间戳
                if iv > 10_000_000_000:
                    return datetime.fromtimestamp(iv / 1000.0).strftime("%H:%M:%S")
                # 秒时间戳
                if iv > 1_000_000_000:
                    return datetime.fromtimestamp(iv).strftime("%H:%M:%S")
                # yyyymmddHHMMSS
                s = str(iv)
                if len(s) >= 14 and s[:8].isdigit():
                    dt = datetime.strptime(s[:14], "%Y%m%d%H%M%S")
                    return dt.strftime("%H:%M:%S")
            s = str(val).strip()
            if not s:
                return None
            if s.isdigit():
                iv = int(s)
                if len(s) >= 13:
                    return datetime.fromtimestamp(iv / 1000.0).strftime("%H:%M:%S")
                if len(s) == 10:
                    return datetime.fromtimestamp(iv).strftime("%H:%M:%S")
                if len(s) >= 14:
                    dt = datetime.strptime(s[:14], "%Y%m%d%H%M%S")
                    return dt.strftime("%H:%M:%S")
        except Exception:
            return None
        return None

    if limit_n is None:
        use_df = raw_df
    else:
        use_df = raw_df.tail(max(1, int(limit_n)))
    data_cols = [str(c) for c in list(getattr(use_df, "columns", []))]
    headers = ["tick_index"] + data_cols
    rows: List[List[str]] = []

    for idx, row in use_df.iterrows():
        row_vals: List[str] = [str(idx)]
        for c in data_cols:
            v = row.get(c, "")
            if v is None:
                row_vals.append("")
            else:
                cl = c.lower()
                if cl == "time" or cl.endswith("_time") or cl.endswith("timestamp"):
                    t_fmt = _format_time_like(v)
                    row_vals.append(t_fmt if t_fmt is not None else str(v))
                else:
                    row_vals.append(str(v))
        rows.append(row_vals)
    return headers, rows


def _calc_close_order_amount_yi(raw_df) -> Optional[float]:
    """按收盘附近 tick 估算封单金额（亿）：优先买一金额，其次买一价*买一量。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        # 有些接口把五档字段返回成列表/元组，买一取第 1 个
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        # 兼容 numpy array 等
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    def _pick(row, names: List[str]):
        for n in names:
            if n in row:
                v = row.get(n)
                if v is not None:
                    return v
        return None

    # 常见 tick 字段别名：兼容不同券商/接口
    amount_names = [
        "bidAmount1", "bidAmt1", "buyAmount1", "bid_amount1", "b1_amt",
        "bidAmount", "buyAmount",
    ]
    price_names = [
        "bidPrice1", "bid1", "buyPrice1", "b1_p",
        "bidPrice", "buyPrice",
    ]
    vol_names = [
        "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1",
        "bidVol", "buyVol", "bidVolume",
    ]

    def _infer_bidvol_to_shares_multiplier(df) -> float:
        """
        根据 tick 的 amount/volume/lastPrice 自动判断 volume 口径：
        - 若 amount 增量 / volume 增量 ≈ price，则 volume 是“股”，乘数=1
        - 若 amount 增量 / volume 增量 ≈ price*100，则 volume 是“手”，乘数=100
        """
        if df is None or getattr(df, "empty", True):
            return 100.0
        need_cols = {"amount", "volume", "lastPrice"}
        cols = set(str(c) for c in list(getattr(df, "columns", [])))
        if not need_cols.issubset(cols):
            return 100.0
        try:
            probe = df.tail(400)
            samples: List[float] = []
            prev_amt = None
            prev_vol = None
            for _, row in probe.iterrows():
                amt = _to_float(row.get("amount"))
                vol = _to_float(row.get("volume"))
                px = _to_float(row.get("lastPrice"))
                if amt is None or vol is None or px is None or px <= 0:
                    prev_amt, prev_vol = amt, vol
                    continue
                if prev_amt is None or prev_vol is None:
                    prev_amt, prev_vol = amt, vol
                    continue
                d_amt = amt - prev_amt
                d_vol = vol - prev_vol
                prev_amt, prev_vol = amt, vol
                if d_amt <= 0 or d_vol <= 0:
                    continue
                ratio = (d_amt / d_vol) / px
                if ratio > 0:
                    samples.append(ratio)
            if not samples:
                return 100.0
            samples.sort()
            mid = samples[len(samples) // 2]
            # 离 1 更近 -> 股；离 100 更近 -> 手
            return 1.0 if abs(mid - 1.0) <= abs(mid - 100.0) else 100.0
        except Exception:
            return 100.0

    vol_to_shares = _infer_bidvol_to_shares_multiplier(raw_df)

    # 收盘最后一笔可能是集合竞价/空档，回溯最后 200 笔找最近可用值
    probe_df = raw_df.tail(200)
    try:
        iter_rows = list(probe_df.iterrows())
    except Exception:
        return None

    for _, row in reversed(iter_rows):
        # 1) 直接金额字段
        amt_f = _to_float(_pick(row, amount_names))
        if amt_f is not None and amt_f > 0:
            return round(amt_f / 1e8, 4)

        # 2) 买一价 * 买一量
        p_f = _to_float(_pick(row, price_names))
        v_f = _to_float(_pick(row, vol_names))
        if p_f is None or v_f is None or p_f <= 0 or v_f <= 0:
            continue

        # 动态按股/手口径换算为“股”
        shares = v_f * vol_to_shares
        amt_yuan = p_f * shares
        if amt_yuan > 0:
            return round(amt_yuan / 1e8, 4)

    return None


def _calc_day_turnover_yuan(raw_df) -> Optional[float]:
    """从当日 tick 估算当日成交额（元）。优先累计成交额字段，其次逐笔金额求和。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    cols = [str(c) for c in list(getattr(raw_df, "columns", []))]
    # 候选：累计成交额（通常最后一笔就是全天累计）
    cum_cols = ["amount", "totalAmount", "turnover", "accAmount", "cumAmount", "成交额"]
    # 候选：逐笔成交额（需要求和）
    per_trade_cols = ["tradeAmount", "lastAmount", "成交金额", "单笔成交额"]

    def _pick_existing(candidates: List[str]) -> Optional[str]:
        lower_map = {c.lower(): c for c in cols}
        for name in candidates:
            if name in cols:
                return name
            k = name.lower()
            if k in lower_map:
                return lower_map[k]
        return None

    ccol = _pick_existing(cum_cols)
    if ccol:
        try:
            vals = raw_df[ccol]
            # 累计字段取最大值更稳，防止尾笔回落/空值
            mx = None
            for v in vals:
                fv = _to_float(v)
                if fv is None:
                    continue
                if mx is None or fv > mx:
                    mx = fv
            if mx is not None and mx > 0:
                return mx
        except Exception:
            pass

    pcol = _pick_existing(per_trade_cols)
    if pcol:
        try:
            s = 0.0
            ok = False
            for v in raw_df[pcol]:
                fv = _to_float(v)
                if fv is None:
                    continue
                s += fv
                ok = True
            if ok and s > 0:
                return s
        except Exception:
            pass

    return None


def _calc_float_market_cap_yuan(stock_code_6: str, raw_df) -> Optional[float]:
    """估算流通市值（元）= FloatVolume * 最新价。"""
    if xtdata is None:
        return None

    full_code = _to_full_stock_code(stock_code_6)
    try:
        info = xtdata.get_instrument_detail(full_code)
    except Exception:
        info = None
    if not isinstance(info, dict):
        return None

    def _to_float(x):
        try:
            return float(x)
        except Exception:
            return None

    float_volume = _to_float(info.get("FloatVolume"))
    if float_volume is None or float_volume <= 0:
        return None

    # 优先用当日 tick 最新价，缺失时回退前收
    px = None
    if raw_df is not None and not getattr(raw_df, "empty", True):
        try:
            probe = raw_df.tail(200)
            for _, row in reversed(list(probe.iterrows())):
                v = _to_float(row.get("lastPrice"))
                if v is not None and v > 0:
                    px = v
                    break
        except Exception:
            px = None
    if px is None:
        px = _to_float(info.get("PreClose"))
    if px is None or px <= 0:
        return None

    cap = float_volume * px
    return cap if cap > 0 else None


def _calc_order_volatility_coeff(raw_df) -> Optional[float]:
    """
    封单波动系数（全天 3 秒快照买一封单量序列）：
    coeff = std(bid1_vol_series) / mean(bid1_vol_series)
    """
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    # 兼容不同 tick 字段命名；若是五档数组，默认取第 1 档
    bidvol_fields = ["bidVol", "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1"]
    cols = set(str(c) for c in list(getattr(raw_df, "columns", [])))
    use_field = None
    for f in bidvol_fields:
        if f in cols:
            use_field = f
            break
    if use_field is None:
        return None

    vols: List[float] = []
    try:
        for v in raw_df[use_field]:
            fv = _to_float(v)
            if fv is None or fv <= 0:
                continue
            vols.append(fv)
    except Exception:
        return None

    n = len(vols)
    if n < 2:
        return None
    mean_v = sum(vols) / n
    if mean_v <= 0:
        return None
    var = sum((x - mean_v) ** 2 for x in vols) / n
    std = var ** 0.5
    return std / mean_v


def _classify_order_stability(vol_coeff: Optional[float]) -> Tuple[str, Optional[int]]:
    """
    根据封单波动系数划分封单稳定性（3档）并给分：
    - < 0.40  -> 封单极度稳定，10分
    - 0.40~0.80 -> 封单整体平稳，5分
    - > 0.80 -> 封单剧烈波动，0分
    """
    if vol_coeff is None:
        return "", None
    if vol_coeff < 0.40:
        return "封单极度稳定", 10
    if vol_coeff <= 0.80:
        return "封单整体平稳", 5
    return "封单剧烈波动", 0


def _score_seal_hardness(seal_ratio: Optional[float]) -> int:
    """
    封板硬度得分（0~50）：
    seal_ratio = 收盘封单金额 / 当日成交额（比值）
    """
    if seal_ratio is None or seal_ratio <= 0:
        return 0
    if seal_ratio < 0.03:
        return 0
    if seal_ratio <= 0.08:
        return 15
    if seal_ratio <= 0.15:
        return 35
    return 50


def _score_rush_intensity(rush_ratio: Optional[float]) -> int:
    """
    抢筹烈度得分（0~30）：
    rush_ratio = 收盘封单金额 / 流通市值（比值）
    """
    if rush_ratio is None or rush_ratio <= 0:
        return 0
    if rush_ratio < 0.002:
        return 0
    if rush_ratio <= 0.005:
        return 15
    return 30


def _score_close_order_amount(close_amt_yi: Optional[float]) -> int:
    """
    收盘封单绝对金额得分（0~20）：
    close_amt_yi 单位为亿。
    """
    if close_amt_yi is None or close_amt_yi <= 0:
        return 0
    if close_amt_yi < 0.3:  # 3000万
        return 0
    if close_amt_yi <= 1.0:  # 1亿
        return 10
    return 20


def _label_order_rating(total_score: int) -> str:
    """按综合得分映射封单评级。"""
    s = int(total_score)
    # 阈值微调：压缩中档混杂，增强高档/低档可区分性（便于后续回测验证）
    if s <= 18:
        return "🔴 虚封高危"
    if s <= 45:
        return "🟠 弱势封板"
    if s <= 68:
        return "🟡 中等封板"
    if s <= 92:
        return "🟢 强势封板"
    return "🔥 超强极致封板"


def _rating_score_from_label(label: str) -> int:
    m = {
        "🔴 虚封高危": 1,
        "🟠 弱势封板": 2,
        "🟡 中等封板": 3,
        "🟢 强势封板": 4,
        "🔥 超强极致封板": 5,
    }
    return int(m.get(label, 0))


def _confidence_tag(total_score: int, stability_score: int, trend_score: int) -> str:
    """
    置信度（用于快速筛选）：
    - 高：总分高且稳定+趋势均不弱
    - 中：总分中高
    - 低：其余
    """
    # 结合当前加权参数重标定阈值，避免高置信样本过少
    if total_score >= 78 and stability_score >= 5 and trend_score >= 5:
        return "高"
    if total_score >= 55:
        return "中"
    return "低"


#
# 评分主公式参数（用于与次日收盘表现相关性调优）
# 思路：提升“日内过程型”特征权重（稳定性/趋势），降低纯静态快照项的相对权重。
# 当前参数来源：tools/optimize_structure_weights.py
# 优化批次：20260417_194655
# 目标口径：次日收盘涨幅 > -1%
# 样本区间：--start 20260318
#
SCORE_WEIGHT_SEAL_HARDNESS = 0.40
SCORE_WEIGHT_RUSH_INTENSITY = 0.40
SCORE_WEIGHT_CLOSE_AMOUNT = 0.30
SCORE_WEIGHT_STABILITY = 1.50
SCORE_WEIGHT_TREND = 1.80
SCORE_BASE_BIAS = -6.0


def _calc_weighted_total_score(
    score_hard: int,
    score_rush: int,
    score_amt: int,
    score_stability: int,
    score_trend: int,
) -> int:
    """
    加权总分（用于评级）：
    - 保留原有各子项打分逻辑
    - 通过可调权重实现参数优化
    """
    raw = (
        SCORE_BASE_BIAS
        + SCORE_WEIGHT_SEAL_HARDNESS * float(score_hard or 0)
        + SCORE_WEIGHT_RUSH_INTENSITY * float(score_rush or 0)
        + SCORE_WEIGHT_CLOSE_AMOUNT * float(score_amt or 0)
        + SCORE_WEIGHT_STABILITY * float(score_stability or 0)
        + SCORE_WEIGHT_TREND * float(score_trend or 0)
    )
    return int(max(0, min(120, round(raw))))


def _rank01(values: List[Optional[float]]) -> List[Optional[float]]:
    """把数值列表转为 0~1 分位（None 保留）。"""
    valid = [(idx, float(v)) for idx, v in enumerate(values) if v is not None]
    out: List[Optional[float]] = [None] * len(values)
    if not valid:
        return out
    if len(valid) == 1:
        out[valid[0][0]] = 0.5
        return out
    sorted_valid = sorted(valid, key=lambda x: x[1])
    n = len(sorted_valid)
    for rk, (idx, _) in enumerate(sorted_valid):
        out[idx] = rk / (n - 1)
    return out


def _apply_cross_sectional_score_adjustment(rows: List[dict]) -> None:
    """
    在单日横截面上做轻量分位修正，缓解固定阈值在不同行情温度下失真。
    仅当样本>=8时启用，避免小样本过拟合。
    """
    use_adjust = len(rows) >= 8
    amt_rank = _rank01([r.get("_close_amt_yi") for r in rows]) if use_adjust else [None] * len(rows)
    hard_rank = _rank01([r.get("_seal_ratio") for r in rows]) if use_adjust else [None] * len(rows)
    rush_rank = _rank01([r.get("_rush_ratio") for r in rows]) if use_adjust else [None] * len(rows)
    for i, row in enumerate(rows):
        base = int(row.get("_base_total_score") or 0)
        st_score = int(row.get("_stability_score") or 0)
        tr_score = int(row.get("_trend_score") or 0)
        if amt_rank[i] is None or hard_rank[i] is None or rush_rank[i] is None:
            adj_total = base
        else:
            bonus = int(round((amt_rank[i] - 0.5) * 12))
            bonus += int(round((hard_rank[i] - 0.5) * 16))
            bonus += int(round((rush_rank[i] - 0.5) * 12))
            adj_total = base + bonus
        # 强化“日内过程型”特征权重：稳定性/趋势高时给正向加分，低时惩罚更明确
        if st_score >= 10:
            adj_total += 8
        elif st_score >= 5:
            adj_total += 3
        elif row.get("order_stability"):
            adj_total -= 5

        if tr_score >= 8:
            adj_total += 8
        elif tr_score >= 5:
            adj_total += 3
        elif row.get("order_trend"):
            adj_total -= 4
        adj_total = max(0, min(120, int(adj_total)))
        row["_total_score"] = adj_total
        row["order_rating"] = _label_order_rating(adj_total)
        row["rating_score"] = _rating_score_from_label(row["order_rating"])
        row["confidence_tag"] = _confidence_tag(
            adj_total,
            st_score,
            tr_score,
        )


def _calc_order_trend_and_score(raw_df) -> Tuple[str, Optional[int]]:
    """
    封单运行趋势（基于买一封单量时序）：
    - Fstart: 上午前30分钟均值（09:30-10:00）
    - Fmid:   全天中段均值（10:00-14:30）
    - Fend:   尾盘最后30分钟均值（14:30-15:00）
    并返回对应趋势分：10/7/5/3/0
    """
    if raw_df is None or getattr(raw_df, "empty", True):
        return "", None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    def _parse_hhmm_from_time_val(v) -> Optional[int]:
        try:
            s = str(int(float(v)))
        except Exception:
            s = str(v)
        digits = "".join(ch for ch in s if ch.isdigit())
        # 1) yyyymmddHHMMSS...
        if len(digits) >= 12:
            hh = int(digits[-6:-4])
            mm = int(digits[-4:-2])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return hh * 100 + mm
        # 2) epoch(ms/s)
        try:
            iv = int(float(v))
            if iv > 1_000_000_000_000:  # ms
                dt = datetime.fromtimestamp(iv / 1000.0)
            elif iv > 1_000_000_000:  # s
                dt = datetime.fromtimestamp(iv)
            else:
                return None
            return dt.hour * 100 + dt.minute
        except Exception:
            return None

    # 兼容不同字段命名；若是五档数组，取买一
    bidvol_fields = ["bidVol", "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1"]
    cols = set(str(c) for c in list(getattr(raw_df, "columns", [])))
    use_field = None
    for f in bidvol_fields:
        if f in cols:
            use_field = f
            break
    if use_field is None:
        return "", None

    start_vals: List[float] = []
    mid_vals: List[float] = []
    end_vals: List[float] = []

    try:
        for idx, row in raw_df.iterrows():
            fv = _to_float(row.get(use_field))
            if fv is None or fv <= 0:
                continue
            tval = row.get("time") if "time" in row else idx
            hhmm = _parse_hhmm_from_time_val(tval)
            if hhmm is None:
                continue
            if 930 <= hhmm < 1000:
                start_vals.append(fv)
            elif 1000 <= hhmm < 1430:
                mid_vals.append(fv)
            elif 1430 <= hhmm <= 1500:
                end_vals.append(fv)
    except Exception:
        return "", None

    # 若交易时段切片为空，按序列三等分兜底
    if not start_vals or not mid_vals or not end_vals:
        all_vals: List[float] = []
        try:
            for v in raw_df[use_field]:
                fv = _to_float(v)
                if fv is not None and fv > 0:
                    all_vals.append(fv)
        except Exception:
            return "", None
        n = len(all_vals)
        if n < 6:
            return "", None
        a = max(1, n // 6)  # 约前1/6 ~ 30分钟窗口
        b = max(1, n // 3)
        start_vals = all_vals[:a]
        end_vals = all_vals[-a:]
        mid_vals = all_vals[a:n - a] if (n - 2 * a) > 0 else all_vals[b:2 * b]

    if not start_vals or not mid_vals or not end_vals:
        return "", None

    f_start = sum(start_vals) / len(start_vals)
    f_mid = sum(mid_vals) / len(mid_vals)
    f_end = sum(end_vals) / len(end_vals)
    if f_start <= 0 or f_mid <= 0:
        return "", None

    chg_mid = (f_mid - f_start) / f_start * 100.0
    chg_end = (f_end - f_mid) / f_mid * 100.0

    # 五档判定（按强特征优先）
    if chg_mid >= 5.0 and chg_end >= 5.0:
        return "趋势持续加强", 10
    if chg_mid <= -5.0 and chg_end <= -5.0:
        return "趋势持续减弱", 0
    if 0.0 <= chg_mid < 5.0 and 0.0 <= chg_end < 5.0:
        return "震荡小幅加强", 7
    if -5.0 < chg_mid <= 0.0 and -5.0 < chg_end <= 0.0:
        return "震荡小幅减弱", 3
    if -5.0 <= chg_mid <= 5.0 and -5.0 <= chg_end <= 5.0:
        return "封单整体平稳", 5

    # 跨档混合时兜底为平稳
    return "封单整体平稳", 5


class TodayLimitUpFinderThread(QThread):
    # today_rows, first_board_rows, streak_rows, limitup_date(yyyymmdd), debug_msg, finish_kind: "list_only"|"full"
    # detail_by_code: { code -> [ {date, open, close, remark}, ... ] }  remark: 首板涨停 / N连板 / 跌停
    finished = pyqtSignal(list, list, list, str, str, str, object)
    error_occurred = pyqtSignal(str)
    debug_info = pyqtSignal(str)
    tick_preview = pyqtSignal(str, str, object)  # code, name, {"headers": [...], "rows": [...]}

    def __init__(
        self,
        *,
        history_dir: str,
        all_a_stocks_csv: str,
        source_mode: str,
        stock_list_filepath: Optional[str],
        manual_codes: Optional[List[str]],
        analysis_date_yyyymmdd: Optional[str],
        limitup_csv_path: Optional[str],
        test_only: bool,
        test_limit_n: int,
        run_mode: str = "full",
        parent=None,
    ):
        super().__init__(parent)
        self.history_dir = history_dir
        self.all_a_stocks_csv = all_a_stocks_csv
        self.source_mode = source_mode
        self.stock_list_filepath = stock_list_filepath
        self.manual_codes = manual_codes or []
        self.analysis_date_yyyymmdd = (analysis_date_yyyymmdd or "").strip()
        self.limitup_csv_path = (limitup_csv_path or "").strip()
        self.test_only = test_only
        self.test_limit_n = test_limit_n
        self.run_mode = run_mode  # "list_only" | "full"

    def run(self):
        try:
            query_date = self.analysis_date_yyyymmdd or datetime.now().strftime("%Y%m%d")
            if self.limitup_csv_path and os.path.isfile(self.limitup_csv_path):
                limitup_path = self.limitup_csv_path
                m = re.search(r"(20\d{6})", os.path.basename(limitup_path))
                limitup_date = m.group(1) if m else query_date
            else:
                limitup_path, limitup_date = _find_latest_limitup_file(self.history_dir, query_date)
                if not limitup_path:
                    self.error_occurred.emit(
                        f"未找到<={query_date}的涨停板数据文件（history_dir={self.history_dir}）"
                    )
                    return

            limitup_code_to_name = _load_limitup_codes(limitup_path)
            total_limitup = len(limitup_code_to_name)

            def _has_st_tag(name: str) -> bool:
                n = (name or "").strip()
                trans = str.maketrans({"Ｓ": "S", "Ｔ": "T", "＊": "*", "－": "-", "％": "%"})
                n2 = n.translate(trans).upper()
                return ("ST" in n2) or ("*ST" in n2) or n2.startswith("ST")

            if self.source_mode == "today_limitup":
                # 全部今日涨停：以涨停板 CSV 为准（剔除 ST），不再与 all_a 求交
                matched_codes = list(limitup_code_to_name.keys())
                code_to_name: Dict[str, str] = dict(limitup_code_to_name)
                st_excluded = []
                for c in matched_codes:
                    nm1 = limitup_code_to_name.get(c, "")
                    nm2 = code_to_name.get(c, "")
                    if _has_st_tag(nm1) or _has_st_tag(nm2):
                        st_excluded.append(c)
                st_excluded_cnt = len(st_excluded)
                if st_excluded_cnt:
                    st_set = set(st_excluded)
                    matched_codes = [c for c in matched_codes if c not in st_set]
                debug_msg = (
                    f"limitup_file={os.path.basename(limitup_path)}; "
                    f"来源=全部今日涨停; 涨停CSV总数={total_limitup}; "
                    f"待分析={len(matched_codes)}; ST已兜底剔除={st_excluded_cnt}"
                )
            else:
                # 文件或手动输入：分析指定列表，不要求当日涨停
                if self.source_mode == "file":
                    if not self.stock_list_filepath:
                        self.error_occurred.emit("未选择股票列表文件")
                        return
                    universe_list = _read_codes_from_txt_or_csv(self.stock_list_filepath)
                else:
                    universe_list = list(self.manual_codes)
                    if not universe_list:
                        self.error_occurred.emit("请输入至少一只股票的代码（逗号分隔）")
                        return

                all_a_map: Dict[str, str] = {}
                if os.path.isfile(self.all_a_stocks_csv):
                    try:
                        all_a_map = _load_all_a_stocks(self.all_a_stocks_csv)
                    except Exception:
                        all_a_map = {}
                matched_codes = list(universe_list)
                code_to_name = {
                    c: (all_a_map.get(c, "").strip() or limitup_code_to_name.get(c, "").strip())
                    for c in matched_codes
                }
                in_limitup_cnt = sum(1 for c in matched_codes if c in limitup_code_to_name)
                debug_msg = (
                    f"limitup_file={os.path.basename(limitup_path)}; "
                    f"来源={'指定文件' if self.source_mode == 'file' else '输入代码'}; "
                    f"待分析={len(matched_codes)}; 其中在今日涨停CSV中={in_limitup_cnt}/{len(matched_codes)}"
                )

            if self.test_only and self.test_limit_n > 0:
                matched_codes = sorted(set(matched_codes))[: self.test_limit_n]

            n_before_valid = len(matched_codes)
            matched_codes = [c for c in matched_codes if _is_valid_stock_code_6(c)]
            if not matched_codes:
                self.error_occurred.emit("有效股票代码为空（已剔除无效占位如 000000）")
                return
            if len(matched_codes) < n_before_valid:
                debug_msg += f"；剔除无效占位代码={n_before_valid - len(matched_codes)}；有效待分析={len(matched_codes)}"

            self.debug_info.emit(debug_msg)

            rows: List[Tuple[str, str]] = []
            for code in matched_codes:
                name = limitup_code_to_name.get(code, "") if code in limitup_code_to_name else ""
                final_name = name or code_to_name.get(code, "")
                rows.append((code, final_name))

            rows.sort(key=lambda x: x[0])
            today_codes = [c for c, _ in rows]

            if self.run_mode == "list_only":
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "list_only", {})
                return

            # 第2步：逐只拉取 tick，做涨停板结构预览（每只股票单独输出）
            if xtdata is None:
                self.error_occurred.emit("xtdata 不可用，无法拉取 tick 数据")
                return

            # 尽量减少 xtdata 欢迎信息
            try:
                xtdata.enable_hello = False
            except Exception:
                pass

            tick_preview_limit = 200
            last_tick_rows: List[dict] = []
            streak_metric_rows: List[dict] = []
            for idx, (code, name) in enumerate(rows, start=1):
                tick_headers: List[str] = []
                tick_rows: List[List[str]] = []
                try:
                    tick_df = _load_or_fetch_tick_df(code, limitup_date)
                    tick_headers, tick_rows = _extract_tick_preview_rows(tick_df, limit_n=tick_preview_limit)
                except Exception:
                    tick_headers, tick_rows = [], []
                    tick_df = None

                last_tick_rows = [{"tick_index": r[0]} for r in tick_rows]  # 占位，兼容 finished 旧签名
                close_amt_yi = _calc_close_order_amount_yi(tick_df)
                day_turnover_yuan = _calc_day_turnover_yuan(tick_df)
                float_cap_yuan = _calc_float_market_cap_yuan(code, tick_df)
                seal_ratio = None
                seal_hardness = ""
                if close_amt_yi is not None and day_turnover_yuan is not None and day_turnover_yuan > 0:
                    # 封板硬度 = 封单金额 / 当日成交额（比值）
                    seal_ratio = (close_amt_yi * 1e8) / day_turnover_yuan
                    seal_hardness = f"{seal_ratio * 100.0:.2f}%"
                rush_ratio = None
                rush_intensity = ""
                if close_amt_yi is not None and float_cap_yuan is not None and float_cap_yuan > 0:
                    # 抢筹烈度 = 封单金额 / 流通市值（比值）
                    rush_ratio = (close_amt_yi * 1e8) / float_cap_yuan
                    rush_intensity = f"{rush_ratio * 100.0:.2f}%"
                order_stability = ""
                stability_score = None
                vol_coeff = _calc_order_volatility_coeff(tick_df)
                if vol_coeff is not None:
                    stability_label, stability_score = _classify_order_stability(vol_coeff)
                    order_stability = stability_label
                order_trend = ""
                trend_score = None
                order_trend, trend_score = _calc_order_trend_and_score(tick_df)
                # 综合得分：前3项新规则 + 稳定性分 + 趋势分
                score_hard = _score_seal_hardness(seal_ratio)
                score_rush = _score_rush_intensity(rush_ratio)
                score_amt = _score_close_order_amount(close_amt_yi)
                score_stability = int(stability_score or 0)
                score_trend = int(trend_score or 0)
                total_score_val = _calc_weighted_total_score(
                    score_hard,
                    score_rush,
                    score_amt,
                    score_stability,
                    score_trend,
                )
                streak_metric_rows.append(
                    {
                        "code": code,
                        "name": name or "",
                        "close_order_amount_yi": "" if close_amt_yi is None else f"{close_amt_yi:.4f}",
                        "seal_hardness": seal_hardness,
                        "rush_intensity": rush_intensity,
                        "order_stability": order_stability,
                        "order_trend": order_trend,
                        "order_rating": "",
                        "rating_score": 0,
                        "confidence_tag": "",
                        "_close_amt_yi": close_amt_yi,
                        "_seal_ratio": seal_ratio,
                        "_rush_ratio": rush_ratio,
                        "_stability_score": score_stability,
                        "_trend_score": score_trend,
                        "_base_total_score": total_score_val,
                        "_total_score": total_score_val,
                    }
                )
                self.debug_info.emit(
                    f"{debug_msg}；tick进度={idx}/{len(rows)}；{code} {name or ''}；tick条数={len(tick_rows)}；字段数={len(tick_headers)}"
                )
                self.tick_preview.emit(code, name or "", {"headers": tick_headers, "rows": tick_rows})

            _apply_cross_sectional_score_adjustment(streak_metric_rows)
            for r in streak_metric_rows:
                r.pop("_close_amt_yi", None)
                r.pop("_seal_ratio", None)
                r.pop("_rush_ratio", None)
                r.pop("_stability_score", None)
                r.pop("_trend_score", None)
                r.pop("_base_total_score", None)
                r.pop("_total_score", None)

            # tick_only：UI 不再套用首板/连板统计渲染逻辑
            self.finished.emit(rows, last_tick_rows, streak_metric_rows, limitup_date, debug_msg, "tick_only", {})
            return

            # 1) 构造交易日历（用于“上一交易日是否涨停/次一交易日收益”口径）
            analysis_today = limitup_date
            today_year = int(analysis_today[:4])
            # 近一年（日历年）：上一自然年（从 YYYY-01-01 到本年的昨日）
            year_start_calendar = f"{today_year - 1}0101"

            from datetime import timedelta as _td

            year_start_dt = datetime.strptime(year_start_calendar, "%Y%m%d")
            # 往前多取一段交易日，保证 year_start 的前一交易日也在范围内
            # 用足够长的回溯期，避免跨年连板/首板在统计起点被截断
            trading_start_dt = year_start_dt - _td(days=400)
            trading_start = trading_start_dt.strftime("%Y%m%d")
            trading_end = analysis_today

            trading_dates_ts = xtdata.get_trading_dates("SH", start_time=trading_start, end_time=trading_end)
            trade_dates: List[str] = []
            for ts in trading_dates_ts or []:
                if isinstance(ts, (int, float)):
                    dt = datetime.fromtimestamp(float(ts) / 1000.0)
                    trade_dates.append(dt.strftime("%Y%m%d"))
                elif isinstance(ts, str):
                    trade_dates.append(ts.replace("-", "")[:8])
                else:
                    trade_dates.append(str(ts)[:8])
            trade_dates = sorted({d for d in trade_dates if d and len(d) == 8})

            if analysis_today not in trade_dates:
                # 容错：用 <= analysis_today 的最后一个交易日
                trade_dates = [d for d in trade_dates if d <= analysis_today]
                if not trade_dates:
                    self.error_occurred.emit("无法构造交易日历（trade_dates 为空）")
                    return
                analysis_today = trade_dates[-1]

            idx_today = trade_dates.index(analysis_today)
            if idx_today <= 0:
                self.error_occurred.emit("交易日历不足以排除今天与计算首板口径")
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "full", {})
                return

            # 统计结果的结束日：昨日（排除今天）
            end_idx = idx_today - 1
            end_date = trade_dates[end_idx]

            # year_start 的第一个交易日：>= year_start_calendar 的最早交易日
            year_start_idx = None
            for i in range(0, len(trade_dates)):
                if trade_dates[i] >= year_start_calendar:
                    year_start_idx = i
                    break
            if year_start_idx is None:
                self.error_occurred.emit("无法定位 year_start 在交易日历中的位置")
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "full", {})
                return

            # 统计时仍按 d < year_start_td 或 d > end_date 过滤；
            # consec 连续性从起点开始计算，避免跨年时低估连板长度。
            start_needed_idx = 0
            year_start_td = trade_dates[year_start_idx]
            debug_msg += f"；统计区间={year_start_td}~{end_date}"

            # 2) 用 QMT 日K 自己判断涨停（不再依赖 history_data 的涨停CSV）
            #    注意：次日收益统计会用到“今天”的开盘/收盘，所以 end 需要包含 today。
            ohlc_codes_needed = sorted(set(today_codes))

            # 为了计算起始日的涨停，需要用到上一交易日的收盘价，因此向前多取 1 个交易日
            ohlc_start_idx = max(0, start_needed_idx - 1)
            ohlc_start_date = trade_dates[ohlc_start_idx]
            ohlc_end_date = analysis_today  # 包含 today

            startdate_full = datetime.strptime(ohlc_start_date, "%Y%m%d").strftime("%Y%m%d") + "000000"
            enddate_full = datetime.strptime(ohlc_end_date, "%Y%m%d").strftime("%Y%m%d") + "235959"

            ohlc_map_by_code: Dict[str, Dict[str, Tuple[float, float]]] = {}
            for idx, code in enumerate(ohlc_codes_needed):
                if idx % 10 == 0:
                    pass
                full_code = _to_full_stock_code(code)
                try:
                    xtdata.download_history_data(full_code, "1d", startdate_full, enddate_full)
                    df = xtdata.get_market_data_ex(
                        [],
                        [full_code],
                        period="1d",
                        start_time=startdate_full,
                        end_time=enddate_full,
                        count=-1,
                    )
                except Exception:
                    continue

                if df is None or full_code not in df or df[full_code] is None or df[full_code].empty:
                    continue
                try:
                    ohlc_map_by_code[code] = _extract_open_close_map(df[full_code])
                except Exception:
                    ohlc_map_by_code[code] = {}

            # 2.1) 构造涨停状态：date -> set(codes)
            def _limit_ratio_for_code(code_6: str) -> float:
                # ST 已通过“兜底剔除”从今日集合中移除，这里按板块规则定涨停比例即可
                if code_6.startswith(("300", "301", "688", "689")):
                    return 0.20
                if code_6.startswith(("8", "4", "920")):
                    return 0.30
                return 0.10

            codes_by_date: Dict[str, Set[str]] = {d: set() for d in trade_dates}
            for code in ohlc_codes_needed:
                m = ohlc_map_by_code.get(code) or {}
                limit_ratio = _limit_ratio_for_code(code)
                for i in range(1, idx_today + 1):
                    d = trade_dates[i]
                    prev_d = trade_dates[i - 1]
                    prev_close = (m.get(prev_d) or (None, None, None, None))[3]
                    close_price = (m.get(d) or (None, None, None, None))[3]
                    if prev_close is None or close_price is None or prev_close == 0:
                        continue

                    limit_up_price = round(prev_close * (1 + limit_ratio), 2)
                    price_diff = abs(close_price - limit_up_price)
                    inc_ratio = close_price / prev_close - 1
                    # 兼容两种判定：接近涨停价（容差） 或 涨幅达到预期（99%）
                    is_limit_up = (price_diff <= 0.02) or (inc_ratio >= limit_ratio * 0.99)
                    if is_limit_up:
                        codes_by_date.setdefault(d, set()).add(code)

            # 3) 计算今天的首板/连板归类（基于“上一交易日是否涨停”）
            today_streak_map: Dict[str, int] = {}
            first_today_codes: List[str] = []
            streak_today_codes: List[str] = []
            for code in today_codes:
                consec = 0
                for j in range(idx_today, -1, -1):
                    d = trade_dates[j]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        break
                # consec==0 说明 QMT 未能识别“今天涨停”（容差/数据缺失等）
                # 这种情况下不纳入首板/连板统计，避免用错误首板口径。
                today_streak_map[code] = consec
                if today_streak_map[code] == 1:
                    first_today_codes.append(code)
                elif today_streak_map[code] >= 2:
                    streak_today_codes.append(code)

            debug_msg += f"；QMT识别今日涨停={len(first_today_codes) + len(streak_today_codes)}/{len(today_codes)}"

            # 4) 首板统计表：含「今日首板」+「今日未涨停（连板数记 0）」
            code_name_map_today = {code: name for code, name in rows}

            def _build_first_board_row(code: str, today_streak_val: int) -> dict:
                consec = 0
                first_board_count = 0
                next_day_limitup_count = 0
                next_day_up_count = 0
                next_day_down_count = 0

                open_ret_sum = 0.0
                close_ret_sum = 0.0
                open_valid_cnt = 0
                close_valid_cnt = 0

                ohlc_map = ohlc_map_by_code.get(code) or {}
                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        consec = 0

                    if d < year_start_td or d > end_date:
                        continue
                    if consec != 1:
                        continue

                    first_board_count += 1

                    if i + 1 <= end_idx:
                        nd = trade_dates[i + 1]
                        if code in codes_by_date.get(nd, set()):
                            next_day_limitup_count += 1

                        close_d = ohlc_map.get(d, (None, None, None, None))[3]
                        close_nd = ohlc_map.get(nd, (None, None, None, None))[3]
                        open_nd = ohlc_map.get(nd, (None, None, None, None))[0]

                        if close_d is not None and close_nd is not None and close_d != 0:
                            if close_nd > close_d:
                                next_day_up_count += 1
                            elif close_nd < close_d:
                                next_day_down_count += 1

                            close_ret_sum += (close_nd - close_d) / close_d
                            close_valid_cnt += 1

                            if open_nd is not None:
                                open_ret_sum += (open_nd - close_d) / close_d
                                open_valid_cnt += 1

                avg_open_ret_pct = (open_ret_sum / open_valid_cnt * 100.0) if open_valid_cnt > 0 else 0.0
                avg_close_ret_pct = (close_ret_sum / close_valid_cnt * 100.0) if close_valid_cnt > 0 else 0.0

                return {
                    "code": code,
                    "name": code_name_map_today.get(code, ""),
                    "today_streak": today_streak_val,
                    "first_board_count_1y": first_board_count,
                    "next_day_limitup_count": next_day_limitup_count,
                    "next_day_up_count": next_day_up_count,
                    "next_day_down_count": next_day_down_count,
                    "avg_open_return_pct": round(avg_open_ret_pct, 2),
                    "avg_close_return_pct": round(avg_close_ret_pct, 2),
                }

            first_board_rows: List[dict] = []
            for code in first_today_codes:
                first_board_rows.append(_build_first_board_row(code, 1))
            non_limitup_today_codes = [c for c in today_codes if today_streak_map.get(c, 0) == 0]
            for code in non_limitup_today_codes:
                first_board_rows.append(_build_first_board_row(code, 0))

            # 6) 统计“今日连板”对应股票的近一年 k连板次数与晋级比例
            #    - k连板次数：任意一段连续涨停，仅在「该段最后一个涨停日」记一次，长度为 N 则
            #      只计 N连板 1 次（不再把同一段里的 2~(N-1) 连板逐日各计一次）。
            #    - 连板晋级比例：仍用逐日 consec≥2 的口径，并排除「以昨日为结束日的当前这段连板」
            #      在统计区间内的逐日样本（与改口径前的晋级算法一致）；若改用分段次数去算
            #      k 进 k+1，会变成互斥集合之比，语义不再成立。
            streak_rows: List[dict] = []
            for code in streak_today_codes:
                consec = 0
                counts_by_k: Dict[int, int] = {}
                max_k = 0

                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        if consec >= 2:
                            last_idx = i - 1
                            last_d = trade_dates[last_idx]
                            if year_start_td <= last_d <= end_date:
                                counts_by_k[consec] = counts_by_k.get(consec, 0) + 1
                                if consec > max_k:
                                    max_k = consec
                        consec = 0

                if consec >= 2:
                    last_d = trade_dates[end_idx]
                    if year_start_td <= last_d <= end_date:
                        counts_by_k[consec] = counts_by_k.get(consec, 0) + 1
                        if consec > max_k:
                            max_k = consec

                if counts_by_k:
                    counts_str = "\n".join(
                        f"{k}连板{counts_by_k.get(k, 0)}次"
                        for k in sorted(counts_by_k.keys())
                    )
                else:
                    counts_str = ""

                j = end_idx
                while j >= 0 and code in codes_by_date.get(trade_dates[j], set()):
                    j -= 1
                exclude_streak_begin = j + 1

                consec_u = 0
                counts_upgrade: Dict[int, int] = {}
                max_k_u = 0
                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec_u += 1
                    else:
                        consec_u = 0
                    if d < year_start_td or d > end_date:
                        continue
                    if consec_u >= 2:
                        if exclude_streak_begin <= i <= end_idx:
                            continue
                        counts_upgrade[consec_u] = counts_upgrade.get(consec_u, 0) + 1
                        if consec_u > max_k_u:
                            max_k_u = consec_u

                upgrade_parts: List[str] = []
                for k in range(2, max_k_u):
                    denom = counts_upgrade.get(k, 0)
                    num = counts_upgrade.get(k + 1, 0)
                    if denom > 0:
                        ratio = num / denom
                        desc = _ratio_to_chinese_desc(ratio)
                        upgrade_parts.append(f"{k}进{k + 1}：{desc}")

                upgrade_str = "\n".join(upgrade_parts)

                streak_rows.append(
                    {
                        "code": code,
                        "name": code_name_map_today.get(code, ""),
                        "today_streak": today_streak_map.get(code, 2),
                        "max_streak_in_year": max_k,
                        "counts_by_k": counts_str,
                        "upgrade_ratios": upgrade_str,
                    }
                )

            streak_rows.sort(
                key=lambda r: (
                    -int(r.get("today_streak") or 0),
                    str(r.get("code") or ""),
                )
            )

            # 7) 每只股票：按交易日列出「日期、开盘、收盘、备注」
            #    - 涨停日：首板涨停 / N连板（连板从 ohlc 起点连续累计）
            #    - 涨停结束后的第一个非涨停日：首板后次日 / N连板后次日（便于看板后盈亏）；若该日跌停则只列一行「跌停」
            #    - 统计区间末日若仍涨停，补列再下一交易日（可为涨停 CSV 当日），规则同上
            detail_by_code: Dict[str, List[dict]] = {c: [] for c in today_codes}
            for code in today_codes:
                m = ohlc_map_by_code.get(code) or {}
                limit_ratio = _limit_ratio_for_board(code)
                streak = 0
                for i in range(1, end_idx + 1):
                    d = trade_dates[i]
                    prev_d = trade_dates[i - 1]
                    prev_close = (m.get(prev_d) or (None, None, None, None))[3]
                    open_d, high_d, low_d, close_d = m.get(d, (None, None, None, None))
                    is_up = code in codes_by_date.get(d, set())

                    if prev_close is None or (isinstance(prev_close, (int, float)) and float(prev_close) <= 0):
                        streak = 0
                        continue
                    if close_d is None:
                        if not is_up:
                            streak = 0
                        continue

                    try:
                        pc = float(prev_close)
                        cl = float(close_d)
                    except (TypeError, ValueError):
                        streak = 0 if not is_up else streak
                        continue

                    limit_dn = round(pc * (1.0 - limit_ratio), 2)
                    price_diff_dn = abs(cl - limit_dn)
                    dec_ratio = 1.0 - cl / pc
                    is_dn = (price_diff_dn <= 0.02) or (dec_ratio >= limit_ratio * 0.99 - 1e-12)

                    in_win = year_start_td <= d <= end_date

                    if is_up:
                        streak += 1
                        if in_win:
                            remark = "首板涨停" if streak == 1 else f"{streak}连板"
                            detail_by_code[code].append(
                                {
                                    "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                    "open": open_d,
                                    "high": high_d,
                                    "low": low_d,
                                    "close": close_d,
                                    "remark": remark,
                                }
                            )
                    else:
                        ended_streak = streak
                        streak = 0
                        if in_win:
                            if is_dn:
                                detail_by_code[code].append(
                                    {
                                        "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                        "open": open_d,
                                        "high": high_d,
                                        "low": low_d,
                                        "close": close_d,
                                        "remark": "跌停",
                                    }
                                )
                            elif ended_streak > 0:
                                post_rm = (
                                    "首板后次日"
                                    if ended_streak == 1
                                    else f"{ended_streak}连板后次日"
                                )
                                detail_by_code[code].append(
                                    {
                                        "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                        "open": open_d,
                                        "high": high_d,
                                        "low": low_d,
                                        "close": close_d,
                                        "remark": post_rm,
                                    }
                                )

                # 区间最后一天（end_date）仍涨停时，补出再下一日（常为「今日」），便于看最后一棒之后盈亏
                if streak > 0 and end_idx + 1 < len(trade_dates):
                    d_next = trade_dates[end_idx + 1]
                    open_n, high_n, low_n, close_n = m.get(d_next, (None, None, None, None))
                    if close_n is not None:
                        prev_ke = trade_dates[end_idx]
                        prev_c = (m.get(prev_ke) or (None, None, None, None))[3]
                        try:
                            if prev_c is None or float(prev_c) <= 0:
                                raise ValueError
                            pc2 = float(prev_c)
                            cl2 = float(close_n)
                        except (TypeError, ValueError):
                            pass
                        else:
                            next_is_up = code in codes_by_date.get(d_next, set())
                            limit_dn2 = round(pc2 * (1.0 - limit_ratio), 2)
                            diff2 = abs(cl2 - limit_dn2)
                            dec2 = 1.0 - cl2 / pc2
                            is_dn_n = (diff2 <= 0.02) or (dec2 >= limit_ratio * 0.99 - 1e-12)
                            if not next_is_up:
                                if is_dn_n:
                                    detail_by_code[code].append(
                                        {
                                            "date": f"{d_next[:4]}-{d_next[4:6]}-{d_next[6:8]}",
                                            "open": open_n,
                                            "high": high_n,
                                            "low": low_n,
                                            "close": close_n,
                                            "remark": "跌停",
                                        }
                                    )
                                else:
                                    post_rm = (
                                        "首板后次日"
                                        if streak == 1
                                        else f"{streak}连板后次日"
                                    )
                                    detail_by_code[code].append(
                                        {
                                            "date": f"{d_next[:4]}-{d_next[4:6]}-{d_next[6:8]}",
                                            "open": open_n,
                                            "high": high_n,
                                            "low": low_n,
                                            "close": close_n,
                                            "remark": post_rm,
                                        }
                                    )

            # 完成
            self.finished.emit(
                rows, first_board_rows, streak_rows, limitup_date, debug_msg, "full", detail_by_code
            )
        except Exception as e:
            self.error_occurred.emit(f"{type(e).__name__}: {e}")


class StockLimitUpDetailDialog(QDialog):
    """单只股票：按日一行展示日期、开盘、最高、最低、收盘、备注（首板涨停 / N连板 / 跌停）。"""

    def __init__(self, code: str, name: str, details: List[dict], parent=None):
        super().__init__(parent)
        title_name = (name or "").strip()
        self.setWindowTitle(f"走势明细 - {code} {title_name}".strip())
        self.resize(640, 520)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        lay = QVBoxLayout(self)
        note = QLabel(
            "说明：含「日K 涨停」日（首板涨停 / 2连板…）、「日K 跌停」日，以及每次涨停结束后的下一交易日"
            "（首板后次日 / N连板后次日，便于对照前一日涨停收盘看盈亏；若次日跌停则只标「跌停」）。"
            "若统计区间最后一天仍涨停，会多列再下一交易日。价格来自 QMT 日K。"
        )
        note.setWordWrap(True)
        note.setStyleSheet("color: #666; font-size: 11px;")
        lay.addWidget(note)
        info = QLabel(
            f"<b>{code}</b>　{title_name}　"
            f"共 <b>{len(details)}</b> 行（按日期顺序）"
        )
        lay.addWidget(info)

        table = QTableWidget(len(details), 6)
        table.setHorizontalHeaderLabels(["日期", "开盘", "最高", "最低", "收盘", "备注"])
        table.setAlternatingRowColors(True)
        color_a = QColor("#2f5597")
        color_b = QColor("#7a4f01")
        seg_idx = 0
        in_cycle = False
        for r, rec in enumerate(details):
            remark = str(rec.get("remark") or "")
            if "首板涨停" in remark:
                in_cycle = True
            fg = color_a if (seg_idx % 2 == 0) else color_b

            d_item = QTableWidgetItem(str(rec.get("date") or ""))
            o_item = QTableWidgetItem(_fmt_px(rec.get("open")))
            h_item = QTableWidgetItem(_fmt_px(rec.get("high")))
            l_item = QTableWidgetItem(_fmt_px(rec.get("low")))
            c_item = QTableWidgetItem(_fmt_px(rec.get("close")))
            rm = QTableWidgetItem(remark)
            rm.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            for it in (d_item, o_item, h_item, l_item, c_item, rm):
                it.setForeground(fg)
            table.setItem(r, 0, d_item)
            table.setItem(r, 1, o_item)
            table.setItem(r, 2, h_item)
            table.setItem(r, 3, l_item)
            table.setItem(r, 4, c_item)
            table.setItem(r, 5, rm)

            if in_cycle and (("后次日" in remark) or ("跌停" in remark)):
                in_cycle = False
                seg_idx += 1
        hdr = table.horizontalHeader()
        hdr.setStretchLastSection(True)
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        lay.addWidget(table, 1)

        btn = QPushButton("关闭")
        btn.clicked.connect(self.accept)
        lay.addWidget(btn)


class LimitUpGeneAnalysisDialog(QDialog):
    def __init__(self, parent=None, auto_run: bool = False):
        super().__init__(parent)
        self._auto_run = bool(auto_run)
        self._auto_export_done = False
        self._auto_last_export_xlsx = ""
        self._auto_last_export_png = ""
        self.setWindowTitle("蚂蚁量化-涨停板结构分析")
        self.resize(1000, 700)
        # 允许窗口最大化/最小化（标题栏通常会自动提供按钮，但有些窗口标志会禁用）
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        layout = QVBoxLayout(self)

        # 输入选择
        top_row = QHBoxLayout()
        layout.addLayout(top_row)

        self.radio_today_limitup = QRadioButton("全部今日涨停")
        self.radio_today_limitup.setChecked(True)
        top_row.addWidget(self.radio_today_limitup)

        self.radio_file = QRadioButton("从指定股票列表")
        top_row.addWidget(self.radio_file)

        self.radio_manual = QRadioButton("直接输入股票代码（逗号分隔）")
        top_row.addWidget(self.radio_manual)

        top_row.addStretch()

        file_row = QHBoxLayout()
        layout.addLayout(file_row)
        file_row.addWidget(QLabel("列表文件："))
        self.file_edit = QLabel("未选择文件")
        file_row.addWidget(self.file_edit, 1)
        self.file_btn = QPushButton("选择文件")
        self.file_btn.setEnabled(False)
        file_row.addWidget(self.file_btn)

        date_row = QHBoxLayout()
        layout.addLayout(date_row)
        date_row.addWidget(QLabel("分析日期："))
        self.analysis_date_edit = QDateEdit()
        self.analysis_date_edit.setCalendarPopup(True)
        self.analysis_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.analysis_date_edit.setDate(QDate.currentDate())
        date_row.addWidget(self.analysis_date_edit)
        date_row.addSpacing(12)
        date_row.addWidget(QLabel("涨停文件(可选)："))
        self.limitup_file_edit = QLabel("自动按分析日期匹配")
        date_row.addWidget(self.limitup_file_edit, 1)
        self.limitup_file_btn = QPushButton("选择涨停文件")
        self.limitup_file_btn.clicked.connect(self._choose_limitup_file)
        date_row.addWidget(self.limitup_file_btn)
        self.limitup_file_clear_btn = QPushButton("清除")
        self.limitup_file_clear_btn.clicked.connect(lambda: self.limitup_file_edit.setText("自动按分析日期匹配"))
        date_row.addWidget(self.limitup_file_clear_btn)

        manual_row = QHBoxLayout()
        layout.addLayout(manual_row)
        manual_row.addWidget(QLabel("股票代码："))
        self.code_edit = QLineEdit()
        self.code_edit.setPlaceholderText("例：600000,000001,300750")
        self.code_edit.setEnabled(False)
        manual_row.addWidget(self.code_edit, 1)

        self.radio_today_limitup.toggled.connect(self._on_radio_changed)
        self.radio_file.toggled.connect(self._on_radio_changed)
        self.radio_manual.toggled.connect(self._on_radio_changed)
        self.file_btn.clicked.connect(self._choose_file)

        # 测试模式（可选）
        test_row = QHBoxLayout()
        layout.addLayout(test_row)
        self.test_cb = QCheckBox("测试：只取股票池前N只")
        self.test_cb.setChecked(False)
        test_row.addWidget(self.test_cb)

        self.test_spin = QSpinBox()
        self.test_spin.setRange(1, 500000)
        self.test_spin.setValue(100)
        self.test_spin.setEnabled(False)
        test_row.addWidget(self.test_spin)

        self.test_cb.toggled.connect(self.test_spin.setEnabled)

        btn_row = QHBoxLayout()
        layout.addLayout(btn_row)
        self.preview_list_btn = QPushButton("查看选择的股票")
        self.analyze_btn = QPushButton("开始分析")
        btn_row.addWidget(self.preview_list_btn)
        btn_row.addWidget(self.analyze_btn)
        btn_row.addStretch()

        self.status_label = QLabel("待开始")
        layout.addWidget(self.status_label)

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["股票代码", "股票名称"])
        layout.addWidget(self.table, 1)

        # Tick 输出（复用原首板区域）
        first_head_row = QHBoxLayout()
        first_head_row.addWidget(QLabel("Tick结构输出（每只股票分析前会清空）"))
        first_head_row.addStretch()
        self.export_first_board_btn = QPushButton("导出 Excel…")
        self.export_first_board_btn.setEnabled(False)
        self.export_first_board_btn.clicked.connect(self._export_first_board_excel)
        first_head_row.addWidget(self.export_first_board_btn)
        layout.addLayout(first_head_row)

        self.first_board_table = QTableWidget()
        self.first_board_table.setColumnCount(0)
        self.first_board_table.setRowCount(0)
        layout.addWidget(self.first_board_table, 1)

        # 今日连板
        streak_head_row = QHBoxLayout()
        streak_head_row.addWidget(QLabel("封单结构分析（封板结构指标，后续会填充）"))
        streak_head_row.addStretch()
        self.export_streak_btn = QPushButton("导出 Excel…")
        self.export_streak_btn.setEnabled(False)
        self.export_streak_btn.clicked.connect(self._export_streak_excel)
        streak_head_row.addWidget(self.export_streak_btn)
        layout.addLayout(streak_head_row)

        self.streak_table = QTableWidget()
        self.streak_table.setColumnCount(0)
        self.streak_table.setRowCount(0)
        layout.addWidget(self.streak_table, 1)

        detail_head = QHBoxLayout()
        detail_head.addWidget(
            QLabel(
                "走势明细：双击「股票列表」或「首板统计」「连板统计」中的行，"
                "查看涨停/连板/跌停及「首板后次日、N连板后次日」（板后第一天的开收，便于算盈亏）。"
            )
        )
        detail_head.addStretch()
        self.export_wechat_html_btn = QPushButton("导出公众号 HTML 源码…")
        self.export_wechat_html_btn.setEnabled(False)
        self.export_wechat_html_btn.setToolTip(
            "生成首板统计、连板统计的公众号 HTML；该股走势明细已全部展开，便于直接粘贴发布。"
        )
        self.export_wechat_html_btn.clicked.connect(self._export_wechat_html)
        detail_head.addWidget(self.export_wechat_html_btn)
        self.export_wechat_overview_html_btn = QPushButton("导出总体情况 HTML（不展开）…")
        self.export_wechat_overview_html_btn.setEnabled(False)
        self.export_wechat_overview_html_btn.setToolTip(
            "生成首板/连板总体情况 HTML，不含每只股票下方的走势明细展开区。"
        )
        self.export_wechat_overview_html_btn.clicked.connect(self._export_wechat_overview_html)
        detail_head.addWidget(self.export_wechat_overview_html_btn)
        self.export_limitup_detail_btn = QPushButton("导出全部走势明细 Excel…")
        self.export_limitup_detail_btn.setEnabled(False)
        self.export_limitup_detail_btn.clicked.connect(self._export_limitup_detail_excel)
        detail_head.addWidget(self.export_limitup_detail_btn)
        layout.addLayout(detail_head)

        self._detail_by_code: Dict[str, List[dict]] = {}
        self._detail_flat: List[dict] = []
        self._code_to_name: Dict[str, str] = {}
        self._current_tick_code: str = ""
        self._current_tick_name: str = ""

        self.table.itemDoubleClicked.connect(self._on_stock_table_double_clicked)
        self.first_board_table.itemDoubleClicked.connect(self._on_first_board_double_clicked)
        self.streak_table.itemDoubleClicked.connect(self._on_streak_double_clicked)

        self.preview_list_btn.clicked.connect(self.show_selected_stocks)
        self.analyze_btn.clicked.connect(self.start_full_analysis)

    def _on_radio_changed(self):
        use_file = self.radio_file.isChecked()
        use_manual = self.radio_manual.isChecked()
        self.file_btn.setEnabled(use_file)
        self.code_edit.setEnabled(use_manual)

    def _choose_file(self):
        fp, _ = QFileDialog.getOpenFileName(self, "选择股票列表文件", "", "Text Files (*.txt *.csv);;All Files (*)")
        if fp:
            self.file_edit.setText(fp)

    def _choose_limitup_file(self):
        fp, _ = QFileDialog.getOpenFileName(
            self,
            "选择涨停板数据文件",
            "",
            "CSV Files (*.csv);;All Files (*)",
        )
        if fp:
            self.limitup_file_edit.setText(fp)

    def _render_tick_preview(self, code: str, name: str, tick_payload: object) -> None:
        payload = tick_payload if isinstance(tick_payload, dict) else {}
        tick_headers = list(payload.get("headers") or [])
        tick_rows = list(payload.get("rows") or [])

        # 展示全部 tick 字段，并额外补上代码/名称两列
        headers = ["股票代码", "股票名称"] + [str(h) for h in tick_headers]
        self.first_board_table.setColumnCount(len(headers))
        self.first_board_table.setHorizontalHeaderLabels(headers)
        self.first_board_table.setRowCount(len(tick_rows))

        for r_idx, row_vals in enumerate(tick_rows):
            vals = [code, name or ""] + [str(v) for v in list(row_vals)]
            for c_idx, txt in enumerate(vals):
                item = QTableWidgetItem(txt)
                if c_idx >= 2:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.first_board_table.setItem(r_idx, c_idx, item)
        self.first_board_table.resizeRowsToContents()

    def _on_tick_preview(self, code: str, name: str, tick_payload: object) -> None:
        # 每只股票输出前先清空显示
        self.first_board_table.setRowCount(0)
        self.first_board_table.setColumnCount(0)
        self._current_tick_code = (code or "").strip()
        self._current_tick_name = (name or "").strip()
        self._render_tick_preview(code, name, tick_payload)

    def _export_qtable_to_excel(
        self,
        table: QTableWidget,
        default_name: str,
        *,
        column_alignments: Optional[List[str]] = None,
        column_widths: Optional[List[Optional[float]]] = None,
        excel_text_format_columns: Optional[Set[int]] = None,
        sheet_zoom: Optional[int] = None,
        show_dialogs: bool = True,
    ) -> bool:
        """将 QTableWidget 导出为 .xlsx（需 openpyxl）。成功返回 True。

        column_alignments: 可选，每列水平对齐：left / center / right，
        长度与列数一致；为 None 时数据格仅在含换行时设置左对齐自动换行（连板表等）。
        column_widths: 可选，每列 Excel 列宽（约等于字符数）；None 表示该列用默认宽度。
        excel_text_format_columns: 对这些列（0 起）设置 Excel 单元格格式为「文本」( @ )；
        对 6 位数字代码前加零宽空格、并尽量设置 quote_prefix，配合 ignoredErrors，
        减轻 Excel/WPS 绿色智能标记；默认仅第 0 列。
        sheet_zoom: 可选，工作表显示缩放百分比（如 90），使表格在 Excel 中更紧凑。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if show_dialogs:
                QMessageBox.warning(
                    self,
                    "缺少依赖",
                    "导出 Excel 需要安装 openpyxl：\npip install openpyxl",
                )
            return False
        if table.rowCount() == 0 or table.columnCount() == 0:
            if show_dialogs:
                QMessageBox.information(self, "提示", "表格为空，无内容可导出。")
            return False
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        if not default_name.lower().endswith(".xlsx"):
            default_name = os.path.splitext(default_name)[0] + ".xlsx"
        if show_dialogs:
            suggested = os.path.join(out_dir, default_name)
            path, _ = QFileDialog.getSaveFileName(
                self,
                "导出 Excel",
                suggested,
                "Excel 工作簿 (*.xlsx);;所有文件 (*.*)",
            )
            if not path:
                return False
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
        else:
            if os.path.isabs(default_name):
                path = default_name
            else:
                path = os.path.join(out_dir, default_name)
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            os.makedirs(os.path.dirname(path), exist_ok=True)
        text_fmt_cols: Set[int] = (
            set(excel_text_format_columns)
            if excel_text_format_columns is not None
            else {0}
        )
        wb = Workbook()
        ws = wb.active
        sheet_title = os.path.splitext(os.path.basename(default_name))[0] or "数据"
        ws.title = sheet_title[:31]
        header_font = Font(bold=True)
        header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        ncols = table.columnCount()
        nrows = table.rowCount()
        for c in range(ncols):
            hi = table.horizontalHeaderItem(c)
            text = hi.text() if hi else ""
            cell = ws.cell(row=1, column=c + 1, value=text)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = cell_border
            if c in text_fmt_cols:
                cell.number_format = "@"
        data_wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for r in range(nrows):
            for c in range(ncols):
                it = table.item(r, c)
                val = it.text() if it else ""
                # 第一列强制文本；其余列若为纯数字则写入数值，便于 Excel 计算/排序
                out_val = val
                if c == 0:
                    out_val = str(val)
                else:
                    sv = str(val).strip()
                    if re.fullmatch(r"[+-]?\d+", sv):
                        try:
                            out_val = int(sv)
                        except Exception:
                            out_val = val
                    elif re.fullmatch(r"[+-]?(?:\d+\.\d*|\.\d+)", sv):
                        try:
                            out_val = float(sv)
                        except Exception:
                            out_val = val
                if c in text_fmt_cols and isinstance(out_val, str):
                    out_val = _excel_stock_cell_str_for_export(out_val)
                cell = ws.cell(row=r + 2, column=c + 1, value=out_val)
                cell.border = cell_border
                if c in text_fmt_cols:
                    cell.number_format = "@"
                    qp = getattr(cell, "quote_prefix", None)
                    if qp is not None:
                        try:
                            cell.quote_prefix = True
                        except Exception:
                            pass
                if column_alignments is not None and c < len(column_alignments):
                    h = column_alignments[c]
                    if h not in ("left", "center", "right"):
                        h = "left"
                    if "\n" in str(val):
                        cell.alignment = Alignment(
                            wrap_text=True, vertical="top", horizontal=h
                        )
                    else:
                        cell.alignment = Alignment(vertical="center", horizontal=h)
                elif "\n" in str(val):
                    cell.alignment = data_wrap
        if column_widths:
            for c in range(ncols):
                if c >= len(column_widths):
                    break
                w = column_widths[c]
                if w is not None and float(w) > 0:
                    ws.column_dimensions[get_column_letter(c + 1)].width = float(w)
        if sheet_zoom is not None:
            try:
                z = int(sheet_zoom)
                if 10 <= z <= 400:
                    ws.sheet_view.zoomScale = z
            except Exception:
                pass
        # Excel 仍可能对「长得像数字的文本」显示绿色角标；在工作表上声明忽略该检查（OOXML ignoredErrors）
        if text_fmt_cols:
            try:
                from openpyxl.worksheet.cell_range import CellRange
                from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

                last_row = nrows + 1
                ign_list = []
                for c in sorted(text_fmt_cols):
                    letter = get_column_letter(c + 1)
                    rng = f"{letter}1:{letter}{last_row}"
                    ign_list.append(
                        IgnoredError(sqref=CellRange(rng), numberStoredAsText=True)
                    )
                if ign_list:
                    ws.ignored_errors = IgnoredErrors(ignoredError=tuple(ign_list))
            except Exception:
                pass
        try:
            wb.save(path)
        except OSError as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", str(e))
            return False
        except Exception as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", f"{type(e).__name__}: {e}")
            return False
        if show_dialogs:
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        return True

    def _export_first_board_excel(self):
        date = getattr(self, "_limitup_date", "") or "export"
        code = (getattr(self, "_current_tick_code", "") or "").strip()
        name = (getattr(self, "_current_tick_name", "") or "").strip()
        full_headers: List[str] = []
        full_rows: List[List[str]] = []
        if xtdata is not None and code and re.fullmatch(r"\d{6}", code):
            try:
                tick_df = _load_or_fetch_tick_df(code, date)
                full_headers, full_rows = _extract_tick_preview_rows(tick_df, limit_n=None)
            except Exception:
                full_headers, full_rows = [], []

        if full_headers and full_rows:
            t = QTableWidget(len(full_rows), 2 + len(full_headers))
            t.setHorizontalHeaderLabels(["股票代码", "股票名称"] + [str(h) for h in full_headers])
            for r_idx, row_vals in enumerate(full_rows):
                vals = [code, name] + [str(v) for v in list(row_vals)]
                for c_idx, txt in enumerate(vals):
                    item = QTableWidgetItem(txt)
                    if c_idx >= 2:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    t.setItem(r_idx, c_idx, item)
            n = t.columnCount()
            aligns: List[str] = []
            for c in range(n):
                if c == 0:
                    aligns.append("center")
                elif c == 1:
                    aligns.append("left")
                else:
                    aligns.append("right")
            self._export_qtable_to_excel(
                t,
                f"ticks_{date}.xlsx",
                column_alignments=aligns,
            )
            return

        n = self.first_board_table.columnCount()
        aligns: List[str] = []
        for c in range(n):
            if c == 0:
                aligns.append("center")
            elif c == 1:
                aligns.append("left")
            else:
                aligns.append("right")
        self._export_qtable_to_excel(
            self.first_board_table,
            f"ticks_{date}.xlsx",
            column_alignments=aligns,
        )

    def _export_streak_excel(self):
        date = getattr(self, "_limitup_date", "") or "export"
        n = self.streak_table.columnCount()
        # 股票代码居中；今日连板数、近一年最大连板数右对齐；名称、各连板次数分布、连板晋级比例左对齐
        streak_col_align = {
            0: "center",
            1: "left",
            2: "right",
            3: "right",
            4: "left",
            5: "left",
        }
        aligns = [streak_col_align.get(c, "left") for c in range(n)]
        # 列宽略收紧（仍支持换行）；工作表缩放 90% 更紧凑（openpyxl 列宽约等于字符宽度）
        streak_widths: List[Optional[float]] = [
            10.5,  # 股票代码（+2 字符）
            9.5,  # 名称
            9.0,  # 今日连板数（+2 字符）
            10.0,  # 近一年最大连板数
            13.0,  # 各连板次数分布（原约一半）
            19.0,  # 连板晋级比例（+2 字符）
        ]
        while len(streak_widths) < n:
            streak_widths.append(None)
        self._export_qtable_to_excel(
            self.streak_table,
            f"封单结构_{date}.xlsx",
            column_alignments=aligns,
            column_widths=streak_widths[:n],
            sheet_zoom=90,
        )

    def _export_qtable_to_png(self, table: QTableWidget, png_path: str) -> bool:
        """将表格按自适应列宽导出为 PNG 图片。"""
        try:
            nrows, ncols = table.rowCount(), table.columnCount()
            if nrows <= 0 or ncols <= 0:
                return False
            tmp = QTableWidget(nrows, ncols)
            headers = []
            for c in range(ncols):
                hi = table.horizontalHeaderItem(c)
                headers.append(hi.text() if hi else "")
            tmp.setHorizontalHeaderLabels(headers)
            for r in range(nrows):
                for c in range(ncols):
                    it = table.item(r, c)
                    tmp.setItem(r, c, QTableWidgetItem(it.text() if it else ""))

            tmp.resizeColumnsToContents()
            tmp.resizeRowsToContents()
            total_w = tmp.verticalHeader().width() + 2 * tmp.frameWidth()
            total_h = tmp.horizontalHeader().height() + 2 * tmp.frameWidth()
            for c in range(ncols):
                total_w += tmp.columnWidth(c)
            for r in range(nrows):
                total_h += tmp.rowHeight(r)
            total_w = max(total_w, 80)
            total_h = max(total_h, 60)
            tmp.resize(total_w, total_h)
            # 强制白底，避免透明背景在部分查看器中显示为黑底
            tmp.setStyleSheet("QTableWidget { background: #ffffff; }")
            tmp.viewport().setStyleSheet("background: #ffffff;")
            QApplication.processEvents()

            pix = QPixmap(tmp.size())
            pix.fill(Qt.white)
            tmp.render(pix)
            return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
        except Exception:
            return False

    def _auto_adjust_excel_column_widths(self, xlsx_path: str) -> bool:
        """自动按内容调列宽，避免导图时列宽拥挤。"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            return False
        if not os.path.exists(xlsx_path):
            return False
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, max_row + 1):
                    val = ws.cell(row=r, column=c).value
                    text = "" if val is None else str(val)
                    text = text.replace("\n", " ")
                    if len(text) > max_len:
                        max_len = len(text)
                # 宽度限制在合理范围，防止超长文本导致巨宽图片
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)
            wb.save(xlsx_path)
            return True
        except Exception:
            return False

    def _excel_to_png_via_wps(self, xlsx_path: str, png_path: str) -> bool:
        """优先通过 WPS 渲染 Excel 后复制为图片。"""
        if not os.path.exists(xlsx_path):
            return False
        try:
            import pythoncom
            import win32com.client as win32
            from PIL import Image, ImageGrab
            import time
        except Exception:
            return False

        app = None
        wb = None
        inited = False
        try:
            pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
            inited = True
            for progid in ("ket.Application", "et.Application", "KET.Application", "ET.Application"):
                try:
                    app = win32.DispatchEx(progid)
                    if app is not None:
                        break
                except Exception:
                    app = None
            if app is None:
                return False
            app.Visible = False
            app.DisplayAlerts = False
            wb = app.Workbooks.Open(os.path.abspath(xlsx_path))
            ws = wb.Worksheets(1)
            ws.Activate()
            used = ws.UsedRange
            used.EntireColumn.AutoFit()
            used.CopyPicture()
            time.sleep(0.25)
            img = ImageGrab.grabclipboard()
            if img is None:
                return False
            if isinstance(img, list):
                return False
            if getattr(img, "mode", "") in ("RGBA", "LA"):
                white_bg = Image.new("RGB", img.size, "white")
                alpha = img.split()[-1]
                white_bg.paste(img, mask=alpha)
                img = white_bg
            elif getattr(img, "mode", "") != "RGB":
                img = img.convert("RGB")
            img.save(png_path, "PNG")
            return os.path.exists(png_path)
        except Exception:
            return False
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=True)
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            if inited:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _auto_export_streak_excel_and_png(self) -> bool:
        """自动模式：导出封单结构 Excel，再导出表格图片。"""
        if self.streak_table.rowCount() <= 0 or self.streak_table.columnCount() <= 0:
            self.status_label.setText("自动运行完成，但封单结构表为空，未导出。")
            return False
        date = getattr(self, "_limitup_date", "") or datetime.now().strftime("%Y%m%d")
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history_data")
        os.makedirs(out_dir, exist_ok=True)
        xlsx_path = os.path.join(out_dir, f"封单结构_{date}.xlsx")
        png_path = os.path.join(out_dir, f"封单结构_{date}.png")

        n = self.streak_table.columnCount()
        streak_col_align = {0: "center", 1: "left", 2: "right", 3: "right", 4: "left", 5: "left"}
        aligns = [streak_col_align.get(c, "left") for c in range(n)]

        ok_xlsx = self._export_qtable_to_excel(
            self.streak_table,
            xlsx_path,
            column_alignments=aligns,
            column_widths=None,
            sheet_zoom=90,
            show_dialogs=False,
        )
        if not ok_xlsx:
            self.status_label.setText("自动导出失败：Excel 导出失败。")
            return False

        self._auto_adjust_excel_column_widths(xlsx_path)
        ok_png = self._excel_to_png_via_wps(xlsx_path, png_path)
        if not ok_png:
            ok_png = self._export_qtable_to_png(self.streak_table, png_path)
        if not ok_png:
            self.status_label.setText("自动导出失败：图片导出失败。")
            return False

        self._auto_last_export_xlsx = xlsx_path
        self._auto_last_export_png = png_path
        self.status_label.setText(
            f"自动运行完成，已导出：{os.path.basename(xlsx_path)}、{os.path.basename(png_path)}；10秒后退出。"
        )
        return True

    def _export_wechat_html(self) -> None:
        """导出公众号用 HTML：首板表 + 连板表；走势明细默认全部展开（无折叠控件）。"""
        if self.first_board_table.rowCount() == 0 and self.streak_table.rowCount() == 0:
            QMessageBox.information(self, "提示", "首板统计与连板统计均为空，请先完成「开始分析」。")
            return
        date = getattr(self, "_limitup_date", "") or "export"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        default_name = f"涨停统计_公众号_{date}.html"
        suggested = os.path.join(out_dir, default_name)
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出公众号 HTML 源码",
            suggested,
            "HTML 文件 (*.html);;所有文件 (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".html"):
            path += ".html"
        detail_map = getattr(self, "_detail_by_code", None) or {}
        doc = _build_wechat_limitup_article_html(
            date,
            self.first_board_table,
            self.streak_table,
            detail_map,
        )
        try:
            with open(path, "w", encoding="utf-8", newline="\n") as f:
                f.write(doc)
        except OSError as e:
            QMessageBox.warning(self, "导出失败", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"{type(e).__name__}: {e}")
            return
        QMessageBox.information(
            self,
            "导出成功",
            "已保存：\n"
            f"{path}\n\n"
            "使用说明：该 HTML 走势明细已全部展开；你可以直接复制粘贴到微信公众号编辑器发布。",
        )

    def _export_wechat_overview_html(self) -> None:
        """导出公众号用 HTML：仅总体情况（不含每只股票下方的明细表）。"""
        if self.first_board_table.rowCount() == 0 and self.streak_table.rowCount() == 0:
            QMessageBox.information(self, "提示", "首板统计与连板统计均为空，请先完成「开始分析」。")
            return
        date = getattr(self, "_limitup_date", "") or "export"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        default_name = f"涨停统计_公众号_总体情况_{date}.html"
        suggested = os.path.join(out_dir, default_name)
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出总体情况 HTML（不展开）",
            suggested,
            "HTML 文件 (*.html);;所有文件 (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".html"):
            path += ".html"
        detail_map = getattr(self, "_detail_by_code", None) or {}
        doc = _build_wechat_limitup_article_html(
            date,
            self.first_board_table,
            self.streak_table,
            detail_map,
            include_details=False,
        )
        try:
            with open(path, "w", encoding="utf-8", newline="\n") as f:
                f.write(doc)
        except OSError as e:
            QMessageBox.warning(self, "导出失败", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"{type(e).__name__}: {e}")
            return
        QMessageBox.information(
            self,
            "导出成功",
            "已保存：\n"
            f"{path}\n\n"
            "使用说明：该版本仅包含总体统计，不含每只股票的走势明细展开内容。",
        )

    def _clear_stat_tables(self) -> None:
        self.first_board_table.setRowCount(0)
        self.first_board_table.setColumnCount(0)
        self.streak_table.setRowCount(0)
        self.streak_table.setColumnCount(0)
        self._detail_by_code = {}
        self._detail_flat = []
        if hasattr(self, "export_limitup_detail_btn"):
            self.export_limitup_detail_btn.setEnabled(False)
        if hasattr(self, "export_wechat_html_btn"):
            self.export_wechat_html_btn.setEnabled(False)
        if hasattr(self, "export_wechat_overview_html_btn"):
            self.export_wechat_overview_html_btn.setEnabled(False)

    def _code_from_table_row(self, table: QTableWidget, row: int) -> str:
        it = table.item(row, 0)
        return (it.text() or "").strip() if it else ""

    def _open_limitup_detail(self, code: str) -> None:
        code = (code or "").strip()
        if not code:
            return
        mp = getattr(self, "_detail_by_code", None) or {}
        details = list(mp.get(code) or [])
        if not details:
            QMessageBox.information(
                self,
                "走势明细",
                f"股票 {code} 在统计区间内无涨停/跌停明细记录，或尚未完成「开始分析」。",
            )
            return
        name = getattr(self, "_code_to_name", {}).get(code, "")
        dlg = StockLimitUpDetailDialog(code, name, details, self)
        dlg.exec_()

    def _on_stock_table_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.table, item.row())
        self._open_limitup_detail(code)

    def _on_first_board_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.first_board_table, item.row())
        self._open_limitup_detail(code)

    def _on_streak_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.streak_table, item.row())
        self._open_limitup_detail(code)

    def _export_limitup_detail_excel(self) -> None:
        flat = getattr(self, "_detail_flat", None) or []
        if not flat:
            QMessageBox.information(self, "提示", "没有可导出的走势明细，请先完成「开始分析」。")
            return
        t = QTableWidget(len(flat), 6)
        t.setHorizontalHeaderLabels(
            ["股票代码", "股票名称", "日期", "开盘", "收盘", "备注"]
        )
        for r, row in enumerate(flat):
            t.setItem(r, 0, QTableWidgetItem(str(row.get("code", ""))))
            t.setItem(r, 1, QTableWidgetItem(str(row.get("name", ""))))
            t.setItem(r, 2, QTableWidgetItem(str(row.get("date", ""))))
            t.setItem(r, 3, QTableWidgetItem(_fmt_px(row.get("open"))))
            t.setItem(r, 4, QTableWidgetItem(_fmt_px(row.get("close"))))
            t.setItem(r, 5, QTableWidgetItem(str(row.get("remark", ""))))
        date = getattr(self, "_limitup_date", "") or "export"
        aligns = ["center", "left", "center", "right", "right", "left"]
        self._export_qtable_to_excel(
            t,
            f"走势明细全量_{date}.xlsx",
            column_alignments=aligns,
        )

    def _enable_action_buttons(self, enabled: bool) -> None:
        self.preview_list_btn.setEnabled(enabled)
        self.analyze_btn.setEnabled(enabled)

    def _build_thread_params(self) -> Optional[dict]:
        """校验输入并返回线程构造参数；失败返回 None。"""
        script_dir = os.path.dirname(os.path.abspath(__file__))
        history_dir = os.path.join(script_dir, "history_data")
        all_a_csv = os.path.join(script_dir, "data", "all_a_stocks.csv")

        source_mode = "today_limitup"
        stock_list_filepath: Optional[str] = None
        manual_codes: List[str] = []

        if self.radio_today_limitup.isChecked():
            source_mode = "today_limitup"
        elif self.radio_file.isChecked():
            source_mode = "file"
            stock_list_filepath = self.file_edit.text().strip()
            if stock_list_filepath == "未选择文件" or not os.path.exists(stock_list_filepath):
                self.status_label.setText("请先选择股票列表文件")
                return None
        else:
            source_mode = "manual"
            manual_codes = _parse_codes_from_comma_text(self.code_edit.text())
            if not manual_codes:
                self.status_label.setText("请在文本框中输入至少一只股票代码（逗号分隔）")
                return None

        test_only = bool(self.test_cb.isChecked())
        test_limit_n = int(self.test_spin.value())
        analysis_date_yyyymmdd = self.analysis_date_edit.date().toString("yyyyMMdd")
        limitup_csv_path = self.limitup_file_edit.text().strip()
        if limitup_csv_path == "自动按分析日期匹配":
            limitup_csv_path = ""

        return {
            "history_dir": history_dir,
            "all_a_stocks_csv": all_a_csv,
            "source_mode": source_mode,
            "stock_list_filepath": stock_list_filepath,
            "manual_codes": manual_codes,
            "analysis_date_yyyymmdd": analysis_date_yyyymmdd,
            "limitup_csv_path": limitup_csv_path,
            "test_only": test_only,
            "test_limit_n": test_limit_n,
        }

    def show_selected_stocks(self) -> None:
        self._enable_action_buttons(False)
        self.table.setRowCount(0)
        self._clear_stat_tables()
        self.export_first_board_btn.setEnabled(False)
        self.export_streak_btn.setEnabled(False)
        if hasattr(self, "export_wechat_html_btn"):
            self.export_wechat_html_btn.setEnabled(False)
        if hasattr(self, "export_wechat_overview_html_btn"):
            self.export_wechat_overview_html_btn.setEnabled(False)
        self.status_label.setText("正在生成股票列表，请稍候...")

        params = self._build_thread_params()
        if params is None:
            self._enable_action_buttons(True)
            return

        self.thread = TodayLimitUpFinderThread(
            run_mode="list_only",
            parent=self,
            **params,
        )
        self.thread.finished.connect(self._on_finished)
        self.thread.error_occurred.connect(self._on_error)
        self.thread.tick_preview.connect(self._on_tick_preview)
        self.thread.start()

    def start_full_analysis(self) -> None:
        self._enable_action_buttons(False)
        self._clear_stat_tables()
        self.export_first_board_btn.setEnabled(False)
        self.export_streak_btn.setEnabled(False)
        if hasattr(self, "export_wechat_html_btn"):
            self.export_wechat_html_btn.setEnabled(False)
        if hasattr(self, "export_wechat_overview_html_btn"):
            self.export_wechat_overview_html_btn.setEnabled(False)
        self.status_label.setText("正在逐只分析 tick 结构，请稍候（可能较慢）...")

        params = self._build_thread_params()
        if params is None:
            self._enable_action_buttons(True)
            return

        self.thread = TodayLimitUpFinderThread(
            run_mode="full",
            parent=self,
            **params,
        )
        self.thread.finished.connect(self._on_finished)
        self.thread.error_occurred.connect(self._on_error)
        self.thread.tick_preview.connect(self._on_tick_preview)
        self.thread.start()

    def _on_finished(
        self,
        today_rows: List[Tuple[str, str]],
        first_board_rows: List[dict],
        streak_rows: List[dict],
        limitup_date: str,
        debug_msg: str,
        finish_kind: str,
        detail_by_code: object = None,
    ):
        self._limitup_date = limitup_date
        if not isinstance(detail_by_code, dict):
            detail_by_code = {}

        self._code_to_name = dict(today_rows)
        self._detail_by_code = dict(detail_by_code) if finish_kind != "list_only" else {}
        self._detail_flat = []
        for c, lst in sorted(self._detail_by_code.items(), key=lambda x: x[0]):
            nm = self._code_to_name.get(c, "")
            for rec in lst:
                self._detail_flat.append({"code": c, "name": nm, **rec})

        # 要分析的股票列表
        self.table.setRowCount(len(today_rows))
        for i, (code, name) in enumerate(today_rows):
            self.table.setItem(i, 0, QTableWidgetItem(code))
            self.table.setItem(i, 1, QTableWidgetItem(name or ""))

        if finish_kind == "list_only":
            self._clear_stat_tables()
            self.export_first_board_btn.setEnabled(False)
            self.export_streak_btn.setEnabled(False)
            self.export_limitup_detail_btn.setEnabled(False)
            if hasattr(self, "export_wechat_html_btn"):
                self.export_wechat_html_btn.setEnabled(False)
            if hasattr(self, "export_wechat_overview_html_btn"):
                self.export_wechat_overview_html_btn.setEnabled(False)
        elif finish_kind == "tick_only":
            # 首板区域已在 _on_tick_preview 中按“逐只清空+输出”更新；连板区展示结构指标结果
            structure_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("close_order_amount_yi", "收盘封单金额(亿)"),
                ("seal_hardness", "封板硬度"),
                ("rush_intensity", "抢筹烈度"),
                ("order_stability", "封单稳定性"),
                ("order_trend", "封单运行趋势"),
                ("order_rating", "封单评级"),
                ("confidence_tag", "置信度"),
            ]
            self.streak_table.setColumnCount(len(structure_headers))
            self.streak_table.setHorizontalHeaderLabels([h[1] for h in structure_headers])
            self.streak_table.setRowCount(len(streak_rows))
            for r_idx, row in enumerate(streak_rows):
                for c_idx, (key, _) in enumerate(structure_headers):
                    txt = "" if row.get(key) is None else str(row.get(key))
                    item = QTableWidgetItem(txt)
                    if key == "confidence_tag":
                        item.setTextAlignment(Qt.AlignCenter)
                    self.streak_table.setItem(r_idx, c_idx, item)
            self.streak_table.resizeRowsToContents()

            self.export_first_board_btn.setEnabled(self.first_board_table.rowCount() > 0)
            self.export_streak_btn.setEnabled(len(streak_rows) > 0)
            self.export_limitup_detail_btn.setEnabled(False)
            if hasattr(self, "export_wechat_html_btn"):
                self.export_wechat_html_btn.setEnabled(False)
            if hasattr(self, "export_wechat_overview_html_btn"):
                self.export_wechat_overview_html_btn.setEnabled(False)
        else:
            # 首板表（列键与后台 dict 一致，表头为中文）
            first_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("today_streak", "今日连板数"),
                ("first_board_count_1y", "首板次数"),
                ("next_day_limitup_count", "次日涨停次数"),
                ("next_day_up_count", "次日收涨次数"),
                ("next_day_down_count", "次日收跌次数"),
                ("avg_open_return_pct", "次日开盘收益均值(%)"),
                ("avg_close_return_pct", "次日收盘收益均值(%)"),
            ]
            self.first_board_table.setColumnCount(len(first_headers))
            self.first_board_table.setHorizontalHeaderLabels([h[1] for h in first_headers])
            self.first_board_table.setRowCount(len(first_board_rows))
            for r_idx, row in enumerate(first_board_rows):
                for c_idx, (key, _) in enumerate(first_headers):
                    val = row.get(key, "")
                    if key in ("avg_open_return_pct", "avg_close_return_pct"):
                        try:
                            cell_text = (
                                f"{float(val):.2f}"
                                if val is not None and val != ""
                                else ""
                            )
                        except (TypeError, ValueError):
                            cell_text = str(val)
                    else:
                        cell_text = str(val if val is not None else "")
                    self.first_board_table.setItem(r_idx, c_idx, QTableWidgetItem(cell_text))

            # 连板表
            streak_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("today_streak", "今日连板数"),
                ("max_streak_in_year", "近一年最大连板数"),
                ("counts_by_k", "各连板次数分布"),
                ("upgrade_ratios", "连板晋级比例"),
            ]
            self.streak_table.setColumnCount(len(streak_headers))
            self.streak_table.setHorizontalHeaderLabels([h[1] for h in streak_headers])
            self.streak_table.setRowCount(len(streak_rows))
            for r_idx, row in enumerate(streak_rows):
                for c_idx, (key, _) in enumerate(streak_headers):
                    v = row.get(key, "")
                    text = "" if v is None else str(v)
                    item = QTableWidgetItem(text)
                    if key in ("counts_by_k", "upgrade_ratios") and "\n" in text:
                        item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)
                    self.streak_table.setItem(r_idx, c_idx, item)
            self.streak_table.resizeRowsToContents()

            self.export_first_board_btn.setEnabled(len(first_board_rows) > 0)
            self.export_streak_btn.setEnabled(len(streak_rows) > 0)
            self.export_limitup_detail_btn.setEnabled(len(self._detail_flat) > 0)
            if hasattr(self, "export_wechat_html_btn"):
                self.export_wechat_html_btn.setEnabled(
                    len(first_board_rows) > 0 or len(streak_rows) > 0
                )
            if hasattr(self, "export_wechat_overview_html_btn"):
                self.export_wechat_overview_html_btn.setEnabled(
                    len(first_board_rows) > 0 or len(streak_rows) > 0
                )

        # 保存一个 txt，方便你核对
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"要分析的股票列表_{limitup_date}.txt")
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                for code, name in today_rows:
                    f.write(f"{code}\t{name}\n")
        except Exception:
            out_path = ""

        suffix = f"；已保存：{out_path}" if out_path else ""
        if finish_kind == "list_only":
            self.status_label.setText(
                f"已列出 {len(today_rows)} 只股票；{suffix}\n可点击「开始分析」生成首板/连板统计。\n{debug_msg}"
            )
        elif finish_kind == "tick_only":
            self.status_label.setText(
                f"tick结构分析完成：共 {len(today_rows)} 只股票；{suffix}\n{debug_msg}"
            )
        else:
            self.status_label.setText(f"分析完成：共 {len(today_rows)} 只股票；{suffix}\n{debug_msg}")

        self._enable_action_buttons(True)
        if self._auto_run and finish_kind != "list_only" and (not self._auto_export_done):
            self._auto_export_done = True
            self._auto_export_streak_excel_and_png()
            QTimer.singleShot(10000, self.accept)

    def _on_error(self, msg: str):
        self.status_label.setText(f"出错：{msg}")
        self._enable_action_buttons(True)


def main():
    parser = argparse.ArgumentParser(description="涨停板结构分析")
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="启动后自动分析，分析完成后自动导出封单结构 Excel 和图片，并在10秒后退出",
    )
    args, _unknown = parser.parse_known_args()

    app = QApplication(sys.argv)
    dlg = LimitUpGeneAnalysisDialog(auto_run=bool(args.auto_run))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg.start_full_analysis)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()


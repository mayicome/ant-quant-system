#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
涨停基因分析：按所选股票池统计历史涨停基因（与当日是否涨停无必然关系）。

输入来源：指定股票列表文件（txt/csv/xls/xlsx；代码列优先「代码」否则首列；列名「选股日」可重复，同一代码多个日期在主列表中为多行），分析列表内全部股票（不要求当日涨停）。
涨停对照表自动取 history_data 下「涨停板数据_*」中不晚于当日的最新文件；分析日期按本机当天。
tick 分析：若列表含「选股日」，则拉取该自然日之后首个交易日的 tick；无选股日时仍用涨停表日期。

若指定股票当日未涨停，仍参与分析，在「首板统计」表中展示，今日连板数为 0。
"""

import csv
import json
import os
import re
import sys
import argparse
import bisect
import warnings
from datetime import date as date_cls, datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

from PyQt5.QtCore import QThread, QTimer, pyqtSignal
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QDoubleSpinBox,
    QGroupBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

# PyQt5 与新版 sip 在子类化 QObject 时会触发 sipPyTypeDict 弃用提示，属绑定层内部，应用代码无法消除根因
warnings.filterwarnings(
    "ignore",
    category=DeprecationWarning,
    message=r".*sipPyTypeDict.*",
)

try:
    import xtquant.xtdata as xtdata
except Exception:
    xtdata = None

try:
    from core.utils.security_type import SecurityTypeUtil as _StockPricePrecisionUtil
except Exception:
    _StockPricePrecisionUtil = None


def _format_pressure_p_by_stock_code(stock_code: str, p_value: float) -> str:
    """压力位 P 展示：按代码证券类型保留小数（与 SecurityTypeUtil 一致：基金 3 位，其余 2 位）。"""
    code = (stock_code or "").strip()
    if _StockPricePrecisionUtil is not None:
        prec = _StockPricePrecisionUtil.get_price_precision(code)
        val = _StockPricePrecisionUtil.round_price(code, float(p_value))
    else:
        base = code.split(".")[0] if "." in code else code
        prec = 3 if (len(base) == 6 and base.startswith("5")) else 2
        val = round(float(p_value), prec)
    return f"{val:.{prec}f}"


def _round_price_like_display(stock_code: str, p: float) -> float:
    """与压力位/最新价展示同一套舍入，用于突破判定（触及同价不算突破）。"""
    code = (stock_code or "").strip()
    if _StockPricePrecisionUtil is not None:
        return float(_StockPricePrecisionUtil.round_price(code, float(p)))
    base = code.split(".")[0] if "." in code else code
    prec = 3 if (len(base) == 6 and base.startswith("5")) else 2
    return round(float(p), prec)


def _pressure_effective_update_threshold(stock_code: str, p_old: float) -> float:
    """
    实际沿用压力位是否用新 P 的判据：|P_new - P_old| 大于本阈值才更新。
    ETF：固定价差（元）；股票：按 P_old 的百分比。
    """
    code = (stock_code or "").strip()
    is_fund = False
    if _StockPricePrecisionUtil is not None:
        is_fund = _StockPricePrecisionUtil.is_fund(code)
    else:
        base = code.split(".")[0] if "." in code else code
        is_fund = len(base) == 6 and base.startswith("5")
    if is_fund:
        return float(_PRESSURE_EFFECTIVE_UPDATE_ABS_ETF)
    po = abs(float(p_old))
    return max(po * float(_PRESSURE_EFFECTIVE_UPDATE_PCT), 1e-12)


def _to_full_stock_code(stock_code_6: str) -> str:
    """把 6 位股票代码映射为 xtdata 常用的全代码（.SH/.SZ/.BJ）。"""
    s = str(stock_code_6).strip()
    if "." in s:
        return s
    s = _zfill_6(s)
    if not s:
        return s
    if s.startswith(("0", "1", "3")):
        return f"{s}.SZ"
    if s.startswith("6"):
        return f"{s}.SH"
    if s.startswith(("8", "4")):
        return f"{s}.BJ"
    if s.startswith("920"):
        return f"{s}.BJ"
    return s


def _extract_open_close_map(stock_df) -> Dict[str, Tuple[float, float, float, float]]:
    """
    stock_df: xtdata.get_market_data_ex 返回的单只股票 DataFrame（列至少 open/high/low/close，index为时间戳或字符串）。
    返回：date_yyyymmdd -> (open, high, low, close)
    """
    out: Dict[str, Tuple[float, float, float, float]] = {}
    if stock_df is None or getattr(stock_df, "empty", True):
        return out
    if (
        "open" not in stock_df.columns
        or "high" not in stock_df.columns
        or "low" not in stock_df.columns
        or "close" not in stock_df.columns
    ):
        return out

    for idx in stock_df.index:
        try:
            if isinstance(idx, (int, float)):
                dt = datetime.fromtimestamp(float(idx) / 1000.0)
                ds = dt.strftime("%Y%m%d")
            elif isinstance(idx, str):
                ds = idx.replace("-", "")[:8]
            else:
                ds = str(idx).replace("-", "")[:8]
            if not ds or len(ds) != 8:
                continue

            o = float(stock_df.loc[idx, "open"])
            h = float(stock_df.loc[idx, "high"])
            l = float(stock_df.loc[idx, "low"])
            c = float(stock_df.loc[idx, "close"])
            out[ds] = (o, h, l, c)
        except Exception:
            continue
    return out


def _parse_pick_day_to_yyyymmdd(s: str) -> Optional[str]:
    """将「选股日」单元格/表字符串解析为 YYYYMMDD；无法解析返回 None。"""
    raw = (s or "").strip()
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 8 and digits[:8].isdigit():
        return digits[:8]
    m = re.search(r"(20\d{2})\D?(\d{1,2})\D?(\d{1,2})", raw)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        try:
            return f"{y}{int(mo):02d}{int(d):02d}"
        except ValueError:
            return None
    return None


def _yyyymmdd_to_display(s: str) -> str:
    """YYYYMMDD -> YYYY-MM-DD 展示；非法则原样返回去首尾空格。"""
    t = (s or "").strip()
    if len(t) == 8 and t.isdigit():
        return f"{t[:4]}-{t[4:6]}-{t[6:8]}"
    return t


def _next_trading_day_after_yyyymmdd(after_yyyymmdd: str) -> Optional[str]:
    """严格晚于 after_yyyymmdd（自然日）的首个交易日 YYYYMMDD；失败返回 None。"""
    if not after_yyyymmdd or len(after_yyyymmdd) != 8 or not after_yyyymmdd.isdigit():
        return None
    if xtdata is None:
        return None
    try:
        start_dt = datetime.strptime(after_yyyymmdd, "%Y%m%d")
        end_dt = start_dt + timedelta(days=400)
        start_s = after_yyyymmdd
        end_s = end_dt.strftime("%Y%m%d")
        ts_list = xtdata.get_trading_dates("SH", start_time=start_s, end_time=end_s)
        trade_dates: List[str] = []
        for ts in ts_list or []:
            if isinstance(ts, (int, float)):
                dt = datetime.fromtimestamp(float(ts) / 1000.0)
                trade_dates.append(dt.strftime("%Y%m%d"))
            elif isinstance(ts, str):
                trade_dates.append(ts.replace("-", "")[:8])
            else:
                trade_dates.append(str(ts)[:8])
        trade_dates = sorted({d for d in trade_dates if d and len(d) == 8})
        for d in trade_dates:
            if d > after_yyyymmdd:
                return d
    except Exception:
        return None
    return None


def _load_or_fetch_tick_df(stock_code_6: str, trade_date_yyyymmdd: str):
    """调用 QMT 拉取指定交易日 trade_date_yyyymmdd 的 tick 数据。"""
    if xtdata is None:
        return None

    full_code = _to_full_stock_code(stock_code_6)
    tick_start = f"{trade_date_yyyymmdd}000000"
    tick_end = f"{trade_date_yyyymmdd}235959"
    try:
        xtdata.download_history_data(full_code, "tick", tick_start, tick_end)
        tick_df_map = xtdata.get_market_data_ex(
            [],
            [full_code],
            period="tick",
            start_time=tick_start,
            end_time=tick_end,
            count=-1,
        )
        tick_df = tick_df_map.get(full_code) if isinstance(tick_df_map, dict) else None
    except Exception:
        tick_df = None

    return tick_df


def _zfill_6(code: str) -> str:
    s = "".join(ch for ch in str(code or "").strip() if ch.isdigit())
    if not s:
        return ""
    return s[:6].zfill(6) if len(s) >= 6 else s.zfill(6)


def _is_valid_stock_code_6(code: str) -> bool:
    """六位数字代码；排除占位/异常产生的 000000（_zfill_6('0') 会得到该值）。"""
    if not code or len(code) != 6 or not code.isdigit():
        return False
    return code != "000000"


# 导出 xlsx 时加在 6 位代码前，打断 Excel/WPS「纯数字当文本」的绿色角标；单元格显示仍为 6 位数字
_EXCEL_STOCK_TEXT_PREFIX = "\u200b"


def _excel_stock_cell_str_for_export(s: str) -> str:
    """仅用于写入 Excel：6 位全数字则前加零宽空格，避免智能标记；其它字符串原样。"""
    t = (s or "").strip()
    if len(t) == 6 and t.isdigit():
        return _EXCEL_STOCK_TEXT_PREFIX + t
    return t


def _excel_cell_to_trimmed_str(x: Any) -> str:
    """Excel 单元格转字符串；处理 NaN、float 代码等。"""
    if x is None:
        return ""
    if isinstance(x, float) and x != x:  # NaN
        return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = re.sub(r"\.0+$", "", s)
    return s


def _col_name_is_stock_code(name: str) -> bool:
    n = (name or "").strip()
    if n in ("代码", "证券代码", "股票代码"):
        return True
    nl = n.lower()
    return nl in ("code", "symbol")


def _col_name_is_pick_date(name: str) -> bool:
    n = (name or "").strip().replace(" ", "")
    if n in ("选股日", "选股日期"):
        return True
    if "选股日" in n:
        return True
    return "选股" in n and "日期" in n


def _format_pick_date_cell(x: Any) -> str:
    """选股日列展示为短字符串。"""
    if x is None:
        return ""
    if isinstance(x, float) and x != x:
        return ""
    if hasattr(x, "strftime"):
        try:
            return x.strftime("%Y-%m-%d")
        except Exception:
            pass
    return _excel_cell_to_trimmed_str(x)


def _stock_list_header_looks_like_header(parts: List[str]) -> bool:
    cells = [(p or "").strip() for p in parts]
    joined = "|".join(cells)
    if "选股日" in joined or "选股日期" in joined:
        return True
    if "代码" in joined and ("名称" in joined or "选股" in joined):
        return True
    if not cells or not cells[0]:
        return False
    first = cells[0].split(",")[0].strip()
    if not any(ch.isdigit() for ch in first):
        return True
    code = _zfill_6(first)
    return not (code and _is_valid_stock_code_6(code))


def _read_stock_list_from_excel(filepath: str) -> List[Tuple[str, str]]:
    """xls/xlsx：代码列优先「代码」否则首列；可选「选股日」列（兼容「选股日期」）。返回 [(代码, 选股日), ...] 保留文件行顺序，同一代码多日期即多条。"""
    try:
        import pandas as pd
    except ImportError as e:
        raise RuntimeError("读取 Excel 股票列表需要 pandas：pip install pandas openpyxl") from e
    suf = os.path.splitext(filepath)[1].lower()
    try:
        if suf == ".xlsx":
            df = pd.read_excel(filepath, dtype=object, engine="openpyxl")
        elif suf == ".xls":
            try:
                df = pd.read_excel(filepath, dtype=object, engine="xlrd")
            except Exception as e:
                raise RuntimeError("读取 .xls 需安装 xlrd：pip install xlrd==1.2.0") from e
        else:
            raise ValueError(f"不支持的 Excel 扩展名: {suf}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"读取股票列表 Excel 失败: {filepath}; {type(e).__name__}: {e}") from e
    if df is None or getattr(df, "empty", True):
        raise RuntimeError(f"股票列表 Excel 为空: {filepath}")
    df = df.rename(columns={c: str(c).strip() for c in df.columns})
    cols = list(df.columns)
    code_col = None
    for c in cols:
        if _col_name_is_stock_code(str(c)):
            code_col = c
            break
    if code_col is None and cols:
        code_col = cols[0]
    date_col = next((c for c in cols if _col_name_is_pick_date(str(c))), None)

    out: List[Tuple[str, str]] = []
    for _, row in df.iterrows():
        raw = _excel_cell_to_trimmed_str(row.get(code_col)) if code_col is not None else ""
        if not raw:
            continue
        first = raw.split(",")[0].strip()
        if not first:
            continue
        if not any(ch.isdigit() for ch in first):
            continue
        code = _zfill_6(first)
        if not code or not _is_valid_stock_code_6(code):
            continue
        dtxt = ""
        if date_col is not None:
            dtxt = _format_pick_date_cell(row.get(date_col))
        out.append((code, dtxt))
    if not out:
        raise RuntimeError(f"股票列表 Excel 中未解析到有效代码: {filepath}")
    return out


def _read_stock_list_from_text_csv(filepath: str) -> List[Tuple[str, str]]:
    """txt/csv：代码列同上；表头含「选股日」时读该列（兼容「选股日期」），否则若每行多列则第二列作日期原文。保留行顺序，同一代码多日期即多条。"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
                except Exception:

                    class _D:
                        delimiter = ","

                    dialect = _D()
                reader = csv.reader(f, dialect)
                grid = [row for row in reader if any((c or "").strip() for c in row)]
            if not grid:
                last_err = RuntimeError("空文件")
                continue

            if _stock_list_header_looks_like_header(grid[0]):
                header = [(c or "").strip() for c in grid[0]]
                code_idx = next((i for i, h in enumerate(header) if _col_name_is_stock_code(h)), 0)
                date_idx = next((i for i, h in enumerate(header) if _col_name_is_pick_date(h)), None)
                data_rows = grid[1:]
            else:
                code_idx = 0
                date_idx = 1 if len(grid[0]) > 1 else None
                data_rows = grid

            out: List[Tuple[str, str]] = []
            for parts in data_rows:
                if code_idx >= len(parts):
                    continue
                raw = (parts[code_idx] or "").strip()
                if not raw:
                    continue
                first = raw.split(",")[0].strip()
                if not first:
                    continue
                if not any(ch.isdigit() for ch in first):
                    continue
                code = _zfill_6(first)
                if not code or not _is_valid_stock_code_6(code):
                    continue
                dtxt = ""
                if date_idx is not None and date_idx < len(parts):
                    dtxt = (parts[date_idx] or "").strip()
                out.append((code, dtxt))

            if not out:
                last_err = RuntimeError("未解析到有效代码")
                continue
            return out
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(
        f"读取股票列表失败: {filepath}; {type(last_err).__name__}: {last_err}" if last_err else f"读取股票列表失败: {filepath}"
    )


def _read_stock_list_from_file(filepath: str) -> List[Tuple[str, str]]:
    suf = os.path.splitext(filepath)[1].lower()
    if suf in (".xls", ".xlsx"):
        return _read_stock_list_from_excel(filepath)
    return _read_stock_list_from_text_csv(filepath)


def _ratio_to_chinese_desc(ratio: float) -> str:
    """将 0~1 的晋级比例转为中文：优先常见分数（如三分之一），否则「百分之整数」。"""
    try:
        r = float(ratio)
    except (TypeError, ValueError):
        return str(ratio)
    if r <= 1e-9:
        return "百分之零"
    if r >= 1.0 - 1e-9:
        return "百分之百"
    candidates = [
        (1.0 / 2, "百分之五十"),
        (1.0 / 3, "三分之一"),
        (2.0 / 3, "三分之二"),
        (1.0 / 4, "四分之一"),
        (3.0 / 4, "四分之三"),
        (1.0 / 5, "五分之一"),
        (2.0 / 5, "五分之二"),
        (3.0 / 5, "五分之三"),
        (4.0 / 5, "五分之四"),
        (1.0 / 6, "六分之一"),
        (5.0 / 6, "六分之五"),
        (1.0 / 8, "八分之一"),
    ]
    tol = 0.008
    for val, label in candidates:
        if abs(r - val) < tol:
            return label
    pct = int(round(r * 100))
    pct = max(0, min(100, pct))
    if pct == 100:
        return "百分之百"
    if pct == 0:
        return "百分之零"
    return f"百分之{pct}"


def _fmt_px(x) -> str:
    """表格中展示价格，空/无效则空串。"""
    if x is None:
        return ""
    try:
        return f"{float(x):.2f}"
    except (TypeError, ValueError):
        return str(x)


def _limit_ratio_for_board(code_6: str) -> float:
    """与统计线程内日K涨停判定一致的板块涨跌停比例（按代码前缀，不含 ST 名称）。"""
    s = _zfill_6(code_6)[:6]
    if s.startswith(("300", "301", "688", "689")):
        return 0.20
    if s.startswith(("8", "4", "920")):
        return 0.30
    return 0.10


def _load_all_a_stocks(csv_path: str) -> Dict[str, str]:
    """
    加载 data/all_a_stocks.csv：
    返回 code -> name
    只保留主A股 code 前缀：0/3/4/6/8/9
    """
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None

    for enc in encodings:
        try:
            with open(csv_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                code_col = "证券代码" if "证券代码" in (reader.fieldnames or []) else (reader.fieldnames or [None])[0]
                name_col = "证券简称" if "证券简称" in (reader.fieldnames or []) else None
                out: Dict[str, str] = {}
                for row in reader:
                    raw_code = row.get(code_col) if code_col else ""
                    code = _zfill_6(raw_code)
                    if not code:
                        continue
                    if not code.startswith(("0", "3", "4", "6", "8", "9")):
                        continue
                    name = ""
                    if name_col:
                        name = (row.get(name_col) or "").strip()
                    out[code] = name
                return out
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"加载 all_a_stocks.csv 失败: {csv_path}; {type(last_err).__name__}: {last_err}")


def _find_latest_limitup_file(history_dir: str, today_yyyymmdd: str) -> Tuple[str, str]:
    """
    找到 history_dir 下 <= today_yyyymmdd 的最新文件：
    返回 (filepath, yyyymmdd)
    """
    prefix = "涨停板数据_"
    valid_suffixes = (".csv", ".xlsx", ".xls")
    latest_date = ""
    latest_path = ""

    if not os.path.exists(history_dir):
        return "", ""

    for fname in os.listdir(history_dir):
        if not fname.startswith(prefix):
            continue
        suffix = next((s for s in valid_suffixes if fname.endswith(s)), "")
        if not suffix:
            continue
        date_part = fname[len(prefix) : -len(suffix)].strip()  # 2026-03-20
        date_norm = date_part.replace("-", "")
        if len(date_norm) != 8 or not date_norm.isdigit():
            continue
        if date_norm > today_yyyymmdd:
            continue
        if date_norm >= latest_date:
            latest_date = date_norm
            latest_path = os.path.join(history_dir, fname)

    return latest_path, latest_date


def _load_limitup_codes_from_excel(filepath: str) -> Dict[str, str]:
    """读取单日涨停板数据 xls/xlsx：返回 code -> name（列名优先「代码」「名称」）。"""
    try:
        import pandas as pd
    except ImportError as e:
        raise RuntimeError("读取 Excel 涨停数据需要 pandas：pip install pandas openpyxl") from e
    suf = os.path.splitext(filepath)[1].lower()
    try:
        if suf == ".xlsx":
            df = pd.read_excel(filepath, dtype=object, engine="openpyxl")
        elif suf == ".xls":
            try:
                df = pd.read_excel(filepath, dtype=object, engine="xlrd")
            except Exception as e:
                raise RuntimeError("读取 .xls 需安装 xlrd：pip install xlrd==1.2.0") from e
        else:
            raise ValueError(f"不支持的 Excel 扩展名: {suf}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"读取涨停板 Excel 失败: {filepath}; {type(e).__name__}: {e}") from e
    if df is None or getattr(df, "empty", True):
        return {}
    df = df.rename(columns={c: str(c).strip() for c in df.columns})
    cols = list(df.columns)
    code_col: Optional[str] = None
    name_col: Optional[str] = None
    if "代码" in df.columns:
        code_col = "代码"
    else:
        for c in cols:
            cs, cl = str(c), str(c).lower()
            if "代码" in cs or "code" in cl or cl in ("symbol", "股票代码", "证券代码"):
                code_col = c
                break
        if code_col is None and cols:
            code_col = cols[0]
    if "名称" in df.columns:
        name_col = "名称"
    else:
        for c in cols:
            if c == code_col:
                continue
            cs, cl = str(c), str(c).lower()
            if "名称" in cs or "name" in cl or cl in ("股票名称", "证券简称"):
                name_col = c
                break
    out: Dict[str, str] = {}
    for _, row in df.iterrows():
        code_raw = _excel_cell_to_trimmed_str(row.get(code_col)) if code_col else ""
        code = _zfill_6(code_raw)
        if not code or not _is_valid_stock_code_6(code):
            continue
        name = _excel_cell_to_trimmed_str(row.get(name_col)) if name_col else ""
        out[code] = name
    return out


def _load_limitup_codes(filepath: str) -> Dict[str, str]:
    """
    读取单日涨停板数据（csv / xls / xlsx）：
    返回 code -> name
    """
    suf = os.path.splitext(filepath)[1].lower()
    if suf in (".xls", ".xlsx"):
        return _load_limitup_codes_from_excel(filepath)

    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                code_col = "代码" if "代码" in (reader.fieldnames or []) else (reader.fieldnames or [None])[0]
                name_col = "名称" if "名称" in (reader.fieldnames or []) else None
                out: Dict[str, str] = {}
                for row in reader:
                    code = _zfill_6(row.get(code_col) or "")
                    if not code or not _is_valid_stock_code_6(code):
                        continue
                    name = (row.get(name_col) or "").strip() if name_col else ""
                    out[code] = name
                return out
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"读取涨停板数据失败: {filepath}; {type(last_err).__name__}: {last_err}")


def _load_limitup_codes_set(filepath: str) -> Set[str]:
    """只返回某日涨停股票 set（用于快速判断是否涨停）"""
    d = _load_limitup_codes(filepath)
    return set(d.keys())


def _extract_tick_preview_rows(raw_df, limit_n: Optional[int] = 200) -> Tuple[List[str], List[List[str]]]:
    """从 xtdata tick DataFrame 提取全部可用字段（含索引列）。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return [], []

    def _format_time_like(val) -> Optional[str]:
        if val is None:
            return None
        try:
            if isinstance(val, (int, float)):
                iv = int(val)
                # 毫秒时间戳
                if iv > 10_000_000_000:
                    return datetime.fromtimestamp(iv / 1000.0).strftime("%H:%M:%S")
                # 秒时间戳
                if iv > 1_000_000_000:
                    return datetime.fromtimestamp(iv).strftime("%H:%M:%S")
                # yyyymmddHHMMSS
                s = str(iv)
                if len(s) >= 14 and s[:8].isdigit():
                    dt = datetime.strptime(s[:14], "%Y%m%d%H%M%S")
                    return dt.strftime("%H:%M:%S")
            s = str(val).strip()
            if not s:
                return None
            if s.isdigit():
                iv = int(s)
                if len(s) >= 13:
                    return datetime.fromtimestamp(iv / 1000.0).strftime("%H:%M:%S")
                if len(s) == 10:
                    return datetime.fromtimestamp(iv).strftime("%H:%M:%S")
                if len(s) >= 14:
                    dt = datetime.strptime(s[:14], "%Y%m%d%H%M%S")
                    return dt.strftime("%H:%M:%S")
        except Exception:
            return None
        return None

    if limit_n is None:
        use_df = raw_df
    else:
        use_df = raw_df.tail(max(1, int(limit_n)))
    data_cols = [str(c) for c in list(getattr(use_df, "columns", []))]
    headers = ["tick_index"] + data_cols
    rows: List[List[str]] = []

    for idx, row in use_df.iterrows():
        row_vals: List[str] = [str(idx)]
        for c in data_cols:
            v = row.get(c, "")
            if v is None:
                row_vals.append("")
            else:
                cl = c.lower()
                if cl == "time" or cl.endswith("_time") or cl.endswith("timestamp"):
                    t_fmt = _format_time_like(v)
                    row_vals.append(t_fmt if t_fmt is not None else str(v))
                else:
                    row_vals.append(str(v))
        rows.append(row_vals)
    return headers, rows


def _tick_scalar_to_float(x) -> Optional[float]:
    """tick 单元格转 float；列表/数组取首元素。"""
    if x is None:
        return None
    if isinstance(x, (list, tuple)):
        if not x:
            return None
        return _tick_scalar_to_float(x[0])
    if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
        try:
            if len(x) > 0:
                return _tick_scalar_to_float(x[0])
        except Exception:
            pass
    try:
        return float(x)
    except Exception:
        return None


def _infer_tick_vol_to_shares_multiplier(raw_df) -> float:
    """与封单估算一致：根据 amount/volume/lastPrice 推断量为股或手。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return 100.0
    need_cols = {"amount", "volume", "lastPrice"}
    cols = set(str(c) for c in list(getattr(raw_df, "columns", [])))
    if not need_cols.issubset(cols):
        return 100.0
    try:
        probe = raw_df.tail(400)
        samples: List[float] = []
        prev_amt = None
        prev_vol = None
        for _, row in probe.iterrows():
            amt = _tick_scalar_to_float(row.get("amount"))
            vol = _tick_scalar_to_float(row.get("volume"))
            px = _tick_scalar_to_float(row.get("lastPrice"))
            if amt is None or vol is None or px is None or px <= 0:
                prev_amt, prev_vol = amt, vol
                continue
            if prev_amt is None or prev_vol is None:
                prev_amt, prev_vol = amt, vol
                continue
            d_amt = amt - prev_amt
            d_vol = vol - prev_vol
            prev_amt, prev_vol = amt, vol
            if d_amt <= 0 or d_vol <= 0:
                continue
            ratio = (d_amt / d_vol) / px
            if ratio > 0:
                samples.append(ratio)
        if not samples:
            return 100.0
        samples.sort()
        mid = samples[len(samples) // 2]
        return 1.0 if abs(mid - 1.0) <= abs(mid - 100.0) else 100.0
    except Exception:
        return 100.0


def _tick_time_value_to_seconds_from_midnight(v) -> Optional[float]:
    """tick 时间字段 -> 当日从 00:00 起的秒数（可小数）。"""
    if v is None:
        return None
    try:
        if isinstance(v, (int, float)):
            iv = int(v)
            if iv > 10_000_000_000:
                dt = datetime.fromtimestamp(iv / 1000.0)
                return (
                    dt.hour * 3600
                    + dt.minute * 60
                    + dt.second
                    + dt.microsecond / 1_000_000.0
                )
            if iv > 1_000_000_000:
                dt = datetime.fromtimestamp(float(iv))
                return dt.hour * 3600 + dt.minute * 60 + dt.second
            s = str(iv)
            if len(s) >= 14 and s[:8].isdigit():
                hh = int(s[8:10])
                mm = int(s[10:12])
                ss = int(s[12:14]) if len(s) >= 14 else 0
                if 0 <= hh <= 23 and 0 <= mm <= 59:
                    return hh * 3600 + mm * 60 + float(ss)
        s = str(v).strip()
        if not s:
            return None
        if s.isdigit():
            iv = int(s)
            if len(s) >= 13:
                dt = datetime.fromtimestamp(iv / 1000.0)
                return (
                    dt.hour * 3600
                    + dt.minute * 60
                    + dt.second
                    + dt.microsecond / 1_000_000.0
                )
            if len(s) == 10:
                dt = datetime.fromtimestamp(iv)
                return dt.hour * 3600 + dt.minute * 60 + dt.second
            if len(s) >= 14:
                hh = int(s[8:10])
                mm = int(s[10:12])
                ss = int(s[12:14]) if len(s) >= 14 else 0
                if 0 <= hh <= 23 and 0 <= mm <= 59:
                    return hh * 3600 + mm * 60 + float(ss)
    except Exception:
        return None
    return None


def _pressure_anchor_seconds_list() -> List[int]:
    """
    压力位锚点（当日 00:00 起秒数）：
    09:35 一次（开盘 5 分钟窗口）；09:47 起每 3 分钟至 11:30；13:00 起每 3 分钟至 15:00。
    每个锚点对应一行输出；窗口均为锚点前 300 秒。
    """
    out: List[int] = []
    out.append(9 * 3600 + 35 * 60)
    t = datetime(2000, 1, 1, 9, 47, 0)
    while t <= datetime(2000, 1, 1, 11, 30, 0):
        out.append(t.hour * 3600 + t.minute * 60 + t.second)
        t += timedelta(minutes=3)
    t = datetime(2000, 1, 1, 13, 0, 0)
    while t <= datetime(2000, 1, 1, 15, 0, 0):
        out.append(t.hour * 3600 + t.minute * 60 + t.second)
        t += timedelta(minutes=3)
    out.sort()
    dedup: List[int] = []
    for s in out:
        if not dedup or s != dedup[-1]:
            dedup.append(s)
    return dedup


def _row_ask_price_vol_pairs(row, vol_to_shares: float) -> List[Tuple[float, float]]:
    pairs: List[Tuple[float, float]] = []
    mul = float(vol_to_shares or 100.0)

    def add_px_vol(px: Optional[float], vol: Optional[float]) -> None:
        if px is None or vol is None or px <= 0 or vol < 0:
            return
        pairs.append((float(px), float(vol) * mul))

    ap = row.get("askPrice")
    av = row.get("askVol")
    if isinstance(ap, (list, tuple)) and isinstance(av, (list, tuple)):
        n = min(5, len(ap), len(av))
        for i in range(n):
            add_px_vol(_tick_scalar_to_float(ap[i]), _tick_scalar_to_float(av[i]))
        if pairs:
            return pairs
    for i in range(1, 6):
        px = None
        vol = None
        for pn in (f"askPrice{i}", f"ask{i}", f"sellPrice{i}", f"a{i}_p"):
            if pn in row:
                px = _tick_scalar_to_float(row.get(pn))
                break
        for vn in (f"askVol{i}", f"askVolume{i}", f"sellVol{i}", f"a{i}_v", f"ask_size{i}"):
            if vn in row:
                vol = _tick_scalar_to_float(row.get(vn))
                break
        add_px_vol(px, vol)
    return pairs


def _row_bid_price_vol_pairs(row, vol_to_shares: float) -> List[Tuple[float, float]]:
    """买一至买五（与卖档字段解析口径一致，量按 vol_mul 换为股）。"""
    pairs: List[Tuple[float, float]] = []
    mul = float(vol_to_shares or 100.0)

    def add_px_vol(px: Optional[float], vol: Optional[float]) -> None:
        if px is None or vol is None or px <= 0 or vol < 0:
            return
        pairs.append((float(px), float(vol) * mul))

    bp = row.get("bidPrice")
    bv = row.get("bidVol")
    if isinstance(bp, (list, tuple)) and isinstance(bv, (list, tuple)):
        n = min(5, len(bp), len(bv))
        for i in range(n):
            add_px_vol(_tick_scalar_to_float(bp[i]), _tick_scalar_to_float(bv[i]))
        if pairs:
            return pairs
    for i in range(1, 6):
        px = None
        vol = None
        for pn in (f"bidPrice{i}", f"bid{i}", f"buyPrice{i}", f"b{i}_p"):
            if pn in row:
                px = _tick_scalar_to_float(row.get(pn))
                break
        for vn in (f"bidVol{i}", f"bidVolume{i}", f"buyVol{i}", f"b{i}_v", f"bid_size{i}"):
            if vn in row:
                vol = _tick_scalar_to_float(row.get(vn))
                break
        add_px_vol(px, vol)
    return pairs


def _depth_vol_sum_bid_1_to_5(row, vol_mul: float) -> float:
    return sum(v for _, v in _row_bid_price_vol_pairs(row, vol_mul))


def _depth_vol_sum_ask_1_to_5(row, vol_mul: float) -> float:
    return sum(v for _, v in _row_ask_price_vol_pairs(row, vol_mul))


def _row_best_ask_price_vol(row, vol_mul: float) -> Tuple[Optional[float], Optional[float]]:
    """卖一：最优卖档（五档中最低价）及对应量（与列表顺序不一致时用最低价兜底）。"""
    pairs = _row_ask_price_vol_pairs(row, vol_mul)
    if not pairs:
        return None, None
    px, v = min(pairs, key=lambda x: x[0])
    return float(px), float(v)


def _tick_trade_price_primary(row: Any) -> Optional[float]:
    """本笔成交价：优先 lastPrice，其次常见别名。"""
    for k in ("lastPrice", "tradePrice", "matchPrice", "price", "last"):
        v = _tick_scalar_to_float(row.get(k))
        if v is not None and float(v) > 0:
            return float(v)
    return None


def _pick_tick_time_column(raw_df) -> Optional[str]:
    if raw_df is None or getattr(raw_df, "empty", True):
        return None
    cols = [str(c) for c in list(getattr(raw_df, "columns", []))]
    lower_map = {c.lower(): c for c in cols}
    for name in ("time", "Time", "timestamp", "tickTime", "quoteTime"):
        if name in cols:
            return name
        k = name.lower()
        if k in lower_map:
            return lower_map[k]
    for c in cols:
        cl = c.lower()
        if cl == "time" or cl.endswith("_time") or cl.endswith("timestamp"):
            return c
    return None


# 冲高回落 B 类：阈值与高点邻域（写死）
_PULLBACK_MIN_UP_RATIO = 0.003  # 拉升 ≥ 0.3%
_PULLBACK_MIN_DOWN_RATIO = 0.002  # 回落 ≥ 0.2%
_PULLBACK_PEAK_RADIUS = 5  # 高点在邻域内为明显峰值（tick 根数）

# C 类 压单驻留：写死
_DWELL_MIN_LOT_HANDS = 5  # 低于视为碎单，不计入「在档」
_DWELL_MIN_VALID_SEC = 10.0  # 最大驻留低于此秒的价位不参与 Top1
_DWELL_LONG_GAP_SEC = 30.0  # 相邻有效 tick 间隔超过此：视为断档，各价驻留清零且不累计该间隔

# 挂单 top3 主压力位 P：仅 A 类挂量参与加权；B/C 为同价情绪系数（不增加手数项）
_PRESSURE_P_W_A = 1.0
_PRESSURE_P_W_B = 0.3  # 冲高回落 Top2 任一高点价命中
_PRESSURE_P_W_C = 0.4  # 压单驻留 Top1 价位命中
# 实际沿用压力位：当日首锚点 P 为初值；之后仅当 |P_new-P_old| 超阈值才刷新 P_old
_PRESSURE_EFFECTIVE_UPDATE_PCT = 0.005  # 股票：相对当前沿用价的 0.5%
_PRESSURE_EFFECTIVE_UPDATE_ABS_ETF = 0.02  # ETF：固定 0.02 元

# 突破检测：开盘首个压力位锚点 9:35 结束之前，不检查是否突破
_FIRST_PRESSURE_BREAK_CHECK_SEC = 9 * 3600 + 35 * 60

# 真突破三条规则（数值判断用统一「股」口径；委卖/委买比为五档手数之比无量纲）
_TRUE_BREAKTHROUGH_MIN_BREAK_VS_AVG = 1.35
_TRUE_BREAKTHROUGH_MAX_ASK_OVER_BID = 1.8
_TRUE_BREAKTHROUGH_MIN_BREAK_VS_PREV_ASK1 = 0.7


def _tick_skip_c_class_sell(row: Any, vol_mul: float) -> bool:
    """无卖档或卖档总量为 0（常见于涨停等）时跳过本帧 C 类统计。"""
    pairs = _row_ask_price_vol_pairs(row, vol_mul)
    if not pairs:
        return True
    if sum(v for _, v in pairs) < 1e-9:
        return True
    return False


def _dwell_top1_from_timed_slice(
    timed_rows: List[Tuple[float, Any]], li: int, ri: int, vol_mul: float
) -> Tuple[str, Optional[float]]:
    """
    C 类：卖 1~5 内各卖价「连续出现在档上」的最大驻留时长（秒），取 Top1 价位。
    单 tick 该价合计量 < 5 手视为不在档；间隔过长视为断档全体清零；涨停无卖档帧跳过。
    返回 (展示文案, Top1 价位或 None)，供挂单 P 加权命中 C 时使用。
    """
    min_shares = float(_DWELL_MIN_LOT_HANDS) * float(vol_mul or 100.0)
    dwell: Dict[float, float] = {}
    best: Dict[float, float] = {}
    universe: Set[float] = set()
    prev_t: Optional[float] = None

    for j in range(li, ri):
        tsec, row = timed_rows[j][0], timed_rows[j][1]
        t = float(tsec)
        if _tick_skip_c_class_sell(row, vol_mul):
            # 不累计本帧；推进时间锚点，避免下一帧 gap 把涨停/无卖档间隔误算进驻留
            prev_t = t
            continue
        gap = (t - prev_t) if prev_t is not None else 0.0
        if prev_t is not None and gap > _DWELL_LONG_GAP_SEC:
            for pk in list(universe):
                if dwell.get(pk, 0.0) > best.get(pk, 0.0):
                    best[pk] = dwell[pk]
                dwell[pk] = 0.0
            gap = 0.0
        prev_t = t

        curr: Dict[float, float] = {}
        for px, sh in _row_ask_price_vol_pairs(row, vol_mul):
            pk = round(float(px), 4)
            if sh >= min_shares:
                curr[pk] = curr.get(pk, 0.0) + float(sh)

        check = universe | set(curr.keys())
        for pk in check:
            if pk in curr:
                dwell[pk] = dwell.get(pk, 0.0) + gap
                if dwell[pk] > best.get(pk, 0.0):
                    best[pk] = dwell[pk]
            else:
                if dwell.get(pk, 0.0) > best.get(pk, 0.0):
                    best[pk] = dwell[pk]
                dwell[pk] = 0.0
            universe.add(pk)

    for pk in list(universe):
        if dwell.get(pk, 0.0) > best.get(pk, 0.0):
            best[pk] = dwell[pk]

    cand = [(p, s) for p, s in best.items() if s >= _DWELL_MIN_VALID_SEC]
    if not cand:
        return "", None
    # 驻留并列全场最长时：优先取价格更高者
    cand.sort(key=lambda x: (-x[1], -x[0]))
    top_p, top_s = cand[0]
    return f"价={top_p:.4f} 驻留={top_s:.1f}秒", float(top_p)


def _pullback_top2_from_lastprices(prices: List[float]) -> List[Tuple[float, float]]:
    """
    从 5 分钟窗口 lastPrice 序列得到冲高回落 Top2：(高点 H, Score)，Score 降序，同价合并保留最大 Score。
    Score = (H-L)/L * (H-D)/H * 10000，L/H/D 为独立波段低点、高点、回落低点。
    """
    n = len(prices)
    if n < 5:
        return []
    pr = [float(p) for p in prices if p and p > 0]
    n = len(pr)
    if n < 5:
        return []
    R = _PULLBACK_PEAK_RADIUS
    MIN_UP = _PULLBACK_MIN_UP_RATIO
    MIN_DN = _PULLBACK_MIN_DOWN_RATIO

    # 严格局部高点 + 邻域内为峰值（明显峰顶）
    peak_idx: List[int] = []
    for i in range(1, n - 1):
        if pr[i] <= pr[i - 1] or pr[i] <= pr[i + 1]:
            continue
        lo = max(0, i - R)
        hi2 = min(n - 1, i + R)
        mx = max(pr[lo : hi2 + 1])
        if pr[i] + 1e-12 < mx:
            continue
        peak_idx.append(i)

    by_h_score: Dict[float, float] = {}
    prev_lo_idx = 0
    for hi in sorted(peak_idx):
        if hi < prev_lo_idx:
            continue
        H = pr[hi]
        if H <= 0:
            continue
        seg_l = pr[prev_lo_idx : hi + 1]
        if not seg_l:
            continue
        L = min(seg_l)
        sub = pr[hi:n]
        if not sub:
            prev_lo_idx = hi
            continue
        D = min(sub)
        d_idx = hi + int(sub.index(D))
        if L <= 0:
            prev_lo_idx = min(d_idx + 1, n - 1)
            continue
        up_r = (H - L) / L
        dn_r = (H - D) / H if H > 0 else 0.0
        if up_r < MIN_UP or dn_r < MIN_DN:
            prev_lo_idx = min(d_idx + 1, n - 1)
            continue
        score = up_r * dn_r * 10000.0
        hk = round(H, 4)
        prev = by_h_score.get(hk)
        if prev is None or score > prev:
            by_h_score[hk] = score
        prev_lo_idx = min(d_idx + 1, n - 1)

    ranked = sorted(by_h_score.items(), key=lambda kv: kv[1], reverse=True)
    out: List[Tuple[float, float]] = []
    for hk, sc in ranked[:2]:
        out.append((float(hk), float(sc)))
    return out


def _format_pullback_top2(pairs: List[Tuple[float, float]]) -> str:
    if not pairs:
        return ""
    parts: List[str] = []
    for rank, (hx, sc) in enumerate(pairs, start=1):
        parts.append(f"第{rank}价={hx:.4f} score={int(round(sc))}")
    return " | ".join(parts)


def _build_tick_timed_rows_sorted(raw_df) -> List[Tuple[float, Any]]:
    """交易时段内 tick，(当日秒, row)，按时间升序（与压力位计算口径一致）。"""
    out: List[Tuple[float, Any]] = []
    if raw_df is None or getattr(raw_df, "empty", True):
        return out
    time_col = _pick_tick_time_column(raw_df)
    if not time_col:
        return out
    try:
        for _, row in raw_df.iterrows():
            tsec = _tick_time_value_to_seconds_from_midnight(row.get(time_col))
            if tsec is None:
                continue
            if not (
                (9 * 3600 + 30 * 60) <= tsec <= (11 * 3600 + 30 * 60)
                or (13 * 3600) <= tsec <= (15 * 3600)
            ):
                continue
            out.append((float(tsec), row))
    except Exception:
        return []
    out.sort(key=lambda x: x[0])
    return out


def _seconds_from_midnight_to_hhmmss(tsec: float) -> str:
    s = int(round(float(tsec))) % 86400
    hh = s // 3600
    mm = (s % 3600) // 60
    ss = s % 60
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def _parse_price_text_to_float(s: str) -> Optional[float]:
    t = (s or "").strip()
    if not t:
        return None
    try:
        return float(t)
    except (TypeError, ValueError):
        return None


def _format_tick_volume_display(v: Optional[float]) -> str:
    if v is None:
        return ""
    x = float(v)
    if abs(x - round(x)) < 1e-6:
        return str(int(round(x)))
    return f"{x:.2f}"


def _breakthrough_vol_display_in_lots(v: Optional[float], vol_mul: float) -> str:
    """
    突破表中与五档对照的成交量展示：五档原始多为「手」，逐笔 lastVol 多为「股」。
    vol_mul 为累计量推断的手→股乘数（通常为 100）；将「股」除以 vol_mul 得到「手」，与五档手数同量级。
    vol_mul≈1 时表示数据源已为股口径，不再换算。
    """
    if v is None:
        return ""
    vm = float(vol_mul or 1.0)
    if vm > 1.01:
        return _format_tick_volume_display(float(v) / vm)
    return _format_tick_volume_display(float(v))


def _per_tick_trade_volumes(
    timed_rows: List[Tuple[float, Any]], vol_mul: float = 100.0
) -> List[Optional[float]]:
    """
    单笔成交量序列（统一为「股」口径，便于与压力位模块一致）：
    优先逐笔量字段（多为股，原样保留）；否则累计 volume 相邻差分为「手」时差分 * vol_mul → 股。
    """
    n = len(timed_rows)
    if n == 0:
        return []
    rows_list = [r for _, r in timed_rows]
    inc_names = (
        "lastVol",
        "tradeVol",
        "tradeVolume",
        "tickVol",
        "singleVol",
        "matchQty",
        "qty",
        "volume_delta",
    )
    best_name: Optional[str] = None
    best_cnt = 0
    for name in inc_names:
        cnt = sum(
            1
            for r in rows_list
            if _tick_scalar_to_float(r.get(name)) is not None
        )
        if cnt > best_cnt:
            best_cnt = cnt
            best_name = name

    if best_name is not None and best_cnt >= max(1, (n + 1) // 2):
        out: List[Optional[float]] = []
        for r in rows_list:
            v = _tick_scalar_to_float(r.get(best_name))
            if v is None or float(v) < 0:
                out.append(None)
            else:
                out.append(float(v))
        return out

    cum_names = ("volume", "cumVol", "totalVol", "cum_volume", "dealVol")
    vm = float(vol_mul or 100.0)
    out = []
    prev: Optional[float] = None
    for r in rows_list:
        cv: Optional[float] = None
        for cn in cum_names:
            cv = _tick_scalar_to_float(r.get(cn))
            if cv is not None:
                break
        if cv is None:
            out.append(None)
            continue
        fv = float(cv)
        if prev is None:
            out.append(max(0.0, fv) * vm)
            prev = fv
            continue
        d = fv - prev
        if d < -1e-3:
            out.append(max(0.0, fv) * vm)
            prev = fv
        else:
            out.append(max(0.0, d) * vm)
            prev = fv
    return out


def _true_breakthrough_remark_text(
    cond1: bool,
    cond2: bool,
    cond3: bool,
    v_break_sh: Optional[float],
    avg_shares: Optional[float],
    a_dep: float,
    b_dep: float,
    prev_a1_shares: Optional[float],
) -> str:
    """真突破为 False 时列出未满足条件及实测比值；全满足时为空字符串。"""
    if cond1 and cond2 and cond3:
        return ""
    parts: List[str] = []
    if not cond1:
        if (
            v_break_sh is not None
            and avg_shares is not None
            and avg_shares > 1e-12
        ):
            r = v_break_sh / avg_shares
            parts.append(
                f"①突破tick成交量为突破前均量的{r:.2f}倍（需≥{_TRUE_BREAKTHROUGH_MIN_BREAK_VS_AVG}）"
            )
        else:
            parts.append(
                "①缺少本笔成交量或突破前均量，无法满足≥突破前均量×"
                f"{_TRUE_BREAKTHROUGH_MIN_BREAK_VS_AVG}"
            )
    if not cond2:
        if b_dep <= 1e-9:
            parts.append(
                f"②委买量为0，无法计算委卖/委买（需≤{_TRUE_BREAKTHROUGH_MAX_ASK_OVER_BID}）"
            )
        else:
            r2 = a_dep / b_dep
            parts.append(
                f"②委卖/委买={r2:.2f}（需≤{_TRUE_BREAKTHROUGH_MAX_ASK_OVER_BID}）"
            )
    if not cond3:
        if (
            v_break_sh is not None
            and prev_a1_shares is not None
            and prev_a1_shares > 1e-12
        ):
            r3 = v_break_sh / prev_a1_shares
            parts.append(
                f"③突破tick成交量为前一tick卖一的{r3:.2f}倍（需>"
                f"{_TRUE_BREAKTHROUGH_MIN_BREAK_VS_PREV_ASK1}）"
            )
        else:
            parts.append(
                "③缺少本笔成交量或前一tick卖一量，无法满足>前一卖一×"
                f"{_TRUE_BREAKTHROUGH_MIN_BREAK_VS_PREV_ASK1}"
            )
    return "；".join(parts)


def _mark_buy_signal_first_true_per_code_buyday(rows: List[dict]) -> None:
    """
    每只股票、每个买入日、每个选股日：若存在真突破，仅将时间最早的一条真突破标为买入信号「是」，其余行置空。
    """
    if not rows:
        return
    by_key: Dict[Tuple[str, str, str], List[int]] = {}
    for i, r in enumerate(rows):
        c = (str(r.get("code") or "")).strip()
        bd = (str(r.get("buy_trade_day") or "")).strip()
        sel = (str(r.get("选股日") or "")).strip()
        k = (c, bd, sel)
        if k not in by_key:
            by_key[k] = []
        by_key[k].append(i)
    for _k, idxs in by_key.items():
        idxs.sort(key=lambda i: (rows[i].get("breakthrough_time") or "99:99:99"))
        placed = False
        for i in idxs:
            r = rows[i]
            if (not placed) and r.get("true_breakthrough") == "True":
                r["buy_signal"] = "是"
                placed = True
            else:
                r["buy_signal"] = ""


def _breakthrough_events_first_per_effective_regime(
    timed_rows: List[Tuple[float, Any]],
    stock_code: str,
    name: str,
    buy_trade_day_disp: str,
    pressure_rows: List[Tuple[str, str, str, str, str, str, int]],
    vol_mul: float = 100.0,
    selection_day_disp: str = "",
) -> List[Dict[str, str]]:
    """
    逐 tick：先按时间应用「锚点结束」后的实际压力位；9:35:00 之前不检突破。
    每个实际压力位数值 regime（自上次数值变化以来的沿用段）只记录第一次突破：
    一行 = 该 regime 下首次「展示精度舍入后的 lastPrice **严格大于** 舍入后的压力位」；
    触及同价（如 5.10 与 5.10）不算突破。突破时刻 + 当时压力位 + 该 tick 最新价；
    另附：截至突破前一 tick 的当日单笔成交量均值、突破 tick 单笔成交量（股；展示时与五档统一为「手」见下）；
    委买/委卖、前 tick 卖一量：五档原始为「手」（乘数 1 求和）；逐笔成交量多为「股」，展示为股/vol_mul 以与五档同量级。
    另：突破前一 tick 的卖一价/量；本 tick 成交价、本 tick 成交量（展示口径同上）。
    真突破：三条同时成立——①本 tick 成交量(股)≥突破前日均(股)×1.35；②委卖/委买(五档手数比)≤1.8；
    ③本 tick 成交量(股)>前一 tick 卖一量(股)×0.7。
    备注：未全真突破时列出未满足项及实测比值（全真突破备注为空）。
    """
    code = (stock_code or "").strip()
    nm = name or ""
    buyd = buy_trade_day_disp or ""
    sel_day = (selection_day_disp or "").strip()
    events: List[Dict[str, str]] = []
    if not timed_rows:
        return events

    per_vols = _per_tick_trade_volumes(timed_rows, vol_mul)
    prefix_sum: List[float] = []
    prefix_cnt: List[int] = []
    rs = 0.0
    rc = 0
    for i in range(len(timed_rows)):
        prefix_sum.append(rs)
        prefix_cnt.append(rc)
        v = per_vols[i] if i < len(per_vols) else None
        if v is not None:
            rs += float(v)
            rc += 1

    timeline: List[Tuple[int, Optional[float], str]] = []
    for row in pressure_rows or []:
        if len(row) < 7:
            continue
        _, _, _, _, _, weff_txt, asec = row
        a = int(asec)
        wf = _parse_price_text_to_float(str(weff_txt))
        timeline.append((a, wf, str(weff_txt or "").strip()))
    timeline.sort(key=lambda x: x[0])

    i_a = 0
    current_p: Optional[float] = None
    current_disp = ""
    regime_id = 0
    broken_regimes: Set[int] = set()

    for idx, (tsec, row) in enumerate(timed_rows):
        while i_a < len(timeline) and float(timeline[i_a][0]) <= float(tsec) + 1e-9:
            wf, wtx = timeline[i_a][1], timeline[i_a][2]
            if wf is not None:
                if current_p is None or abs(float(wf) - float(current_p)) > 1e-6:
                    regime_id += 1
                current_p = float(wf)
                current_disp = (
                    wtx if wtx else _format_pressure_p_by_stock_code(code, current_p)
                )
            i_a += 1

        if float(tsec) < float(_FIRST_PRESSURE_BREAK_CHECK_SEC):
            continue

        lp = _tick_scalar_to_float(row.get("lastPrice"))
        if current_p is None or lp is None or float(lp) <= 0:
            continue
        if regime_id <= 0 or regime_id in broken_regimes:
            continue
        r_lp = _round_price_like_display(code, float(lp))
        r_ref = _round_price_like_display(code, float(current_p))
        if r_lp <= r_ref:
            continue
        broken_regimes.add(regime_id)
        pc = prefix_cnt[idx] if idx < len(prefix_cnt) else 0
        ps = prefix_sum[idx] if idx < len(prefix_sum) else 0.0
        avg_before = (
            _breakthrough_vol_display_in_lots(ps / float(pc), vol_mul) if pc > 0 else ""
        )
        break_vol = ""
        if idx < len(per_vols) and per_vols[idx] is not None:
            break_vol = _breakthrough_vol_display_in_lots(per_vols[idx], vol_mul)
        b_dep = max(0.0, float(_depth_vol_sum_bid_1_to_5(row, 1.0)))
        a_dep = max(0.0, float(_depth_vol_sum_ask_1_to_5(row, 1.0)))
        bid_depth_txt = _format_tick_volume_display(b_dep)
        ask_depth_txt = _format_tick_volume_display(a_dep)

        prev_a1p_txt = ""
        prev_a1v_txt = ""
        if idx > 0:
            prow = timed_rows[idx - 1][1]
            apx, avo = _row_best_ask_price_vol(prow, 1.0)
            if apx is not None:
                prev_a1p_txt = _format_pressure_p_by_stock_code(code, apx)
            if avo is not None:
                prev_a1v_txt = _format_tick_volume_display(avo)

        trade_px = _tick_trade_price_primary(row)
        trade_px_txt = (
            _format_pressure_p_by_stock_code(code, trade_px)
            if trade_px is not None
            else ""
        )

        v_break_sh: Optional[float] = None
        if idx < len(per_vols) and per_vols[idx] is not None:
            v_break_sh = float(per_vols[idx])
        avg_shares: Optional[float] = (
            float(ps) / float(pc) if pc > 0 else None
        )
        prev_a1_shares: Optional[float] = None
        if idx > 0:
            _, pav_s = _row_best_ask_price_vol(timed_rows[idx - 1][1], vol_mul)
            if pav_s is not None:
                prev_a1_shares = float(pav_s)

        cond1 = (
            v_break_sh is not None
            and avg_shares is not None
            and v_break_sh >= avg_shares * float(_TRUE_BREAKTHROUGH_MIN_BREAK_VS_AVG)
        )
        cond2 = b_dep > 1e-9 and (a_dep / b_dep) <= float(
            _TRUE_BREAKTHROUGH_MAX_ASK_OVER_BID
        )
        cond3 = (
            v_break_sh is not None
            and prev_a1_shares is not None
            and v_break_sh > prev_a1_shares * float(_TRUE_BREAKTHROUGH_MIN_BREAK_VS_PREV_ASK1)
        )
        true_break = bool(cond1 and cond2 and cond3)
        remark_txt = _true_breakthrough_remark_text(
            cond1,
            cond2,
            cond3,
            v_break_sh,
            avg_shares,
            a_dep,
            b_dep,
            prev_a1_shares,
        )

        events.append(
            {
                "code": code,
                "name": nm,
                "buy_trade_day": buyd,
                "选股日": sel_day,
                "breakthrough_time": _seconds_from_midnight_to_hhmmss(float(tsec)),
                "pressure_effective_used": current_disp
                or _format_pressure_p_by_stock_code(code, float(current_p)),
                "last_price": _format_pressure_p_by_stock_code(code, float(lp)),
                "volume_avg_before_break": avg_before,
                "volume_break_tick": break_vol,
                "depth_bid_vol": bid_depth_txt,
                "depth_ask_vol": ask_depth_txt,
                "prev_tick_ask1_price": prev_a1p_txt,
                "prev_tick_ask1_vol": prev_a1v_txt,
                "break_tick_trade_price": trade_px_txt,
                "break_tick_trade_vol": break_vol,
                "true_breakthrough": "True" if true_break else "False",
                "true_breakthrough_remark": remark_txt,
            }
        )
    return events


def _compute_pressure_level_snapshot_rows(
    raw_df,
    stock_code: str = "",
    timed_rows: Optional[List[Tuple[float, Any]]] = None,
) -> List[Tuple[str, str, str, str, str, str, int]]:
    """
    每个锚点一条：…, 压力位 P_new, 实际使用的压力位 P_old, 锚点结束时刻(当日秒)。
    P 仅对 A 类 top3 挂量加权；B/C 为同价情绪系数。
    """
    out: List[Tuple[str, str, str, str, str, str, int]] = []
    if raw_df is None or getattr(raw_df, "empty", True):
        return out
    vol_mul = _infer_tick_vol_to_shares_multiplier(raw_df)
    anchors = _pressure_anchor_seconds_list()
    if timed_rows is None:
        timed_rows = _build_tick_timed_rows_sorted(raw_df)
    else:
        timed_rows = list(timed_rows)
    if not timed_rows:
        return out
    times_only = [t for t, _ in timed_rows]
    p_effective_float: Optional[float] = None  # 当日沿用的实际压力位（数值）
    for anchor in anchors:
        low = float(anchor - 300)
        hi = float(anchor)
        li = bisect.bisect_left(times_only, low)
        ri = bisect.bisect_right(times_only, hi)
        if li >= ri:
            continue
        window_prices: List[float] = []
        for j in range(li, ri):
            _, row = timed_rows[j]
            lp = _tick_scalar_to_float(row.get("lastPrice"))
            if lp is not None and lp > 0:
                window_prices.append(float(lp))
        stat: Dict[float, List[float]] = {}
        for j in range(li, ri):
            _, row = timed_rows[j]
            for px, sh in _row_ask_price_vol_pairs(row, vol_mul):
                pk = round(float(px), 4)
                rec = stat.get(pk)
                if rec is None:
                    stat[pk] = [1.0, sh]
                else:
                    rec[0] += 1.0
                    rec[1] += sh
        pb_pairs = _pullback_top2_from_lastprices(window_prices)
        pb_txt = _format_pullback_top2(pb_pairs)
        dwell_txt, dwell_c_px = _dwell_top1_from_timed_slice(timed_rows, li, ri, vol_mul)
        b_price_set = {round(float(h), 4) for h, _ in pb_pairs}

        parts: List[str] = []
        w_txt = ""
        p_new: Optional[float] = None
        if stat:
            ranked = sorted(stat.items(), key=lambda kv: kv[1][1], reverse=True)
            top3 = ranked[:3]
            w_num = 0.0
            w_den = 0.0
            for rank, (px, (cnt, tot)) in enumerate(top3, start=1):
                avg = tot / cnt if cnt > 0 else 0.0
                parts.append(
                    f"第{rank}价={px:.4f} 总挂量={tot:.0f} 次数={int(cnt)} 均量={avg:.1f}"
                )
                k = float(_PRESSURE_P_W_A)
                if px in b_price_set:
                    k += float(_PRESSURE_P_W_B)
                if dwell_c_px is not None and round(float(dwell_c_px), 4) == float(px):
                    k += float(_PRESSURE_P_W_C)
                tot_f = float(tot)
                w_num += float(px) * tot_f * k
                w_den += tot_f * k
            if w_den > 0:
                p_new = float(w_num / w_den)
                w_txt = _format_pressure_p_by_stock_code(stock_code, p_new)
        body = " | ".join(parts) if parts else ""
        if not body and not pb_txt and not dwell_txt:
            continue
        w_eff_txt = ""
        if p_new is not None:
            if p_effective_float is None:
                p_effective_float = p_new
            else:
                th = _pressure_effective_update_threshold(stock_code, p_effective_float)
                if abs(p_new - p_effective_float) > th:
                    p_effective_float = p_new
            w_eff_txt = _format_pressure_p_by_stock_code(stock_code, p_effective_float)
        elif p_effective_float is not None:
            w_eff_txt = _format_pressure_p_by_stock_code(stock_code, p_effective_float)
        hh = anchor // 3600
        mm = (anchor % 3600) // 60
        ss = anchor % 60
        t_label = f"{hh:02d}:{mm:02d}:{ss:02d}"
        win_note = "开盘5分钟" if anchor == 9 * 3600 + 35 * 60 else "前5分钟"
        time_cell = f"{t_label}({win_note})"
        out.append((time_cell, body, w_txt, pb_txt, dwell_txt, w_eff_txt, int(anchor)))
    return out


def _calc_close_order_amount_yi(raw_df) -> Optional[float]:
    """按收盘附近 tick 估算封单金额（亿）：优先买一金额，其次买一价*买一量。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        # 有些接口把五档字段返回成列表/元组，买一取第 1 个
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        # 兼容 numpy array 等
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    def _pick(row, names: List[str]):
        for n in names:
            if n in row:
                v = row.get(n)
                if v is not None:
                    return v
        return None

    # 常见 tick 字段别名：兼容不同券商/接口
    amount_names = [
        "bidAmount1", "bidAmt1", "buyAmount1", "bid_amount1", "b1_amt",
        "bidAmount", "buyAmount",
    ]
    price_names = [
        "bidPrice1", "bid1", "buyPrice1", "b1_p",
        "bidPrice", "buyPrice",
    ]
    vol_names = [
        "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1",
        "bidVol", "buyVol", "bidVolume",
    ]

    def _infer_bidvol_to_shares_multiplier(df) -> float:
        """
        根据 tick 的 amount/volume/lastPrice 自动判断 volume 口径：
        - 若 amount 增量 / volume 增量 ≈ price，则 volume 是“股”，乘数=1
        - 若 amount 增量 / volume 增量 ≈ price*100，则 volume 是“手”，乘数=100
        """
        if df is None or getattr(df, "empty", True):
            return 100.0
        need_cols = {"amount", "volume", "lastPrice"}
        cols = set(str(c) for c in list(getattr(df, "columns", [])))
        if not need_cols.issubset(cols):
            return 100.0
        try:
            probe = df.tail(400)
            samples: List[float] = []
            prev_amt = None
            prev_vol = None
            for _, row in probe.iterrows():
                amt = _to_float(row.get("amount"))
                vol = _to_float(row.get("volume"))
                px = _to_float(row.get("lastPrice"))
                if amt is None or vol is None or px is None or px <= 0:
                    prev_amt, prev_vol = amt, vol
                    continue
                if prev_amt is None or prev_vol is None:
                    prev_amt, prev_vol = amt, vol
                    continue
                d_amt = amt - prev_amt
                d_vol = vol - prev_vol
                prev_amt, prev_vol = amt, vol
                if d_amt <= 0 or d_vol <= 0:
                    continue
                ratio = (d_amt / d_vol) / px
                if ratio > 0:
                    samples.append(ratio)
            if not samples:
                return 100.0
            samples.sort()
            mid = samples[len(samples) // 2]
            # 离 1 更近 -> 股；离 100 更近 -> 手
            return 1.0 if abs(mid - 1.0) <= abs(mid - 100.0) else 100.0
        except Exception:
            return 100.0

    vol_to_shares = _infer_bidvol_to_shares_multiplier(raw_df)

    # 收盘最后一笔可能是集合竞价/空档，回溯最后 200 笔找最近可用值
    probe_df = raw_df.tail(200)
    try:
        iter_rows = list(probe_df.iterrows())
    except Exception:
        return None

    for _, row in reversed(iter_rows):
        # 1) 直接金额字段
        amt_f = _to_float(_pick(row, amount_names))
        if amt_f is not None and amt_f > 0:
            return round(amt_f / 1e8, 4)

        # 2) 买一价 * 买一量
        p_f = _to_float(_pick(row, price_names))
        v_f = _to_float(_pick(row, vol_names))
        if p_f is None or v_f is None or p_f <= 0 or v_f <= 0:
            continue

        # 动态按股/手口径换算为“股”
        shares = v_f * vol_to_shares
        amt_yuan = p_f * shares
        if amt_yuan > 0:
            return round(amt_yuan / 1e8, 4)

    return None


def _calc_day_turnover_yuan(raw_df) -> Optional[float]:
    """从当日 tick 估算当日成交额（元）。优先累计成交额字段，其次逐笔金额求和。"""
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    cols = [str(c) for c in list(getattr(raw_df, "columns", []))]
    # 候选：累计成交额（通常最后一笔就是全天累计）
    cum_cols = ["amount", "totalAmount", "turnover", "accAmount", "cumAmount", "成交额"]
    # 候选：逐笔成交额（需要求和）
    per_trade_cols = ["tradeAmount", "lastAmount", "成交金额", "单笔成交额"]

    def _pick_existing(candidates: List[str]) -> Optional[str]:
        lower_map = {c.lower(): c for c in cols}
        for name in candidates:
            if name in cols:
                return name
            k = name.lower()
            if k in lower_map:
                return lower_map[k]
        return None

    ccol = _pick_existing(cum_cols)
    if ccol:
        try:
            vals = raw_df[ccol]
            # 累计字段取最大值更稳，防止尾笔回落/空值
            mx = None
            for v in vals:
                fv = _to_float(v)
                if fv is None:
                    continue
                if mx is None or fv > mx:
                    mx = fv
            if mx is not None and mx > 0:
                return mx
        except Exception:
            pass

    pcol = _pick_existing(per_trade_cols)
    if pcol:
        try:
            s = 0.0
            ok = False
            for v in raw_df[pcol]:
                fv = _to_float(v)
                if fv is None:
                    continue
                s += fv
                ok = True
            if ok and s > 0:
                return s
        except Exception:
            pass

    return None


def _calc_float_market_cap_yuan(stock_code_6: str, raw_df) -> Optional[float]:
    """估算流通市值（元）= FloatVolume * 最新价。"""
    if xtdata is None:
        return None

    full_code = _to_full_stock_code(stock_code_6)
    try:
        info = xtdata.get_instrument_detail(full_code)
    except Exception:
        info = None
    if not isinstance(info, dict):
        return None

    def _to_float(x):
        try:
            return float(x)
        except Exception:
            return None

    float_volume = _to_float(info.get("FloatVolume"))
    if float_volume is None or float_volume <= 0:
        return None

    # 优先用当日 tick 最新价，缺失时回退前收
    px = None
    if raw_df is not None and not getattr(raw_df, "empty", True):
        try:
            probe = raw_df.tail(200)
            for _, row in reversed(list(probe.iterrows())):
                v = _to_float(row.get("lastPrice"))
                if v is not None and v > 0:
                    px = v
                    break
        except Exception:
            px = None
    if px is None:
        px = _to_float(info.get("PreClose"))
    if px is None or px <= 0:
        return None

    cap = float_volume * px
    return cap if cap > 0 else None


def _calc_order_volatility_coeff(raw_df) -> Optional[float]:
    """
    封单波动系数（全天 3 秒快照买一封单量序列）：
    coeff = std(bid1_vol_series) / mean(bid1_vol_series)
    """
    if raw_df is None or getattr(raw_df, "empty", True):
        return None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    # 兼容不同 tick 字段命名；若是五档数组，默认取第 1 档
    bidvol_fields = ["bidVol", "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1"]
    cols = set(str(c) for c in list(getattr(raw_df, "columns", [])))
    use_field = None
    for f in bidvol_fields:
        if f in cols:
            use_field = f
            break
    if use_field is None:
        return None

    vols: List[float] = []
    try:
        for v in raw_df[use_field]:
            fv = _to_float(v)
            if fv is None or fv <= 0:
                continue
            vols.append(fv)
    except Exception:
        return None

    n = len(vols)
    if n < 2:
        return None
    mean_v = sum(vols) / n
    if mean_v <= 0:
        return None
    var = sum((x - mean_v) ** 2 for x in vols) / n
    std = var ** 0.5
    return std / mean_v


def _classify_order_stability(vol_coeff: Optional[float]) -> Tuple[str, Optional[int]]:
    """
    根据封单波动系数划分封单稳定性（3档）并给分：
    - < 0.40  -> 封单极度稳定，10分
    - 0.40~0.80 -> 封单整体平稳，5分
    - > 0.80 -> 封单剧烈波动，0分
    """
    if vol_coeff is None:
        return "", None
    if vol_coeff < 0.40:
        return "封单极度稳定", 10
    if vol_coeff <= 0.80:
        return "封单整体平稳", 5
    return "封单剧烈波动", 0


def _score_seal_hardness(seal_ratio: Optional[float]) -> int:
    """
    封板硬度得分（0~50）：
    seal_ratio = 收盘封单金额 / 当日成交额（比值）
    """
    if seal_ratio is None or seal_ratio <= 0:
        return 0
    if seal_ratio < 0.03:
        return 0
    if seal_ratio <= 0.08:
        return 15
    if seal_ratio <= 0.15:
        return 35
    return 50


def _score_rush_intensity(rush_ratio: Optional[float]) -> int:
    """
    抢筹烈度得分（0~30）：
    rush_ratio = 收盘封单金额 / 流通市值（比值）
    """
    if rush_ratio is None or rush_ratio <= 0:
        return 0
    if rush_ratio < 0.002:
        return 0
    if rush_ratio <= 0.005:
        return 15
    return 30


def _score_close_order_amount(close_amt_yi: Optional[float]) -> int:
    """
    收盘封单绝对金额得分（0~20）：
    close_amt_yi 单位为亿。
    """
    if close_amt_yi is None or close_amt_yi <= 0:
        return 0
    if close_amt_yi < 0.3:  # 3000万
        return 0
    if close_amt_yi <= 1.0:  # 1亿
        return 10
    return 20


def _label_order_rating(total_score: int) -> str:
    """按综合得分映射封单评级。"""
    s = int(total_score)
    # 阈值微调：压缩中档混杂，增强高档/低档可区分性（便于后续回测验证）
    if s <= 18:
        return "🔴 虚封高危"
    if s <= 45:
        return "🟠 弱势封板"
    if s <= 68:
        return "🟡 中等封板"
    if s <= 92:
        return "🟢 强势封板"
    return "🔥 超强极致封板"


def _rating_score_from_label(label: str) -> int:
    m = {
        "🔴 虚封高危": 1,
        "🟠 弱势封板": 2,
        "🟡 中等封板": 3,
        "🟢 强势封板": 4,
        "🔥 超强极致封板": 5,
    }
    return int(m.get(label, 0))


def _confidence_tag(total_score: int, stability_score: int, trend_score: int) -> str:
    """
    置信度（用于快速筛选）：
    - 高：总分高且稳定+趋势均不弱
    - 中：总分中高
    - 低：其余
    """
    # 结合当前加权参数重标定阈值，避免高置信样本过少
    if total_score >= 78 and stability_score >= 5 and trend_score >= 5:
        return "高"
    if total_score >= 55:
        return "中"
    return "低"


#
# 评分主公式参数（用于与次日收盘表现相关性调优）
# 思路：提升“日内过程型”特征权重（稳定性/趋势），降低纯静态快照项的相对权重。
# 当前参数来源：tools/optimize_structure_weights.py
# 优化批次：20260417_194655
# 目标口径：次日收盘涨幅 > -1%
# 样本区间：--start 20260318
#
SCORE_WEIGHT_SEAL_HARDNESS = 0.40
SCORE_WEIGHT_RUSH_INTENSITY = 0.40
SCORE_WEIGHT_CLOSE_AMOUNT = 0.30
SCORE_WEIGHT_STABILITY = 1.50
SCORE_WEIGHT_TREND = 1.80
SCORE_BASE_BIAS = -6.0


def _calc_weighted_total_score(
    score_hard: int,
    score_rush: int,
    score_amt: int,
    score_stability: int,
    score_trend: int,
) -> int:
    """
    加权总分（用于评级）：
    - 保留原有各子项打分逻辑
    - 通过可调权重实现参数优化
    """
    raw = (
        SCORE_BASE_BIAS
        + SCORE_WEIGHT_SEAL_HARDNESS * float(score_hard or 0)
        + SCORE_WEIGHT_RUSH_INTENSITY * float(score_rush or 0)
        + SCORE_WEIGHT_CLOSE_AMOUNT * float(score_amt or 0)
        + SCORE_WEIGHT_STABILITY * float(score_stability or 0)
        + SCORE_WEIGHT_TREND * float(score_trend or 0)
    )
    return int(max(0, min(120, round(raw))))


def _rank01(values: List[Optional[float]]) -> List[Optional[float]]:
    """把数值列表转为 0~1 分位（None 保留）。"""
    valid = [(idx, float(v)) for idx, v in enumerate(values) if v is not None]
    out: List[Optional[float]] = [None] * len(values)
    if not valid:
        return out
    if len(valid) == 1:
        out[valid[0][0]] = 0.5
        return out
    sorted_valid = sorted(valid, key=lambda x: x[1])
    n = len(sorted_valid)
    for rk, (idx, _) in enumerate(sorted_valid):
        out[idx] = rk / (n - 1)
    return out


def _apply_cross_sectional_score_adjustment(rows: List[dict]) -> None:
    """
    在单日横截面上做轻量分位修正，缓解固定阈值在不同行情温度下失真。
    仅当样本>=8时启用，避免小样本过拟合。
    """
    use_adjust = len(rows) >= 8
    amt_rank = _rank01([r.get("_close_amt_yi") for r in rows]) if use_adjust else [None] * len(rows)
    hard_rank = _rank01([r.get("_seal_ratio") for r in rows]) if use_adjust else [None] * len(rows)
    rush_rank = _rank01([r.get("_rush_ratio") for r in rows]) if use_adjust else [None] * len(rows)
    for i, row in enumerate(rows):
        base = int(row.get("_base_total_score") or 0)
        st_score = int(row.get("_stability_score") or 0)
        tr_score = int(row.get("_trend_score") or 0)
        if amt_rank[i] is None or hard_rank[i] is None or rush_rank[i] is None:
            adj_total = base
        else:
            bonus = int(round((amt_rank[i] - 0.5) * 12))
            bonus += int(round((hard_rank[i] - 0.5) * 16))
            bonus += int(round((rush_rank[i] - 0.5) * 12))
            adj_total = base + bonus
        # 强化“日内过程型”特征权重：稳定性/趋势高时给正向加分，低时惩罚更明确
        if st_score >= 10:
            adj_total += 8
        elif st_score >= 5:
            adj_total += 3
        elif row.get("order_stability"):
            adj_total -= 5

        if tr_score >= 8:
            adj_total += 8
        elif tr_score >= 5:
            adj_total += 3
        elif row.get("order_trend"):
            adj_total -= 4
        adj_total = max(0, min(120, int(adj_total)))
        row["_total_score"] = adj_total
        row["order_rating"] = _label_order_rating(adj_total)
        row["rating_score"] = _rating_score_from_label(row["order_rating"])
        row["confidence_tag"] = _confidence_tag(
            adj_total,
            st_score,
            tr_score,
        )


def _calc_order_trend_and_score(raw_df) -> Tuple[str, Optional[int]]:
    """
    封单运行趋势（基于买一封单量时序）：
    - Fstart: 上午前30分钟均值（09:30-10:00）
    - Fmid:   全天中段均值（10:00-14:30）
    - Fend:   尾盘最后30分钟均值（14:30-15:00）
    并返回对应趋势分：10/7/5/3/0
    """
    if raw_df is None or getattr(raw_df, "empty", True):
        return "", None

    def _to_float(x):
        if x is None:
            return None
        if isinstance(x, (list, tuple)):
            if not x:
                return None
            return _to_float(x[0])
        if hasattr(x, "__len__") and hasattr(x, "__getitem__") and not isinstance(x, (str, bytes, dict)):
            try:
                if len(x) > 0:
                    return _to_float(x[0])
            except Exception:
                pass
        try:
            return float(x)
        except Exception:
            return None

    def _parse_hhmm_from_time_val(v) -> Optional[int]:
        try:
            s = str(int(float(v)))
        except Exception:
            s = str(v)
        digits = "".join(ch for ch in s if ch.isdigit())
        # 1) yyyymmddHHMMSS...
        if len(digits) >= 12:
            hh = int(digits[-6:-4])
            mm = int(digits[-4:-2])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return hh * 100 + mm
        # 2) epoch(ms/s)
        try:
            iv = int(float(v))
            if iv > 1_000_000_000_000:  # ms
                dt = datetime.fromtimestamp(iv / 1000.0)
            elif iv > 1_000_000_000:  # s
                dt = datetime.fromtimestamp(iv)
            else:
                return None
            return dt.hour * 100 + dt.minute
        except Exception:
            return None

    # 兼容不同字段命名；若是五档数组，取买一
    bidvol_fields = ["bidVol", "bidVol1", "bidVolume1", "buyVol1", "b1_v", "bid_size1"]
    cols = set(str(c) for c in list(getattr(raw_df, "columns", [])))
    use_field = None
    for f in bidvol_fields:
        if f in cols:
            use_field = f
            break
    if use_field is None:
        return "", None

    start_vals: List[float] = []
    mid_vals: List[float] = []
    end_vals: List[float] = []

    try:
        for idx, row in raw_df.iterrows():
            fv = _to_float(row.get(use_field))
            if fv is None or fv <= 0:
                continue
            tval = row.get("time") if "time" in row else idx
            hhmm = _parse_hhmm_from_time_val(tval)
            if hhmm is None:
                continue
            if 930 <= hhmm < 1000:
                start_vals.append(fv)
            elif 1000 <= hhmm < 1430:
                mid_vals.append(fv)
            elif 1430 <= hhmm <= 1500:
                end_vals.append(fv)
    except Exception:
        return "", None

    # 若交易时段切片为空，按序列三等分兜底
    if not start_vals or not mid_vals or not end_vals:
        all_vals: List[float] = []
        try:
            for v in raw_df[use_field]:
                fv = _to_float(v)
                if fv is not None and fv > 0:
                    all_vals.append(fv)
        except Exception:
            return "", None
        n = len(all_vals)
        if n < 6:
            return "", None
        a = max(1, n // 6)  # 约前1/6 ~ 30分钟窗口
        b = max(1, n // 3)
        start_vals = all_vals[:a]
        end_vals = all_vals[-a:]
        mid_vals = all_vals[a:n - a] if (n - 2 * a) > 0 else all_vals[b:2 * b]

    if not start_vals or not mid_vals or not end_vals:
        return "", None

    f_start = sum(start_vals) / len(start_vals)
    f_mid = sum(mid_vals) / len(mid_vals)
    f_end = sum(end_vals) / len(end_vals)
    if f_start <= 0 or f_mid <= 0:
        return "", None

    chg_mid = (f_mid - f_start) / f_start * 100.0
    chg_end = (f_end - f_mid) / f_mid * 100.0

    # 五档判定（按强特征优先）
    if chg_mid >= 5.0 and chg_end >= 5.0:
        return "趋势持续加强", 10
    if chg_mid <= -5.0 and chg_end <= -5.0:
        return "趋势持续减弱", 0
    if 0.0 <= chg_mid < 5.0 and 0.0 <= chg_end < 5.0:
        return "震荡小幅加强", 7
    if -5.0 < chg_mid <= 0.0 and -5.0 < chg_end <= 0.0:
        return "震荡小幅减弱", 3
    if -5.0 <= chg_mid <= 5.0 and -5.0 <= chg_end <= 5.0:
        return "封单整体平稳", 5

    # 跨档混合时兜底为平稳
    return "封单整体平稳", 5


def _vol_for_buy_amount(amt: float, price: float, min_order_amount: float) -> int:
    """与策略生成器默认策略一致的取整股（百元一手）。"""
    if price <= 0:
        return 0
    v = max(100, (int(amt / price / 100) * 100))
    if v * price < min_order_amount:
        return 0
    return int(v)


def _parse_yyyy_mm_dd_to_date(s: str) -> Optional[date_cls]:
    t = (s or "").strip()
    if len(t) >= 10 and t[4] == "-" and t[7] == "-":
        try:
            return date_cls.fromisoformat(t[:10])
        except ValueError:
            return None
    py = _parse_pick_day_to_yyyymmdd(t)
    if py and len(py) == 8:
        try:
            return date_cls(int(py[:4]), int(py[4:6]), int(py[6:8]))
        except ValueError:
            return None
    return None


def _try_backtest_window_from_selection_day():
    root = os.path.dirname(os.path.abspath(__file__))
    if root not in sys.path:
        sys.path.insert(0, root)
    try:
        from strategy_generator_app.trading_calendar import (  # type: ignore
            backtest_window_from_selection_day,
        )

        return backtest_window_from_selection_day
    except Exception:
        pass
    try:
        from trading_calendar import backtest_window_from_selection_day  # type: ignore

        return backtest_window_from_selection_day
    except Exception:
        return None


def _fetch_close_prices_on_trade_date(
    codes_6: List[str], as_of: date_cls
) -> Dict[str, float]:
    """取各股票在 as_of 日线收盘价，用于 last_mark_prices；失败则跳过该股。"""
    out: Dict[str, float] = {}
    if xtdata is None or not codes_6:
        return out
    day_s = as_of.strftime("%Y%m%d")
    try:
        xtdata.enable_hello = False  # type: ignore[attr-defined]
    except Exception:
        pass
    for code in codes_6:
        c6 = (code or "").strip()
        if len(c6) < 6:
            c6 = c6.zfill(6) if c6 else ""
        if len(c6) != 6:
            continue
        full = _to_full_stock_code(c6)
        try:
            xtdata.download_history_data(full, "1d", day_s, day_s)
            raw = xtdata.get_market_data_ex(
                [],
                [full],
                period="1d",
                start_time=day_s,
                end_time=day_s,
                count=-1,
            )
            if not isinstance(raw, dict) or full not in raw:
                continue
            df = raw[full]
            if df is None or getattr(df, "empty", True):
                continue
            if "close" in df.columns:
                v = float(df["close"].iloc[-1])
            else:
                continue
            if v > 0:
                out[c6] = round(v, 4)
        except Exception:
            continue
    return out


def _simulate_buys_for_rows(
    ordered_rows: List[dict],
    *,
    initial_cash: float,
    buy_amount_per_stock: float,
    min_order_amount: float,
    start_date_s: str,
    end_date_s: str,
) -> Tuple[float, Dict[str, Dict[str, Any]], List[dict]]:
    """
    按顺序模拟买入，返回 (期末现金, 持仓 {volume,cost}, 成交列表)。
    成交列表元素字段与策略生成器回测 trades + CSV 列兼容。
    """
    cash = float(initial_cash)
    positions: Dict[str, Dict[str, Any]] = {}
    trades: List[dict] = []

    def _sort_key(r: dict) -> Tuple[str, str, str]:
        return (
            str(r.get("buy_trade_day") or ""),
            str(r.get("breakthrough_time") or ""),
            str(r.get("code") or ""),
        )

    for r in sorted(ordered_rows, key=_sort_key):
        if str(r.get("buy_signal") or "").strip() != "是":
            continue
        code = (str(r.get("code") or "")).strip()
        if len(code) < 6:
            code = code.zfill(6) if code else ""
        if len(code) != 6:
            continue
        name = str(r.get("name") or "").strip()
        px = _parse_price_text_to_float(str(r.get("break_tick_trade_price") or ""))
        if px is None or float(px) <= 0:
            continue
        px = float(px)
        vol = _vol_for_buy_amount(
            float(buy_amount_per_stock), px, float(min_order_amount)
        )
        if vol <= 0:
            continue
        amt = round(px * vol, 2)
        if cash + 1e-6 < amt:
            continue
        cash = round(cash - amt, 2)
        pos = positions.setdefault(code, {"volume": 0, "cost": 0.0})
        ov = int(pos["volume"])
        oc = float(pos["cost"])
        nv = ov + vol
        pos["volume"] = nv
        pos["cost"] = round((ov * oc + px * vol) / nv, 2) if nv else 0.0

        sel_disp = (str(r.get("选股日") or "")).strip()
        trig = (
            f"[选股日 {sel_disp}] 智能突破买入 压力位={r.get('pressure_effective_used') or ''}"
        ).strip()
        trades.append(
            {
                "date": str(r.get("buy_trade_day") or ""),
                "time": str(r.get("breakthrough_time") or ""),
                "code": code,
                "stock_name": name,
                "选股日": sel_disp,
                "side": "buy",
                "price": round(px, 4),
                "volume": vol,
                "amount": amt,
                "position_after": int(positions[code]["volume"]),
                "trigger_info": trig,
                "start_date": start_date_s,
                "end_date": end_date_s,
            }
        )
    return cash, positions, trades


def build_strategy_generator_exports(
    breakthrough_rows: List[dict],
    *,
    initial_cash: float,
    buy_amount_per_stock: float,
    min_order_amount: float,
    hold_trading_days: int,
    start_next_trading_day: bool,
    source_strategy_id: str = "intelligentbuy_export",
    source_strategy_name: str = "智能突破买入",
) -> Dict[str, Any]:
    """
    构造与策略生成器「导出上次回测结果 / 导出批量回测 JSON」兼容的数据：
    - csv_rows: 写入 CSV 的二维行列表（含表头）
    - export_v1: version=1 根对象（合并全部买入为同一资金账户）
    - export_v2: version=2 kind=batch_backtest（按选股日分档，每档独立初始资金）
    """
    win_fn = _try_backtest_window_from_selection_day()
    hold_n = max(1, int(hold_trading_days))

    buy_rows = [
        r
        for r in (breakthrough_rows or [])
        if str(r.get("buy_signal") or "").strip() == "是"
    ]
    if not buy_rows:
        raise ValueError(
            "当前突破表中没有「买入信号=是」的记录；请确认已完成 tick 分析且存在真突破。"
        )

    by_sel: Dict[str, List[dict]] = {}
    for r in buy_rows:
        sel = (str(r.get("选股日") or "")).strip()
        by_sel.setdefault(sel, []).append(r)

    segment_payloads: List[dict] = []
    all_trades_flat: List[dict] = []

    exported_at = datetime.now().isoformat(timespec="seconds")

    for sel_disp, seg_rows in sorted(by_sel.items(), key=lambda x: x[0]):
        sel_date = _parse_yyyy_mm_dd_to_date(sel_disp)
        start_d: Optional[date_cls] = None
        end_d: Optional[date_cls] = None
        if win_fn is not None and sel_date is not None:
            start_d, end_d, _hint = win_fn(
                sel_date,
                start_next_trading_day=start_next_trading_day,
                hold_trading_days=hold_n,
            )
        if start_d is None or end_d is None:
            bd0 = _parse_yyyy_mm_dd_to_date(str(seg_rows[0].get("buy_trade_day") or ""))
            start_d = bd0 or date_cls.today()
            end_d = start_d

        start_s = start_d.isoformat()
        end_s = end_d.isoformat()
        fcash, pos_map, trades_seg = _simulate_buys_for_rows(
            seg_rows,
            initial_cash=float(initial_cash),
            buy_amount_per_stock=float(buy_amount_per_stock),
            min_order_amount=float(min_order_amount),
            start_date_s=start_s,
            end_date_s=end_s,
        )
        all_trades_flat.extend(trades_seg)

        codes_mark = [c for c, p in pos_map.items() if int((p or {}).get("volume") or 0) > 0]
        last_px = _fetch_close_prices_on_trade_date(codes_mark, end_d)
        for c in codes_mark:
            if c not in last_px:
                try:
                    last_px[c] = round(float(pos_map[c]["cost"]), 4)
                except Exception:
                    last_px[c] = 0.0

        payload = {
            "version": 1,
            "exported_at": exported_at,
            "source_strategy_id": source_strategy_id,
            "source_strategy_name": source_strategy_name,
            "backtest_range": {
                "start": start_s,
                "end": end_s,
                "last_equity_date": end_s,
            },
            "initial_cash_used": float(initial_cash),
            "skipped_initial_positions_not_in_pool": 0,
            "dual_segment": False,
            "seg2_strategy_name": None,
            "final_cash": float(fcash),
            "positions": {
                k: {
                    "volume": int((v or {}).get("volume") or 0),
                    "cost": float((v or {}).get("cost") or 0),
                }
                for k, v in pos_map.items()
                if int((v or {}).get("volume") or 0) > 0
            },
            "last_mark_prices": last_px,
            "batch_selection_date": (sel_date.isoformat() if sel_date else sel_disp),
            "batch_mode": True,
        }
        segment_payloads.append(payload)

    # 合并全部买入（同一初始资金）的 v1
    all_rows_sorted = sorted(buy_rows, key=lambda r: (
        str(r.get("buy_trade_day") or ""),
        str(r.get("breakthrough_time") or ""),
        str(r.get("code") or ""),
    ))
    min_start: Optional[date_cls] = None
    max_end: Optional[date_cls] = None
    for sel_disp in by_sel.keys():
        sel_date = _parse_yyyy_mm_dd_to_date(sel_disp)
        if win_fn is not None and sel_date is not None:
            sd, ed, _ = win_fn(
                sel_date,
                start_next_trading_day=start_next_trading_day,
                hold_trading_days=hold_n,
            )
            if sd is not None:
                min_start = sd if min_start is None else min(min_start, sd)
            if ed is not None:
                max_end = ed if max_end is None else max(max_end, ed)
    if min_start is None:
        min_start = _parse_yyyy_mm_dd_to_date(str(all_rows_sorted[0].get("buy_trade_day") or "")) or date_cls.today()
    if max_end is None:
        max_end = min_start

    fc_all, pos_all, trades_all = _simulate_buys_for_rows(
        buy_rows,
        initial_cash=float(initial_cash),
        buy_amount_per_stock=float(buy_amount_per_stock),
        min_order_amount=float(min_order_amount),
        start_date_s=min_start.isoformat(),
        end_date_s=max_end.isoformat(),
    )
    codes_all = [c for c, p in pos_all.items() if int((p or {}).get("volume") or 0) > 0]
    last_all = _fetch_close_prices_on_trade_date(codes_all, max_end)
    for c in codes_all:
        if c not in last_all:
            try:
                last_all[c] = round(float(pos_all[c]["cost"]), 4)
            except Exception:
                last_all[c] = 0.0

    export_v1 = {
        "version": 1,
        "exported_at": exported_at,
        "source_strategy_id": source_strategy_id,
        "source_strategy_name": source_strategy_name,
        "backtest_range": {
            "start": min_start.isoformat(),
            "end": max_end.isoformat(),
            "last_equity_date": max_end.isoformat(),
        },
        "initial_cash_used": float(initial_cash),
        "skipped_initial_positions_not_in_pool": 0,
        "dual_segment": False,
        "seg2_strategy_name": None,
        "final_cash": float(fc_all),
        "positions": {
            k: {
                "volume": int((v or {}).get("volume") or 0),
                "cost": float((v or {}).get("cost") or 0),
            }
            for k, v in pos_all.items()
            if int((v or {}).get("volume") or 0) > 0
        },
        "last_mark_prices": last_all,
    }

    batch_rule = (
        f"智能突破买入导出；{'T+1 起' if start_next_trading_day else 'T 当日起'}，"
        f"持有 {hold_n} 个交易日（与策略生成器批量回测口径一致）"
    )
    export_v2 = {
        "version": 2,
        "kind": "batch_backtest",
        "exported_at": exported_at,
        "source_strategy_id": source_strategy_id,
        "source_strategy_name": source_strategy_name,
        "batch_rule": batch_rule,
        "segments": [
            {k: v for k, v in seg.items() if not str(k).startswith("_")}
            for seg in segment_payloads
        ],
    }

    # CSV：与主窗口「成交明细」表头一致（UTF-8 BOM）
    csv_headers = [
        "日期",
        "时间",
        "代码",
        "股票名称",
        "选股日",
        "方向",
        "价格",
        "数量",
        "金额",
        "交易后持仓",
        "触发信息",
        "start_date",
        "end_date",
    ]
    csv_rows: List[List[Any]] = [csv_headers]
    for t in sorted(
        all_trades_flat,
        key=lambda x: (str(x.get("date") or ""), str(x.get("time") or ""), str(x.get("code") or "")),
    ):
        csv_rows.append(
            [
                t.get("date", ""),
                t.get("time", ""),
                t.get("code", ""),
                t.get("stock_name", ""),
                t.get("选股日", ""),
                "买入",
                t.get("price", ""),
                t.get("volume", ""),
                t.get("amount", ""),
                t.get("position_after", ""),
                t.get("trigger_info", ""),
                t.get("start_date", ""),
                t.get("end_date", ""),
            ]
        )

    return {
        "csv_rows": csv_rows,
        "export_v1": export_v1,
        "export_v2": export_v2,
        "trades": all_trades_flat,
    }


class TodayLimitUpFinderThread(QThread):
    # today_rows: (code, name, 选股日), first_board_rows, streak_rows, limitup_date(yyyymmdd), debug_msg, finish_kind
    # detail_by_code: { code -> [ {date, open, close, remark}, ... ] }  remark: 首板涨停 / N连板 / 跌停
    finished = pyqtSignal(list, list, list, str, str, str, object)
    error_occurred = pyqtSignal(str)
    debug_info = pyqtSignal(str)
    tick_preview = pyqtSignal(str, str, object)  # code, name, {"headers": [...], "rows": [...]}

    def __init__(
        self,
        *,
        history_dir: str,
        all_a_stocks_csv: str,
        source_mode: str,
        stock_list_filepath: Optional[str],
        analysis_date_yyyymmdd: Optional[str],
        limitup_csv_path: Optional[str],
        test_only: bool,
        test_limit_n: int,
        run_mode: str = "full",
        parent=None,
    ):
        super().__init__(parent)
        self.history_dir = history_dir
        self.all_a_stocks_csv = all_a_stocks_csv
        self.source_mode = source_mode
        self.stock_list_filepath = stock_list_filepath
        self.analysis_date_yyyymmdd = (analysis_date_yyyymmdd or "").strip()
        self.limitup_csv_path = (limitup_csv_path or "").strip()
        self.test_only = test_only
        self.test_limit_n = test_limit_n
        self.run_mode = run_mode  # "list_only" | "full"

    def run(self):
        try:
            query_date = self.analysis_date_yyyymmdd or datetime.now().strftime("%Y%m%d")
            if self.limitup_csv_path and os.path.isfile(self.limitup_csv_path):
                limitup_path = self.limitup_csv_path
                m = re.search(r"(20\d{6})", os.path.basename(limitup_path))
                limitup_date = m.group(1) if m else query_date
            else:
                limitup_path, limitup_date = _find_latest_limitup_file(self.history_dir, query_date)
                if not limitup_path:
                    self.error_occurred.emit(
                        f"未找到<={query_date}的涨停板数据文件（history_dir={self.history_dir}）"
                    )
                    return

            limitup_code_to_name = _load_limitup_codes(limitup_path)
            total_limitup = len(limitup_code_to_name)

            def _has_st_tag(name: str) -> bool:
                n = (name or "").strip()
                trans = str.maketrans({"Ｓ": "S", "Ｔ": "T", "＊": "*", "－": "-", "％": "%"})
                n2 = n.translate(trans).upper()
                return ("ST" in n2) or ("*ST" in n2) or n2.startswith("ST")

            code_entries: List[Tuple[str, str]] = []

            if self.source_mode == "today_limitup":
                # 全部今日涨停：以涨停板 CSV 为准（剔除 ST），不再与 all_a 求交
                matched_codes = list(limitup_code_to_name.keys())
                code_to_name: Dict[str, str] = dict(limitup_code_to_name)
                st_excluded = []
                for c in matched_codes:
                    nm1 = limitup_code_to_name.get(c, "")
                    nm2 = code_to_name.get(c, "")
                    if _has_st_tag(nm1) or _has_st_tag(nm2):
                        st_excluded.append(c)
                st_excluded_cnt = len(st_excluded)
                if st_excluded_cnt:
                    st_set = set(st_excluded)
                    matched_codes = [c for c in matched_codes if c not in st_set]
                code_entries = [(c, "") for c in matched_codes]
                debug_msg = (
                    f"limitup_file={os.path.basename(limitup_path)}; "
                    f"来源=全部今日涨停; 涨停CSV总数={total_limitup}; "
                    f"待分析={len(code_entries)}; ST已兜底剔除={st_excluded_cnt}"
                )
            else:
                # 指定列表文件：分析文件内代码，不要求当日涨停（同一代码多「选股日」则多行）
                if not self.stock_list_filepath:
                    self.error_occurred.emit("未选择股票列表文件")
                    return
                code_entries = _read_stock_list_from_file(self.stock_list_filepath)

                all_a_map: Dict[str, str] = {}
                if os.path.isfile(self.all_a_stocks_csv):
                    try:
                        all_a_map = _load_all_a_stocks(self.all_a_stocks_csv)
                    except Exception:
                        all_a_map = {}
                unique_codes = sorted({c for c, _ in code_entries})
                code_to_name = {
                    c: (all_a_map.get(c, "").strip() or limitup_code_to_name.get(c, "").strip())
                    for c in unique_codes
                }
                in_limitup_cnt = sum(1 for c, _ in code_entries if c in limitup_code_to_name)
                debug_msg = (
                    f"limitup_file={os.path.basename(limitup_path)}; "
                    f"来源=指定文件; "
                    f"待分析={len(code_entries)}; 其中在今日涨停CSV中={in_limitup_cnt}/{len(code_entries)}"
                )

            if self.test_only and self.test_limit_n > 0:
                code_entries = code_entries[: self.test_limit_n]

            n_before_valid = len(code_entries)
            code_entries = [(c, d) for c, d in code_entries if _is_valid_stock_code_6(c)]
            if not code_entries:
                self.error_occurred.emit("有效股票代码为空（已剔除无效占位如 000000）")
                return
            if len(code_entries) < n_before_valid:
                debug_msg += f"；剔除无效占位代码={n_before_valid - len(code_entries)}；有效待分析={len(code_entries)}"

            self.debug_info.emit(debug_msg)

            rows: List[Tuple[str, str, str]] = []
            for code, pick_d in code_entries:
                name = limitup_code_to_name.get(code, "") if code in limitup_code_to_name else ""
                final_name = name or code_to_name.get(code, "")
                rows.append((code, final_name, pick_d))

            rows.sort(key=lambda x: (x[0], x[2]))
            today_codes = sorted({c for c, _, _ in rows})

            if self.run_mode == "list_only":
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "list_only", {})
                return

            # 第2步：逐只拉取 tick，做涨停板结构预览（每只股票单独输出）
            if xtdata is None:
                self.error_occurred.emit("xtdata 不可用，无法拉取 tick 数据")
                return

            # 尽量减少 xtdata 欢迎信息
            try:
                xtdata.enable_hello = False
            except Exception:
                pass

            tick_preview_limit = 200
            last_tick_rows: List[dict] = []
            streak_metric_rows: List[dict] = []
            breakthrough_metric_rows: List[dict] = []
            for idx, (code, name, pick_d) in enumerate(rows, start=1):
                tick_headers: List[str] = []
                tick_rows: List[List[str]] = []
                tick_trade_date = limitup_date
                py = _parse_pick_day_to_yyyymmdd(pick_d)
                if py:
                    nd = _next_trading_day_after_yyyymmdd(py)
                    if nd:
                        tick_trade_date = nd
                try:
                    tick_df = _load_or_fetch_tick_df(code, tick_trade_date)
                    tick_headers, tick_rows = _extract_tick_preview_rows(tick_df, limit_n=tick_preview_limit)
                except Exception:
                    tick_headers, tick_rows = [], []
                    tick_df = None

                last_tick_rows = [{"tick_index": r[0]} for r in tick_rows]  # 占位，兼容 finished 旧签名
                buy_disp = _yyyymmdd_to_display(tick_trade_date)
                timed_sorted = _build_tick_timed_rows_sorted(tick_df)
                pressure_rows = _compute_pressure_level_snapshot_rows(
                    tick_df, code, timed_sorted
                )
                vol_mul_bt = _infer_tick_vol_to_shares_multiplier(tick_df)
                py_pick = _parse_pick_day_to_yyyymmdd(pick_d)
                sel_disp = (
                    _yyyymmdd_to_display(py_pick)
                    if py_pick
                    else (str(pick_d).strip() if pick_d else "")
                )
                breakthrough_metric_rows.extend(
                    _breakthrough_events_first_per_effective_regime(
                        timed_sorted,
                        code,
                        name or "",
                        buy_disp,
                        pressure_rows,
                        vol_mul_bt,
                        selection_day_disp=sel_disp,
                    )
                )
                if not pressure_rows:
                    streak_metric_rows.append(
                        {
                            "code": code,
                            "name": name or "",
                            "buy_trade_day": buy_disp,
                            "pressure_time": "",
                            "pressure_detail": "",
                            "pressure_weighted_avg": "",
                            "pressure_effective_used": "",
                            "pullback_top2": "",
                            "dwell_top1": "",
                        }
                    )
                else:
                    for tlab, body, wavg, pbtxt, dwell_txt, weff, _anch in pressure_rows:
                        streak_metric_rows.append(
                            {
                                "code": code,
                                "name": name or "",
                                "buy_trade_day": buy_disp,
                                "pressure_time": tlab,
                                "pressure_detail": body,
                                "pressure_weighted_avg": wavg,
                                "pressure_effective_used": weff,
                                "pullback_top2": pbtxt,
                                "dwell_top1": dwell_txt,
                            }
                        )
                self.debug_info.emit(
                    f"{debug_msg}；tick进度={idx}/{len(rows)}；{code} {name or ''}；"
                    f"tick日={tick_trade_date}；选股日={pick_d or '-'}；tick条数={len(tick_rows)}；字段数={len(tick_headers)}"
                )
                self.tick_preview.emit(
                    code,
                    name or "",
                    {
                        "headers": tick_headers,
                        "rows": tick_rows,
                        "tick_trade_date": tick_trade_date,
                    },
                )

            # tick_only：第三表每锚点一行；突破详情为「每个实际压力位 regime 首次突破」可多行，经 detail 传递
            _mark_buy_signal_first_true_per_code_buyday(breakthrough_metric_rows)
            self.finished.emit(
                rows,
                last_tick_rows,
                streak_metric_rows,
                limitup_date,
                debug_msg,
                "tick_only",
                {"breakthrough_rows": breakthrough_metric_rows},
            )
            return

            # 1) 构造交易日历（用于“上一交易日是否涨停/次一交易日收益”口径）
            analysis_today = limitup_date
            today_year = int(analysis_today[:4])
            # 近一年（日历年）：上一自然年（从 YYYY-01-01 到本年的昨日）
            year_start_calendar = f"{today_year - 1}0101"

            from datetime import timedelta as _td

            year_start_dt = datetime.strptime(year_start_calendar, "%Y%m%d")
            # 往前多取一段交易日，保证 year_start 的前一交易日也在范围内
            # 用足够长的回溯期，避免跨年连板/首板在统计起点被截断
            trading_start_dt = year_start_dt - _td(days=400)
            trading_start = trading_start_dt.strftime("%Y%m%d")
            trading_end = analysis_today

            trading_dates_ts = xtdata.get_trading_dates("SH", start_time=trading_start, end_time=trading_end)
            trade_dates: List[str] = []
            for ts in trading_dates_ts or []:
                if isinstance(ts, (int, float)):
                    dt = datetime.fromtimestamp(float(ts) / 1000.0)
                    trade_dates.append(dt.strftime("%Y%m%d"))
                elif isinstance(ts, str):
                    trade_dates.append(ts.replace("-", "")[:8])
                else:
                    trade_dates.append(str(ts)[:8])
            trade_dates = sorted({d for d in trade_dates if d and len(d) == 8})

            if analysis_today not in trade_dates:
                # 容错：用 <= analysis_today 的最后一个交易日
                trade_dates = [d for d in trade_dates if d <= analysis_today]
                if not trade_dates:
                    self.error_occurred.emit("无法构造交易日历（trade_dates 为空）")
                    return
                analysis_today = trade_dates[-1]

            idx_today = trade_dates.index(analysis_today)
            if idx_today <= 0:
                self.error_occurred.emit("交易日历不足以排除今天与计算首板口径")
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "full", {})
                return

            # 统计结果的结束日：昨日（排除今天）
            end_idx = idx_today - 1
            end_date = trade_dates[end_idx]

            # year_start 的第一个交易日：>= year_start_calendar 的最早交易日
            year_start_idx = None
            for i in range(0, len(trade_dates)):
                if trade_dates[i] >= year_start_calendar:
                    year_start_idx = i
                    break
            if year_start_idx is None:
                self.error_occurred.emit("无法定位 year_start 在交易日历中的位置")
                self.finished.emit(rows, [], [], limitup_date, debug_msg, "full", {})
                return

            # 统计时仍按 d < year_start_td 或 d > end_date 过滤；
            # consec 连续性从起点开始计算，避免跨年时低估连板长度。
            start_needed_idx = 0
            year_start_td = trade_dates[year_start_idx]
            debug_msg += f"；统计区间={year_start_td}~{end_date}"

            # 2) 用 QMT 日K 自己判断涨停（不再依赖 history_data 的涨停CSV）
            #    注意：次日收益统计会用到“今天”的开盘/收盘，所以 end 需要包含 today。
            ohlc_codes_needed = sorted(set(today_codes))

            # 为了计算起始日的涨停，需要用到上一交易日的收盘价，因此向前多取 1 个交易日
            ohlc_start_idx = max(0, start_needed_idx - 1)
            ohlc_start_date = trade_dates[ohlc_start_idx]
            ohlc_end_date = analysis_today  # 包含 today

            startdate_full = datetime.strptime(ohlc_start_date, "%Y%m%d").strftime("%Y%m%d") + "000000"
            enddate_full = datetime.strptime(ohlc_end_date, "%Y%m%d").strftime("%Y%m%d") + "235959"

            ohlc_map_by_code: Dict[str, Dict[str, Tuple[float, float]]] = {}
            for idx, code in enumerate(ohlc_codes_needed):
                if idx % 10 == 0:
                    pass
                full_code = _to_full_stock_code(code)
                try:
                    xtdata.download_history_data(full_code, "1d", startdate_full, enddate_full)
                    df = xtdata.get_market_data_ex(
                        [],
                        [full_code],
                        period="1d",
                        start_time=startdate_full,
                        end_time=enddate_full,
                        count=-1,
                    )
                except Exception:
                    continue

                if df is None or full_code not in df or df[full_code] is None or df[full_code].empty:
                    continue
                try:
                    ohlc_map_by_code[code] = _extract_open_close_map(df[full_code])
                except Exception:
                    ohlc_map_by_code[code] = {}

            # 2.1) 构造涨停状态：date -> set(codes)
            def _limit_ratio_for_code(code_6: str) -> float:
                # ST 已通过“兜底剔除”从今日集合中移除，这里按板块规则定涨停比例即可
                if code_6.startswith(("300", "301", "688", "689")):
                    return 0.20
                if code_6.startswith(("8", "4", "920")):
                    return 0.30
                return 0.10

            codes_by_date: Dict[str, Set[str]] = {d: set() for d in trade_dates}
            for code in ohlc_codes_needed:
                m = ohlc_map_by_code.get(code) or {}
                limit_ratio = _limit_ratio_for_code(code)
                for i in range(1, idx_today + 1):
                    d = trade_dates[i]
                    prev_d = trade_dates[i - 1]
                    prev_close = (m.get(prev_d) or (None, None, None, None))[3]
                    close_price = (m.get(d) or (None, None, None, None))[3]
                    if prev_close is None or close_price is None or prev_close == 0:
                        continue

                    limit_up_price = round(prev_close * (1 + limit_ratio), 2)
                    price_diff = abs(close_price - limit_up_price)
                    inc_ratio = close_price / prev_close - 1
                    # 兼容两种判定：接近涨停价（容差） 或 涨幅达到预期（99%）
                    is_limit_up = (price_diff <= 0.02) or (inc_ratio >= limit_ratio * 0.99)
                    if is_limit_up:
                        codes_by_date.setdefault(d, set()).add(code)

            # 3) 计算今天的首板/连板归类（基于“上一交易日是否涨停”）
            today_streak_map: Dict[str, int] = {}
            first_today_codes: List[str] = []
            streak_today_codes: List[str] = []
            for code in today_codes:
                consec = 0
                for j in range(idx_today, -1, -1):
                    d = trade_dates[j]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        break
                # consec==0 说明 QMT 未能识别“今天涨停”（容差/数据缺失等）
                # 这种情况下不纳入首板/连板统计，避免用错误首板口径。
                today_streak_map[code] = consec
                if today_streak_map[code] == 1:
                    first_today_codes.append(code)
                elif today_streak_map[code] >= 2:
                    streak_today_codes.append(code)

            debug_msg += f"；QMT识别今日涨停={len(first_today_codes) + len(streak_today_codes)}/{len(today_codes)}"

            # 4) 首板统计表：含「今日首板」+「今日未涨停（连板数记 0）」
            code_name_map_today = {code: name for code, name, _ in rows}

            def _build_first_board_row(code: str, today_streak_val: int) -> dict:
                consec = 0
                first_board_count = 0
                next_day_limitup_count = 0
                next_day_up_count = 0
                next_day_down_count = 0

                open_ret_sum = 0.0
                close_ret_sum = 0.0
                open_valid_cnt = 0
                close_valid_cnt = 0

                ohlc_map = ohlc_map_by_code.get(code) or {}
                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        consec = 0

                    if d < year_start_td or d > end_date:
                        continue
                    if consec != 1:
                        continue

                    first_board_count += 1

                    if i + 1 <= end_idx:
                        nd = trade_dates[i + 1]
                        if code in codes_by_date.get(nd, set()):
                            next_day_limitup_count += 1

                        close_d = ohlc_map.get(d, (None, None, None, None))[3]
                        close_nd = ohlc_map.get(nd, (None, None, None, None))[3]
                        open_nd = ohlc_map.get(nd, (None, None, None, None))[0]

                        if close_d is not None and close_nd is not None and close_d != 0:
                            if close_nd > close_d:
                                next_day_up_count += 1
                            elif close_nd < close_d:
                                next_day_down_count += 1

                            close_ret_sum += (close_nd - close_d) / close_d
                            close_valid_cnt += 1

                            if open_nd is not None:
                                open_ret_sum += (open_nd - close_d) / close_d
                                open_valid_cnt += 1

                avg_open_ret_pct = (open_ret_sum / open_valid_cnt * 100.0) if open_valid_cnt > 0 else 0.0
                avg_close_ret_pct = (close_ret_sum / close_valid_cnt * 100.0) if close_valid_cnt > 0 else 0.0

                return {
                    "code": code,
                    "name": code_name_map_today.get(code, ""),
                    "today_streak": today_streak_val,
                    "first_board_count_1y": first_board_count,
                    "next_day_limitup_count": next_day_limitup_count,
                    "next_day_up_count": next_day_up_count,
                    "next_day_down_count": next_day_down_count,
                    "avg_open_return_pct": round(avg_open_ret_pct, 2),
                    "avg_close_return_pct": round(avg_close_ret_pct, 2),
                }

            first_board_rows: List[dict] = []
            for code in first_today_codes:
                first_board_rows.append(_build_first_board_row(code, 1))
            non_limitup_today_codes = [c for c in today_codes if today_streak_map.get(c, 0) == 0]
            for code in non_limitup_today_codes:
                first_board_rows.append(_build_first_board_row(code, 0))

            # 6) 统计“今日连板”对应股票的近一年 k连板次数与晋级比例
            #    - k连板次数：任意一段连续涨停，仅在「该段最后一个涨停日」记一次，长度为 N 则
            #      只计 N连板 1 次（不再把同一段里的 2~(N-1) 连板逐日各计一次）。
            #    - 连板晋级比例：仍用逐日 consec≥2 的口径，并排除「以昨日为结束日的当前这段连板」
            #      在统计区间内的逐日样本（与改口径前的晋级算法一致）；若改用分段次数去算
            #      k 进 k+1，会变成互斥集合之比，语义不再成立。
            streak_rows: List[dict] = []
            for code in streak_today_codes:
                consec = 0
                counts_by_k: Dict[int, int] = {}
                max_k = 0

                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec += 1
                    else:
                        if consec >= 2:
                            last_idx = i - 1
                            last_d = trade_dates[last_idx]
                            if year_start_td <= last_d <= end_date:
                                counts_by_k[consec] = counts_by_k.get(consec, 0) + 1
                                if consec > max_k:
                                    max_k = consec
                        consec = 0

                if consec >= 2:
                    last_d = trade_dates[end_idx]
                    if year_start_td <= last_d <= end_date:
                        counts_by_k[consec] = counts_by_k.get(consec, 0) + 1
                        if consec > max_k:
                            max_k = consec

                if counts_by_k:
                    counts_str = "\n".join(
                        f"{k}连板{counts_by_k.get(k, 0)}次"
                        for k in sorted(counts_by_k.keys())
                    )
                else:
                    counts_str = ""

                j = end_idx
                while j >= 0 and code in codes_by_date.get(trade_dates[j], set()):
                    j -= 1
                exclude_streak_begin = j + 1

                consec_u = 0
                counts_upgrade: Dict[int, int] = {}
                max_k_u = 0
                for i in range(start_needed_idx, end_idx + 1):
                    d = trade_dates[i]
                    if code in codes_by_date.get(d, set()):
                        consec_u += 1
                    else:
                        consec_u = 0
                    if d < year_start_td or d > end_date:
                        continue
                    if consec_u >= 2:
                        if exclude_streak_begin <= i <= end_idx:
                            continue
                        counts_upgrade[consec_u] = counts_upgrade.get(consec_u, 0) + 1
                        if consec_u > max_k_u:
                            max_k_u = consec_u

                upgrade_parts: List[str] = []
                for k in range(2, max_k_u):
                    denom = counts_upgrade.get(k, 0)
                    num = counts_upgrade.get(k + 1, 0)
                    if denom > 0:
                        ratio = num / denom
                        desc = _ratio_to_chinese_desc(ratio)
                        upgrade_parts.append(f"{k}进{k + 1}：{desc}")

                upgrade_str = "\n".join(upgrade_parts)

                streak_rows.append(
                    {
                        "code": code,
                        "name": code_name_map_today.get(code, ""),
                        "today_streak": today_streak_map.get(code, 2),
                        "max_streak_in_year": max_k,
                        "counts_by_k": counts_str,
                        "upgrade_ratios": upgrade_str,
                    }
                )

            streak_rows.sort(
                key=lambda r: (
                    -int(r.get("today_streak") or 0),
                    str(r.get("code") or ""),
                )
            )

            # 7) 每只股票：按交易日列出「日期、开盘、收盘、备注」
            #    - 涨停日：首板涨停 / N连板（连板从 ohlc 起点连续累计）
            #    - 涨停结束后的第一个非涨停日：首板后次日 / N连板后次日（便于看板后盈亏）；若该日跌停则只列一行「跌停」
            #    - 统计区间末日若仍涨停，补列再下一交易日（可为涨停 CSV 当日），规则同上
            detail_by_code: Dict[str, List[dict]] = {c: [] for c in today_codes}
            for code in today_codes:
                m = ohlc_map_by_code.get(code) or {}
                limit_ratio = _limit_ratio_for_board(code)
                streak = 0
                for i in range(1, end_idx + 1):
                    d = trade_dates[i]
                    prev_d = trade_dates[i - 1]
                    prev_close = (m.get(prev_d) or (None, None, None, None))[3]
                    open_d, high_d, low_d, close_d = m.get(d, (None, None, None, None))
                    is_up = code in codes_by_date.get(d, set())

                    if prev_close is None or (isinstance(prev_close, (int, float)) and float(prev_close) <= 0):
                        streak = 0
                        continue
                    if close_d is None:
                        if not is_up:
                            streak = 0
                        continue

                    try:
                        pc = float(prev_close)
                        cl = float(close_d)
                    except (TypeError, ValueError):
                        streak = 0 if not is_up else streak
                        continue

                    limit_dn = round(pc * (1.0 - limit_ratio), 2)
                    price_diff_dn = abs(cl - limit_dn)
                    dec_ratio = 1.0 - cl / pc
                    is_dn = (price_diff_dn <= 0.02) or (dec_ratio >= limit_ratio * 0.99 - 1e-12)

                    in_win = year_start_td <= d <= end_date

                    if is_up:
                        streak += 1
                        if in_win:
                            remark = "首板涨停" if streak == 1 else f"{streak}连板"
                            detail_by_code[code].append(
                                {
                                    "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                    "open": open_d,
                                    "high": high_d,
                                    "low": low_d,
                                    "close": close_d,
                                    "remark": remark,
                                }
                            )
                    else:
                        ended_streak = streak
                        streak = 0
                        if in_win:
                            if is_dn:
                                detail_by_code[code].append(
                                    {
                                        "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                        "open": open_d,
                                        "high": high_d,
                                        "low": low_d,
                                        "close": close_d,
                                        "remark": "跌停",
                                    }
                                )
                            elif ended_streak > 0:
                                post_rm = (
                                    "首板后次日"
                                    if ended_streak == 1
                                    else f"{ended_streak}连板后次日"
                                )
                                detail_by_code[code].append(
                                    {
                                        "date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                                        "open": open_d,
                                        "high": high_d,
                                        "low": low_d,
                                        "close": close_d,
                                        "remark": post_rm,
                                    }
                                )

                # 区间最后一天（end_date）仍涨停时，补出再下一日（常为「今日」），便于看最后一棒之后盈亏
                if streak > 0 and end_idx + 1 < len(trade_dates):
                    d_next = trade_dates[end_idx + 1]
                    open_n, high_n, low_n, close_n = m.get(d_next, (None, None, None, None))
                    if close_n is not None:
                        prev_ke = trade_dates[end_idx]
                        prev_c = (m.get(prev_ke) or (None, None, None, None))[3]
                        try:
                            if prev_c is None or float(prev_c) <= 0:
                                raise ValueError
                            pc2 = float(prev_c)
                            cl2 = float(close_n)
                        except (TypeError, ValueError):
                            pass
                        else:
                            next_is_up = code in codes_by_date.get(d_next, set())
                            limit_dn2 = round(pc2 * (1.0 - limit_ratio), 2)
                            diff2 = abs(cl2 - limit_dn2)
                            dec2 = 1.0 - cl2 / pc2
                            is_dn_n = (diff2 <= 0.02) or (dec2 >= limit_ratio * 0.99 - 1e-12)
                            if not next_is_up:
                                if is_dn_n:
                                    detail_by_code[code].append(
                                        {
                                            "date": f"{d_next[:4]}-{d_next[4:6]}-{d_next[6:8]}",
                                            "open": open_n,
                                            "high": high_n,
                                            "low": low_n,
                                            "close": close_n,
                                            "remark": "跌停",
                                        }
                                    )
                                else:
                                    post_rm = (
                                        "首板后次日"
                                        if streak == 1
                                        else f"{streak}连板后次日"
                                    )
                                    detail_by_code[code].append(
                                        {
                                            "date": f"{d_next[:4]}-{d_next[4:6]}-{d_next[6:8]}",
                                            "open": open_n,
                                            "high": high_n,
                                            "low": low_n,
                                            "close": close_n,
                                            "remark": post_rm,
                                        }
                                    )

            # 完成
            self.finished.emit(
                rows, first_board_rows, streak_rows, limitup_date, debug_msg, "full", detail_by_code
            )
        except Exception as e:
            self.error_occurred.emit(f"{type(e).__name__}: {e}")


class StockLimitUpDetailDialog(QDialog):
    """单只股票：按日一行展示日期、开盘、最高、最低、收盘、备注（首板涨停 / N连板 / 跌停）。"""

    def __init__(self, code: str, name: str, details: List[dict], parent=None):
        super().__init__(parent)
        title_name = (name or "").strip()
        self.setWindowTitle(f"走势明细 - {code} {title_name}".strip())
        self.resize(640, 520)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        lay = QVBoxLayout(self)
        note = QLabel(
            "说明：含「日K 涨停」日（首板涨停 / 2连板…）、「日K 跌停」日，以及每次涨停结束后的下一交易日"
            "（首板后次日 / N连板后次日，便于对照前一日涨停收盘看盈亏；若次日跌停则只标「跌停」）。"
            "若统计区间最后一天仍涨停，会多列再下一交易日。价格来自 QMT 日K。"
        )
        note.setWordWrap(True)
        note.setStyleSheet("color: #666; font-size: 11px;")
        lay.addWidget(note)
        info = QLabel(
            f"<b>{code}</b>　{title_name}　"
            f"共 <b>{len(details)}</b> 行（按日期顺序）"
        )
        lay.addWidget(info)

        table = QTableWidget(len(details), 6)
        table.setHorizontalHeaderLabels(["日期", "开盘", "最高", "最低", "收盘", "备注"])
        table.setAlternatingRowColors(True)
        color_a = QColor("#2f5597")
        color_b = QColor("#7a4f01")
        seg_idx = 0
        in_cycle = False
        for r, rec in enumerate(details):
            remark = str(rec.get("remark") or "")
            if "首板涨停" in remark:
                in_cycle = True
            fg = color_a if (seg_idx % 2 == 0) else color_b

            d_item = QTableWidgetItem(str(rec.get("date") or ""))
            o_item = QTableWidgetItem(_fmt_px(rec.get("open")))
            h_item = QTableWidgetItem(_fmt_px(rec.get("high")))
            l_item = QTableWidgetItem(_fmt_px(rec.get("low")))
            c_item = QTableWidgetItem(_fmt_px(rec.get("close")))
            rm = QTableWidgetItem(remark)
            rm.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            for it in (d_item, o_item, h_item, l_item, c_item, rm):
                it.setForeground(fg)
            table.setItem(r, 0, d_item)
            table.setItem(r, 1, o_item)
            table.setItem(r, 2, h_item)
            table.setItem(r, 3, l_item)
            table.setItem(r, 4, c_item)
            table.setItem(r, 5, rm)

            if in_cycle and (("后次日" in remark) or ("跌停" in remark)):
                in_cycle = False
                seg_idx += 1
        hdr = table.horizontalHeader()
        hdr.setStretchLastSection(True)
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        lay.addWidget(table, 1)

        btn = QPushButton("关闭")
        btn.clicked.connect(self.accept)
        lay.addWidget(btn)


class LimitUpGeneAnalysisDialog(QDialog):
    def __init__(self, parent=None, auto_run: bool = False):
        super().__init__(parent)
        self._auto_run = bool(auto_run)
        self._auto_export_done = False
        self._auto_last_export_xlsx = ""
        self._auto_last_export_png = ""
        self._breakthrough_rows: List[dict] = []
        self.setWindowTitle("蚂蚁量化-智能突破买入小程序")
        self.resize(1000, 700)
        # 允许窗口最大化/最小化（标题栏通常会自动提供按钮，但有些窗口标志会禁用）
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        layout = QVBoxLayout(self)

        file_row = QHBoxLayout()
        layout.addLayout(file_row)
        file_row.addWidget(QLabel("股票列表文件："))
        self.file_edit = QLabel("未选择文件")
        file_row.addWidget(self.file_edit, 1)
        self.file_btn = QPushButton("选择文件…")
        file_row.addWidget(self.file_btn)

        self.file_btn.clicked.connect(self._choose_file)

        # 调试：若存在则预填股票列表（失败不提示）
        try:
            _dbg_stock_list = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "history_data",
                "选股结果_模式一_5-07.xls",
            )
            if os.path.isfile(_dbg_stock_list):
                self.file_edit.setText(_dbg_stock_list)
        except Exception:
            pass

        # 测试模式（可选）
        test_row = QHBoxLayout()
        layout.addLayout(test_row)
        self.test_cb = QCheckBox("测试：只取股票池前N只")
        test_row.addWidget(self.test_cb)

        self.test_spin = QSpinBox()
        self.test_spin.setRange(1, 500000)
        self.test_spin.setValue(1)
        test_row.addWidget(self.test_spin)

        self.test_cb.toggled.connect(self.test_spin.setEnabled)
        self.test_cb.setChecked(True)

        btn_row = QHBoxLayout()
        layout.addLayout(btn_row)
        self.preview_list_btn = QPushButton("查看选择的股票")
        self.analyze_btn = QPushButton("开始分析")
        btn_row.addWidget(self.preview_list_btn)
        btn_row.addWidget(self.analyze_btn)
        btn_row.addStretch()

        self.status_label = QLabel("待开始")
        layout.addWidget(self.status_label)

        # 接续导出放在各数据表之上：下方多张表均带垂直 stretch，若导出区放在最底易被挤出可视区域
        exp_box = QGroupBox("接续策略生成器：导出 JSON / 成交明细 CSV（与「蚂蚁量化策略生成系统」回测导入格式一致）")
        exp_form = QVBoxLayout()
        exp_box.setLayout(exp_form)
        exp_params = QHBoxLayout()
        exp_params.addWidget(QLabel("初始资金(元)："))
        self.export_ib_initial_cash = QDoubleSpinBox()
        self.export_ib_initial_cash.setRange(1000, 999999999)
        self.export_ib_initial_cash.setDecimals(0)
        self.export_ib_initial_cash.setValue(1_000_000)
        exp_params.addWidget(self.export_ib_initial_cash)
        exp_params.addWidget(QLabel("单股买入预算(元)："))
        self.export_ib_buy_amt = QDoubleSpinBox()
        self.export_ib_buy_amt.setRange(1000, 999999999)
        self.export_ib_buy_amt.setDecimals(0)
        self.export_ib_buy_amt.setValue(50_000)
        exp_params.addWidget(self.export_ib_buy_amt)
        exp_params.addWidget(QLabel("最小成交额(元)："))
        self.export_ib_min_ord = QDoubleSpinBox()
        self.export_ib_min_ord.setRange(0, 9999999)
        self.export_ib_min_ord.setDecimals(0)
        self.export_ib_min_ord.setValue(5000)
        exp_params.addWidget(self.export_ib_min_ord)
        exp_params.addWidget(QLabel("持有交易日数："))
        self.export_ib_hold_days = QSpinBox()
        self.export_ib_hold_days.setRange(1, 250)
        self.export_ib_hold_days.setValue(5)
        exp_params.addWidget(self.export_ib_hold_days)
        self.export_ib_t1_cb = QCheckBox("选股日后 T+1 起算区间（与批量回测默认一致）")
        self.export_ib_t1_cb.setChecked(True)
        exp_params.addWidget(self.export_ib_t1_cb)
        exp_params.addStretch()
        exp_form.addLayout(exp_params)
        exp_btns = QHBoxLayout()
        self.export_ib_csv_btn = QPushButton("导出成交明细(CSV)…")
        self.export_ib_csv_btn.setToolTip(
            "列与策略生成器回测页「导出成交明细」相同，可供合并汇总或留档。"
        )
        self.export_ib_csv_btn.setEnabled(False)
        self.export_ib_csv_btn.clicked.connect(self._on_export_ib_trades_csv)
        exp_btns.addWidget(self.export_ib_csv_btn)
        self.export_ib_json_v1_btn = QPushButton("导出接续 JSON(v1)…")
        self.export_ib_json_v1_btn.setToolTip(
            "单次接续快照（version=1）：对应策略生成器「从回测导出文件导入」。"
            "不能用于「载入上一轮批量回测→本策略」；批量接续请用 v2。"
        )
        self.export_ib_json_v1_btn.setEnabled(False)
        self.export_ib_json_v1_btn.clicked.connect(self._on_export_ib_json_v1)
        exp_btns.addWidget(self.export_ib_json_v1_btn)
        self.export_ib_json_v2_btn = QPushButton("导出批量接续 JSON(v2)…")
        self.export_ib_json_v2_btn.setToolTip(
            "version=2 + kind=batch_backtest + segments[]，与策略生成器「导出批量回测(JSON)」同形。"
            "「载入上一轮批量回测→本策略」时必须用本按钮保存的文件。"
        )
        self.export_ib_json_v2_btn.setEnabled(False)
        self.export_ib_json_v2_btn.clicked.connect(self._on_export_ib_json_v2)
        exp_btns.addWidget(self.export_ib_json_v2_btn)
        exp_btns.addStretch()
        exp_form.addLayout(exp_btns)
        layout.addWidget(exp_box)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["股票代码", "股票名称", "选股日"])
        layout.addWidget(self.table, 1)

        # Tick 输出（复用原首板区域）
        first_head_row = QHBoxLayout()
        first_head_row.addWidget(QLabel("Tick结构输出（每只股票分析前会清空）"))
        first_head_row.addStretch()
        self.export_first_board_btn = QPushButton("导出 Excel…")
        self.export_first_board_btn.setEnabled(False)
        self.export_first_board_btn.clicked.connect(self._export_first_board_excel)
        first_head_row.addWidget(self.export_first_board_btn)
        layout.addLayout(first_head_row)

        self.first_board_table = QTableWidget()
        self.first_board_table.setColumnCount(0)
        self.first_board_table.setRowCount(0)
        layout.addWidget(self.first_board_table, 1)

        # 今日连板
        streak_head_row = QHBoxLayout()
        streak_head_row.addWidget(
            QLabel(
                "挂单快照（挂单top3、冲高回落top2、压单驻留top1；"
                "压力位/实际使用的压力位在末两列；同股多买入日多行）"
            )
        )
        streak_head_row.addStretch()
        self.export_streak_btn = QPushButton("导出 Excel…")
        self.export_streak_btn.setEnabled(False)
        self.export_streak_btn.clicked.connect(self._export_streak_excel)
        streak_head_row.addWidget(self.export_streak_btn)
        layout.addLayout(streak_head_row)

        self.streak_table = QTableWidget()
        self.streak_table.setColumnCount(0)
        self.streak_table.setRowCount(0)
        layout.addWidget(self.streak_table, 1)

        breakthrough_head_row = QHBoxLayout()
        breakthrough_head_row.addWidget(
            QLabel("突破详情（逐 tick 检测；9:35 前不检；每个实际压力位仅记首次突破，可多行）")
        )
        breakthrough_head_row.addStretch()
        layout.addLayout(breakthrough_head_row)
        self.breakthrough_table = QTableWidget()
        self.breakthrough_table.setColumnCount(0)
        self.breakthrough_table.setRowCount(0)
        layout.addWidget(self.breakthrough_table, 1)

        self._detail_by_code: Dict[str, List[dict]] = {}
        self._detail_flat: List[dict] = []
        self._code_to_name: Dict[str, str] = {}
        self._current_tick_code: str = ""
        self._current_tick_name: str = ""
        self._current_tick_trade_date: str = ""

        self.table.itemDoubleClicked.connect(self._on_stock_table_double_clicked)
        self.first_board_table.itemDoubleClicked.connect(self._on_first_board_double_clicked)
        self.streak_table.itemDoubleClicked.connect(self._on_streak_double_clicked)

        self.preview_list_btn.clicked.connect(self.show_selected_stocks)
        self.analyze_btn.clicked.connect(self.start_full_analysis)

    def _refresh_ib_export_buttons(self) -> None:
        rows = getattr(self, "_breakthrough_rows", None) or []
        ok = any(str(r.get("buy_signal") or "").strip() == "是" for r in rows)
        self.export_ib_csv_btn.setEnabled(ok)
        self.export_ib_json_v1_btn.setEnabled(ok)
        self.export_ib_json_v2_btn.setEnabled(ok)

    def _run_ib_export_build(self) -> Dict[str, Any]:
        return build_strategy_generator_exports(
            getattr(self, "_breakthrough_rows", None) or [],
            initial_cash=float(self.export_ib_initial_cash.value()),
            buy_amount_per_stock=float(self.export_ib_buy_amt.value()),
            min_order_amount=float(self.export_ib_min_ord.value()),
            hold_trading_days=int(self.export_ib_hold_days.value()),
            start_next_trading_day=bool(self.export_ib_t1_cb.isChecked()),
        )

    def _default_ib_export_dir(self) -> str:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        d = os.path.join(script_dir, "history_data")
        os.makedirs(d, exist_ok=True)
        return d

    def _on_export_ib_trades_csv(self) -> None:
        try:
            data = self._run_ib_export_build()
        except Exception as e:
            QMessageBox.warning(self, "导出成交明细", str(e))
            return
        rows = data.get("csv_rows") or []
        if len(rows) <= 1:
            QMessageBox.information(self, "导出", "没有可导出的买入成交行。")
            return
        default_fn = os.path.join(
            self._default_ib_export_dir(),
            f"智能突破买入_成交明细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出成交明细 CSV",
            default_fn,
            "CSV (*.csv);;所有文件 (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerows(rows)
            QMessageBox.information(self, "已导出", path)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _on_export_ib_json_v1(self) -> None:
        try:
            data = self._run_ib_export_build()
        except Exception as e:
            QMessageBox.warning(self, "导出 JSON", str(e))
            return
        payload = data.get("export_v1")
        if not isinstance(payload, dict):
            QMessageBox.warning(self, "导出", "内部错误：缺少 export_v1")
            return
        default_fn = os.path.join(
            self._default_ib_export_dir(),
            f"ib_backtest_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出接续 JSON (version=1)",
            default_fn,
            "JSON (*.json);;所有文件 (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "已导出", path)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _on_export_ib_json_v2(self) -> None:
        try:
            data = self._run_ib_export_build()
        except Exception as e:
            QMessageBox.warning(self, "导出 JSON", str(e))
            return
        payload = data.get("export_v2")
        if not isinstance(payload, dict):
            QMessageBox.warning(self, "导出", "内部错误：缺少 export_v2")
            return
        segs = payload.get("segments") or []
        if not segs:
            QMessageBox.information(self, "导出", "批量 JSON 无分档数据。")
            return
        default_fn = os.path.join(
            self._default_ib_export_dir(),
            f"ib_batch_backtest_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出批量接续 JSON (version=2)",
            default_fn,
            "JSON (*.json);;所有文件 (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            QMessageBox.information(
                self,
                "已导出",
                f"共 {len(segs)} 档（选股日分档）：\n{path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _choose_file(self):
        fp, _ = QFileDialog.getOpenFileName(
            self,
            "选择股票列表文件",
            "",
            "列表 (*.txt *.csv *.xls *.xlsx);;Text (*.txt);;CSV (*.csv);;Excel (*.xls *.xlsx);;所有文件 (*.*)",
        )
        if fp:
            self.file_edit.setText(fp)

    def _render_tick_preview(self, code: str, name: str, tick_payload: object) -> None:
        payload = tick_payload if isinstance(tick_payload, dict) else {}
        tick_headers = list(payload.get("headers") or [])
        tick_rows = list(payload.get("rows") or [])

        # 展示全部 tick 字段，并额外补上代码/名称两列
        headers = ["股票代码", "股票名称"] + [str(h) for h in tick_headers]
        self.first_board_table.setColumnCount(len(headers))
        self.first_board_table.setHorizontalHeaderLabels(headers)
        self.first_board_table.setRowCount(len(tick_rows))

        for r_idx, row_vals in enumerate(tick_rows):
            vals = [code, name or ""] + [str(v) for v in list(row_vals)]
            for c_idx, txt in enumerate(vals):
                item = QTableWidgetItem(txt)
                if c_idx >= 2:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.first_board_table.setItem(r_idx, c_idx, item)
        self.first_board_table.resizeRowsToContents()

    def _on_tick_preview(self, code: str, name: str, tick_payload: object) -> None:
        # 每只股票输出前先清空显示
        self.first_board_table.setRowCount(0)
        self.first_board_table.setColumnCount(0)
        self._current_tick_code = (code or "").strip()
        self._current_tick_name = (name or "").strip()
        pl = tick_payload if isinstance(tick_payload, dict) else {}
        self._current_tick_trade_date = str(pl.get("tick_trade_date") or "").strip()
        self._render_tick_preview(code, name, tick_payload)

    def _export_qtable_to_excel(
        self,
        table: QTableWidget,
        default_name: str,
        *,
        column_alignments: Optional[List[str]] = None,
        column_widths: Optional[List[Optional[float]]] = None,
        excel_text_format_columns: Optional[Set[int]] = None,
        sheet_zoom: Optional[int] = None,
        show_dialogs: bool = True,
    ) -> bool:
        """将 QTableWidget 导出为 .xlsx（需 openpyxl）。成功返回 True。

        column_alignments: 可选，每列水平对齐：left / center / right，
        长度与列数一致；为 None 时数据格仅在含换行时设置左对齐自动换行（连板表等）。
        column_widths: 可选，每列 Excel 列宽（约等于字符数）；None 表示该列用默认宽度。
        excel_text_format_columns: 对这些列（0 起）设置 Excel 单元格格式为「文本」( @ )；
        对 6 位数字代码前加零宽空格、并尽量设置 quote_prefix，配合 ignoredErrors，
        减轻 Excel/WPS 绿色智能标记；默认仅第 0 列。
        sheet_zoom: 可选，工作表显示缩放百分比（如 90），使表格在 Excel 中更紧凑。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if show_dialogs:
                QMessageBox.warning(
                    self,
                    "缺少依赖",
                    "导出 Excel 需要安装 openpyxl：\npip install openpyxl",
                )
            return False
        if table.rowCount() == 0 or table.columnCount() == 0:
            if show_dialogs:
                QMessageBox.information(self, "提示", "表格为空，无内容可导出。")
            return False
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        if not default_name.lower().endswith(".xlsx"):
            default_name = os.path.splitext(default_name)[0] + ".xlsx"
        if show_dialogs:
            suggested = os.path.join(out_dir, default_name)
            path, _ = QFileDialog.getSaveFileName(
                self,
                "导出 Excel",
                suggested,
                "Excel 工作簿 (*.xlsx);;所有文件 (*.*)",
            )
            if not path:
                return False
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
        else:
            if os.path.isabs(default_name):
                path = default_name
            else:
                path = os.path.join(out_dir, default_name)
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            os.makedirs(os.path.dirname(path), exist_ok=True)
        text_fmt_cols: Set[int] = (
            set(excel_text_format_columns)
            if excel_text_format_columns is not None
            else {0}
        )
        wb = Workbook()
        ws = wb.active
        sheet_title = os.path.splitext(os.path.basename(default_name))[0] or "数据"
        ws.title = sheet_title[:31]
        header_font = Font(bold=True)
        header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        ncols = table.columnCount()
        nrows = table.rowCount()
        for c in range(ncols):
            hi = table.horizontalHeaderItem(c)
            text = hi.text() if hi else ""
            cell = ws.cell(row=1, column=c + 1, value=text)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = cell_border
            if c in text_fmt_cols:
                cell.number_format = "@"
        data_wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for r in range(nrows):
            for c in range(ncols):
                it = table.item(r, c)
                val = it.text() if it else ""
                # 第一列强制文本；其余列若为纯数字则写入数值，便于 Excel 计算/排序
                out_val = val
                if c == 0:
                    out_val = str(val)
                else:
                    sv = str(val).strip()
                    if re.fullmatch(r"[+-]?\d+", sv):
                        try:
                            out_val = int(sv)
                        except Exception:
                            out_val = val
                    elif re.fullmatch(r"[+-]?(?:\d+\.\d*|\.\d+)", sv):
                        try:
                            out_val = float(sv)
                        except Exception:
                            out_val = val
                if c in text_fmt_cols and isinstance(out_val, str):
                    out_val = _excel_stock_cell_str_for_export(out_val)
                cell = ws.cell(row=r + 2, column=c + 1, value=out_val)
                cell.border = cell_border
                if c in text_fmt_cols:
                    cell.number_format = "@"
                    qp = getattr(cell, "quote_prefix", None)
                    if qp is not None:
                        try:
                            cell.quote_prefix = True
                        except Exception:
                            pass
                if column_alignments is not None and c < len(column_alignments):
                    h = column_alignments[c]
                    if h not in ("left", "center", "right"):
                        h = "left"
                    if "\n" in str(val):
                        cell.alignment = Alignment(
                            wrap_text=True, vertical="top", horizontal=h
                        )
                    else:
                        cell.alignment = Alignment(vertical="center", horizontal=h)
                elif "\n" in str(val):
                    cell.alignment = data_wrap
        if column_widths:
            for c in range(ncols):
                if c >= len(column_widths):
                    break
                w = column_widths[c]
                if w is not None and float(w) > 0:
                    ws.column_dimensions[get_column_letter(c + 1)].width = float(w)
        if sheet_zoom is not None:
            try:
                z = int(sheet_zoom)
                if 10 <= z <= 400:
                    ws.sheet_view.zoomScale = z
            except Exception:
                pass
        # Excel 仍可能对「长得像数字的文本」显示绿色角标；在工作表上声明忽略该检查（OOXML ignoredErrors）
        if text_fmt_cols:
            try:
                from openpyxl.worksheet.cell_range import CellRange
                from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

                last_row = nrows + 1
                ign_list = []
                for c in sorted(text_fmt_cols):
                    letter = get_column_letter(c + 1)
                    rng = f"{letter}1:{letter}{last_row}"
                    ign_list.append(
                        IgnoredError(sqref=CellRange(rng), numberStoredAsText=True)
                    )
                if ign_list:
                    ws.ignored_errors = IgnoredErrors(ignoredError=tuple(ign_list))
            except Exception:
                pass
        try:
            wb.save(path)
        except OSError as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", str(e))
            return False
        except Exception as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", f"{type(e).__name__}: {e}")
            return False
        if show_dialogs:
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        return True

    def _export_first_board_excel(self):
        date = (
            (getattr(self, "_current_tick_trade_date", "") or "").strip()
            or getattr(self, "_limitup_date", "")
            or "export"
        )
        code = (getattr(self, "_current_tick_code", "") or "").strip()
        name = (getattr(self, "_current_tick_name", "") or "").strip()
        full_headers: List[str] = []
        full_rows: List[List[str]] = []
        if xtdata is not None and code and re.fullmatch(r"\d{6}", code):
            try:
                tick_df = _load_or_fetch_tick_df(code, date)
                full_headers, full_rows = _extract_tick_preview_rows(tick_df, limit_n=None)
            except Exception:
                full_headers, full_rows = [], []

        if full_headers and full_rows:
            t = QTableWidget(len(full_rows), 2 + len(full_headers))
            t.setHorizontalHeaderLabels(["股票代码", "股票名称"] + [str(h) for h in full_headers])
            for r_idx, row_vals in enumerate(full_rows):
                vals = [code, name] + [str(v) for v in list(row_vals)]
                for c_idx, txt in enumerate(vals):
                    item = QTableWidgetItem(txt)
                    if c_idx >= 2:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    t.setItem(r_idx, c_idx, item)
            n = t.columnCount()
            aligns: List[str] = []
            for c in range(n):
                if c == 0:
                    aligns.append("center")
                elif c == 1:
                    aligns.append("left")
                else:
                    aligns.append("right")
            self._export_qtable_to_excel(
                t,
                f"ticks_{date}.xlsx",
                column_alignments=aligns,
            )
            return

        n = self.first_board_table.columnCount()
        aligns: List[str] = []
        for c in range(n):
            if c == 0:
                aligns.append("center")
            elif c == 1:
                aligns.append("left")
            else:
                aligns.append("right")
        self._export_qtable_to_excel(
            self.first_board_table,
            f"ticks_{date}.xlsx",
            column_alignments=aligns,
        )

    def _export_streak_excel(self):
        date = getattr(self, "_limitup_date", "") or "export"
        n = self.streak_table.columnCount()
        if n == 9:
            streak_col_align = {
                0: "center",
                1: "left",
                2: "center",
                3: "center",
                4: "left",
                5: "left",
                6: "left",
                7: "right",
                8: "right",
            }
            streak_widths: List[Optional[float]] = [
                10.5,
                12.0,
                12.0,
                20.0,
                44.0,
                26.0,
                22.0,
                10.5,
                13.0,
            ]
            out_name = f"压力位_{date}.xlsx"
        else:
            streak_col_align = {
                0: "center",
                1: "left",
                2: "right",
                3: "right",
                4: "left",
                5: "left",
            }
            streak_widths = [
                10.5,
                9.5,
                9.0,
                10.0,
                13.0,
                19.0,
            ]
            out_name = f"封单结构_{date}.xlsx"
        aligns = [streak_col_align.get(c, "left") for c in range(n)]
        while len(streak_widths) < n:
            streak_widths.append(None)
        self._export_qtable_to_excel(
            self.streak_table,
            out_name,
            column_alignments=aligns,
            column_widths=streak_widths[:n],
            sheet_zoom=90,
        )

    def _export_qtable_to_png(self, table: QTableWidget, png_path: str) -> bool:
        """将表格按自适应列宽导出为 PNG 图片。"""
        try:
            nrows, ncols = table.rowCount(), table.columnCount()
            if nrows <= 0 or ncols <= 0:
                return False
            tmp = QTableWidget(nrows, ncols)
            headers = []
            for c in range(ncols):
                hi = table.horizontalHeaderItem(c)
                headers.append(hi.text() if hi else "")
            tmp.setHorizontalHeaderLabels(headers)
            for r in range(nrows):
                for c in range(ncols):
                    it = table.item(r, c)
                    tmp.setItem(r, c, QTableWidgetItem(it.text() if it else ""))

            tmp.resizeColumnsToContents()
            tmp.resizeRowsToContents()
            total_w = tmp.verticalHeader().width() + 2 * tmp.frameWidth()
            total_h = tmp.horizontalHeader().height() + 2 * tmp.frameWidth()
            for c in range(ncols):
                total_w += tmp.columnWidth(c)
            for r in range(nrows):
                total_h += tmp.rowHeight(r)
            # 不要用过大下限撑宽画布，否则 PNG 右侧会出现大块空白
            total_w = max(total_w, 80)
            total_h = max(total_h, 60)
            tmp.resize(total_w, total_h)
            # 强制白底，避免透明背景在部分查看器中显示为黑底
            tmp.setStyleSheet("QTableWidget { background: #ffffff; }")
            tmp.viewport().setStyleSheet("background: #ffffff;")
            QApplication.processEvents()

            pix = QPixmap(tmp.size())
            pix.fill(Qt.white)
            tmp.render(pix)
            return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
        except Exception:
            return False

    def _auto_adjust_excel_column_widths(self, xlsx_path: str) -> bool:
        """自动按内容调列宽，避免导图时列宽拥挤。"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            return False
        if not os.path.exists(xlsx_path):
            return False
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, max_row + 1):
                    val = ws.cell(row=r, column=c).value
                    text = "" if val is None else str(val)
                    text = text.replace("\n", " ")
                    if len(text) > max_len:
                        max_len = len(text)
                # 宽度限制在合理范围，防止超长文本导致巨宽图片
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)
            wb.save(xlsx_path)
            return True
        except Exception:
            return False

    def _excel_to_png_via_wps(self, xlsx_path: str, png_path: str) -> bool:
        """通过 WPS 将首个工作表导出为 PNG（按有值区域收缩，避免 UsedRange 右侧空白）。"""
        if not os.path.exists(xlsx_path):
            return False
        try:
            root = os.path.dirname(os.path.abspath(__file__))
            if root not in sys.path:
                sys.path.insert(0, root)
            from tools.wps_excel_to_png import excel_sheet_to_png_via_wps

            ok, _msg = excel_sheet_to_png_via_wps(
                xlsx_path,
                png_path,
                sheet_index=1,
                sheet_name=None,
                close_workbook_save=True,
            )
            return bool(ok)
        except Exception:
            return False

    def _auto_export_streak_excel_and_png(self) -> bool:
        """自动模式：导出第三表 Excel（压力位或连板统计），再导出表格图片。"""
        if self.streak_table.rowCount() <= 0 or self.streak_table.columnCount() <= 0:
            self.status_label.setText("自动运行完成，但第三表为空，未导出。")
            return False
        date = getattr(self, "_limitup_date", "") or datetime.now().strftime("%Y%m%d")
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history_data")
        os.makedirs(out_dir, exist_ok=True)
        n = self.streak_table.columnCount()
        if n == 9:
            base = f"压力位_{date}"
        else:
            base = f"封单结构_{date}"
        xlsx_path = os.path.join(out_dir, f"{base}.xlsx")
        png_path = os.path.join(out_dir, f"{base}.png")

        if n == 9:
            streak_col_align = {
                0: "center",
                1: "left",
                2: "center",
                3: "center",
                4: "left",
                5: "left",
                6: "left",
                7: "right",
                8: "right",
            }
        else:
            streak_col_align = {0: "center", 1: "left", 2: "right", 3: "right", 4: "left", 5: "left"}
        aligns = [streak_col_align.get(c, "left") for c in range(n)]

        ok_xlsx = self._export_qtable_to_excel(
            self.streak_table,
            xlsx_path,
            column_alignments=aligns,
            column_widths=None,
            sheet_zoom=90,
            show_dialogs=False,
        )
        if not ok_xlsx:
            self.status_label.setText("自动导出失败：Excel 导出失败。")
            return False

        self._auto_adjust_excel_column_widths(xlsx_path)
        ok_png = self._excel_to_png_via_wps(xlsx_path, png_path)
        if not ok_png:
            ok_png = self._export_qtable_to_png(self.streak_table, png_path)
        if not ok_png:
            self.status_label.setText("自动导出失败：图片导出失败。")
            return False

        self._auto_last_export_xlsx = xlsx_path
        self._auto_last_export_png = png_path
        self.status_label.setText(
            f"自动运行完成，已导出：{os.path.basename(xlsx_path)}、{os.path.basename(png_path)}；10秒后退出。"
        )
        return True

    def _clear_stat_tables(self) -> None:
        self.first_board_table.setRowCount(0)
        self.first_board_table.setColumnCount(0)
        self.streak_table.setRowCount(0)
        self.streak_table.setColumnCount(0)
        if hasattr(self, "breakthrough_table"):
            self.breakthrough_table.setRowCount(0)
            self.breakthrough_table.setColumnCount(0)
        self._breakthrough_rows = []
        self._detail_by_code = {}
        self._detail_flat = []
        if hasattr(self, "export_ib_csv_btn"):
            self._refresh_ib_export_buttons()

    def _code_from_table_row(self, table: QTableWidget, row: int) -> str:
        it = table.item(row, 0)
        return (it.text() or "").strip() if it else ""

    def _open_limitup_detail(self, code: str) -> None:
        code = (code or "").strip()
        if not code:
            return
        mp = getattr(self, "_detail_by_code", None) or {}
        details = list(mp.get(code) or [])
        if not details:
            QMessageBox.information(
                self,
                "走势明细",
                f"股票 {code} 在统计区间内无涨停/跌停明细记录，或尚未完成「开始分析」。",
            )
            return
        name = getattr(self, "_code_to_name", {}).get(code, "")
        dlg = StockLimitUpDetailDialog(code, name, details, self)
        dlg.exec_()

    def _on_stock_table_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.table, item.row())
        self._open_limitup_detail(code)

    def _on_first_board_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.first_board_table, item.row())
        self._open_limitup_detail(code)

    def _on_streak_double_clicked(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        code = self._code_from_table_row(self.streak_table, item.row())
        self._open_limitup_detail(code)

    def _enable_action_buttons(self, enabled: bool) -> None:
        self.preview_list_btn.setEnabled(enabled)
        self.analyze_btn.setEnabled(enabled)

    def _build_thread_params(self) -> Optional[dict]:
        """校验输入并返回线程构造参数；失败返回 None。"""
        script_dir = os.path.dirname(os.path.abspath(__file__))
        history_dir = os.path.join(script_dir, "history_data")
        all_a_csv = os.path.join(script_dir, "data", "all_a_stocks.csv")

        stock_list_filepath = self.file_edit.text().strip()
        if stock_list_filepath == "未选择文件" or not os.path.exists(stock_list_filepath):
            self.status_label.setText("请先选择股票列表文件")
            return None

        test_only = bool(self.test_cb.isChecked())
        test_limit_n = int(self.test_spin.value())

        return {
            "history_dir": history_dir,
            "all_a_stocks_csv": all_a_csv,
            "source_mode": "file",
            "stock_list_filepath": stock_list_filepath,
            "analysis_date_yyyymmdd": "",
            "limitup_csv_path": "",
            "test_only": test_only,
            "test_limit_n": test_limit_n,
        }

    def show_selected_stocks(self) -> None:
        self._enable_action_buttons(False)
        self.table.setRowCount(0)
        self._clear_stat_tables()
        self.export_first_board_btn.setEnabled(False)
        self.export_streak_btn.setEnabled(False)
        self.status_label.setText("正在生成股票列表，请稍候...")

        params = self._build_thread_params()
        if params is None:
            self._enable_action_buttons(True)
            return

        self.thread = TodayLimitUpFinderThread(
            run_mode="list_only",
            parent=self,
            **params,
        )
        self.thread.finished.connect(self._on_finished)
        self.thread.error_occurred.connect(self._on_error)
        self.thread.tick_preview.connect(self._on_tick_preview)
        self.thread.start()

    def start_full_analysis(self) -> None:
        self._enable_action_buttons(False)
        self._clear_stat_tables()
        self.export_first_board_btn.setEnabled(False)
        self.export_streak_btn.setEnabled(False)
        self.status_label.setText("正在逐只分析 tick 结构，请稍候（可能较慢）...")

        params = self._build_thread_params()
        if params is None:
            self._enable_action_buttons(True)
            return

        self.thread = TodayLimitUpFinderThread(
            run_mode="full",
            parent=self,
            **params,
        )
        self.thread.finished.connect(self._on_finished)
        self.thread.error_occurred.connect(self._on_error)
        self.thread.tick_preview.connect(self._on_tick_preview)
        self.thread.start()

    def _on_finished(
        self,
        today_rows: List[Tuple[str, str, str]],
        first_board_rows: List[dict],
        streak_rows: List[dict],
        limitup_date: str,
        debug_msg: str,
        finish_kind: str,
        detail_by_code: object = None,
    ):
        self._limitup_date = limitup_date
        if not isinstance(detail_by_code, dict):
            detail_by_code = {}

        raw_detail = dict(detail_by_code)
        if finish_kind == "tick_only":
            self._breakthrough_rows = list(raw_detail.pop("breakthrough_rows", []))
        else:
            self._breakthrough_rows = []
            raw_detail.pop("breakthrough_rows", None)

        self._code_to_name = {c: n for c, n, _ in today_rows}
        self._detail_by_code = raw_detail if finish_kind != "list_only" else {}
        self._detail_flat = []
        for c, lst in sorted(self._detail_by_code.items(), key=lambda x: x[0]):
            nm = self._code_to_name.get(c, "")
            for rec in lst:
                self._detail_flat.append({"code": c, "name": nm, **rec})

        # 要分析的股票列表
        self.table.setRowCount(len(today_rows))
        for i, (code, name, pick_date) in enumerate(today_rows):
            self.table.setItem(i, 0, QTableWidgetItem(code))
            self.table.setItem(i, 1, QTableWidgetItem(name or ""))
            self.table.setItem(i, 2, QTableWidgetItem(pick_date or ""))

        if finish_kind == "list_only":
            self._clear_stat_tables()
            self.export_first_board_btn.setEnabled(False)
            self.export_streak_btn.setEnabled(False)
        elif finish_kind == "tick_only":
            # 首板区域已在 _on_tick_preview 中按“逐只清空+输出”更新；第三表为挂单 top3 多行文本
            structure_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("buy_trade_day", "买入日"),
                ("pressure_time", "时间点"),
                ("pressure_detail", "挂单top3"),
                ("pullback_top2", "冲高回落top2"),
                ("dwell_top1", "压单驻留top1"),
                ("pressure_weighted_avg", "压力位"),
                ("pressure_effective_used", "实际使用的压力位"),
            ]
            self.streak_table.setColumnCount(len(structure_headers))
            self.streak_table.setHorizontalHeaderLabels([h[1] for h in structure_headers])
            self.streak_table.setRowCount(len(streak_rows))
            for r_idx, row in enumerate(streak_rows):
                for c_idx, (key, _) in enumerate(structure_headers):
                    txt = "" if row.get(key) is None else str(row.get(key))
                    item = QTableWidgetItem(txt)
                    if key in ("pressure_detail", "pullback_top2", "dwell_top1"):
                        item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)
                    elif key in ("pressure_weighted_avg", "pressure_effective_used"):
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    elif key in ("code", "buy_trade_day", "pressure_time"):
                        item.setTextAlignment(Qt.AlignCenter)
                    self.streak_table.setItem(r_idx, c_idx, item)
            self.streak_table.resizeRowsToContents()

            bt_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("buy_trade_day", "买入日"),
                ("选股日", "选股日"),
                ("breakthrough_time", "突破时刻"),
                ("pressure_effective_used", "实际使用的压力位"),
                ("last_price", "最新价"),
                ("volume_avg_before_break", "突破前均量(当日)"),
                ("volume_break_tick", "突破tick成交量"),
                ("depth_bid_vol", "委买量"),
                ("depth_ask_vol", "委卖量"),
                ("prev_tick_ask1_price", "前tick卖一价"),
                ("prev_tick_ask1_vol", "前tick卖一量"),
                ("break_tick_trade_price", "本tick成交价"),
                ("break_tick_trade_vol", "本tick成交量"),
                ("true_breakthrough", "真突破"),
                ("true_breakthrough_remark", "备注"),
                ("buy_signal", "买入信号"),
            ]
            bt_rows = getattr(self, "_breakthrough_rows", None) or []
            self.breakthrough_table.setColumnCount(len(bt_headers))
            self.breakthrough_table.setHorizontalHeaderLabels([h[1] for h in bt_headers])
            self.breakthrough_table.setRowCount(len(bt_rows))
            for r_idx, brow in enumerate(bt_rows):
                for c_idx, (key, _) in enumerate(bt_headers):
                    txt = "" if brow.get(key) is None else str(brow.get(key))
                    item = QTableWidgetItem(txt)
                    if key in (
                        "code",
                        "buy_trade_day",
                        "选股日",
                        "breakthrough_time",
                        "true_breakthrough",
                        "buy_signal",
                    ):
                        item.setTextAlignment(Qt.AlignCenter)
                    elif key in (
                        "pressure_effective_used",
                        "last_price",
                        "volume_avg_before_break",
                        "volume_break_tick",
                        "depth_bid_vol",
                        "depth_ask_vol",
                        "prev_tick_ask1_price",
                        "prev_tick_ask1_vol",
                        "break_tick_trade_price",
                        "break_tick_trade_vol",
                    ):
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    elif key == "true_breakthrough_remark":
                        item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)
                    else:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    self.breakthrough_table.setItem(r_idx, c_idx, item)
            self.breakthrough_table.resizeRowsToContents()

            self.export_first_board_btn.setEnabled(self.first_board_table.rowCount() > 0)
            self.export_streak_btn.setEnabled(len(streak_rows) > 0)
            self._refresh_ib_export_buttons()
        else:
            # 首板表（列键与后台 dict 一致，表头为中文）
            first_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("today_streak", "今日连板数"),
                ("first_board_count_1y", "首板次数"),
                ("next_day_limitup_count", "次日涨停次数"),
                ("next_day_up_count", "次日收涨次数"),
                ("next_day_down_count", "次日收跌次数"),
                ("avg_open_return_pct", "次日开盘收益均值(%)"),
                ("avg_close_return_pct", "次日收盘收益均值(%)"),
            ]
            self.first_board_table.setColumnCount(len(first_headers))
            self.first_board_table.setHorizontalHeaderLabels([h[1] for h in first_headers])
            self.first_board_table.setRowCount(len(first_board_rows))
            for r_idx, row in enumerate(first_board_rows):
                for c_idx, (key, _) in enumerate(first_headers):
                    val = row.get(key, "")
                    if key in ("avg_open_return_pct", "avg_close_return_pct"):
                        try:
                            cell_text = (
                                f"{float(val):.2f}"
                                if val is not None and val != ""
                                else ""
                            )
                        except (TypeError, ValueError):
                            cell_text = str(val)
                    else:
                        cell_text = str(val if val is not None else "")
                    self.first_board_table.setItem(r_idx, c_idx, QTableWidgetItem(cell_text))

            # 连板表
            streak_headers = [
                ("code", "股票代码"),
                ("name", "股票名称"),
                ("today_streak", "今日连板数"),
                ("max_streak_in_year", "近一年最大连板数"),
                ("counts_by_k", "各连板次数分布"),
                ("upgrade_ratios", "连板晋级比例"),
            ]
            self.streak_table.setColumnCount(len(streak_headers))
            self.streak_table.setHorizontalHeaderLabels([h[1] for h in streak_headers])
            self.streak_table.setRowCount(len(streak_rows))
            for r_idx, row in enumerate(streak_rows):
                for c_idx, (key, _) in enumerate(streak_headers):
                    v = row.get(key, "")
                    text = "" if v is None else str(v)
                    item = QTableWidgetItem(text)
                    if key in ("counts_by_k", "upgrade_ratios") and "\n" in text:
                        item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)
                    self.streak_table.setItem(r_idx, c_idx, item)
            self.streak_table.resizeRowsToContents()

            self.export_first_board_btn.setEnabled(len(first_board_rows) > 0)
            self.export_streak_btn.setEnabled(len(streak_rows) > 0)
            self._refresh_ib_export_buttons()

        # 保存一个 txt，方便你核对
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"要分析的股票列表_{limitup_date}.txt")
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                for code, name, pick_date in today_rows:
                    f.write(f"{code}\t{name}\t{pick_date}\n")
        except Exception:
            out_path = ""

        suffix = f"；已保存：{out_path}" if out_path else ""
        if finish_kind == "list_only":
            self.status_label.setText(
                f"已列出 {len(today_rows)} 只股票；{suffix}\n可点击「开始分析」生成首板/连板统计。\n{debug_msg}"
            )
        elif finish_kind == "tick_only":
            self.status_label.setText(
                f"tick 压力位分析完成：共 {len(today_rows)} 只股票；{suffix}\n"
                f"接续导出（成交明细 / JSON v1·v2）在窗口上方「接续策略生成器」分组。\n{debug_msg}"
            )
        else:
            self.status_label.setText(f"分析完成：共 {len(today_rows)} 只股票；{suffix}\n{debug_msg}")

        self._enable_action_buttons(True)
        if self._auto_run and finish_kind != "list_only" and (not self._auto_export_done):
            self._auto_export_done = True
            self._auto_export_streak_excel_and_png()
            QTimer.singleShot(10000, self.accept)

    def _on_error(self, msg: str):
        self.status_label.setText(f"出错：{msg}")
        self._enable_action_buttons(True)


def main():
    parser = argparse.ArgumentParser(description="蚂蚁量化-智能突破买入小程序")
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="启动后自动分析，分析完成后自动导出第三表（压力位）Excel 和图片，并在10秒后退出",
    )
    args, _unknown = parser.parse_known_args()

    app = QApplication(sys.argv)
    dlg = LimitUpGeneAnalysisDialog(auto_run=bool(args.auto_run))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg.start_full_analysis)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()


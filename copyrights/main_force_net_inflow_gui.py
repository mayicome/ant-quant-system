#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主力净流入统计（独立版）

读取 history_data 下「个股主力净流入_YYYYMMDD.csv」，按用户指定截止日统计近 10 个有数据交易日内的模式。
"""

from __future__ import annotations

import csv
import argparse
import io
import json
import os
import re
import sys
import warnings
from typing import Dict, List, Optional, Set, Tuple

_TOOLS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools")
if _TOOLS_DIR not in sys.path:
    sys.path.insert(0, _TOOLS_DIR)
import trim_export_png as _trim_png_mod  # noqa: E402

_trim_export_png_margins = _trim_png_mod.trim_export_png_margins

from PyQt5.QtCore import QBuffer, QByteArray, QDate, QIODevice, Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtGui import QColor, QBrush, QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QDateEdit,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

warnings.filterwarnings(
    "ignore",
    category=DeprecationWarning,
    message=r".*sipPyTypeDict.*",
)

FLOW_FILE_PREFIX = "个股主力净流入_"
FLOW_FILE_SUFFIX = ".csv"
CONFIG_FILENAME = "main_force_net_inflow_types.json"
# 与 _window_dates_upto_anchor(..., 10) 一致：最多 10 个交易日；N 日统计不得超过该窗口
MAX_N_DAYS_ALLOWED = 10

# 主结果表按「统计类型」分组交替行背景（便于区分相邻类型块）
_ROW_BG_TYPE_A = QColor(210, 220, 235)  # 略深
_ROW_BG_TYPE_B = QColor(245, 247, 250)  # 略浅


def _qt_color_to_excel_pattern_fill(qc: QColor):
    """与界面 QColor 一致的实心填充，供导出 xlsx。"""
    from openpyxl.styles import PatternFill

    hx = f"{qc.red():02X}{qc.green():02X}{qc.blue():02X}"
    return PatternFill(fill_type="solid", fgColor=hx)

# 统一一行：前 4 列固定 (统计类型, 代码, 名称, 起始日期)，之后 k 列依次为 0日…-(k-1)日净流入（k = 配置中 N 的最大值，且 ≤ 窗口长度）
UnifiedRow = Tuple[str, ...]


def _zfill_code(raw: str) -> str:
    s = "".join(ch for ch in str(raw or "").strip() if ch.isdigit())
    if not s:
        return ""
    return s[:6].zfill(6) if len(s) >= 6 else s.zfill(6)


# 与 limit_up_gene_analysis_gui 导出一致：6 位代码前加零宽空格，减轻 Excel/WPS 绿色智能标记
_EXCEL_STOCK_TEXT_PREFIX = "\u200b"


def _excel_stock_cell_str_for_export(s: str) -> str:
    t = (s or "").strip()
    if len(t) == 6 and t.isdigit():
        return _EXCEL_STOCK_TEXT_PREFIX + t
    return t


def _parse_main_flow_to_yuan(text: object) -> float:
    """将「21.52亿」「2432.77万」等转为元（元）；负数支持。"""
    if text is None:
        return 0.0
    if isinstance(text, (int, float)):
        return float(text)
    s = str(text).strip().replace(",", "").replace(" ", "")
    if not s or s in ("--", "-", "nan", "None"):
        return 0.0
    neg = 1.0
    if s.startswith("-"):
        neg = -1.0
        s = s[1:].strip()
    # 亿
    m = re.match(r"^(\d+\.?\d*)\s*亿", s)
    if m:
        return neg * float(m.group(1)) * 1e8
    m = re.match(r"^(\d+\.?\d*)\s*万", s)
    if m:
        return neg * float(m.group(1)) * 1e4
    m = re.match(r"^(\d+\.?\d*)$", s)
    if m:
        return neg * float(m.group(1))
    return 0.0


def _yuan_to_display(yuan: float) -> str:
    a = abs(yuan)
    sign = "-" if yuan < 0 else ""
    if a >= 1e8:
        return f"{sign}{a / 1e8:.2f}亿"
    if a >= 1e4:
        # 「万」单位只保留整数，不显示小数
        return f"{sign}{int(a / 1e4)}万"
    if a > 0:
        return f"{sign}{a:.2f}元"
    return "0"


def _amount_to_yuan(amount: float, unit: str) -> float:
    u = (unit or "").strip()
    if u == "亿":
        return float(amount) * 1e8
    # 默认按千万
    return float(amount) * 1e7


def _build_type_label(n_days: int, m_days: Optional[int], amount: float, unit: str) -> str:
    u = (unit or "").strip() or "千万"
    amt_txt = f"{float(amount):g}{u}"
    if m_days is None or int(m_days) == int(n_days):
        return f"连续{int(n_days)}日＞{amt_txt}"
    return f"{int(n_days)}日内{int(m_days)}日＞{amt_txt}"


def _list_flow_file_dates(history_dir: str) -> List[str]:
    dates: List[str] = []
    if not os.path.isdir(history_dir):
        return dates
    for fn in os.listdir(history_dir):
        if not fn.startswith(FLOW_FILE_PREFIX) or not fn.endswith(FLOW_FILE_SUFFIX):
            continue
        part = fn[len(FLOW_FILE_PREFIX) : -len(FLOW_FILE_SUFFIX)]
        if len(part) == 8 and part.isdigit():
            dates.append(part)
    return sorted(set(dates))


def _find_anchor_date(user_yyyymmdd: str, available: List[str]) -> Optional[str]:
    if user_yyyymmdd in available:
        return user_yyyymmdd
    earlier = [d for d in available if d <= user_yyyymmdd]
    if not earlier:
        return None
    return max(earlier)


def _window_dates_upto_anchor(available: List[str], anchor: str, n: int = 10) -> List[str]:
    """anchor 及之前、在 available 中的连续序列里，取最多 n 个交易日（含 anchor）。"""
    idx = available.index(anchor)
    start = max(0, idx - (n - 1))
    return available[start : idx + 1]


def _read_flow_csv(filepath: str) -> Dict[str, Tuple[float, str, str]]:
    """
    返回 code6 -> (净流入元, 原始显示字符串, 证券名称)
    """
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            out: Dict[str, Tuple[float, str, str]] = {}
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue
                for row in reader:
                    code = _zfill_code(row.get("代码") or "")
                    if not code:
                        continue
                    raw = row.get("今日主力净流入")
                    yuan = _parse_main_flow_to_yuan(raw)
                    name = (row.get("名称") or "").strip()
                    if code not in out:
                        # 统一使用格式化结果，保证“万”为整数显示（无小数）
                        out[code] = (yuan, _yuan_to_display(yuan), name)
            return out
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"读取失败: {filepath}; {last_err}")


def _date_fmt(ds: str) -> str:
    if len(ds) == 8:
        return f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
    return ds


def _date_mmdd(ds: str) -> str:
    if len(ds) == 8:
        return f"{ds[4:6]}-{ds[6:8]}"
    if len(ds) == 10 and "-" in ds:
        return ds[5:10]
    return ds


def _max_n_days_in_configs(type_configs: List[Dict[str, object]]) -> int:
    mx = 0
    for cfg in type_configs:
        n = int(cfg.get("n_days") or 0)
        if n > mx:
            mx = n
    return mx


def analyze_window(
    history_dir: str, window_dates: List[str], type_configs: List[Dict[str, object]]
) -> List[UnifiedRow]:
    """
    载入 window_dates 内各日文件，按配置统计并合并为统一行列表。
    每行: (统计类型, 代码, 名称, 起始日期, 之后 k 列净流入)；k = min(配置中 N 的最大值, len(window_dates))，
    列依次为 0 日、-1 日…-(k-1) 日（0 日为最近交易日）。
    """
    by_code_date: Dict[str, Dict[str, Tuple[float, str]]] = {}
    code_to_name: Dict[str, str] = {}
    all_codes: Set[str] = set()

    for d in window_dates:
        path = os.path.join(history_dir, f"{FLOW_FILE_PREFIX}{d}{FLOW_FILE_SUFFIX}")
        if not os.path.isfile(path):
            continue
        day_map = _read_flow_csv(path)
        for code, (yuan, disp, name) in day_map.items():
            all_codes.add(code)
            by_code_date.setdefault(code, {})[d] = (yuan, disp)
            if name:
                code_to_name[code] = name  # 按窗口顺序，后读到的日覆盖（靠近锚定日）

    def flow_on(code: str, d: str) -> Optional[Tuple[float, str]]:
        return by_code_date.get(code, {}).get(d)

    out: List[UnifiedRow] = []
    W = window_dates

    if not W:
        return out
    latest_day = W[-1]

    max_n_cfg = _max_n_days_in_configs(type_configs)
    if max_n_cfg < 1:
        max_n_cfg = 1
    flow_cols = min(max_n_cfg, len(W))

    for code in sorted(all_codes):
        nm = code_to_name.get(code, "")
        
        def _flow_disp_by_offset(offset: int) -> str:
            idx = len(W) - 1 - offset
            if idx < 0:
                return ""
            d = W[idx]
            fx = flow_on(code, d)
            return fx[1] if fx else ""

        recent_flows = tuple(_flow_disp_by_offset(i) for i in range(flow_cols))

        # 按配置逐项判断，且最后一个交易日必须是最近交易日（即窗口尾部）
        for t_idx, cfg in enumerate(type_configs):
            n_days = int(cfg.get("n_days") or 0)
            m_days_raw = cfg.get("m_days")
            m_days = int(m_days_raw) if m_days_raw is not None else n_days
            threshold_yuan = float(cfg.get("threshold_yuan") or 0.0)
            label = str(cfg.get("label") or "")

            if n_days <= 0 or n_days > len(W) or m_days <= 0 or m_days > n_days:
                continue
            if len(W) < n_days or latest_day != W[-1]:
                continue

            sub = W[-n_days:]
            hit_count = 0
            all_hit = True
            for d in sub:
                fx = flow_on(code, d)
                hit = bool(fx and fx[0] > threshold_yuan)
                if hit:
                    hit_count += 1
                else:
                    all_hit = False

            is_continuous = (m_days_raw is None) or (m_days == n_days)
            ok = all_hit if is_continuous else (hit_count >= m_days)
            if not ok:
                continue

            out.append((label, code, nm, _date_mmdd(sub[0])) + recent_flows)

    # 按配置顺序排序，同类按股票代码
    label_order = {str(cfg.get("label") or ""): i for i, cfg in enumerate(type_configs)}
    out.sort(key=lambda r: (label_order.get(r[0], 999), r[1]))
    return out


class MainForceAnalyzeThread(QThread):
    finished_ok = pyqtSignal(list, str, str)
    # rows, anchor_yyyymmdd, debug_msg
    failed = pyqtSignal(str)

    def __init__(
        self,
        history_dir: str,
        user_date: str,
        type_configs: List[Dict[str, object]],
        parent=None,
    ):
        super().__init__(parent)
        self.history_dir = history_dir
        self.user_date = user_date
        self.type_configs = type_configs

    def run(self) -> None:
        try:
            available = _list_flow_file_dates(self.history_dir)
            if not available:
                self.failed.emit("history_data 下未找到「个股主力净流入_*.csv」文件。")
                return
            anchor = _find_anchor_date(self.user_date, available)
            if anchor is None:
                self.failed.emit(
                    f"在 {self.user_date} 及之前没有可用的主力净流入文件（请检查日期与文件命名）。"
                )
                return
            window = _window_dates_upto_anchor(available, anchor, 10)
            rows = analyze_window(self.history_dir, window, self.type_configs)
            win_str = ",".join(_date_fmt(d) for d in window)
            dbg = (
                f"截止输入日={self.user_date}；锚定交易日={_date_fmt(anchor)}；"
                f"共载入 {len(window)} 个交易日：{win_str}"
            )
            self.finished_ok.emit(rows, anchor, dbg)
        except Exception as e:
            self.failed.emit(f"{type(e).__name__}: {e}")


class MainForceNetInflowDialog(QDialog):
    def __init__(self, parent=None, auto_run: bool = False):
        super().__init__(parent)
        self._auto_run = bool(auto_run)
        self._auto_export_done = False
        self._auto_last_export_xlsx = ""
        self._auto_last_export_png = ""
        self.setWindowTitle("蚂蚁量化 - 主力净流入统计（独立版）")
        self.resize(1040, 620)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )

        self._rows: List[UnifiedRow] = []
        self._anchor = ""
        self._thread: Optional[MainForceAnalyzeThread] = None
        self._type_configs: List[Dict[str, object]] = []
        self._n_flow_cols: int = 6
        self.headers: List[str] = []
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self._config_path = os.path.join(script_dir, "data", CONFIG_FILENAME)

        root = QVBoxLayout(self)

        row = QHBoxLayout()
        root.addLayout(row)
        row.addWidget(QLabel("截止日期："))
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        row.addWidget(self.date_edit)

        self.btn_run = QPushButton("统计")
        row.addWidget(self.btn_run)
        self.btn_export = QPushButton("导出 Excel…")
        self.btn_export.setEnabled(False)
        row.addWidget(self.btn_export)
        row.addStretch()

        self.status = QLabel("请选择日期后点击「统计」。")
        self.status.setWordWrap(True)
        root.addWidget(self.status)

        # 统计类型配置
        type_head = QHBoxLayout()
        root.addLayout(type_head)
        type_head.addWidget(QLabel("统计类型配置："))
        self.btn_add_type = QPushButton("添加统计类型")
        self.btn_del_type = QPushButton("删除选中")
        self.btn_save_type = QPushButton("保存配置")
        type_head.addWidget(self.btn_add_type)
        type_head.addWidget(self.btn_del_type)
        type_head.addWidget(self.btn_save_type)
        type_head.addStretch()

        self.type_headers = ["N日", "M日(可空)", "金额", "单位", "预览"]
        self.type_table = QTableWidget()
        self.type_table.setColumnCount(len(self.type_headers))
        self.type_table.setHorizontalHeaderLabels(self.type_headers)
        self.type_table.setMaximumHeight(190)
        self._configure_type_table_appearance()
        type_row = QHBoxLayout()
        type_row.addWidget(self.type_table, alignment=Qt.AlignLeft)
        type_row.addStretch()
        root.addLayout(type_row)

        self.table = QTableWidget()
        self._apply_flow_headers(6)
        root.addWidget(self.table, 1)

        self.btn_run.clicked.connect(self._on_run)
        self.btn_export.clicked.connect(self._on_export)
        self.btn_add_type.clicked.connect(self._on_add_type)
        self.btn_del_type.clicked.connect(self._on_delete_type)
        self.btn_save_type.clicked.connect(self._on_save_types)
        self.type_table.itemChanged.connect(self._refresh_type_preview)

        self._load_type_configs()

    def _configure_type_table_appearance(self) -> None:
        """统计类型表不拉满整窗宽度：限制总宽并固定各列。"""
        self.type_table.setMaximumWidth(500)
        th = self.type_table.horizontalHeader()
        th.setSectionResizeMode(QHeaderView.Fixed)
        self.type_table.setColumnWidth(0, 36)
        self.type_table.setColumnWidth(1, 52)
        self.type_table.setColumnWidth(2, 64)
        self.type_table.setColumnWidth(3, 44)
        self.type_table.setColumnWidth(4, 260)

    def _apply_flow_headers(self, n_flow: int) -> None:
        n_flow = max(1, min(int(n_flow), MAX_N_DAYS_ALLOWED))
        self._n_flow_cols = n_flow
        flow_names = ["0日净流入" if i == 0 else f"-{i}日净流入" for i in range(n_flow)]
        self.headers = ["统计类型", "股票代码", "股票名称", "起始日期"] + flow_names
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)

    def _peek_max_n_from_type_table(self) -> int:
        mx = 0
        for r in range(self.type_table.rowCount()):
            it = self.type_table.item(r, 0)
            if not it:
                continue
            try:
                n = int((it.text() or "").strip() or 0)
                if n > mx:
                    mx = n
            except ValueError:
                pass
        if mx <= 0:
            return 6
        return min(mx, MAX_N_DAYS_ALLOWED)

    def _sync_result_table_headers_with_types(self) -> None:
        new_flow = self._peek_max_n_from_type_table()
        if self._rows and len(self._rows[0]) - 4 != new_flow:
            self._rows = []
            self.table.setRowCount(0)
            self.btn_export.setEnabled(False)
            self.status.setText("统计类型列数已变化，请重新点击「统计」。")
        self._apply_flow_headers(new_flow)

    def _fill_unified_table(self, rows: List[UnifiedRow]) -> None:
        if rows:
            self._apply_flow_headers(len(rows[0]) - 4)
        self.table.setRowCount(len(rows))
        self.table.setAlternatingRowColors(False)
        prev_label: Optional[str] = None
        use_alt_palette = False  # False -> A 色，True -> B 色；类型切换时交替
        for r, tup in enumerate(rows):
            label = str(tup[0]) if tup else ""
            if r > 0 and label != prev_label:
                use_alt_palette = not use_alt_palette
            prev_label = label
            row_bg = QBrush(_ROW_BG_TYPE_B if use_alt_palette else _ROW_BG_TYPE_A)
            for c, val in enumerate(tup):
                text = val if val is not None else ""
                item = QTableWidgetItem(text)
                item.setBackground(row_bg)
                if c == 0:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                elif c == 1:
                    item.setTextAlignment(Qt.AlignCenter)
                elif c == 2:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                elif c == 3:
                    item.setTextAlignment(Qt.AlignCenter)
                elif c >= 4:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(r, c, item)
        self.table.resizeRowsToContents()

    def _set_type_table_rows(self, cfgs: List[Dict[str, object]]) -> None:
        self.type_table.blockSignals(True)
        self.type_table.setRowCount(len(cfgs))
        for r, cfg in enumerate(cfgs):
            n_days = int(cfg.get("n_days") or 2)
            m_days = cfg.get("m_days")
            amount = float(cfg.get("amount") or 1.0)
            unit = str(cfg.get("unit") or "千万")
            label = str(cfg.get("label") or "")

            self.type_table.setItem(r, 0, QTableWidgetItem(str(n_days)))
            self.type_table.setItem(r, 1, QTableWidgetItem("" if m_days is None else str(int(m_days))))
            self.type_table.setItem(r, 2, QTableWidgetItem(f"{amount:g}"))
            self.type_table.setItem(r, 3, QTableWidgetItem(unit))
            self.type_table.setItem(r, 4, QTableWidgetItem(label))
        self.type_table.blockSignals(False)

    def _collect_type_configs(self) -> List[Dict[str, object]]:
        cfgs: List[Dict[str, object]] = []
        for r in range(self.type_table.rowCount()):
            n_txt = (self.type_table.item(r, 0).text().strip() if self.type_table.item(r, 0) else "")
            m_txt = (self.type_table.item(r, 1).text().strip() if self.type_table.item(r, 1) else "")
            a_txt = (self.type_table.item(r, 2).text().strip() if self.type_table.item(r, 2) else "")
            u_txt = (self.type_table.item(r, 3).text().strip() if self.type_table.item(r, 3) else "")
            n_days = int(n_txt)
            m_days = int(m_txt) if m_txt else None
            amount = float(a_txt)
            unit = u_txt if u_txt in ("千万", "亿") else "千万"

            if n_days <= 0 or n_days > MAX_N_DAYS_ALLOWED:
                raise ValueError(f"第 {r+1} 行：N 日必须在 1~{MAX_N_DAYS_ALLOWED} 之间。")
            if m_days is not None and (m_days <= 0 or m_days > n_days):
                raise ValueError(f"第 {r+1} 行：M 日必须为空或 1~N。")
            if amount <= 0:
                raise ValueError(f"第 {r+1} 行：金额必须大于 0。")

            label = _build_type_label(n_days, m_days, amount, unit)
            cfgs.append(
                {
                    "n_days": n_days,
                    "m_days": m_days,
                    "amount": amount,
                    "unit": unit,
                    "threshold_yuan": _amount_to_yuan(amount, unit),
                    "label": label,
                }
            )
        return cfgs

    def _refresh_type_preview(self) -> None:
        self.type_table.blockSignals(True)
        for r in range(self.type_table.rowCount()):
            try:
                n_txt = (self.type_table.item(r, 0).text().strip() if self.type_table.item(r, 0) else "")
                m_txt = (self.type_table.item(r, 1).text().strip() if self.type_table.item(r, 1) else "")
                a_txt = (self.type_table.item(r, 2).text().strip() if self.type_table.item(r, 2) else "")
                u_txt = (self.type_table.item(r, 3).text().strip() if self.type_table.item(r, 3) else "")
                n_days = int(n_txt) if n_txt else 0
                m_days = int(m_txt) if m_txt else None
                amount = float(a_txt) if a_txt else 0.0
                unit = u_txt if u_txt in ("千万", "亿") else "千万"
                text = _build_type_label(n_days, m_days, amount, unit) if n_days > 0 and amount > 0 else ""
            except Exception:
                text = ""
            self.type_table.setItem(r, 4, QTableWidgetItem(text))
        self.type_table.blockSignals(False)
        self._sync_result_table_headers_with_types()

    def _on_add_type(self) -> None:
        r = self.type_table.rowCount()
        self.type_table.insertRow(r)
        self.type_table.setItem(r, 0, QTableWidgetItem("2"))
        self.type_table.setItem(r, 1, QTableWidgetItem(""))
        self.type_table.setItem(r, 2, QTableWidgetItem("1"))
        self.type_table.setItem(r, 3, QTableWidgetItem("亿"))
        self.type_table.setItem(r, 4, QTableWidgetItem(_build_type_label(2, None, 1, "亿")))
        self._sync_result_table_headers_with_types()

    def _on_delete_type(self) -> None:
        rows = sorted({idx.row() for idx in self.type_table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.type_table.removeRow(r)
        self._sync_result_table_headers_with_types()

    def _load_type_configs(self) -> None:
        cfgs: List[Dict[str, object]] = []
        if os.path.isfile(self._config_path):
            try:
                with open(self._config_path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, list):
                    cfgs = [x for x in raw if isinstance(x, dict)]
            except Exception:
                cfgs = []
        self._set_type_table_rows(cfgs)
        self._sync_result_table_headers_with_types()

    def _on_save_types(self) -> None:
        try:
            cfgs = self._collect_type_configs()
            os.makedirs(os.path.dirname(self._config_path), exist_ok=True)
            to_save = [
                {
                    "n_days": int(c["n_days"]),
                    "m_days": c["m_days"],
                    "amount": float(c["amount"]),
                    "unit": str(c["unit"]),
                }
                for c in cfgs
            ]
            with open(self._config_path, "w", encoding="utf-8") as f:
                json.dump(to_save, f, ensure_ascii=False, indent=2)
            self.status.setText(f"统计类型已保存：{self._config_path}")
        except Exception as e:
            QMessageBox.warning(self, "保存失败", str(e))

    def _on_run(self) -> None:
        qd = self.date_edit.date()
        user_date = f"{qd.year():04d}{qd.month():02d}{qd.day():02d}"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        history_dir = os.path.join(script_dir, "history_data")
        try:
            self._type_configs = self._collect_type_configs()
        except Exception as e:
            if self._auto_run:
                self.status.setText(f"自动运行失败：统计类型错误：{e}")
                QTimer.singleShot(10000, self.accept)
            else:
                QMessageBox.warning(self, "统计类型错误", str(e))
            return
        if not self._type_configs:
            if self._auto_run:
                self.status.setText("自动运行失败：请先添加至少一个统计类型。")
                QTimer.singleShot(10000, self.accept)
            else:
                QMessageBox.information(self, "提示", "请先添加至少一个统计类型。")
            return

        self.btn_run.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.status.setText("正在读取 CSV 并统计，请稍候…")
        self._auto_export_done = False

        self._thread = MainForceAnalyzeThread(
            history_dir, user_date, self._type_configs, parent=self
        )
        self._thread.finished_ok.connect(self._on_done)
        self._thread.failed.connect(self._on_fail)
        self._thread.start()

    def _on_done(self, rows: list, anchor: str, dbg: str) -> None:
        self._rows = rows
        self._anchor = anchor
        if rows:
            self._fill_unified_table(rows)
        else:
            self._apply_flow_headers(self._peek_max_n_from_type_table())
            self.table.setRowCount(0)
        parts: List[str] = []
        for cfg in self._type_configs:
            label = str(cfg.get("label") or "")
            cnt = sum(1 for row in rows if row and row[0] == label)
            parts.append(f"「{label}」{cnt} 条")
        detail = "；".join(parts) if parts else ""
        self.status.setText(
            f"完成。{dbg}\n"
            f"共 {len(rows)} 条。{detail}"
        )
        self.btn_run.setEnabled(True)
        self.btn_export.setEnabled(True)
        if self._auto_run and (not self._auto_export_done):
            self._auto_export_done = True
            self._auto_export_excel_and_png()
            QTimer.singleShot(10000, self.accept)

    def _on_fail(self, msg: str) -> None:
        self.status.setText(f"出错：{msg}")
        self.btn_run.setEnabled(True)
        self.btn_export.setEnabled(False)
        if self._auto_run:
            QTimer.singleShot(10000, self.accept)

    def _export_table_to_png(self, png_path: str) -> bool:
        try:
            nrows, ncols = self.table.rowCount(), self.table.columnCount()
            if nrows <= 0 or ncols <= 0:
                return False
            tmp = QTableWidget(nrows, ncols)
            headers = []
            for c in range(ncols):
                hi = self.table.horizontalHeaderItem(c)
                headers.append(hi.text() if hi else "")
            tmp.setHorizontalHeaderLabels(headers)
            for r in range(nrows):
                for c in range(ncols):
                    it = self.table.item(r, c)
                    new_it = QTableWidgetItem(it.text() if it else "")
                    if it is not None:
                        new_it.setTextAlignment(it.textAlignment())
                        b = it.background()
                        if b is not None:
                            new_it.setBackground(b)
                    tmp.setItem(r, c, new_it)
            tmp.resizeColumnsToContents()
            tmp.resizeRowsToContents()
            total_w = tmp.verticalHeader().width() + 2 * tmp.frameWidth()
            total_h = tmp.horizontalHeader().height() + 2 * tmp.frameWidth()
            for c in range(ncols):
                total_w += tmp.columnWidth(c)
            for r in range(nrows):
                total_h += tmp.rowHeight(r)
            # 原先用 780 下限会把窄表强行拉宽，导出 PNG 右侧一大片空白
            total_w = max(total_w, 80)
            total_h = max(total_h, 60)
            tmp.resize(total_w, total_h)
            tmp.setStyleSheet("QTableWidget { background: #ffffff; }")
            tmp.viewport().setStyleSheet("background: #ffffff;")
            QApplication.processEvents()

            pix = QPixmap(tmp.size())
            pix.fill(Qt.white)
            tmp.render(pix)
            try:
                from PIL import Image

                ba = QByteArray()
                qbuf = QBuffer(ba)
                qbuf.open(QIODevice.WriteOnly)
                if not pix.save(qbuf, "PNG"):
                    qbuf.close()
                    return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
                qbuf.close()
                im = Image.open(io.BytesIO(bytes(ba))).convert("RGB")
                im = _trim_export_png_margins(im)
                im.save(png_path, "PNG")
                return os.path.exists(png_path)
            except Exception:
                return bool(pix.save(png_path, "PNG") and os.path.exists(png_path))
        except Exception:
            return False

    def _export_excel_file(self, *, show_dialogs: bool = True, path: str = "") -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if show_dialogs:
                QMessageBox.warning(
                    self,
                    "缺少依赖",
                    "导出 Excel 需要：pip install openpyxl",
                )
            return ""

        fill_type_a = _qt_color_to_excel_pattern_fill(_ROW_BG_TYPE_A)
        fill_type_b = _qt_color_to_excel_pattern_fill(_ROW_BG_TYPE_B)

        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        default_name = f"主力净流入统计_{self._anchor or 'export'}.xlsx"
        if show_dialogs:
            path, _ = QFileDialog.getSaveFileName(
                self,
                "导出 Excel",
                os.path.join(out_dir, default_name),
                "Excel 工作簿 (*.xlsx);;所有文件 (*.*)",
            )
            if not path:
                return ""
        else:
            if not path:
                path = os.path.join(out_dir, default_name)
            os.makedirs(os.path.dirname(path), exist_ok=True)
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "主力净流入统计"[:31]

        header_font = Font(bold=True)
        header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        # 与界面列一致：类型左、代码中、名称左、起始日期中、各日净流入右
        ncols = len(self.headers)
        col_h_align = ["left", "center", "left", "center"] + ["right"] * (ncols - 4)

        # 第 2 列为「股票代码」，与涨停基因分析导出相同：文本格式 + 零宽空格 + quote_prefix + ignoredErrors
        code_col_1based = 2

        for c, h in enumerate(self.headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = cell_border
            if c == code_col_1based:
                cell.number_format = "@"

        prev_label: Optional[str] = None
        use_alt_palette = False
        for data_idx, tup in enumerate(self._rows):
            r = data_idx + 2
            label = str(tup[0]) if tup else ""
            if data_idx > 0 and label != prev_label:
                use_alt_palette = not use_alt_palette
            prev_label = label
            row_fill = fill_type_b if use_alt_palette else fill_type_a
            for c, val in enumerate(tup, start=1):
                v = val if val is not None else ""
                if c == code_col_1based:
                    v = _excel_stock_cell_str_for_export(str(v))
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = cell_border
                cell.fill = row_fill
                if c == code_col_1based:
                    cell.number_format = "@"
                    qp = getattr(cell, "quote_prefix", None)
                    if qp is not None:
                        try:
                            cell.quote_prefix = True
                        except Exception:
                            pass
                ha = col_h_align[c - 1]
                if "\n" in str(v):
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="top", horizontal=ha
                    )
                else:
                    cell.alignment = Alignment(vertical="center", horizontal=ha)

        n_data = len(self._rows)
        if n_data > 0:
            try:
                from openpyxl.worksheet.cell_range import CellRange
                from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

                letter = get_column_letter(code_col_1based)
                last_row = n_data + 1
                rng = f"{letter}1:{letter}{last_row}"
                ws.ignored_errors = IgnoredErrors(
                    ignoredError=(
                        IgnoredError(sqref=CellRange(rng), numberStoredAsText=True),
                    )
                )
            except Exception:
                pass

        # 第 1 列「统计类型」列宽（openpyxl 列宽约等于字符宽度）
        ws.column_dimensions[get_column_letter(1)].width = 20.0

        try:
            wb.save(path)
        except Exception as e:
            if show_dialogs:
                QMessageBox.warning(self, "导出失败", str(e))
            return ""
        if show_dialogs:
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        return path

    def _on_export(self) -> None:
        self._export_excel_file(show_dialogs=True)

    def _auto_export_excel_and_png(self) -> bool:
        if not self._rows:
            self.status.setText("自动运行完成，但无统计结果，未导出。")
            return False
        script_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(script_dir, "history_data")
        os.makedirs(out_dir, exist_ok=True)
        base = f"主力净流入统计_{self._anchor or 'export'}"
        xlsx_path = os.path.join(out_dir, f"{base}.xlsx")
        png_path = os.path.join(out_dir, f"{base}.png")
        out_xlsx = self._export_excel_file(show_dialogs=False, path=xlsx_path)
        if not out_xlsx:
            self.status.setText("自动导出失败：Excel 导出失败。")
            return False
        ok_png = self._export_table_to_png(png_path)
        if not ok_png:
            self.status.setText("自动导出失败：图片导出失败。")
            return False
        self._auto_last_export_xlsx = out_xlsx
        self._auto_last_export_png = png_path
        self.status.setText(
            f"自动运行完成，已导出：{os.path.basename(out_xlsx)}、{os.path.basename(png_path)}；10秒后退出。"
        )
        return True


def main() -> None:
    parser = argparse.ArgumentParser(description="主力净流入统计（独立版）")
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="启动后自动统计并导出 Excel 和图片，10秒后自动退出",
    )
    args, _unknown = parser.parse_known_args()
    app = QApplication(sys.argv)
    dlg = MainForceNetInflowDialog(auto_run=bool(args.auto_run))
    dlg.show()
    if args.auto_run:
        QTimer.singleShot(200, dlg._on_run)
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

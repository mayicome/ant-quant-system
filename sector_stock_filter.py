#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
板块股票筛选程序
指定一个或多个板块（QMT 申万行业 / 概念等），从这些板块中筛选符合条件的股票
筛选条件：前50个交易日到前10个交易日前有且只有一个涨停，近10个交易日内没有涨停
输出：股票代码、股票名称、所属板块
选股基准日：默认同原逻辑（与“今天”一致时按 15:00 规则）；可指定起止日期，区间内每个交易日各选一次股，每只股票日线只下载一次。
"""

import sys
import argparse
import os
import warnings

# 抑制Qt和log4cplus的警告/错误信息（必须在导入任何Qt或xtquant模块之前设置）
# 设置环境变量抑制Qt Windows版本警告
os.environ.setdefault('QT_LOGGING_RULES', '*.debug=false')

# 抑制Python警告
warnings.filterwarnings("ignore")

# 重定向log4cplus错误（来自第三方库xtquant/QMT）
# 保存原始的stderr
_original_stderr = sys.stderr

class FilteredStderr:
    """过滤stderr，隐藏log4cplus错误和Qt警告"""
    def __init__(self, original_stderr):
        self.original_stderr = original_stderr
    
    def write(self, text):
        # 过滤log4cplus错误信息
        if 'log4cplus' in text.lower() or 'adsyncnamespace' in text.lower():
            return
        # 过滤Qt Windows版本警告
        if 'Qt: Untested Windows version' in text:
            return
        # 过滤 xtdata 连接成功提示及“设置 enable_hello”说明
        if 'xtdata' in text and ('连接成功' in text or 'enable_hello' in text):
            return
        # 其他信息正常输出
        self.original_stderr.write(text)
    
    def flush(self):
        self.original_stderr.flush()
    
    def __getattr__(self, name):
        return getattr(self.original_stderr, name)

# 应用stderr过滤器（必须在导入任何可能产生警告的模块之前）
sys.stderr = FilteredStderr(_original_stderr)

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# 作为脚本启动时：在 pandas/xtdata 等重依赖之前弹出“正在加载”，否则提示出现太晚无意义。
# 作为模块被 import 时不弹。
_EARLY_APP = None
_EARLY_SPLASH = None
_EARLY_SINGLETON_SERVER = None
_SINGLETON_SERVER_NAME = "AntStockFilterSingletonV2"

if __name__ == "__main__":
    try:
        from PyQt5.QtWidgets import QApplication, QLabel
        from PyQt5.QtCore import Qt
        from PyQt5.QtNetwork import QLocalServer, QLocalSocket

        _EARLY_APP = QApplication.instance() or QApplication(sys.argv)

        # 单例：第二实例只激活已有窗口并立即退出，不加载重依赖、不弹 splash
        _sock = QLocalSocket()
        _sock.connectToServer(_SINGLETON_SERVER_NAME)
        if _sock.waitForConnected(200):
            try:
                _sock.write(b"activate")
                _sock.flush()
                _sock.waitForBytesWritten(200)
            except Exception:
                pass
            _sock.disconnectFromServer()
            sys.exit(0)
        _sock.abort()

        _EARLY_SINGLETON_SERVER = QLocalServer()
        try:
            QLocalServer.removeServer(_SINGLETON_SERVER_NAME)
        except Exception:
            pass
        _EARLY_SINGLETON_SERVER.listen(_SINGLETON_SERVER_NAME)

        _EARLY_SPLASH = QLabel("正在加载选股系统")
        _EARLY_SPLASH.setWindowTitle("蚂蚁量化选股系统")
        _EARLY_SPLASH.setAlignment(Qt.AlignCenter)
        _EARLY_SPLASH.setFixedSize(360, 80)
        _EARLY_SPLASH.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.SplashScreen)
        _EARLY_SPLASH.setStyleSheet(
            "background:#FFF8E1; color:#333; font-size:12pt; border:1px solid #FFCC80;"
        )
        _EARLY_SPLASH.show()
        _EARLY_APP.processEvents()
    except Exception:
        _EARLY_APP = None
        _EARLY_SPLASH = None
        _EARLY_SINGLETON_SERVER = None

import pandas as pd
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Tuple, Set
import time
import logging
import json
import csv
import uuid
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
import traceback
import re

# 尽早导入 xtdata 并关闭连接成功提示（必须在首次建立连接前设置，否则会打印“设置xtdata.enable_hello = False可隐藏此消息”）
try:
    import xtquant.xtdata as _xtdata_early
    _xtdata_early.enable_hello = False
except Exception:
    pass

from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QTableWidget, QTableWidgetItem, QLabel, QMessageBox,
                             QProgressBar, QHeaderView, QApplication, QTextEdit, QLineEdit,
                             QListWidget, QListWidgetItem, QCheckBox, QGroupBox, QWidget,
                             QScrollArea, QGridLayout, QSizePolicy, QFileDialog, QDateEdit,
                             QPlainTextEdit, QSplitter, QMenu, QInputDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QDate, QEvent
from PyQt5.QtNetwork import QLocalServer, QLocalSocket
from PyQt5.QtGui import QFont, QIcon

try:
    from utils.stock_info_manager import get_stock_name
except ImportError as e:
    print(f"导入模块失败: {e}")

from utils.daily_cache_reader import get_cache_dir, get_sync_trade_date, load_daily_dataframe
from utils.limit_ratio import is_st_stock, normalize_stock_code
from utils.qmt_sector_store import (
    get_qmt_sector_store,
    load_all_sectors,
    split_sector_tags,
)
from utils.trading_day import (
    get_trading_dates,
    get_trading_dates_in_range_sorted,
    get_trading_dates_set_for_range,
)

# 配置日志（必须在导入其他模块之前配置，以便在导入失败时可以使用logger）
_LOGS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(_LOGS_DIR, exist_ok=True)
logging.basicConfig(
    # 默认 WARNING，显著减少筛选过程中的高频日志写盘；需要详细排查时可设环境变量 SECTOR_FILTER_LOG_LEVEL=INFO
    level=getattr(logging, os.environ.get("SECTOR_FILTER_LOG_LEVEL", "WARNING").upper(), logging.WARNING),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_LOGS_DIR, 'sector_stock_filter.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

try:
    from sector_rule_templates import (
        rule_code_limit_up_after_n_days,
        rule_code_limit_up_with_prior_l_in_l_days,
        rule_code_mode2_old_high,
        rule_code_mode3_volume_reversal,
    )
except ImportError:
    rule_code_limit_up_after_n_days = None  # type: ignore[assignment,misc]
    rule_code_limit_up_with_prior_l_in_l_days = None  # type: ignore[assignment,misc]
    rule_code_mode2_old_high = None  # type: ignore[assignment,misc]
    rule_code_mode3_volume_reversal = None  # type: ignore[assignment,misc]


_RULES_BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
_RULES_DIR = os.path.join(_RULES_BASE_DIR, "sector_rules")
_LEGACY_RULES_PATH = os.path.join(_RULES_BASE_DIR, "sector_stock_rules.json")
_UI_PREFS_PATH = os.path.join(_RULES_BASE_DIR, "sector_filter_ui.json")

_RULE_CODE_UNSAVED_BANNER_STYLE = (
    "QLabel { background-color: #fff3cd; color: #856404; border: 2px solid #ffc107; "
    "padding: 10px 12px; font-weight: bold; font-size: 13px; border-radius: 4px; }"
)
_RULE_CODE_SAVED_BANNER_STYLE = (
    "QLabel { background-color: #d4edda; color: #155724; border: 2px solid #28a745; "
    "padding: 10px 12px; font-weight: bold; font-size: 13px; border-radius: 4px; }"
)
_RULE_SAVE_BTN_DIRTY_STYLE = (
    "QPushButton { background-color: #ff9800; color: white; font-weight: bold; "
    "padding: 8px 16px; border: 2px solid #e65100; border-radius: 4px; }"
    "QPushButton:hover { background-color: #fb8c00; }"
)

# 内置模式一/二/三窗口参数默认值：仅在规则代码顶部未定义 N、M 等常量时使用（见各规则代码前注释）
SECTOR_RULE_DEFAULT_N = 4
SECTOR_RULE_DEFAULT_M = 100
SECTOR_RULE_DEFAULT_N_MODE3 = 4
SECTOR_RULE_DEFAULT_M_MODE3 = 100
SECTOR_RULE_DEFAULT_L_MODE3 = 10


def _rule_param_code_block() -> str:
    """规则代码顶部应出现的窗口参数块（与 SECTOR_RULE_DEFAULT_* 一致，可编辑）。"""
    return (
        "# 本规则独立参数：仅影响本条规则；选股引擎与 ctx['params'] 使用该组数值\n"
        f"N = {SECTOR_RULE_DEFAULT_N}\n"
        f"M = {SECTOR_RULE_DEFAULT_M}\n"
        f"N_MODE3 = {SECTOR_RULE_DEFAULT_N_MODE3}\n"
        f"M_MODE3 = {SECTOR_RULE_DEFAULT_M_MODE3}\n"
        f"L_MODE3 = {SECTOR_RULE_DEFAULT_L_MODE3}\n"
        "\n"
    )


def _rule_code_has_top_level_N(code: str) -> bool:
    """是否已有顶层 N = … 赋值（避免重复插入参数块）。"""
    return bool(re.search(r"^\s*N\s*=\s*", code or "", re.MULTILINE))


def ensure_rule_code_has_param_preamble(code: str) -> str:
    """若规则代码中未定义顶层 N，则在开头插入标准参数块（加载旧规则时用）。"""
    c = (code or "").lstrip("\ufeff")
    if _rule_code_has_top_level_N(c):
        return c
    return _rule_param_code_block() + c


def _rule_file_path(rule_id: str) -> str:
    rid = str(rule_id or "").strip()
    if not rid:
        rid = str(uuid.uuid4())
    return os.path.join(_RULES_DIR, f"{rid}.json")


def _sanitize_rule_filename(name: str) -> str:
    """
    Windows 文件名禁止字符：\\ / : * ? " < > |，以及末尾的空格/点。
    这里做最小替换，保证可读性。
    """
    s = str(name or "").strip()
    if not s:
        return "未命名规则"
    bad = '\\/:*?"<>|'
    for ch in bad:
        s = s.replace(ch, "_")
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = " ".join(s.split())
    s = s.rstrip(" .")
    return s or "未命名规则"


def _rule_file_path_by_rule(rule: Dict[str, object]) -> str:
    rid = str(rule.get("id") or "").strip()
    if not rid:
        rid = str(uuid.uuid4())
        rule["id"] = rid
    name = _sanitize_rule_filename(str(rule.get("name") or "规则"))
    short_id = rid.split("-")[0] if "-" in rid else rid[:8]
    # 用 “名称__短ID.json” 避免同名覆盖，同时文件名仍可读
    return os.path.join(_RULES_DIR, f"{name}__{short_id}.json")


def _normalize_rule_name(name: str) -> str:
    """规则显示名：去首尾空白、合并连续空白。"""
    s = " ".join(str(name or "").strip().split())
    return s or "未命名规则"


def _collect_used_rule_names(
    rules: List[Dict[str, object]], exclude_rule_id: Optional[str] = None
) -> set:
    used = set()
    for r in rules or []:
        rid = str(r.get("id") or "")
        if exclude_rule_id and rid == str(exclude_rule_id):
            continue
        used.add(_normalize_rule_name(str(r.get("name") or "")))
    return used


def rule_name_is_taken(
    name: str, rules: List[Dict[str, object]], exclude_rule_id: Optional[str] = None
) -> bool:
    """系统内是否已有同名规则（比较归一化后的名称）。"""
    return _normalize_rule_name(name) in _collect_used_rule_names(rules, exclude_rule_id)


def make_unique_rule_name(
    desired_name: str, rules: List[Dict[str, object]], exclude_rule_id: Optional[str] = None
) -> str:
    """生成不重复的规则名；重复时自动追加 (2)/(3)/...（用于新增/复制/导入）。"""
    base = _normalize_rule_name(desired_name)
    used = _collect_used_rule_names(rules, exclude_rule_id)
    if base not in used:
        return base
    i = 2
    while True:
        cand = f"{base}({i})"
        if cand not in used:
            return cand
        i += 1


def _ensure_unique_rule_names_in_list(rules: List[Dict[str, object]]) -> bool:
    """加载多条规则后，保证名称互不重复；若有调整返回 True。"""
    changed = False
    used: set = set()
    for r in rules or []:
        if not isinstance(r, dict):
            continue
        name = _normalize_rule_name(str(r.get("name") or ""))
        if name in used:
            name = make_unique_rule_name(name, rules)
            r["name"] = name
            changed = True
        else:
            r["name"] = name
        used.add(name)
    return changed


def _legacy_rule_code_uses_external_bridge(code: str) -> bool:
    """是否仍依赖 ctx['builtin_check'] / mode1_diagnose 等引擎外挂逻辑。"""
    c = str(code or "")
    return (
        "builtin_check" in c
        or "mode1_diagnose" in c
        or "run_mode1" in c
    )


def _upgrade_legacy_rule_code(rule: Dict[str, object]) -> bool:
    """将仍调用 builtin_check 等的旧规则代码替换为自包含实现。返回是否已升级。"""
    code = str(rule.get("code") or "")
    if not _legacy_rule_code_uses_external_bridge(code):
        return False
    if not all(
        callable(x)
        for x in (rule_code_limit_up_after_n_days, rule_code_mode2_old_high, rule_code_mode3_volume_reversal)
    ):
        return False
    name = str(rule.get("name") or "")
    if name == "模式三" or "模式三" in name:
        rule["code"] = rule_code_mode3_volume_reversal()
    elif name == "模式二" or "模式二" in name:
        rule["code"] = rule_code_mode2_old_high()
    else:
        rule["code"] = rule_code_limit_up_after_n_days()
    return True


def _upgrade_prior_l_rule_code(rule: Dict[str, object]) -> bool:
    """为含 L 日前序涨停检查但缺少 REQUIRE_PRIOR_LU_IN_L 的旧规则补丁升级。"""
    code = str(rule.get("code") or "")
    if "has_prior_lu_in_l" not in code or "REQUIRE_PRIOR_LU_IN_L" in code:
        return False
    new_code = code
    if not re.search(r"^L\s*=", new_code, re.MULTILINE):
        return False
    new_code = re.sub(
        r"^(L\s*=.*)$",
        r"\1\nREQUIRE_PRIOR_LU_IN_L = True  "
        r"# True=L日内至少一个涨停；False=L日内全无涨停",
        new_code,
        count=1,
        flags=re.MULTILINE,
    )
    old_block = (
        "            if not has_prior_lu_in_l:\n"
        "                continue"
    )
    new_block = (
        "            if REQUIRE_PRIOR_LU_IN_L:\n"
        "                if not has_prior_lu_in_l:\n"
        "                    continue\n"
        "            else:\n"
        "                if has_prior_lu_in_l:\n"
        "                    continue"
    )
    if old_block not in new_code:
        return False
    rule["code"] = new_code.replace(old_block, new_block, 1)
    return True


def _default_rules() -> List[Dict[str, object]]:
    """
    预置规则：逻辑全部写在规则代码内，不调用引擎外部函数。
    须定义 select(stock_code, stock_name, sectors, daily_data, as_of_date, ctx) -> (bool, extra_dict)。
    """
    if not all(
        callable(x)
        for x in (rule_code_limit_up_after_n_days, rule_code_mode2_old_high, rule_code_mode3_volume_reversal)
    ):
        raise RuntimeError("缺少 sector_rule_templates 模块，无法生成默认规则")
    return [
        {
            "id": str(uuid.uuid4()),
            "name": "涨停后N日",
            "enabled": True,
            "code": rule_code_limit_up_after_n_days(),
        },
        {
            "id": str(uuid.uuid4()),
            "name": "模式二",
            "enabled": True,
            "code": rule_code_mode2_old_high(),
        },
        {
            "id": str(uuid.uuid4()),
            "name": "模式三",
            "enabled": False,
            "code": rule_code_mode3_volume_reversal(),
        },
    ]


def load_sector_rules() -> List[Dict[str, object]]:
    try:
        os.makedirs(_RULES_DIR, exist_ok=True)

        # 1) 若存在旧版单文件，先迁移到每条规则一个文件
        if os.path.exists(_LEGACY_RULES_PATH):
            try:
                with open(_LEGACY_RULES_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                rules = data.get("rules") if isinstance(data, dict) else None
                if isinstance(rules, list) and rules:
                    migrated = []
                    for r in rules:
                        if not isinstance(r, dict):
                            continue
                        rid = str(r.get("id") or uuid.uuid4())
                        rr = {
                            "id": rid,
                            "name": str(r.get("name") or f"规则-{rid[:8]}"),
                            "enabled": bool(r.get("enabled", True)),
                            "code": ensure_rule_code_has_param_preamble(str(r.get("code") or "")),
                        }
                        migrated.append(rr)
                    if migrated:
                        save_sector_rules(migrated)
                        # 迁移完成后保留旧文件为 .bak（避免误删）
                        try:
                            os.replace(_LEGACY_RULES_PATH, _LEGACY_RULES_PATH + ".bak")
                        except Exception:
                            pass
            except Exception as e:
                logger.warning(f"迁移旧规则文件失败，将继续按新目录加载: {e}", exc_info=True)

        # 2) 从目录加载所有规则文件
        out: List[Dict[str, object]] = []
        upgraded_any = False
        try:
            for fn in os.listdir(_RULES_DIR):
                if not fn.lower().endswith(".json"):
                    continue
                p = os.path.join(_RULES_DIR, fn)
                try:
                    with open(p, "r", encoding="utf-8") as f:
                        r = json.load(f)
                    if not isinstance(r, dict):
                        continue
                    rid = str(r.get("id") or os.path.splitext(fn)[0] or uuid.uuid4())
                    rr = {
                        "id": rid,
                        "name": str(r.get("name") or f"规则-{rid[:8]}"),
                        "enabled": bool(r.get("enabled", True)),
                        "code": ensure_rule_code_has_param_preamble(str(r.get("code") or "")),
                    }
                    if _upgrade_legacy_rule_code(rr):
                        upgraded_any = True
                        logger.info(f"已升级旧规则代码为自包含实现：{rr.get('name')}")
                    elif _upgrade_prior_l_rule_code(rr):
                        upgraded_any = True
                        logger.info(f"已为 L 日规则补充 REQUIRE_PRIOR_LU_IN_L 开关：{rr.get('name')}")
                    out.append(rr)
                except Exception:
                    continue
        except Exception:
            out = []

        if out:
            names_changed = _ensure_unique_rule_names_in_list(out)
            if names_changed or upgraded_any:
                save_sector_rules(out)
            # 稳定排序：启用优先，其次按名称
            out.sort(key=lambda r: (not bool(r.get("enabled", True)), str(r.get("name") or "")))
            return out

        # 目录为空：写入默认规则
        rules = _default_rules()
        save_sector_rules(rules)
        return rules
    except Exception as e:
        logger.warning(f"加载选股规则失败，将使用默认规则: {e}", exc_info=True)
        return _default_rules()


def save_sector_rules(rules: List[Dict[str, object]]) -> None:
    try:
        os.makedirs(_RULES_DIR, exist_ok=True)
        keep_ids = set()
        for r in (rules or []):
            if not isinstance(r, dict):
                continue
            rid = str(r.get("id") or uuid.uuid4())
            r["id"] = rid
            keep_ids.add(rid)
            p = _rule_file_path_by_rule(r)
            payload = {
                "id": rid,
                "name": str(r.get("name") or f"规则-{rid[:8]}"),
                "enabled": bool(r.get("enabled", True)),
                "code": str(r.get("code") or ""),
            }
            with open(p, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)

        # 清理目录里不在 keep_ids 的旧规则文件（用户删除规则后同步删除）
        for fn in os.listdir(_RULES_DIR):
            if not fn.lower().endswith(".json"):
                continue
            fp = os.path.join(_RULES_DIR, fn)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    rr = json.load(f)
                fid = str(rr.get("id") or "")
            except Exception:
                fid = ""
            if fid and fid not in keep_ids:
                try:
                    os.remove(fp)
                except Exception:
                    pass
    except Exception as e:
        logger.warning(f"保存选股规则失败: {e}", exc_info=True)


def save_single_sector_rule(rule: Dict[str, object]) -> None:
    """只保存一条规则，并清理同 id 的旧文件名残留。"""
    try:
        os.makedirs(_RULES_DIR, exist_ok=True)
        if not isinstance(rule, dict):
            return
        rid = str(rule.get("id") or uuid.uuid4())
        rule["id"] = rid
        target_path = _rule_file_path_by_rule(rule)

        # 清理同 id 的旧文件（改名会导致文件名变化）
        for fn in os.listdir(_RULES_DIR):
            if not fn.lower().endswith(".json"):
                continue
            fp = os.path.join(_RULES_DIR, fn)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    rr = json.load(f)
                fid = str(rr.get("id") or "")
            except Exception:
                continue
            if fid == rid and os.path.abspath(fp) != os.path.abspath(target_path):
                try:
                    os.remove(fp)
                except Exception:
                    pass

        payload = {
            "id": rid,
            "name": str(rule.get("name") or f"规则-{rid[:8]}"),
            "enabled": bool(rule.get("enabled", True)),
            "code": str(rule.get("code") or ""),
        }
        with open(target_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"保存单条选股规则失败: {e}", exc_info=True)


def delete_sector_rule_files(rule_id: str) -> int:
    """删除某条规则在 sector_rules 目录下的所有文件（兼容旧文件名残留）。返回删除数量。"""
    try:
        rid = str(rule_id or "").strip()
        if not rid:
            return 0
        if not os.path.isdir(_RULES_DIR):
            return 0
        removed = 0
        for fn in os.listdir(_RULES_DIR):
            if not fn.lower().endswith(".json"):
                continue
            fp = os.path.join(_RULES_DIR, fn)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    rr = json.load(f)
                fid = str(rr.get("id") or "")
            except Exception:
                continue
            if fid == rid:
                try:
                    os.remove(fp)
                    removed += 1
                except Exception:
                    pass
        return removed
    except Exception:
        return 0


def _coerce_mode12_screen_return(
    r: object,
) -> Tuple[bool, Optional[date], Optional[date]]:
    """模式一/二共用筛选函数须返回三元组；兼容历史或异常路径误返回二元组的情况。"""
    if isinstance(r, tuple):
        if len(r) == 2:
            return r[0], r[1], None
        if len(r) >= 3:
            return r[0], r[1], r[2]
    return False, None, None


# 交易日历：utils.trading_day（xtdata/akshare 缓存，不依赖 MiniQMT 日线）


# 导入主力净流入添加工具
try:
    from auto_add_inflow import InflowAdder
    logger.info("成功导入主力净流入工具")
except ImportError as e:
    logger.warning(f"无法导入主力净流入工具: {e}")
    InflowAdder = None
except Exception as e:
    logger.error(f"导入主力净流入工具时出错: {e}", exc_info=True)
    InflowAdder = None

# 导入“当日最多涨停概念 / 数量 / 排名”列的工具函数（与 auto_limit_up_filter 复用同一实现）
try:
    from auto_limit_up_filter import add_concept_rank_columns
    logger.info("成功导入概念排名工具 add_concept_rank_columns")
except ImportError as e:
    logger.warning(f"无法导入 add_concept_rank_columns: {e}")
    add_concept_rank_columns = None
except Exception as e:
    logger.error(f"导入 add_concept_rank_columns 时出错: {e}", exc_info=True)
    add_concept_rank_columns = None


def _normalize_import_stock_code(raw) -> Optional[str]:
    """从文件/Excel 单元格解析 6 位 A 股代码。Excel 常把 000628 读成 628，需 zfill。"""
    if raw is None:
        return None
    try:
        if pd.isna(raw):
            return None
    except (TypeError, ValueError):
        pass
    code_str = str(raw).strip()
    if not code_str or code_str.lower() in ("nan", "none"):
        return None
    if re.fullmatch(r"\d+\.0+", code_str):
        code_str = code_str.split(".", 1)[0]
    m = re.search(r'["\']?(\d{6})["\']?', code_str)
    if m:
        digits = m.group(1)
    else:
        digits = "".join(c for c in code_str if c.isdigit())
    if not digits or len(digits) > 6:
        return None
    code = digits.zfill(6)
    if code == "000000":
        return None
    return code


def _stock_code6(stock_code: object) -> str:
    """选股列表代码 → 6 位；失败返回空串。"""
    c = _normalize_import_stock_code(stock_code)
    return c or ""


def _enabled_rules_ctx_needs(enabled_rules: List[Dict[str, object]]) -> Dict[str, object]:
    """根据启用规则代码推断本轮需要哪些 ctx / 东财臂 / Elig 收窄。

    选股阶段只加载规则判断所需上下文；净流入展示等在命中后再补。
    """
    need_inflow = False
    need_hot_theme = False
    need_em = False
    arms: Set[str] = set()
    elig_bands: List[Tuple[int, int]] = []
    em_top_ns: List[int] = []
    em_rs_top_ks: List[int] = []
    em_min_members: List[int] = []
    for r in enabled_rules or []:
        code = str(r.get("code") or "")
        if "inflow_rank" in code:
            need_inflow = True
        if "hot_theme" in code:
            need_hot_theme = True
        if "em_board_hot" not in code:
            continue
        need_em = True
        if (
            "today_pool_codes" in code
            or "today_code_hits" in code
            or "合格榜内序位" in code
        ):
            arms.add("today")
        if "new_only_pool" in code or "new_only_code_hits" in code:
            arms.add("new_only")
        # 连续臂：用裸 pool_codes / 裸 code_hits（非 today_/new_only_ 前缀）
        if re.search(
            r"(?<![\\w])pool_codes(?![\\w])|(?<![\\w])code_hits(?![\\w])",
            code,
        ):
            # 排除 today_pool_codes / today_code_hits / new_only_* 已覆盖的情况：
            # 若仅出现带前缀字段则不算连续臂
            has_bare_pool = bool(
                re.search(r"(?<![a-z_])pool_codes(?![a-z_])", code, flags=re.I)
            )
            has_bare_hits = bool(
                re.search(r"(?<![a-z_])code_hits(?![a-z_])", code, flags=re.I)
            )
            if has_bare_pool or has_bare_hits:
                arms.add("continuous")
        m_lo = re.search(r"ELIG_LO\s*=\s*(\d+)", code)
        m_hi = re.search(r"ELIG_HI\s*=\s*(\d+)", code)
        if m_lo and m_hi:
            lo, hi = int(m_lo.group(1)), int(m_hi.group(1))
            if lo >= 1 and hi >= lo:
                elig_bands.append((lo, hi))
        m_top = re.search(r"TOP_N\s*=\s*(\d+)", code)
        if m_top:
            em_top_ns.append(int(m_top.group(1)))
        m_rs = re.search(r"RS_TOP_K\s*=\s*(\d+)", code)
        if m_rs:
            em_rs_top_ks.append(int(m_rs.group(1)))
        m_rs_hi = re.search(r"RS_HI\s*=\s*(\d+)", code)
        if m_rs_hi:
            em_rs_top_ks.append(int(m_rs_hi.group(1)))
        m_mm = re.search(r"MIN_MEMBERS\s*=\s*(\d+)", code)
        if m_mm:
            em_min_members.append(int(m_mm.group(1)))
    if need_em and not arms:
        arms = {"continuous", "new_only", "today"}
    # 多规则并存：取最宽覆盖（更大 TopN/RS；更低成分门槛），避免漏掉任一规则所需池
    return {
        "inflow": need_inflow,
        "hot_theme": need_hot_theme,
        "em": need_em,
        "em_arms": arms,
        "elig_bands": elig_bands,
        "em_top_n": max(em_top_ns) if em_top_ns else 50,
        "em_rs_top_k": max(em_rs_top_ks) if em_rs_top_ks else 20,
        "em_min_members": min(em_min_members) if em_min_members else 30,
    }


def _em_candidate_codes6(
    emh: Dict[str, object],
    *,
    arms: Set[str],
    elig_bands: List[Tuple[int, int]],
) -> Set[str]:
    """从已加载的东财热门 ctx 取出本轮候选 code6。"""
    out: Set[str] = set()
    if not isinstance(emh, dict) or str(emh.get("error") or "").strip():
        return out
    if "continuous" in arms:
        out |= {str(c).zfill(6) for c in (emh.get("pool_codes") or set()) if c}
    if "new_only" in arms:
        out |= {str(c).zfill(6) for c in (emh.get("new_only_pool_codes") or set()) if c}
    if "today" in arms:
        hits = emh.get("today_code_hits") or {}
        pool = emh.get("today_pool_codes") or set()
        if elig_bands and isinstance(hits, dict) and hits:
            for c6, hit in hits.items():
                if not isinstance(hit, dict):
                    continue
                try:
                    elig = int(hit.get("合格榜内序位") or 0)
                except (TypeError, ValueError):
                    elig = 0
                if any(lo <= elig <= hi for lo, hi in elig_bands):
                    out.add(str(c6).zfill(6))
        else:
            out |= {str(c).zfill(6) for c in pool if c}
    return out


def _safe_rank_int(v: object, default: int = 10**9) -> int:
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except (TypeError, ValueError):
        return default


# 导出综合分：score = Elig * 权重 + 标签内RS（越小越靠前）
# 权重=8：约「差 1 档 Elig ≈ 差 8 名 RS」；若希望 Elig 几乎不被 RS 反超，可调到 50+
EXPORT_ELIG_WEIGHT = 8


def reorder_selection_rows_for_export(
    rows: List[Dict[str, object]],
    *,
    max_per_tag: int = 2,
    elig_weight: int = EXPORT_ELIG_WEIGHT,
) -> List[Dict[str, object]]:
    """导出前按选股日重排，便于策略生成器按池顺序做 clip 截断。

    每个选股日内：
    1) 按综合分 score=Elig*elig_weight+标签内RS 升序（再 Elig、RS、代码 tie-break）；
    2) 同「合格榜对应标签」优先最多保留 max_per_tag 只排在前面，多出的同标签票放到该日末尾。
    无 Elig 字段的规则：保持该日内相对顺序不变。
    """
    if not rows:
        return []
    max_per_tag = max(0, int(max_per_tag))
    try:
        w = max(0, int(elig_weight))
    except (TypeError, ValueError):
        w = int(EXPORT_ELIG_WEIGHT)

    def _as_of(r: Dict[str, object]) -> str:
        return str(r.get("as_of") or r.get("选股日") or "").strip()

    def _code(r: Dict[str, object]) -> str:
        c = str(r.get("code") or r.get("股票代码") or "").strip()
        if c.isdigit():
            return c.zfill(6)
        return c

    def _tag(r: Dict[str, object]) -> str:
        t = str(r.get("合格榜对应标签") or "").strip()
        return t if t else "_none_"

    def _has_elig(r: Dict[str, object]) -> bool:
        return ("合格榜内序位" in r) and r.get("合格榜内序位") not in (None, "")

    def _score(r: Dict[str, object]) -> int:
        elig = _safe_rank_int(r.get("合格榜内序位"))
        rs = _safe_rank_int(r.get("合格榜标签内RS排名"))
        return elig * w + rs

    by_day: Dict[str, List[Dict[str, object]]] = {}
    day_order: List[str] = []
    for r in rows:
        d = _as_of(r) or "_nodate_"
        if d not in by_day:
            by_day[d] = []
            day_order.append(d)
        by_day[d].append(r)

    out: List[Dict[str, object]] = []
    for d in day_order:
        day_rows = by_day[d]
        if not any(_has_elig(r) for r in day_rows):
            out.extend(day_rows)
            continue
        ranked = sorted(
            day_rows,
            key=lambda r: (
                _score(r),
                _safe_rank_int(r.get("合格榜内序位")),
                _safe_rank_int(r.get("合格榜标签内RS排名")),
                _code(r),
            ),
        )
        if max_per_tag <= 0:
            out.extend(ranked)
            continue
        head: List[Dict[str, object]] = []
        tail: List[Dict[str, object]] = []
        tag_n: Dict[str, int] = {}
        for r in ranked:
            tag = _tag(r)
            n = int(tag_n.get(tag) or 0)
            if n < max_per_tag:
                head.append(r)
                tag_n[tag] = n + 1
            else:
                tail.append(r)
        out.extend(head)
        out.extend(tail)
    return out


def save_excel_with_text_code(excel_file_path: str, df: pd.DataFrame):
    """保存Excel文件，确保股票代码列是文本格式（支持列名为 'code' 或 '股票代码'）
    
    Args:
        excel_file_path: Excel文件路径
        df: 要保存的DataFrame
    """
    # 股票代码列名：优先 code，否则 股票代码
    code_col_name = 'code' if 'code' in df.columns else ('股票代码' if '股票代码' in df.columns else None)

    def clean_code(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ''
        code_str = str(val).strip()
        if '.' in code_str:
            code_str = code_str.split('.')[0]
        if code_str.isdigit():
            return code_str.zfill(6)
        return code_str

    try:
        if code_col_name:
            df = df.copy()
            df[code_col_name] = df[code_col_name].apply(clean_code)
            df[code_col_name] = df[code_col_name].astype(str)

        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            code_col_idx = None
            if code_col_name:
                for idx, col_name in enumerate(df.columns, start=1):
                    if col_name == code_col_name:
                        code_col_idx = idx
                        break

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if code_col_idx and c_idx == code_col_idx and r_idx > 1:
                        cell.number_format = '@'
                        cell.value = clean_code(value)
                    else:
                        cell.value = value

            wb.save(excel_file_path)
            wb.close()
        except Exception as e:
            logger.warning(f"使用openpyxl直接写入失败: {str(e)}，尝试回退方法")
            df.to_excel(excel_file_path, index=False, engine='openpyxl')

            if code_col_name:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_file_path)
                    ws = wb.active
                    for col_idx, header in enumerate(df.columns, start=1):
                        if header == code_col_name:
                            for row_idx in range(2, len(df) + 2):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.number_format = '@'
                                cell.value = clean_code(cell.value)
                            break
                    wb.save(excel_file_path)
                    wb.close()
                except Exception as e2:
                    logger.warning(f"设置Excel文本格式时出错: {str(e2)}，但文件已保存")

    except Exception as e:
        raise Exception(f"保存Excel文件失败: {str(e)}")


def save_xls_with_text_code(xls_file_path: str, df: pd.DataFrame):
    """保存 .xls 文件（使用 xlwt 原生写入，避免 pandas 对 xlwt writer 的兼容问题）。"""
    code_col_name = 'code' if 'code' in df.columns else ('股票代码' if '股票代码' in df.columns else None)

    def clean_code(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ''
        code_str = str(val).strip()
        if '.' in code_str:
            code_str = code_str.split('.')[0]
        if code_str.isdigit():
            return code_str.zfill(6)
        return code_str

    try:
        import xlwt
    except Exception as e:
        raise Exception(f"写入 .xls 需要 xlwt：{e}")

    try:
        data = df.copy()
        if code_col_name:
            data[code_col_name] = data[code_col_name].apply(clean_code).astype(str)

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sheet1')
        text_style = xlwt.easyxf(num_format_str='@')

        for c_idx, col in enumerate(list(data.columns)):
            ws.write(0, c_idx, col)

        code_col_idx = None
        if code_col_name and code_col_name in data.columns:
            code_col_idx = list(data.columns).index(code_col_name)

        for r_idx, row in enumerate(data.itertuples(index=False, name=None), start=1):
            for c_idx, val in enumerate(row):
                v = '' if val is None else val
                if code_col_idx is not None and c_idx == code_col_idx:
                    ws.write(r_idx, c_idx, clean_code(v), text_style)
                else:
                    ws.write(r_idx, c_idx, v)

        wb.save(xls_file_path)
    except Exception as e:
        raise Exception(f"保存 .xls 失败: {e}")


def add_stock_info_columns(file_path: str) -> bool:
    """为 CSV 添加 QMT 概念/行业/细分板块列（基于 qmt_sector_index 缓存）。"""
    try:
        store = get_qmt_sector_store()
        store.ensure_inverted_index()
        logger.info("已加载 QMT 板块反查索引，共 %d 只股票", store.indexed_stock_count())
        
        # 检测文件编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']
        encoding = None
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    f.readline()
                encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if encoding is None:
            logger.error("无法检测文件编码")
            return False
        
        logger.info(f"检测到文件编码: {encoding}")
        
        # 读取CSV文件
        rows = []
        fieldnames = None
        
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames) if reader.fieldnames else []
            
            for row in reader:
                rows.append(row)
        
        if not rows:
            logger.warning("文件中没有数据行")
            return False
        
        logger.info(f"读取到 {len(rows)} 行数据")
        
        # 添加新列（如果不存在）
        new_columns = ['概念', '行业', '板块']
        for col in new_columns:
            if col not in fieldnames:
                fieldnames.append(col)
                logger.info(f"添加新列: {col}")
        
        def normalize_code_for_lookup(raw):
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                return ''
            s = str(raw).strip()
            if '.' in s and s.split('.')[0].isdigit():
                s = s.split('.')[0]
            if s.replace('.', '').isdigit():
                s = s.split('.')[0].zfill(6)
            return s

        filled_count = 0
        empty_count = 0

        for row in rows:
            code = normalize_code_for_lookup(row.get('code', ''))
            if code:
                tags = split_sector_tags(store.sectors_for_stock(code))
                row['概念'] = ';'.join(tags['concepts'])
                row['行业'] = ';'.join(tags['industries'])
                row['板块'] = ';'.join(tags['sub_sectors'])
                if tags['concepts'] or tags['industries'] or tags['sub_sectors']:
                    filled_count += 1
                else:
                    empty_count += 1
            else:
                row['概念'] = ''
                row['行业'] = ''
                row['板块'] = ''
                empty_count += 1
        
        logger.info(f"填充完成：成功 {filled_count} 条，未找到 {empty_count} 条")
        
        # 保存回原文件
        logger.info(f"正在保存文件...")
        with open(file_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        logger.info(f"完成！文件已保存: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"添加股票信息列时出错: {str(e)}", exc_info=True)
        return False


class SectorStockFilterThread(QThread):
    """板块股票筛选线程"""
    
    # 信号：更新进度
    progress_updated = pyqtSignal(int, int, str)  # (当前数量, 总数, 当前股票代码)
    # 信号：找到符合条件的股票（规则驱动；emit dict）
    stock_found = pyqtSignal(object)
    # 信号：筛选完成
    finished = pyqtSignal(object)  # dict {rule_id: count}
    # 信号：错误信息
    error_occurred = pyqtSignal(str)  # (错误信息)
    # 信号：调试信息（用于显示统计）
    debug_info = pyqtSignal(str)  # (调试信息)
    
    def __init__(self, stock_list: List[Tuple[str, str, str]], n: int, m_mode1: int,
                 n_mode2: int, m_mode2: int, n_mode3: int, m_mode3: int, l_mode3: int,
                 rules: Optional[List[Dict[str, object]]] = None, as_of_date: Optional[date] = None,
                 as_of_end_date: Optional[date] = None, parent=None):
        super().__init__(parent)
        self.stock_list = stock_list
        self._n = n
        self._m_mode1 = m_mode1
        self._n_mode2 = n_mode2
        self._m_mode2 = m_mode2
        self._n_mode3 = n_mode3
        self._m_mode3 = m_mode3
        self._l_mode3 = l_mode3
        self.rules = rules if rules else []
        # 选股基准日：None 表示按程序内“当前日”逻辑；指定则用于回测（日线与交易日窗口均截至该日）
        self._as_of_date = as_of_date
        # 若指定且与起始日不同，则区间内每个交易日各跑一次（日线只拉取一次）
        self._as_of_end_date = as_of_end_date
        self.trading_dates_50 = []   # 在 run() 中由 get_trading_dates 填充
        self.trading_dates_10 = []  # 兼容保留
        self.trading_dates_n_mode2 = []
        self.trading_dates_n_mode3 = []
        self.trading_days_m_mode2 = m_mode2
        self.trading_days_m_mode3 = m_mode3
        self.trading_days_l_mode3 = l_mode3
        self.trading_days_m_mode1 = m_mode1
        self.is_running = True
        
    def stop(self):
        """停止筛选"""
        self.is_running = False

    _RULE_PARAM_KEYS = ("N", "M", "L", "P", "N_MODE2", "M_MODE2", "N_MODE3", "M_MODE3", "L_MODE3")
    _RULE_EXPORT_PARAM_KEYS = ("P", "N", "M", "L")

    @staticmethod
    def _format_rule_param_for_export(value: object) -> str:
        if isinstance(value, bool):
            return "True" if value else "False"
        return str(value)

    @classmethod
    def _extract_rule_export_params(cls, locals_dict: Dict[str, object]) -> Dict[str, str]:
        """从规则代码命名空间读取 P/N/M/L 配置，供选股结果保存（REQUIRE_* 由规则命中时写入实际 True/False）。"""
        out: Dict[str, str] = {}
        for key in cls._RULE_EXPORT_PARAM_KEYS:
            if key not in locals_dict:
                continue
            out[key] = cls._format_rule_param_for_export(locals_dict[key])
        return out

    @staticmethod
    def _extract_rule_constants(locals_dict: Dict[str, object]) -> Dict[str, int]:
        """从规则 exec 命名空间中读取内置窗口常量（整数 1~500）。"""
        out: Dict[str, int] = {}
        for key in SectorStockFilterThread._RULE_PARAM_KEYS:
            if key not in locals_dict:
                continue
            try:
                v = int(locals_dict[key])  # type: ignore[arg-type]
                if 1 <= v <= 500:
                    out[key] = v
            except (TypeError, ValueError):
                pass
        if "M" not in out and "M_MODE1" in locals_dict:
            try:
                v = int(locals_dict["M_MODE1"])  # type: ignore[arg-type]
                if 1 <= v <= 500:
                    out["M"] = v
            except (TypeError, ValueError):
                pass
        return out

    def _merge_rule_calendar_params(self, extracted: Dict[str, int]) -> Dict[str, int]:
        """合并为线程内部使用的 _n / _m_mode1 / … 字典；未在代码中写的项沿用线程当前默认值。"""
        N = extracted.get("N", self._n)
        M1 = extracted.get("M", self._m_mode1)
        N2 = extracted.get("N_MODE2", N)
        M2 = extracted.get("M_MODE2", M1)
        N3 = extracted.get("N_MODE3", self._n_mode3)
        M3 = extracted.get("M_MODE3", self._m_mode3)
        L3 = extracted.get("L_MODE3", self._l_mode3)
        return {
            "_n": N,
            "_m_mode1": M1,
            "_n_mode2": N2,
            "_m_mode2": M2,
            "_n_mode3": N3,
            "_m_mode3": M3,
            "_l_mode3": L3,
        }

    def _apply_rule_calendar_params(self, p: Dict[str, int]) -> None:
        self._n = p["_n"]
        self._m_mode1 = p["_m_mode1"]
        self._n_mode2 = p["_n_mode2"]
        self._m_mode2 = p["_m_mode2"]
        self._n_mode3 = p["_n_mode3"]
        self._m_mode3 = p["_m_mode3"]
        self._l_mode3 = p["_l_mode3"]
        self.trading_days_m_mode1 = p["_m_mode1"]
        self.trading_days_m_mode2 = p["_m_mode2"]
        self.trading_days_m_mode3 = p["_m_mode3"]
        self.trading_days_l_mode3 = p["_l_mode3"]

    def _current_rule_param_ctx(self) -> Dict[str, int]:
        """代码顶部常量经合并、缺省补全后，当前规则实际使用的窗口参数；与 builtin_check 内部逻辑一致。"""
        return {
            "N": int(self._n),
            "M": int(self._m_mode1),
            "N_MODE2": int(self._n_mode2),
            "M_MODE2": int(self._m_mode2),
            "N_MODE3": int(self._n_mode3),
            "M_MODE3": int(self._m_mode3),
            "L_MODE3": int(self._l_mode3),
        }
    
    def _get_limit_ratio(self, stock_code: str, as_of_date=None) -> float:
        """获取股票的涨停幅度"""
        try:
            from utils.limit_ratio import get_limit_ratio

            stock_name = get_stock_name(stock_code) or ""
            return get_limit_ratio(stock_code, stock_name, as_of_date)
        except Exception:
            return 0.10
    
    def _get_full_stock_code(self, stock_code: str) -> str:
        """获取完整的股票代码（带市场后缀）"""
        if '.' in stock_code:
            return stock_code
        
        if stock_code.startswith(('0', '1', '3')):
            return f"{stock_code}.SZ"
        elif stock_code.startswith('6'):
            return f"{stock_code}.SH"
        elif stock_code.startswith(('8', '4', '920')):
            return f"{stock_code}.BJ"
        else:
            return stock_code
    
    def _get_daily_data(
        self,
        stock_code: str,
        *,
        through_date: Optional[date] = None,
        apply_as_of_slice: bool = True,
    ) -> Optional[pd.DataFrame]:
        """获取日线数据（约3年历史）。

        优先读 data/daily_cache/{code}.csv（大 QMT 内置同步）；缺失时回退 xtdata。
        """
        try:
            if through_date is not None:
                end_date = through_date
            else:
                end_date = self._as_of_date if getattr(self, "_as_of_date", None) is not None else date.today()

            daily_data = load_daily_dataframe(
                stock_code,
                through_date=end_date,
                allow_xtdata_fallback=True,
            )
            if daily_data is None or daily_data.empty:
                return None

            if len(daily_data) > 0:
                d_min, d_max = daily_data["date"].min(), daily_data["date"].max()
                trading_set = get_trading_dates_set_for_range(d_min, d_max)
                if trading_set:
                    daily_data = daily_data[daily_data["date"].isin(trading_set)]

            daily_data = daily_data.sort_values("date")
            if apply_as_of_slice and getattr(self, "_as_of_date", None) is not None:
                daily_data = daily_data[daily_data["date"] <= self._as_of_date]

            return daily_data

        except Exception as e:
            logger.error(f"[{stock_code}] 获取日线数据异常: {str(e)}", exc_info=True)
            return None
    
    def _setup_calendars_for_as_of(self, as_of_d: Optional[date]) -> bool:
        """按某一选股基准日设置交易日窗口；失败返回 False。"""
        self._as_of_date = as_of_d
        self.trading_dates_50 = get_trading_dates(self._n, as_of_d)
        if len(self.trading_dates_50) < self._n:
            return False
        # 为内置筛选骨架预先准备交易日窗口（即使用户规则不使用也不影响）
        self.trading_dates_n_mode2 = list(self.trading_dates_50)
        self.trading_dates_n_mode3 = get_trading_dates(self._n_mode3, as_of_d)
        if len(self.trading_dates_n_mode3) < self._n_mode3:
            # mode3 可能用不到；这里不强制失败，避免影响自定义规则
            self.trading_dates_n_mode3 = list(self.trading_dates_50)
        return True

    def _compile_rule(self, rule: Dict[str, object]):
        code = str(rule.get("code") or "").strip()
        if not code:
            raise ValueError("规则代码为空")
        # 确保项目根在 path 上，供规则 import core.* 共用判定
        _root = os.path.dirname(os.path.abspath(__file__))
        if _root and _root not in sys.path:
            sys.path.insert(0, _root)
        from core.strict_bull import apply_strict_bull_requirement  # noqa: WPS433

        safe_builtins = {
            "__import__": __import__,
            "Exception": Exception,
            "BaseException": BaseException,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
            "AttributeError": AttributeError,
            "StopIteration": StopIteration,
            "len": len,
            "min": min,
            "max": max,
            "sum": sum,
            "any": any,
            "all": all,
            "abs": abs,
            "round": round,
            "float": float,
            "int": int,
            "str": str,
            "bool": bool,
            "range": range,
            "enumerate": enumerate,
            "sorted": sorted,
            "reversed": reversed,
            "set": set,
            "list": list,
            "dict": dict,
            "tuple": tuple,
            "isinstance": isinstance,
            "hasattr": hasattr,
            "getattr": getattr,
            "callable": callable,
            "type": type,
        }
        # 必须用同一 dict 作 globals 与 locals：否则「def run_mode1」在 locals、
        # 「def select」的全局却是 globals，select 内调用 run_mode1 会 NameError。
        ns: Dict[str, object] = {
            "__builtins__": safe_builtins,
            "date": date,
            "timedelta": timedelta,
            # 严多头 / 实盘用等规则共用
            "apply_strict_bull_requirement": apply_strict_bull_requirement,
        }
        exec(code, ns, ns)
        fn = ns.get("select")
        if not callable(fn):
            raise ValueError("规则代码必须定义 select(...) 函数")
        extracted = self._extract_rule_constants(ns)
        params = self._merge_rule_calendar_params(extracted)
        export_params = self._extract_rule_export_params(ns)
        return {"fn": fn, "params": params, "export_params": export_params}
    
    def _check_stock(self, stock_code: str, stock_name: str, mode: int) -> Tuple[bool, Optional[date], Optional[date]]:
        """检查股票是否符合指定模式的条件（内部方法，会获取日线数据）
        
        Args:
            stock_code: 股票代码
            stock_name: 股票名称
            mode: 模式编号 (1, 2, 3)
        
        Returns:
            tuple[bool, date | None, date | None]: (是否符合条件, 涨停日期, 信号触发日期)
        """
        if mode == 1:
            is_valid, limit_up_date = self._check_stock_mode1(stock_code, stock_name)
            return is_valid, limit_up_date, None
        elif mode == 2:
            return self._check_stock_mode2(stock_code, stock_name)
        elif mode == 3:
            return self._check_stock_mode3(stock_code, stock_name)
        else:
            return False, None, None
    
    def _check_stock_with_data(self, stock_code: str, stock_name: str, mode: int, daily_data: pd.DataFrame) -> Tuple[bool, Optional[date], Optional[date]]:
        """检查股票是否符合指定模式的条件（使用已获取的日线数据，避免重复获取）
        
        Args:
            stock_code: 股票代码
            stock_name: 股票名称
            mode: 模式编号 (1, 2, 3)
            daily_data: 已获取的日线数据
        
        Returns:
            tuple[bool, date | None, date | None]: (是否符合条件, 涨停日期, 信号触发日期)
        """
        if mode == 1:
            is_valid, limit_up_date = self._check_stock_mode1_with_data(stock_code, stock_name, daily_data)
            return is_valid, limit_up_date, None
        elif mode == 2:
            return self._check_stock_mode2_with_data(stock_code, stock_name, daily_data)
        elif mode == 3:
            return self._check_stock_mode3_with_data(stock_code, stock_name, daily_data)
        else:
            return False, None, None
    
    def _check_stock_mode1(self, stock_code: str, stock_name: str) -> Tuple[bool, Optional[date]]:
        """检查股票是否符合模式一的条件（内部方法，会获取日线数据）
        
        条件：
        1. 前N个交易日到前M个交易日前有且只有一个涨停
        2. 近M个交易日内没有涨停
        
        Returns:
            tuple[bool, date | None]: (是否符合条件, 涨停日期)
        """
        try:
            # 获取日线数据
            daily_data = self._get_daily_data(stock_code)
            if daily_data is None or daily_data.empty:
                return False, None
            return self._check_stock_mode1_with_data(stock_code, stock_name, daily_data)
        except Exception as e:
            logger.error(f"[{stock_code}] 检查过程中出错: {str(e)}", exc_info=True)
            return False, None
    
    def _check_mode1_style_limit_up_screen(
        self,
        stock_code: str,
        stock_name: str,
        daily_data: pd.DataFrame,
        *,
        require_obvious_new_high_after_limit: bool,
        debug_label: str,
        require_old_high_before_limit: bool = False,
        reject_if_prior_trading_day_also_limit_up: bool = False,
        require_lower_shadow: bool = True,
        require_boll_break: bool = True,
        require_ma_support_after: bool = True,
    ) -> Tuple[bool, Optional[date], Optional[date]]:
        """模式一与模式二共用的涨停筛选骨架（N、M、布林、前高、均线支撑相同）。

        require_old_high_before_limit:
            False — 模式一：涨停前 M 个交易日收盘价须无「旧高」（max_close < 涨停价，容差 0.01）；
            True  — 模式二：与模式一相同，但涨停前 M 个交易日须「有旧高」（与模式一相反）。
        reject_if_prior_trading_day_also_limit_up:
            True — 仅模式一：若某个「涨停候选日」的紧邻前一交易日按同款规则亦为涨停，则淘汰该候选（不要二连板中的第二板作为 L）。
        require_obvious_new_high_after_limit:
            False — 涨停后收盘价不得突破「涨停价×(1+涨停幅度×0.5)」（无明显新高）；
            True  — 涨停后须出现上述意义上的明显新高（当前模式一/二均传 False）。
        require_lower_shadow:
            True — 须下影线（low < 前日 high）；False — 须跳空涨停（low >= 前日 high）。
        require_boll_break:
            True — 涨停价须 >= 布林上轨；False — 须低于布林上轨。
        require_ma_support_after:
            True — 涨停后每日收盘不破 min(MA5,MA10) 与 min(MA20~120)；
            False — 须至少有一日跌破上述均线支撑。

        Returns:
            (是否通过, 涨停日期, 预留第三元；明显新高时为该日，否则为 None)
        """
        try:
            # 默认只打一行摘要，避免海量筛选时日志框刷屏；需要「窗口/逐日/淘汰」全文时设环境变量 SECTOR_FILTER_MODE1_VERBOSE_LOG=1
            verbose_m12 = os.environ.get("SECTOR_FILTER_MODE1_VERBOSE_LOG", "").strip().lower() in ("1", "true", "yes")
            # 本函数内所有出口必须经 _r，保证恒为三元组，避免解包错误
            def _r(
                ok: bool,
                limit_up: Optional[date] = None,
                aux: Optional[date] = None,
            ) -> Tuple[bool, Optional[date], Optional[date]]:
                return (ok, limit_up, aux)

            if not self.trading_dates_50:
                try:
                    ad = getattr(self, "_as_of_date", None)
                    ad_hint = "当前自然日规则" if ad is None else str(ad)
                    if verbose_m12:
                        self.debug_info.emit(
                            f"[{debug_label}不通过] {stock_code} {stock_name}（选股基准日 {ad_hint}）：\n"
                            f"  - 最近N交易日列表为空（日历未就绪）"
                        )
                    else:
                        self.debug_info.emit(
                            f"[{debug_label}不通过] {stock_code} {stock_name}（{ad_hint}）最近N交易日列表为空"
                        )
                except Exception:
                    pass
                return _r(False)
            reject_msgs: List[str] = []

            def _add_reject(msg: str) -> None:
                if len(reject_msgs) >= 16:
                    return
                reject_msgs.append(msg)

            def _emit_failure_block(
                td_window: List,
                rd_dates: List,
                last_td: date,
                raw_hits: List,
                scan_lines: List[str],
            ) -> None:
                try:
                    raw_hit_cnt = len(raw_hits) if raw_hits else 0
                    if raw_hit_cnt <= 0:
                        # 没检测到涨停候选时，不要为海量股票都输出详细失败信息
                        cnt = getattr(self, "_mode1_fail_emit_count", 0)
                        limit = getattr(self, "_mode1_fail_emit_limit", 30)
                        if cnt >= limit:
                            return
                        setattr(self, "_mode1_fail_emit_count", cnt + 1)
                    else:
                        # 检测到涨停候选时：尽量输出，但仍做上限保护
                        cnt = getattr(self, "_mode1_limitup_emit_count", 0)
                        limit = getattr(self, "_mode1_limitup_emit_limit", 200)
                        if cnt >= limit:
                            return
                        setattr(self, "_mode1_limitup_emit_count", cnt + 1)

                    ad = getattr(self, "_as_of_date", None)
                    ad_hint = "当前自然日规则" if ad is None else str(ad)
                    if not verbose_m12:
                        brief = (
                            f"[{debug_label}不通过] {stock_code} {stock_name}（{ad_hint}）"
                            f" 窗口={','.join(str(x) for x in td_window)}"
                            f" 有K线={','.join(str(x) for x in rd_dates)}"
                            f" 排除末根={last_td}"
                            f" 涨停候选={'无' if not raw_hits else '、'.join(str(x) for x in raw_hits)}"
                        )
                        if raw_hits and reject_msgs:
                            brief += f" | 淘汰示例：{reject_msgs[0]}"
                        self.debug_info.emit(brief)
                        sn = getattr(self, "_diag_snippets", None)
                        if sn is not None:
                            sn.append(brief[:1600])
                        return
                    parts: List[str] = [
                        f"窗口交易日(最近N天)：{', '.join(str(x) for x in td_window)}",
                        f"实际有日线数据日期：{', '.join(str(x) for x in rd_dates)}",
                        f"排除最后交易日(不作涨停候选)：{last_td}",
                        "收盘价判定涨停的日期："
                        + ("无" if not raw_hits else "、".join(str(x) for x in raw_hits)),
                    ]
                    if reject_msgs:
                        parts.append("条件淘汰说明：")
                        parts.extend(reject_msgs)
                    if scan_lines:
                        parts.append("各日扫描(prev_close/close/理论涨停价/is_limit_up)：")
                        parts.extend(scan_lines)
                    msg = (
                        f"[{debug_label}不通过] {stock_code} {stock_name}（选股基准日 {ad_hint}）：\n  - "
                        + "\n  - ".join(parts)
                    )
                    self.debug_info.emit(msg)
                    sn = getattr(self, "_diag_snippets", None)
                    if sn is not None:
                        sn.append(msg if len(msg) <= 1600 else (msg[:1600] + "…（已截断）"))
                except Exception:
                    pass
            if require_old_high_before_limit:
                m = getattr(self, "trading_days_m_mode2", None) or getattr(self, "trading_days_m_mode1", None) or 10
            else:
                m = getattr(self, "trading_days_m_mode1", None) or 10
            
            # 计算均线与布林线上轨（20 日）
            dd = daily_data.sort_values('date').copy()
            # 5日、10日均线，供“后续收盘价不跌破短期均线”判断使用
            dd['MA5'] = dd['close'].rolling(window=5, min_periods=1).mean()
            dd['MA10'] = dd['close'].rolling(window=10, min_periods=1).mean()
            # 中长期均线：20/30/60/120 日
            dd['MA20'] = dd['close'].rolling(window=20, min_periods=1).mean()
            dd['MA30'] = dd['close'].rolling(window=30, min_periods=1).mean()
            dd['MA60'] = dd['close'].rolling(window=60, min_periods=1).mean()
            dd['MA120'] = dd['close'].rolling(window=120, min_periods=1).mean()
            # 布林线上轨（基于20日中轨）
            dd['STD20'] = dd['close'].rolling(window=20, min_periods=1).std()
            dd['BOLL_UPPER'] = dd['MA20'] + 2.0 * dd['STD20']
            
            # 最近 N 个交易日（模式一窗口）
            recent_dates_set = set(self.trading_dates_50)
            recent_data = dd[dd['date'].isin(recent_dates_set)]
            td_window_sorted = sorted(list(recent_dates_set))
            if recent_data.empty:
                try:
                    ad = getattr(self, "_as_of_date", None)
                    ad_hint = "当前自然日规则" if ad is None else str(ad)
                    if verbose_m12:
                        self.debug_info.emit(
                            f"[{debug_label}不通过] {stock_code} {stock_name}（选股基准日 {ad_hint}）：\n"
                            f"  - 窗口交易日(最近N天)：{', '.join(str(x) for x in td_window_sorted)}\n"
                            f"  - 上述日期与当前日线无交集（请确认选股基准日、数据是否已下载）"
                        )
                    else:
                        self.debug_info.emit(
                            f"[{debug_label}不通过] {stock_code} {stock_name}（{ad_hint}）"
                            f"窗口{','.join(str(x) for x in td_window_sorted)}与日线无交集"
                        )
                except Exception:
                    pass
                return _r(False)
            
            def _is_close_limit_up_on_day(d_check) -> bool:
                lr = self._get_limit_ratio(stock_code, d_check)
                sub = dd[dd["date"] == d_check]
                if sub.empty:
                    return False
                rw = sub.iloc[-1]
                pvd = dd[dd["date"] < d_check]
                if pvd.empty:
                    return False
                pr = pvd.iloc[-1]
                pc = float(pr["close"])
                cl = float(rw["close"])
                if pc <= 0:
                    return False
                lup = round(pc * (1 + lr), 2)
                pdiff = abs(cl - lup)
                inc = (cl - pc) / pc
                return (pdiff < 0.02) or (inc >= lr * 0.99)
            
            # 最近 N 天里最晚的一个交易日（用于排除最后一天）
            last_trading_date = recent_data['date'].max()
            recent_data_sorted = recent_data.sort_values('date')
            rd_dates_sorted = sorted(list(recent_data_sorted['date'].unique()))
            limitup_scan_lines: List[str] = []
            raw_limit_hit_dates: List[date] = []
            
            # 遍历最近 N 个交易日，查找第一个符合条件的涨停日（排除最后一个交易日）
            for _, row in recent_data_sorted.iterrows():
                trade_date = row['date']
                if trade_date >= last_trading_date:
                    # 排除窗口内最后一个交易日：后面没有数据，无法判断是否“无新高”
                    continue
                
                close_price = row['close']
                high_price = row['high']
                
                # 找到该日之前的最近一个交易日
                prev_trading_days = dd[dd['date'] < trade_date]
                if prev_trading_days.empty:
                    continue
                prev_row = prev_trading_days.iloc[-1]
                prev_close = prev_row['close']
                prev_high = prev_row['high']
                
                limit_ratio = self._get_limit_ratio(stock_code, trade_date)
                
                # 理论涨停价
                limit_up_price = round(prev_close * (1 + limit_ratio), 2)
                
                # 是否涨停（与 auto_limit_up_filter 相同的判定）
                price_diff = abs(close_price - limit_up_price)
                price_increase_ratio = (close_price - prev_close) / prev_close if prev_close > 0 else 0
                expected_increase_ratio = limit_ratio
                is_limit_up = (price_diff < 0.02) or (price_increase_ratio >= expected_increase_ratio * 0.99)
                if is_limit_up:
                    raw_limit_hit_dates.append(trade_date)
                # 调试：记录最近N天逐日涨停判定细节，便于核对为何“未找到涨停日”
                try:
                    limitup_scan_lines.append(
                        f"{trade_date}: prev_close={float(prev_close):.2f}, close={float(close_price):.2f}, "
                        f"limit_up={float(limit_up_price):.2f}, diff={float(price_diff):.4f}, "
                        f"inc={float(price_increase_ratio) * 100:.2f}%, is_limit_up={bool(is_limit_up)}"
                    )
                except Exception:
                    pass
                if not is_limit_up:
                    continue
                
                # 0) 下影线 / 跳空涨停
                has_lower_shadow = float(row['low']) < float(prev_high)
                if require_lower_shadow:
                    if not has_lower_shadow:
                        try:
                            _add_reject(
                                f"{trade_date}：无下影线（low={float(row['low']):.2f} >= prev_high={float(prev_high):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：无下影线（low >= prev_high）")
                        continue
                else:
                    if has_lower_shadow:
                        try:
                            _add_reject(
                                f"{trade_date}：非跳空涨停（low={float(row['low']):.2f} < prev_high={float(prev_high):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：非跳空涨停（low < prev_high）")
                        continue

                if reject_if_prior_trading_day_also_limit_up:
                    prev_d = prev_row["date"]
                    if _is_close_limit_up_on_day(prev_d):
                        _add_reject(
                            f"{trade_date}：前一交易日 {prev_d} 亦为涨停（二连板），排除"
                        )
                        continue
                
                # 1) 涨停价是否不低于布林线上轨
                limit_day_data = dd[dd['date'] == trade_date]
                if limit_day_data.empty:
                    _add_reject(f"{trade_date}：无法取得该日布林线上轨数据")
                    continue
                boll_upper = limit_day_data['BOLL_UPPER'].iloc[0]
                boll_break = limit_up_price >= boll_upper - 0.01
                if require_boll_break:
                    if not boll_break:
                        try:
                            _add_reject(
                                f"{trade_date}：涨停价低于布林上轨（limit_up={float(limit_up_price):.2f} < boll_upper={float(boll_upper):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停价低于布林上轨")
                        continue
                else:
                    if boll_break:
                        try:
                            _add_reject(
                                f"{trade_date}：涨停价不低于布林上轨（limit_up={float(limit_up_price):.2f} >= boll_upper={float(boll_upper):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停价不低于布林上轨")
                        continue
                
                # 2) 涨停日前最近 M 个交易日「旧高」：模式一要求无旧高；模式二要求有旧高（M 可大于 N，按全历史往前取 M 天）
                # 这里用「收盘价」：旧高 = 窗口内收盘价最大值 >= 涨停价（容差 0.01）
                before_limit_data = dd[dd['date'] < trade_date]
                if before_limit_data.empty:
                    if require_old_high_before_limit:
                        _add_reject(f"{trade_date}：涨停前无日线数据，无法判定旧高")
                        continue
                else:
                    prev_dates = sorted(before_limit_data['date'].unique())
                    window_dates = prev_dates[-m:] if len(prev_dates) > m else prev_dates
                    window_data = before_limit_data[before_limit_data['date'].isin(window_dates)]
                    max_close_before = window_data['close'].max()
                    has_old_high = max_close_before >= limit_up_price - 0.01
                    if require_old_high_before_limit:
                        if not has_old_high:
                            try:
                                _add_reject(
                                    f"{trade_date}：涨停前{len(window_dates)}日收盘无旧高（max_close={float(max_close_before):.2f} < limit_up={float(limit_up_price):.2f}）"
                                )
                            except Exception:
                                _add_reject(f"{trade_date}：涨停前{len(window_dates)}日收盘无旧高")
                            continue
                    else:
                        if has_old_high:
                            try:
                                _add_reject(
                                    f"{trade_date}：涨停前{len(window_dates)}日收盘有旧高（max_close={float(max_close_before):.2f} >= limit_up={float(limit_up_price):.2f}）"
                                )
                            except Exception:
                                _add_reject(f"{trade_date}：涨停前{len(window_dates)}日收盘有旧高")
                            continue
                
                # 3) 涨停日之后：要求「无明显新高」（阈值 涨停价×(1+涨停幅度×0.5)，容差0.01）；须明显新高时见 require_obvious_new_high_after_limit
                after_limit_data = dd[dd['date'] > trade_date]
                if after_limit_data.empty:
                    _add_reject(f"{trade_date}：涨停后无数据，无法判断新高/均线条件")
                    continue
                max_allowed_price = limit_up_price * (1 + limit_ratio * 0.5)
                nh_tol = 0.01
                max_high_after = after_limit_data['close'].max()
                new_high_signal_date: Optional[date] = None

                if require_obvious_new_high_after_limit:
                    if max_high_after <= max_allowed_price + nh_tol:
                        try:
                            _add_reject(
                                f"{trade_date}：涨停后未出现明显新高（max_close_after={float(max_high_after):.2f} <= max_allowed={float(max_allowed_price):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停后未出现明显新高")
                        continue
                    after_sorted = after_limit_data.sort_values('date')
                    nh_rows = after_sorted[after_sorted['close'] > max_allowed_price + nh_tol]
                    new_high_signal_date = nh_rows.iloc[0]['date']
                else:
                    if max_high_after > max_allowed_price + nh_tol:
                        try:
                            _add_reject(
                                f"{trade_date}：涨停后收盘出现明显新高（max_close_after={float(max_high_after):.2f} > max_allowed={float(max_allowed_price):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停后收盘出现明显新高")
                        continue

                after_limit_with_ma = after_limit_data.copy()
                after_limit_with_ma['MA_MIN_5_10'] = after_limit_with_ma[['MA5', 'MA10']].min(axis=1, skipna=True)
                after_limit_with_ma['MA_MIN_20_120'] = after_limit_with_ma[
                    ['MA20', 'MA30', 'MA60', 'MA120']
                ].min(axis=1, skipna=True)

                # 4) 涨停后均线支撑（可开关；False 时须至少一日跌破）
                below_support = after_limit_with_ma['close'] < (after_limit_with_ma['MA_MIN_5_10'] - 0.01)
                below_long_ma = after_limit_with_ma['close'] < (after_limit_with_ma['MA_MIN_20_120'] - 0.01)
                breaks_ma = below_support | below_long_ma
                if require_ma_support_after:
                    if below_support.any():
                        try:
                            first_bad = after_limit_with_ma[below_support].iloc[0]
                            _add_reject(
                                f"{trade_date}：涨停后跌破短期支撑（{first_bad['date']} close={float(first_bad['close']):.2f} < min(MA5,MA10)={float(first_bad['MA_MIN_5_10']):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停后跌破短期支撑（min(MA5,MA10)）")
                        continue
                    if below_long_ma.any():
                        try:
                            first_bad = after_limit_with_ma[below_long_ma].iloc[0]
                            _add_reject(
                                f"{trade_date}：涨停后跌破中长期均线（{first_bad['date']} close={float(first_bad['close']):.2f} < min(MA20/30/60/120)={float(first_bad['MA_MIN_20_120']):.2f}）"
                            )
                        except Exception:
                            _add_reject(f"{trade_date}：涨停后跌破中长期均线（min(MA20/30/60/120)）")
                        continue
                else:
                    if not breaks_ma.any():
                        _add_reject(f"{trade_date}：涨停后未跌破均线支撑")
                        continue
                return _r(True, trade_date, new_high_signal_date)
            
            # 最近N天内没有找到“满足全部条件”的涨停日
            if not raw_limit_hit_dates and not reject_msgs:
                _add_reject("最近N天内收盘价未判定为涨停，或涨停后条件未满足（见下方逐日 is_limit_up）")
            
            if raw_limit_hit_dates:
                self._mode1_limitup_detected_checks = getattr(self, "_mode1_limitup_detected_checks", 0) + 1
            _emit_failure_block(
                td_window_sorted,
                rd_dates_sorted,
                last_trading_date,
                raw_limit_hit_dates,
                limitup_scan_lines,
            )
            return _r(False)
        
        except Exception as e:
            logger.error(f"[{stock_code}] {debug_label}检查过程中出错: {str(e)}", exc_info=True)
            try:
                self.debug_info.emit(f"[{debug_label}] {stock_code} {stock_name} 检查异常：{e}")
            except Exception:
                pass
            return False, None, None

    def _check_stock_mode1_with_data(self, stock_code: str, stock_name: str, daily_data: pd.DataFrame) -> Tuple[bool, Optional[date]]:
        """模式一：涨停前 M 天无旧高；涨停后无明显新高 + 不破均线。
        另：作为模式一涨停日 L 的候选日，若其前一交易日亦为涨停（二连板），则该候选排除。"""
        raw = self._check_mode1_style_limit_up_screen(
            stock_code,
            stock_name,
            daily_data,
            require_obvious_new_high_after_limit=False,
            debug_label="模式一",
            require_old_high_before_limit=False,
            reject_if_prior_trading_day_also_limit_up=True,
        )
        ok, lu, _ = _coerce_mode12_screen_return(raw)
        return ok, lu
    
    def _check_stock_mode2(self, stock_code: str, stock_name: str) -> Tuple[bool, Optional[date], Optional[date]]:
        """模式二：与模式一相同，区别为涨停前 M 天须有旧高（收盘价曾达涨停价附近）。"""
        try:
            daily_data = self._get_daily_data(stock_code)
            if daily_data is None or daily_data.empty:
                return False, None, None
            return self._check_stock_mode2_with_data(stock_code, stock_name, daily_data)
        except Exception as e:
            logger.error(f"[{stock_code}] 模式二检查过程中出错: {str(e)}", exc_info=True)
            return False, None, None
    
    def _check_stock_mode2_with_data(self, stock_code: str, stock_name: str, daily_data: pd.DataFrame) -> Tuple[bool, Optional[date], Optional[date]]:
        """模式二：返回 (是否通过, 涨停日, None) — 逻辑同模式一，仅涨停前旧高条件相反。"""
        raw = self._check_mode1_style_limit_up_screen(
            stock_code,
            stock_name,
            daily_data,
            require_obvious_new_high_after_limit=False,
            debug_label="模式二",
            require_old_high_before_limit=True,
        )
        return _coerce_mode12_screen_return(raw)
    
    def _check_stock_mode3(self, stock_code: str, stock_name: str) -> Tuple[bool, Optional[date], Optional[date]]:
        """检查股票是否符合模式三的条件（内部方法，会获取日线数据）
        
        条件：
        1. 从当前日期前一天往前数，N天之内有一个涨停板
        2. 这个涨停板的股价不小于它之前M个交易日的收盘价（涨停板收盘价 >= 之前M个交易日的所有收盘价）
        3. 涨停板之后，至少有一天的阴线。阴线之后的第一个阳线的收盘价超过第一个阴线的开盘价，
           其成交量大于这个或者这些阴线的成交量之和。满足条件的日期为"放量反包日期"（即信号触发日期）
        4. 该涨停板的前L个交易日没有涨停板
        5. 涨停板前一交易日的收盘价在当天MA30的0.7倍到1.3倍之间
        6. 涨停板前一交易日的收盘价在当天MA60的0.7倍到1.3倍之间
        
        Returns:
            tuple[bool, date | None, date | None]: (是否符合条件, 涨停日期, 放量反包日期)
        """
        try:
            # 获取日线数据
            daily_data = self._get_daily_data(stock_code)
            if daily_data is None or daily_data.empty:
                return False, None, None
            return self._check_stock_mode3_with_data(stock_code, stock_name, daily_data)
        except Exception as e:
            logger.error(f"[{stock_code}] 模式三检查过程中出错: {str(e)}", exc_info=True)
            return False, None, None
    
    def _check_stock_mode3_with_data(self, stock_code: str, stock_name: str, daily_data: pd.DataFrame) -> Tuple[bool, Optional[date], Optional[date]]:
        """检查股票是否符合模式三的条件（使用已获取的日线数据）
        
        条件：
        1. 从当前日期前一天往前数，N天之内有一个涨停板
        2. 这个涨停板的股价不小于它之前M个交易日的收盘价（涨停板收盘价 >= 之前M个交易日的所有收盘价）
        3. 涨停板之后，至少有一天的阴线。阴线之后的第一个阳线的收盘价超过第一个阴线的开盘价，
           其成交量大于这个或者这些阴线的成交量之和。满足条件的日期为"放量反包日期"（即信号触发日期）
        4. 该涨停板的前L个交易日没有涨停板
        5. 涨停板前一交易日的收盘价在当天MA30的0.7倍到1.3倍之间
        6. 涨停板前一交易日的收盘价在当天MA60的0.7倍到1.3倍之间
        
        Returns:
            tuple[bool, date | None, date | None]: (是否符合条件, 涨停日期, 放量反包日期)
        """
        try:
            
            # 获取模式三的参数（使用模式三的独立N和M参数）
            # 有选股基准日时：窗口数据截至该日；否则沿用「自然日昨天」排除未完成当日
            ref_as_of = getattr(self, "_as_of_date", None)
            if ref_as_of is not None:
                bar_cutoff = ref_as_of
            else:
                bar_cutoff = date.today() - timedelta(days=1)
            
            # 获取前N个交易日（使用模式三的参数）
            trading_dates_n_set = set(self.trading_dates_n_mode3)  # 前N个交易日（模式三）
            
            # 筛选出前N个交易日的数据，且日期不超过 cutoff
            data_n = daily_data[(daily_data['date'].isin(trading_dates_n_set)) & (daily_data['date'] <= bar_cutoff)]
            if data_n.empty:
                return False, None, None
            
            # 按日期排序
            data_n_sorted = data_n.sort_values('date')
            
            # 获取M和L的值（使用模式三的参数）
            m = self.trading_days_m_mode3
            l = self.trading_days_l_mode3
            if m <= 0 or l <= 0:
                return False, None, None
            
            # 获取所有交易日列表（用于计算间隔和检查前L个交易日）
            all_trading_dates = sorted(daily_data['date'].unique().tolist())
            
            # 遍历前N个交易日，查找符合条件的涨停板
            for idx, row in data_n_sorted.iterrows():
                trade_date = row['date']
                close_price = row['close']
                volume = row['volume']
                
                # 检查是否是涨停板
                prev_trading_days = daily_data[daily_data['date'] < trade_date]
                if prev_trading_days.empty:
                    continue
                
                prev_row = prev_trading_days.iloc[-1]
                prev_close = prev_row['close']
                
                limit_ratio = self._get_limit_ratio(stock_code, trade_date)
                
                # 计算涨停价
                limit_up_price = round(prev_close * (1 + limit_ratio), 2)
                
                # 判断是否涨停
                price_diff = abs(close_price - limit_up_price)
                price_increase_ratio = (close_price - prev_close) / prev_close if prev_close > 0 else 0
                is_limit_up = (price_diff < 0.02) or (price_increase_ratio >= limit_ratio * 0.99)
                
                if not is_limit_up:
                    continue  # 不是涨停板，跳过
                
                # 获取涨停板日期在all_trading_dates中的索引（用于后续条件检查）
                trade_date_idx = all_trading_dates.index(trade_date)
                
                # 新条件：涨停板前一交易日的收盘价在MA30的0.7倍到1.3倍之间
                # 计算前一个交易日当天的MA30 = 前30个交易日的收盘价平均值（包括前一个交易日当天）
                if trade_date_idx < 30:
                    continue  # 之前没有30个交易日，无法计算MA30，跳过
                
                # 获取前30个交易日的收盘价（包括前一个交易日当天）
                # MA30 = 前30个交易日的收盘价之和 / 30
                prev_30_dates = all_trading_dates[trade_date_idx - 30:trade_date_idx]
                prev_30_data = daily_data[daily_data['date'].isin(prev_30_dates)]
                
                if prev_30_data.empty or len(prev_30_data) < 30:
                    continue  # 数据不足，无法计算MA30
                
                # 计算MA30 = 前30个交易日的收盘价之和 / 30
                prev_30_closes = prev_30_data['close'].tolist()
                if len(prev_30_closes) < 30:
                    continue  # 数据不足
                
                # 确保只取30个收盘价
                prev_30_closes = prev_30_closes[-30:] if len(prev_30_closes) > 30 else prev_30_closes
                ma30 = sum(prev_30_closes) / len(prev_30_closes) if len(prev_30_closes) > 0 else 0
                
                if ma30 == 0:
                    continue  # MA30计算失败
                
                # 检查：前一个交易日的收盘价必须在 MA30 * 0.7 到 MA30 * 1.3 之间
                ma30_lower = ma30 * 0.7
                ma30_upper = ma30 * 1.3
                condition_met = prev_close >= ma30_lower and prev_close <= ma30_upper
                
                # 如果不满足条件，跳过
                if not condition_met:
                    continue  # 前一个交易日的收盘价不在MA30的0.7倍到1.3倍之间，不符合条件
                
                # MA60条件：涨停板前一交易日的收盘价在当天MA60的0.7倍到1.3倍之间
                # 计算前一个交易日当天的MA60 = 前60个交易日的收盘价平均值
                if trade_date_idx < 60:
                    continue  # 之前没有60个交易日，无法计算MA60，跳过
                
                # 获取前60个交易日的收盘价
                # MA60 = 前60个交易日的收盘价之和 / 60
                prev_60_dates = all_trading_dates[trade_date_idx - 60:trade_date_idx]
                prev_60_data = daily_data[daily_data['date'].isin(prev_60_dates)]
                
                if prev_60_data.empty or len(prev_60_data) < 60:
                    continue  # 数据不足，无法计算MA60
                
                # 计算MA60 = 前60个交易日的收盘价之和 / 60
                prev_60_closes = prev_60_data['close'].tolist()
                if len(prev_60_closes) < 60:
                    continue  # 数据不足
                
                # 确保只取60个收盘价
                prev_60_closes = prev_60_closes[-60:] if len(prev_60_closes) > 60 else prev_60_closes
                ma60 = sum(prev_60_closes) / len(prev_60_closes) if len(prev_60_closes) > 0 else 0
                
                if ma60 == 0:
                    continue  # MA60计算失败
                
                # 检查：前一个交易日的收盘价必须在 MA60 * 0.7 到 MA60 * 1.3 之间
                ma60_lower = ma60 * 0.7
                ma60_upper = ma60 * 1.3
                ma60_condition_met = prev_close >= ma60_lower and prev_close <= ma60_upper
                
                # 如果不满足条件，跳过
                if not ma60_condition_met:
                    continue  # 前一个交易日的收盘价不在MA60的0.7倍到1.3倍之间，不符合条件
                
                # 条件4：该涨停板的前L个交易日没有涨停板
                if trade_date_idx < l:
                    continue  # 之前没有L个交易日，跳过
                
                prev_l_dates = all_trading_dates[trade_date_idx - l:trade_date_idx]
                prev_l_data = daily_data[daily_data['date'].isin(prev_l_dates)]
                
                # 检查前L个交易日是否有涨停板
                has_limit_up_in_prev_l = False
                if not prev_l_data.empty:
                    for prev_idx, prev_row in prev_l_data.iterrows():
                        prev_trade_date = prev_row['date']
                        prev_close_price = prev_row['close']
                        
                        # 找到前一个交易日
                        prev_prev_trading_days = daily_data[daily_data['date'] < prev_trade_date]
                        if prev_prev_trading_days.empty:
                            continue
                        
                        prev_prev_row = prev_prev_trading_days.iloc[-1]
                        prev_prev_close = prev_prev_row['close']
                        
                        prev_lr = self._get_limit_ratio(stock_code, prev_trade_date)
                        
                        # 计算涨停价
                        prev_limit_up_price = round(prev_prev_close * (1 + prev_lr), 2)
                        
                        # 判断是否涨停
                        prev_price_diff = abs(prev_close_price - prev_limit_up_price)
                        prev_price_increase_ratio = (prev_close_price - prev_prev_close) / prev_prev_close if prev_prev_close > 0 else 0
                        prev_is_limit_up = (prev_price_diff < 0.02) or (prev_price_increase_ratio >= prev_lr * 0.99)
                        
                        if prev_is_limit_up:
                            has_limit_up_in_prev_l = True
                            break  # 前L个交易日有涨停板，不符合条件
                
                if has_limit_up_in_prev_l:
                    continue  # 前L个交易日有涨停板，跳过当前涨停板
                
                # 条件2：涨停板的收盘价不小于它之前M个交易日的收盘价
                # 找到之前M个交易日的数据
                if trade_date_idx < m:
                    continue  # 之前没有M个交易日，跳过
                
                prev_m_dates = all_trading_dates[trade_date_idx - m:trade_date_idx]
                prev_m_data = daily_data[daily_data['date'].isin(prev_m_dates)]
                
                if prev_m_data.empty:
                    continue
                
                # 检查涨停板收盘价是否 >= 之前M个交易日的所有收盘价
                prev_m_closes = prev_m_data['close'].tolist()
                if any(close_price < prev_close for prev_close in prev_m_closes):
                    continue  # 涨停板收盘价小于某个之前M个交易日的收盘价，不符合条件
                
                # 条件3：涨停板之后，紧接着（第一个交易日）必须有至少一个阴线。阴线之后的第一个阳线的收盘价超过第一个阴线的开盘价，
                # 其成交量大于这个或者这些阴线的成交量之和。满足条件的日期为"放量反包日期"（即信号触发日期）
                # 找到之后的数据
                after_data = daily_data[daily_data['date'] > trade_date]
                if after_data.empty:
                    continue  # 之后没有数据，不符合条件
                
                # 按日期排序
                after_data_sorted = after_data.sort_values('date')
                
                # 检查第一个交易日必须是阴线（收盘价 < 开盘价）
                first_row = after_data_sorted.iloc[0]
                first_open = float(first_row['open'])
                first_close = float(first_row['close'])
                
                # 如果第一个交易日不是阴线，不符合条件
                if first_close >= first_open:
                    continue
                
                # 第一个交易日是阴线，开始收集所有连续的阴线
                first_negative_row = first_row
                negative_rows = [first_row]  # 存储所有连续的阴线
                
                # 继续查找连续的阴线（直到遇到阳线）
                for i in range(1, len(after_data_sorted)):
                    after_row = after_data_sorted.iloc[i]
                    after_open = float(after_row['open'])
                    after_close = float(after_row['close'])
                    
                    # 判断是否是阴线
                    if after_close < after_open:
                        negative_rows.append(after_row)
                    else:
                        # 遇到阳线，停止收集阴线
                        break
                
                # 获取第一个阴线的开盘价
                first_negative_open = float(first_negative_row['open'])
                
                # 计算所有阴线的成交量之和
                negative_volumes_sum = sum(float(row['volume']) for row in negative_rows)
                
                # 查找阴线之后的第一个阳线（收盘价 >= 开盘价）
                # 从最后一个阴线的下一个交易日开始查找
                last_negative_date = negative_rows[-1]['date']
                first_positive_row = None
                
                for after_idx, after_row in after_data_sorted.iterrows():
                    after_date = after_row['date']
                    
                    # 只查找最后一个阴线之后的交易日
                    if after_date <= last_negative_date:
                        continue
                    
                    after_open = float(after_row['open'])
                    after_close = float(after_row['close'])
                    
                    # 判断是否是阳线（收盘价 >= 开盘价）
                    if after_close >= after_open:
                        # 找到第一个阳线，停止查找
                        first_positive_row = after_row
                        break
                
                # 如果没有找到阳线，不符合条件
                if first_positive_row is None:
                    continue
                
                # 找到第一个阳线后，判断是否满足条件
                first_positive_close = float(first_positive_row['close'])
                first_positive_volume = float(first_positive_row['volume'])
                
                # 检查条件：
                # 1. 第一个阳线的收盘价 > 第一个阴线的开盘价
                # 2. 第一个阳线的成交量 > 所有阴线的成交量之和
                if first_positive_close > first_negative_open and first_positive_volume > negative_volumes_sum:
                    # 所有条件都满足，返回这个涨停板的日期和放量反包的日期
                    signal_date = first_positive_row['date']
                    return True, trade_date, signal_date
                else:
                    # 第一个阳线不满足条件，结束，不再查找下一个阳线
                    continue
            
            # 没有找到符合条件的涨停板
            return False, None, None
            
        except Exception as e:
            logger.error(f"[{stock_code}] 模式三检查过程中出错: {str(e)}", exc_info=True)
            return False, None, None
    
    def run(self):
        """运行筛选（单日或区间内每个交易日；区间内每只股票只下载一次日线）"""
        rule_counts: Dict[str, int] = {}
        
        try:
            self._diag_snippets: List[str] = []
            # 大批量筛选时，日志/进度信号可能海量触发，导致 Qt 主线程事件队列堆积“卡住”
            # 这里统一做节流：只保留必要的“命中/异常/有限示例”输出。
            self._round_debug_emit_count = 0
            self._round_debug_emit_limit = int(os.environ.get("SECTOR_FILTER_ROUND_DEBUG_LIMIT", "30"))
            self._mode1_fail_emit_count = 0
            self._mode1_fail_emit_limit = int(os.environ.get("SECTOR_FILTER_MODE1_FAIL_EMIT_LIMIT", "30"))
            self._mode1_limitup_emit_count = 0
            self._mode1_limitup_emit_limit = int(os.environ.get("SECTOR_FILTER_MODE1_LIMITUP_EMIT_LIMIT", "200"))
            self._mode1_limitup_detected_checks = 0

            self._last_progress_emit_unit = 0
            self._progress_emit_every = int(os.environ.get("SECTOR_FILTER_PROGRESS_EVERY", "200"))
            # 检查输入参数
            if not self.stock_list:
                self.error_occurred.emit("股票列表为空，无法筛选")
                self.finished.emit(rule_counts)
                return
            
            enabled_rules = [r for r in (self.rules or []) if bool(r.get("enabled", True))]
            if not enabled_rules:
                self.error_occurred.emit("未启用任何选股规则")
                self.finished.emit(rule_counts)
                return

            compiled_rules: Dict[str, object] = {}
            for r in enabled_rules:
                rid = str(r.get("id") or "")
                if not rid:
                    continue
                try:
                    compiled_rules[rid] = self._compile_rule(r)
                    rule_counts[rid] = 0
                except Exception as e:
                    self.debug_info.emit(f"规则「{r.get('name','未命名')}」编译失败：{e}")
            
            self.debug_info.emit("正在准备选股日历与区间…")
            _as = getattr(self, "_as_of_date", None)
            _as_end = getattr(self, "_as_of_end_date", None)
            
            # 单日：与原先一致，[None] 表示「今天」且走 15:00 日历逻辑
            if _as_end is None:
                screen_as_of_list: List[Optional[date]] = [_as]
                fetch_through = date.today() if _as is None else _as
            else:
                if _as is None:
                    self.error_occurred.emit("批量区间选股时，起始日不能为「今天」未指定状态，请选具体起始日期")
                    self.finished.emit(rule_counts)
                    return
                lo, hi = (_as, _as_end) if _as <= _as_end else (_as_end, _as)
                screen_as_of_list = get_trading_dates_in_range_sorted(lo, hi)
                if not screen_as_of_list:
                    self.error_occurred.emit("所选区间内没有交易日，请调整起止日期")
                    self.finished.emit(rule_counts)
                    return
                fetch_through = hi
            
            # 各规则可在代码中定义不同的 N / N_MODE3，预检日历时用「所有已编译规则」中的最大值
            _need_n = int(self._n)
            _need_n3 = int(self._n_mode3)
            for r in enabled_rules:
                rid = str(r.get("id") or "")
                ent = compiled_rules.get(rid)
                if isinstance(ent, dict):
                    p = ent.get("params")
                    if isinstance(p, dict):
                        _need_n = max(_need_n, int(p.get("_n", self._n)))
                        _need_n3 = max(_need_n3, int(p.get("_n_mode3", self._n_mode3)))
            _calendar_need = max(_need_n, _need_n3)

            for as_of_d in screen_as_of_list:
                if len(get_trading_dates(_calendar_need, as_of_d)) < _calendar_need:
                    dlabel = "当前日" if as_of_d is None else str(as_of_d)
                    self.error_occurred.emit(
                        f"无法获取足够的交易日（{dlabel}，需要至少{_calendar_need}个交易日；"
                        f"请检查各规则代码顶部的 N / N_MODE3 是否过大）"
                    )
                    self.finished.emit(rule_counts)
                    return

            # 明确写出本次使用的选股日，避免「只改了起、止仍为今日」误以为在用历史日
            try:
                if len(screen_as_of_list) == 1:
                    d0 = screen_as_of_list[0]
                    if d0 is None:
                        self.debug_info.emit(
                            "选股范围：当前自然日规则（日线拉取至今日；最近 N 个交易日按 15:00 前后是否计入今日收盘）"
                        )
                    else:
                        self.debug_info.emit(
                            f"选股基准日：{d0}（最近 N 个交易日以该日为窗口终点，K 线不含该日之后）"
                        )
                else:
                    parts = [str(x) for x in screen_as_of_list]
                    self.debug_info.emit(
                        f"区间选股：共 {len(screen_as_of_list)} 个交易日将逐一计算 — {', '.join(parts)}"
                    )
            except Exception:
                pass
            
            # 按股票代码分组，避免重复筛选同一只股票
            stock_dict = {}  # {stock_code: {'name': stock_name, 'sectors': [sector1, sector2, ...]}}
            for stock_code, stock_name, sector in self.stock_list:
                if stock_code not in stock_dict:
                    stock_dict[stock_code] = {'name': stock_name, 'sectors': []}
                stock_dict[stock_code]['sectors'].append(sector)
            
            unique_stocks = list(stock_dict.keys())
            total_stocks_universe = len(unique_stocks)
            n_days = len(screen_as_of_list)

            ctx_needs = _enabled_rules_ctx_needs(enabled_rules)
            need_inflow = bool(ctx_needs.get("inflow"))
            need_hot_theme = bool(ctx_needs.get("hot_theme"))
            need_em = bool(ctx_needs.get("em"))
            em_arms: Set[str] = set(ctx_needs.get("em_arms") or [])
            elig_bands: List[Tuple[int, int]] = list(ctx_needs.get("elig_bands") or [])
            em_top_n = int(ctx_needs.get("em_top_n") or 50)
            em_rs_top_k = int(ctx_needs.get("em_rs_top_k") or 20)
            em_min_members = int(ctx_needs.get("em_min_members") or 30)

            self.debug_info.emit(
                "上下文按需加载："
                + f"净流入={'是' if need_inflow else '否(选后补)'}"
                + f" 热门题材={'是' if need_hot_theme else '否'}"
                + f" 东财热门={'是(' + ','.join(sorted(em_arms)) + ')' if need_em else '否'}"
                + (
                    f" TopN={em_top_n} RS_K={em_rs_top_k} 成分≥{em_min_members}"
                    if need_em
                    else ""
                )
                + (
                    f" Elig收窄={elig_bands}"
                    if elig_bands
                    else ""
                )
            )

            processed_units = 0
            skip_calendar = 0
            skip_no_daily = 0
            rule_calls = 0
            # 按选股日缓存主力净流入排名（供 A/B 档等规则 ctx['inflow_rank']）
            inflow_rank_cache: Dict[str, Dict[str, object]] = {}
            # 按选股日缓存十大热门板块/概念成员（供近10日涨停+热门题材规则 ctx['hot_theme']）
            hot_theme_cache: Dict[str, Dict[str, object]] = {}
            # 按选股日缓存东财连续热门 + 组内 RS（供东财热门-连续2日Top50-组内RS前20）
            em_board_hot_cache: Dict[str, Dict[str, object]] = {}
            load_inflow_rank_map = None
            load_hot_theme_map = None
            load_em_board_hot_map = None
            if need_inflow:
                try:
                    from utils.main_force_inflow_selection_ctx import load_inflow_rank_map
                except Exception as e:
                    load_inflow_rank_map = None  # type: ignore[assignment]
                    self.debug_info.emit(f"主力净流入排名模块不可用：{e}")
            if need_hot_theme:
                try:
                    from utils.hot_theme_selection_ctx import load_hot_theme_map
                except Exception as e:
                    load_hot_theme_map = None  # type: ignore[assignment]
                    self.debug_info.emit(f"热门题材模块不可用：{e}")
            if need_em:
                try:
                    from utils.eastmoney_board_rank_ctx import load_em_board_hot_map
                except Exception as e:
                    load_em_board_hot_map = None  # type: ignore[assignment]
                    self.debug_info.emit(f"东财连续热门模块不可用：{e}")
            hist_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history_data")
            code6_to_stock_keys: Dict[str, List[str]] = {}
            for sc in unique_stocks:
                c6 = _stock_code6(sc)
                if c6:
                    code6_to_stock_keys.setdefault(c6, []).append(sc)

            # 按「选股日」交错：当日热门加载完立刻筛池内票，不预加载全部日期
            # 进度按「日」固定总量，避免边跑边加 total 导致前几天就到 99%/100%、加载下一日时假死
            PROGRESS_PER_DAY = 1000
            total_units = max(1, n_days) * PROGRESS_PER_DAY
            self.debug_info.emit(
                f"按日交错筛选：{n_days} 个选股日 × {len(enabled_rules)} 条规则"
                f"（全市场底座 {total_stocks_universe} 只；日线按票缓存）"
            )
            logger.info("按日交错筛选开始…")
            self.progress_updated.emit(0, total_units, "准备中")

            daily_cache: Dict[str, object] = {}
            stocks_touched: Set[str] = set()

            for day_i, as_of_d in enumerate(screen_as_of_list):
                if not self.is_running:
                    break
                key = "today" if as_of_d is None else as_of_d.strftime("%Y-%m-%d")
                screen_str = (
                    date.today().strftime("%Y-%m-%d")
                    if as_of_d is None
                    else as_of_d.strftime("%Y-%m-%d")
                )
                day_base = day_i * PROGRESS_PER_DAY
                day_cands: Set[str] = set()
                self.progress_updated.emit(
                    day_base + 10, total_units, f"加载上下文 {key}"
                )

                if need_inflow:
                    if load_inflow_rank_map is None:
                        inflow_rank_cache[key] = {}
                    else:
                        try:
                            mp = load_inflow_rank_map(as_of_d, history_dir=hist_dir)
                        except Exception as e:
                            mp = {}
                            self.debug_info.emit(f"加载净流入排名失败 {key}: {e}")
                        inflow_rank_cache[key] = mp
                        if not mp:
                            self.debug_info.emit(f"选股日 {key} 无主力净流入 CSV（A/B 档将无命中）")
                        else:
                            self.debug_info.emit(f"选股日 {key} 已加载净流入排名 {len(mp)} 只")
                else:
                    inflow_rank_cache[key] = {}

                if need_hot_theme:
                    if load_hot_theme_map is None:
                        hot_theme_cache[key] = {}
                    else:
                        try:
                            ht = load_hot_theme_map(as_of_d, history_dir=hist_dir, top_n=10)
                        except Exception as e:
                            ht = {}
                            self.debug_info.emit(f"加载热门题材失败 {key}: {e}")
                        hot_theme_cache[key] = ht if isinstance(ht, dict) else {}
                        union_n = len((ht or {}).get("union_codes") or [])
                        src = str((ht or {}).get("source_date") or "")
                        if union_n <= 0:
                            self.debug_info.emit(
                                f"选股日 {key} 无十大热门板块/概念成员（热门题材规则将无命中）"
                            )
                        else:
                            self.debug_info.emit(
                                f"选股日 {key} 已加载热门题材成员 {union_n} 只"
                                + (f"（涨停日={src}）" if src else "")
                            )
                            day_cands |= {
                                str(c).zfill(6)
                                for c in ((ht or {}).get("union_codes") or set())
                                if c
                            }
                else:
                    hot_theme_cache[key] = {}

                if need_em:
                    if load_em_board_hot_map is None:
                        em_board_hot_cache[key] = {}
                    else:
                        try:
                            emh = load_em_board_hot_map(
                                as_of_d,
                                top_n=em_top_n,
                                rs_top_k=em_rs_top_k,
                                min_members=em_min_members,
                                arms=sorted(em_arms) if em_arms else None,
                                elig_bands=elig_bands or None,
                            )
                        except Exception as e:
                            emh = {"error": str(e), "pool_codes": set(), "code_hits": {}}
                            self.debug_info.emit(f"加载东财连续热门失败 {key}: {e}")
                        em_board_hot_cache[key] = emh if isinstance(emh, dict) else {}
                        em_err = str((emh or {}).get("error") or "").strip()
                        pool_n = len((emh or {}).get("pool_codes") or [])
                        new_pool_n = len((emh or {}).get("new_only_pool_codes") or [])
                        today_pool_n = len((emh or {}).get("today_pool_codes") or [])
                        sec_n = len((emh or {}).get("continuous_sectors") or [])
                        con_n = len((emh or {}).get("continuous_concepts") or [])
                        new_sec_n = len((emh or {}).get("new_only_sectors") or [])
                        new_con_n = len((emh or {}).get("new_only_concepts") or [])
                        today_sec_n = len((emh or {}).get("today_sectors") or [])
                        today_con_n = len((emh or {}).get("today_concepts") or [])
                        mv_n = len((emh or {}).get("float_mv_yi") or {})
                        prev_ds = str((emh or {}).get("prev_date") or "")
                        em_day_cands = _em_candidate_codes6(
                            emh if isinstance(emh, dict) else {},
                            arms=em_arms,
                            elig_bands=elig_bands,
                        )
                        day_cands |= em_day_cands
                        if em_err:
                            self.debug_info.emit(f"选股日 {key} 东财热门不可用：{em_err}")
                        else:
                            arm_txt = ",".join(sorted(em_arms)) if em_arms else "all"
                            parts = [f"臂={arm_txt}"]
                            if "continuous" in em_arms:
                                parts.append(f"连续板块{sec_n}/概念{con_n}→RS池{pool_n}")
                            if "new_only" in em_arms:
                                parts.append(
                                    f"仅今日板块{new_sec_n}/概念{new_con_n}→RS池{new_pool_n}"
                                )
                            if "today" in em_arms:
                                if elig_bands:
                                    parts.append(
                                        f"今日Elig标签 板块{today_sec_n}/概念{today_con_n}"
                                        f"→RS池{today_pool_n}"
                                    )
                                else:
                                    parts.append(
                                        f"今日板块{today_sec_n}/概念{today_con_n}→全RS池{today_pool_n}"
                                    )
                            parts.append(f"流通市值{mv_n}只")
                            if prev_ds:
                                parts.append(f"D-1={prev_ds}")
                            self.debug_info.emit(f"选股日 {key} 东财热门：" + "；".join(parts))
                else:
                    em_board_hot_cache[key] = {}

                if need_em or need_hot_theme:
                    day_stocks: List[str] = []
                    seen_sc: Set[str] = set()
                    for c6 in sorted(day_cands):
                        for sc in code6_to_stock_keys.get(c6, []):
                            if sc not in seen_sc:
                                seen_sc.add(sc)
                                day_stocks.append(sc)
                    if not day_stocks and day_cands:
                        self.debug_info.emit(
                            f"选股日 {key} 池内 {len(day_cands)} 只与股票列表无交集，跳过筛选"
                        )
                    elif not day_stocks:
                        self.debug_info.emit(f"选股日 {key} 候选为空，跳过筛选")
                else:
                    day_stocks = list(unique_stocks)

                day_n = len(day_stocks)
                self.debug_info.emit(
                    f"选股日 {key} 就绪 → 立即筛选 {day_n} 只"
                    f"（进度日 {day_i + 1}/{n_days}）"
                )
                self.progress_updated.emit(
                    day_base + 50, total_units, f"筛选 {key} · {day_n}只"
                )

                for stock_i, stock_code in enumerate(day_stocks):
                    if not self.is_running:
                        break
                    stock_name = stock_dict[stock_code]["name"]
                    sectors = stock_dict[stock_code]["sectors"]
                    stocks_touched.add(stock_code)

                    if stock_code not in daily_cache:
                        daily_cache[stock_code] = self._get_daily_data(
                            stock_code, through_date=fetch_through, apply_as_of_slice=False
                        )
                        df0 = daily_cache[stock_code]
                        if df0 is None or getattr(df0, "empty", True):
                            skip_no_daily += 1
                    daily_full = daily_cache[stock_code]

                    processed_units += 1
                    # 日进度：50–999；全日完成顶到 (day_i+1)*1000
                    frac = int(((stock_i + 1) / max(day_n, 1)) * 949)
                    pos = min(day_base + 50 + frac, (day_i + 1) * PROGRESS_PER_DAY - 1)
                    if (
                        stock_i == 0
                        or stock_i + 1 == day_n
                        or (processed_units - self._last_progress_emit_unit)
                        >= self._progress_emit_every
                    ):
                        self.progress_updated.emit(
                            pos,
                            total_units,
                            f"{stock_code} [{screen_str}]",
                        )
                        self._last_progress_emit_unit = processed_units

                    if daily_full is None or getattr(daily_full, "empty", True):
                        continue
                    if as_of_d is None:
                        dd = daily_full
                    else:
                        dd = daily_full[daily_full["date"] <= as_of_d].copy()
                    if dd.empty:
                        continue

                    sectors_str = ";".join(sorted(sectors))
                    for r in enabled_rules:
                        if not self.is_running:
                            break
                        rid = str(r.get("id") or "")
                        entry = compiled_rules.get(rid)
                        if not isinstance(entry, dict):
                            continue
                        fn = entry.get("fn")
                        if not callable(fn):
                            continue
                        self._apply_rule_calendar_params(entry["params"])
                        if not self._setup_calendars_for_as_of(as_of_d):
                            skip_calendar += 1
                            continue
                        ctx = {
                            "params": self._current_rule_param_ctx(),
                            "inflow_rank": inflow_rank_cache.get(key) or {},
                            "hot_theme": hot_theme_cache.get(key) or {},
                            "em_board_hot": em_board_hot_cache.get(key) or {},
                        }
                        try:
                            rule_calls += 1
                            ok, extra = fn(stock_code, stock_name, sectors, dd, as_of_d, ctx)
                        except Exception as e:
                            err = str(e)
                            extra = {"_error": err}
                            if "__import__" in err or "import" in err.lower():
                                try:
                                    rule_code = str(r.get("code") or "")
                                    bad_lines: List[str] = []
                                    for ln in rule_code.splitlines():
                                        lnl = ln.strip()
                                        if (
                                            lnl.startswith("import ")
                                            or lnl.startswith("from ")
                                            or "__import__" in lnl
                                        ):
                                            bad_lines.append(ln.strip())
                                            if len(bad_lines) >= 3:
                                                break
                                    if bad_lines:
                                        extra["_rule_import_snippet"] = "; ".join(bad_lines)
                                    else:
                                        head_lines: List[str] = []
                                        for ln in rule_code.splitlines()[:5]:
                                            head_lines.append(ln.strip())
                                        head = " | ".join(head_lines)
                                        extra["_rule_code_head"] = head[:220]
                                        extra["_rule_code_has_import_like"] = (
                                            ("import " in rule_code)
                                            or ("from " in rule_code)
                                            or ("__import__" in rule_code)
                                        )
                                except Exception:
                                    pass
                            if "__import__" in err:
                                try:
                                    tb_lines = traceback.format_exc().splitlines()
                                    extra["_tb_head"] = " | ".join(tb_lines[:10]).strip()
                                except Exception:
                                    pass
                            ok = False
                        rn = str(r.get("name", "未命名"))
                        st = "通过" if ok else "未通过"
                        if isinstance(extra, dict) and extra.get("_error"):
                            st = f"异常({extra.get('_error')})"
                        import_hint = ""
                        if isinstance(extra, dict) and str(extra.get("_error") or "") == "__import__":
                            import_hint = "（可能包含 import/from，被安全策略禁止）"
                        round_line = (
                            f"[本轮] {screen_str} {stock_code} 规则「{rn}」(id={rid}): {st}"
                        )
                        if import_hint:
                            round_line += import_hint
                        should_emit_round = (
                            ok
                            or (isinstance(extra, dict) and bool(extra.get("_error")))
                            or (self._round_debug_emit_count < self._round_debug_emit_limit)
                        )
                        if should_emit_round:
                            self.debug_info.emit(round_line)
                            self._diag_snippets.append(round_line)
                            self._round_debug_emit_count += 1
                        if isinstance(extra, dict) and extra.get("_rule_import_snippet"):
                            s = str(extra.get("_rule_import_snippet"))
                            short = s if len(s) <= 220 else (s[:220] + "…")
                            emit_line = f"[安全] 该规则代码疑似包含 import/from：{short}"
                            if self._round_debug_emit_count < self._round_debug_emit_limit:
                                self.debug_info.emit(emit_line)
                                self._diag_snippets.append(emit_line)
                                self._round_debug_emit_count += 1
                        if isinstance(extra, dict) and extra.get("_rule_code_head"):
                            head = str(extra.get("_rule_code_head"))
                            has_like = extra.get("_rule_code_has_import_like")
                            emit_line = (
                                f"[安全] 规则代码头部（用于核对触发 __import__ 的实际代码）："
                                f"{head}（has import/from/__import__={has_like}）"
                            )
                            if self._round_debug_emit_count < self._round_debug_emit_limit:
                                self.debug_info.emit(emit_line)
                                self._diag_snippets.append(emit_line)
                                self._round_debug_emit_count += 1
                        if isinstance(extra, dict) and extra.get("_tb_head"):
                            tb = str(extra.get("_tb_head"))
                            emit_line = f"[定位] __import__ 异常调用栈(前几行)：{tb}"
                            if self._round_debug_emit_count < self._round_debug_emit_limit:
                                self.debug_info.emit(emit_line)
                                self._diag_snippets.append(emit_line)
                                self._round_debug_emit_count += 1
                        if ok:
                            hit_extra = dict(extra) if isinstance(extra, dict) else {}
                            export_params = entry.get("export_params")
                            if isinstance(export_params, dict):
                                for pk, pv in export_params.items():
                                    hit_extra.setdefault(pk, pv)
                            rec = {
                                "rule_id": rid,
                                "rule_name": r.get("name", "未命名规则"),
                                "stock_code": stock_code,
                                "stock_name": stock_name,
                                "sectors": sectors_str,
                                "as_of": screen_str,
                                "extra": hit_extra,
                            }
                            self.stock_found.emit(rec)
                            rule_counts[rid] = int(rule_counts.get(rid, 0)) + 1

                total_found = sum(rule_counts.values()) if rule_counts else 0
                if total_found > 0 and total_found % 10 == 0:
                    self.debug_info.emit(f"已找到 {total_found} 条（按规则累计）")
                self.progress_updated.emit(
                    (day_i + 1) * PROGRESS_PER_DAY,
                    total_units,
                    f"{key} 完成",
                )

            total_stocks = len(stocks_touched) if stocks_touched else total_stocks_universe
            self.progress_updated.emit(total_units, total_units, "完成")

            # 如果被停止，发送停止信息
            if not self.is_running:
                self.debug_info.emit("筛选已停止")
            
            total_hits = sum(int(v or 0) for v in rule_counts.values()) if rule_counts else 0
            if total_hits == 0 and self.is_running:
                lines = [
                    "[诊断] 本次筛选未命中任何股票。常见原因：",
                    f"  • 已执行规则判断次数：{rule_calls}（若恒为 0，多为规则未编译成功）",
                    f"  • 因「交易日历不足」跳过的 (股票×选股日) 步数：{skip_calendar}",
                    f"  • 完全无日线数据的股票数：{skip_no_daily} / {total_stocks}",
                    "  • （模式一内置）检测到涨停候选（is_limit_up）的检查次数："
                    f"{getattr(self, '_mode1_limitup_detected_checks', 0)}",
                    "  • 若日历/日线异常：请确认 data/daily_cache 已由大 QMT 同步（manifest.json），"
                    f"目录={get_cache_dir()}；交易日历走 akshare/xtdata 缓存。",
                    "  • 若数据正常：说明当前备选股在设定参数下确实不满足模式条件（属正常）。",
                    "  • 模式一/二详细「窗口/逐日/淘汰」全文：设置环境变量 SECTOR_FILTER_MODE1_VERBOSE_LOG=1 后重跑；"
                    "否则默认仅一行摘要。其它日志见「[本轮]」「[提示]」等。",
                ]
                snip = getattr(self, "_diag_snippets", None) or []
                if snip:
                    lines.append("  • 本轮执行摘录（与上方日志相同，便于复制）：")
                    for s in snip[-12:]:
                        lines.append(f"    {s}")
                if enabled_rules and not compiled_rules:
                    lines.insert(1, "  • 所有启用规则的代码编译失败，请查看上方「编译失败」提示。")
                self.debug_info.emit("\n".join(lines))
            
            self.finished.emit(rule_counts)
            
        except Exception as e:
            self.error_occurred.emit(f"筛选过程出错: {str(e)}")
            self.finished.emit(rule_counts)


class SectorStockFilterDialog(QDialog):
    """板块股票筛选对话框"""

    # 所有规则选股结果统一追加的列（与 auto_limit_up_filter.add_concept_rank_columns、主力净流入文件一致）
    RESULT_CONCEPT_INFLOW_COLUMNS = (
        "当日最多涨停概念",
        "该概念当日涨停数",
        "该概念当日排名",
        "主力净流入",
    )
    RESULT_RULE_CONFIG_COLUMNS = ("P", "N", "M", "L")
    RESULT_CONDITION_OUTCOME_COLUMNS = (
        "REQUIRE_PRIOR_LU_IN_L",
        "REQUIRE_OLD_HIGH",
        "REJECT_PRIOR_LIMIT_UP",
        "REQUIRE_OBVIOUS_NEW_HIGH",
        "REQUIRE_LOWER_SHADOW",
        "REQUIRE_BOLL_BREAK",
        "REQUIRE_MA_SUPPORT_AFTER",
    )
    RESULT_RULE_PARAM_COLUMNS = RESULT_RULE_CONFIG_COLUMNS + RESULT_CONDITION_OUTCOME_COLUMNS

    def _order_result_extra_columns(self, extra_cols: List[str]) -> List[str]:
        """规则配置列靠前，条件实际结果列次之，概念/净流入列靠后。"""
        config_part = [k for k in self.RESULT_RULE_CONFIG_COLUMNS if k in extra_cols]
        outcome_part = [k for k in self.RESULT_CONDITION_OUTCOME_COLUMNS if k in extra_cols]
        tail_part = [k for k in self.RESULT_CONCEPT_INFLOW_COLUMNS if k in extra_cols]
        middle_part = [
            k
            for k in extra_cols
            if k not in config_part and k not in outcome_part and k not in tail_part
        ]
        return config_part + outcome_part + middle_part + tail_part

    @staticmethod
    def _is_first_board_rule_name(rule_name: str) -> bool:
        s = str(rule_name or "")
        return ("模式一" in s) or ("涨停后" in s) or ("首板" in s)

    def _enrich_stock_result_extra(self, stock_code: str, as_of: str, extra: Optional[Dict[str, object]]) -> Dict[str, object]:
        """按涨停日/信号日（无则用选股日）补全概念排名与主力净流入，供结果表与导出使用。"""
        merged: Dict[str, object] = dict(extra) if isinstance(extra, dict) else {}
        for k in self.RESULT_CONCEPT_INFLOW_COLUMNS:
            merged.setdefault(k, "")

        date_ref = (
            merged.get("涨停日期")
            or merged.get("放量反包日期")
            or merged.get("信号日期")
            or as_of
        )
        date_ref = str(date_ref or "").strip()
        if not date_ref:
            date_ref = str(as_of or "").strip()
        if len(date_ref) >= 10 and date_ref[4] == "-":
            date_ref = date_ref[:10]

        code6 = str(stock_code or "").strip()
        if code6.isdigit():
            code6 = code6.zfill(6)

        root = os.path.dirname(os.path.abspath(__file__))
        hist = os.path.join(root, "history_data")

        if add_concept_rank_columns:
            try:
                one = pd.DataFrame([{"code": code6, "limit_date": date_ref}])
                out = add_concept_rank_columns(one, hist)
                if out is not None and not out.empty:
                    r0 = out.iloc[0]
                    merged["当日最多涨停概念"] = str(r0.get("当日最多涨停概念", "") or "")
                    for ck, key in (
                        ("该概念当日涨停数", "该概念当日涨停数"),
                        ("该概念当日排名", "该概念当日排名"),
                    ):
                        v = r0.get(key, "")
                        try:
                            if v == "" or v is None or (isinstance(v, float) and pd.isna(v)):
                                merged[ck] = ""
                            else:
                                merged[ck] = str(int(v))
                        except Exception:
                            merged[ck] = str(v) if v is not None else ""
            except Exception:
                logger.debug("补全概念排名列失败", exc_info=True)

        if InflowAdder:
            try:
                adder = InflowAdder()
                dmap = adder._load_inflow_file(date_ref, base_dir=hist)
                info = dmap.get(code6) if dmap else None
                merged["主力净流入"] = (info or {}).get("display", "") if info else ""
            except Exception:
                logger.debug("补全主力净流入失败", exc_info=True)

        return merged

    def __init__(self, parent=None, initial_as_of: Optional[date] = None, auto_run: bool = False):
        super().__init__(parent)
        # 启动时预设的选股基准日（如命令行 --as-of）；None 表示由界面默认为今天
        self._initial_as_of = initial_as_of
        self._auto_run = bool(auto_run)
        self._auto_run_started = False
        self._auto_run_finished = False
        self._auto_run_pending = bool(auto_run)
        self._auto_run_retry_count = 0
        # 统一命名风格：蚂蚁量化选股系统
        self.setWindowTitle("蚂蚁量化选股系统")
        # 使用与主程序和策略生成器相同的图标 ant.ico（若存在）
        root_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(root_dir, "ant.ico")
        if not os.path.exists(icon_path):
            # 若当前目录没有，则尝试项目根目录
            project_root = os.path.dirname(os.path.abspath(__file__))
            icon_path = os.path.join(project_root, "ant.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.setMinimumSize(800, 600)
        self.resize(1000, 1000)
        
        # 设置窗口标志：显示最大化/最小化按钮，去掉帮助按钮
        flags = (Qt.Window | Qt.WindowMinimizeButtonHint | 
                 Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
        self.setWindowFlags(flags)
        
        self.filter_thread = None
        # 存储筛选结果（已移到setup_ui中，使用多模式结果列表）
        self.countdown_timer = None  # 倒计时定时器
        self.countdown_seconds = 30  # 倒计时秒数
        
        # 存储所有板块数据（需要在setup_ui之前初始化，因为load_sectors_to_list会使用）
        self.all_sectors_list = []  # QMT 申万/概念等板块列表
        self.sector_checkboxes = {}  # {sector_name: QCheckBox}
        
        self._sector_store = get_qmt_sector_store()
        
        # 存储股票列表的板块信息 {stock_code: [sector1, sector2, ...]}
        self._stock_sectors_map = {}
        # 板块生成的完整备选底池（开关过滤前）[(code, name, sectors_str), ...]
        self._stock_list_base: List[Tuple[str, str, str]] = []

        # 选股规则（可增删改）
        self.rules: List[Dict[str, object]] = load_sector_rules()
        self._current_rule_id: Optional[str] = None
        self._rule_code_saved_snapshot: str = ""
        self._rule_code_dirty: bool = False
        self._rule_list_prev_row: int = -1
        self._rule_switch_guard: bool = False
        self._result_tables_by_rule: Dict[str, QTableWidget] = {}
        self._result_rows_by_rule: Dict[str, List[Dict[str, object]]] = {}
        # rule_id -> 结果 Tab 下标（避免仅靠规则名匹配 Tab 标题失败）
        self._rule_tab_index: Dict[str, int] = {}
        
        self.setup_ui()
        # auto-run 在 update_stock_list 完成后再触发，避免早于股票列表加载（原 200ms 定时过短）
    
    def setup_ui(self):
        """设置UI"""
        # 外层纵向滚动：仅包「板块 + 股票列表」；规则代码区单独占一行，避免与 QScrollArea 嵌套导致编辑框内无法滚动
        outer_layout = QVBoxLayout()
        self._main_scroll = QScrollArea()
        self._main_scroll.setWidgetResizable(True)
        self._main_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._main_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        inner = QWidget()
        layout = QVBoxLayout(inner)
        
        # 板块选择区域
        self.sector_group = QGroupBox("板块选择（可多选）")
        sector_layout = QVBoxLayout()
        
        # 搜索框和筛选模式选择
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索板块："))
        self.sector_search = QLineEdit()
        self.sector_search.setPlaceholderText("输入关键词过滤板块...")
        self.sector_search.textChanged.connect(self.filter_sectors)
        self.sector_search.returnPressed.connect(lambda: self.filter_sectors(self.sector_search.text()))
        search_layout.addWidget(self.sector_search)
        
        # 全选/取消全选按钮
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(self.select_all_sectors)
        search_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("取消全选")
        deselect_all_btn.clicked.connect(self.deselect_all_sectors)
        search_layout.addWidget(deselect_all_btn)
        
        search_layout.addStretch()
        
        # 筛选模式选择（并集/交集）
        search_layout.addWidget(QLabel("筛选模式："))
        from PyQt5.QtWidgets import QRadioButton, QButtonGroup
        self.sector_mode_group = QButtonGroup()
        self.union_mode_radio = QRadioButton("并集（属于任意一个板块）")
        self.intersection_mode_radio = QRadioButton("交集（同时属于所有板块）")
        self.union_mode_radio.setChecked(True)  # 默认并集
        self.sector_mode_group.addButton(self.union_mode_radio, 0)
        self.sector_mode_group.addButton(self.intersection_mode_radio, 1)
        search_layout.addWidget(self.union_mode_radio)
        search_layout.addWidget(self.intersection_mode_radio)
        # 当筛选模式改变时，重新更新股票列表
        self.union_mode_radio.toggled.connect(self.on_sector_mode_changed)
        self.intersection_mode_radio.toggled.connect(self.on_sector_mode_changed)
        
        sector_layout.addLayout(search_layout)
        
        # 板块列表（使用QScrollArea + QGridLayout，一行显示多个）
        self.sector_scroll = QScrollArea()
        self.sector_scroll.setWidgetResizable(True)
        self.sector_scroll.setMaximumHeight(300)
        self.sector_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # 创建容器widget和网格布局
        self.sector_container = QWidget()
        self.sector_grid_layout = QGridLayout()
        self.sector_grid_layout.setSpacing(5)  # 设置间距
        self.sector_container.setLayout(self.sector_grid_layout)
        self.sector_scroll.setWidget(self.sector_container)
        
        sector_layout.addWidget(self.sector_scroll)
        
        # 统计标签
        self.sector_count_label = QLabel("已选择: 0 个板块")
        sector_layout.addWidget(self.sector_count_label)
        
        self.sector_group.setLayout(sector_layout)
        layout.addWidget(self.sector_group)
        
        # 选中板块对应的股票列表
        self.stock_list_group = QGroupBox("选中板块对应的股票列表 (0 只)")
        stock_list_layout = QVBoxLayout()
        
        # 股票列表搜索框与操作按钮
        stock_search_layout = QHBoxLayout()
        stock_search_layout.addWidget(QLabel("搜索股票："))
        self.stock_search_input = QLineEdit()
        self.stock_search_input.setPlaceholderText("输入股票代码或名称过滤...")
        self.stock_search_input.textChanged.connect(self.filter_stock_list)
        self.stock_search_input.installEventFilter(self)
        stock_search_layout.addWidget(self.stock_search_input)
        
        # 全选/取消全选按钮
        stock_select_all_btn = QPushButton("全选")
        stock_select_all_btn.clicked.connect(self.select_all_stocks)
        stock_search_layout.addWidget(stock_select_all_btn)
        
        stock_deselect_all_btn = QPushButton("取消全选")
        stock_deselect_all_btn.clicked.connect(self.deselect_all_stocks)
        stock_search_layout.addWidget(stock_deselect_all_btn)
        
        # 从文件选择按钮
        stock_load_file_btn = QPushButton("从文件选择")
        stock_load_file_btn.clicked.connect(self.load_stocks_from_file)
        stock_search_layout.addWidget(stock_load_file_btn)

        # 导出当前股票列表按钮
        stock_export_btn = QPushButton("导出到文件")
        stock_export_btn.clicked.connect(self.export_stock_list_to_file)
        stock_search_layout.addWidget(stock_export_btn)

        self.exclude_st_checkbox = QCheckBox("排除ST股")
        self.exclude_st_checkbox.setToolTip(
            "开关：开启时备选池不含 ST/*ST；关闭后自动加回。状态会保存，下次启动恢复。"
        )
        self.exclude_star_bj_checkbox = QCheckBox("排除科创板和北交所")
        self.exclude_star_bj_checkbox.setToolTip(
            "开关：开启时备选池不含科创板(688/689)与北交所(4/8/920)；关闭后自动加回。"
            "状态会保存，下次启动恢复。"
        )
        self._load_stock_exclude_prefs()
        self.exclude_st_checkbox.stateChanged.connect(self._on_stock_exclude_filter_changed)
        self.exclude_star_bj_checkbox.stateChanged.connect(self._on_stock_exclude_filter_changed)
        stock_search_layout.addWidget(self.exclude_st_checkbox)
        stock_search_layout.addWidget(self.exclude_star_bj_checkbox)
        
        stock_list_layout.addLayout(stock_search_layout)
        
        self.stock_list_table = QTableWidget()
        self.stock_list_table.setColumnCount(4)
        self.stock_list_table.setHorizontalHeaderLabels(["选择", "股票代码", "股票名称", "所属板块"])
        
        # 设置表格属性
        stock_list_header = self.stock_list_table.horizontalHeader()
        stock_list_header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        stock_list_header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        stock_list_header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        stock_list_header.setSectionResizeMode(3, QHeaderView.Stretch)
        
        self.stock_list_table.setMaximumHeight(200)
        stock_list_layout.addWidget(self.stock_list_table)
        
        # 已选择股票数量标签
        self.stock_selected_count_label = QLabel("已选择: 0 只股票")
        stock_list_layout.addWidget(self.stock_selected_count_label)
        
        # 存储股票复选框的引用 {stock_code: QCheckBox}
        self.stock_checkboxes = {}
        
        self.stock_list_group.setLayout(stock_list_layout)
        layout.addWidget(self.stock_list_group)

        # 仅板块/股票放入外层 QScrollArea。规则代码编辑框若放在同一 QScrollArea 内会与外层滚动嵌套，
        # QPlainTextEdit 被文档高度撑满，内部无法用自己的滚动条滚动代码。
        self._main_scroll.setWidget(inner)
        outer_layout.addWidget(self._main_scroll, 0)
        
        # 选股规则区域（替代固定的三种模式）
        rules_group = QGroupBox("选股规则（可增删改、导入导出；每条规则一个 JSON 文件，名称不可重复）")
        rules_layout = QVBoxLayout()

        # 内置模式窗口参数：单行节省纵向空间，完整说明放在悬停提示里（避免规则区总高度过小把左侧按钮挤出可视区）
        params_layout = QHBoxLayout()
        _hint_full = (
            "规则参数与逻辑均写在规则代码内（自包含，不调用 ctx['builtin_check'] 等引擎函数）。"
            "可在代码顶部用常量定义 N、M、L 及 REQUIRE_* 开关（True/False/None）；"
            "True/False 为互斥条件，None 表示忽略该条件；"
            "REQUIRE_LOWER_SHADOW、REQUIRE_BOLL_BREAK、REQUIRE_MA_SUPPORT_AFTER 等为可反相开关；"
            "select(...) 接收 daily_data（截至选股日的日线）、as_of_date；"
            "ctx 含 params（引擎读取 N/M 用于日历预取）、inflow_rank"
            "（当日净流入占流通% 排名 {code6:{rank,pct,name}}，供主力净流入 A/B 档）、"
            "hot_theme（十大热门题材）、em_board_hot（东财连续2日热门+组内RS）。"
        )
        _hint_label = QLabel("规则逻辑全部写在代码内；顶部定义 N、M 等参数（悬停查看说明）")
        _hint_label.setToolTip(_hint_full)
        params_layout.addWidget(_hint_label)
        params_layout.addStretch()
        rules_layout.addLayout(params_layout)

        # 规则代码书写说明（整行），避免压在右侧栏内把「列表 vs 代码框」撑成不同高度
        _rule_code_hint_full = (
            "规则代码须定义 select(stock_code, stock_name, sectors, daily_data, as_of_date, ctx)"
            " → (bool, extra_dict)；全部选股逻辑写在本文件内，勿调用规则代码外的引擎函数。"
        )
        _rule_code_top = QLabel("须定义 select(...)→(bool,extra)；逻辑自包含；详情见悬停说明")
        _rule_code_top.setToolTip(_rule_code_hint_full)
        rules_layout.addWidget(_rule_code_top)

        splitter = QSplitter(Qt.Horizontal)

        # 左侧：规则列表 + 操作按钮（与右侧「规则代码」标题行 + 编辑框 + 底栏对称，便于同高）
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("规则列表："))
        self.rule_list = QListWidget()
        self.rule_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        left_layout.addWidget(self.rule_list, 1)
        rule_btn_bar = QWidget()
        rule_btn_bar.setMinimumHeight(44)
        rule_action_btns = QHBoxLayout(rule_btn_bar)
        rule_action_btns.setContentsMargins(0, 4, 0, 0)
        self.rule_add_btn = QPushButton("新增")
        self.rule_del_btn = QPushButton("删除")
        self.rule_rename_btn = QPushButton("重命名")
        self.rule_dup_btn = QPushButton("复制")
        self.rule_import_btn = QPushButton("导入规则")
        self.rule_export_btn = QPushButton("导出规则")
        rule_action_btns.addWidget(self.rule_add_btn)
        rule_action_btns.addWidget(self.rule_del_btn)
        rule_action_btns.addWidget(self.rule_rename_btn)
        rule_action_btns.addWidget(self.rule_dup_btn)
        rule_action_btns.addWidget(self.rule_import_btn)
        rule_action_btns.addWidget(self.rule_export_btn)
        rule_action_btns.addStretch()
        left_layout.addWidget(rule_btn_bar, 0)
        left_widget.setMinimumWidth(480)
        left_widget.setLayout(left_layout)
        splitter.addWidget(left_widget)

        # 右侧：与左侧同结构 — 单行标题 + 主区域 + 底栏（底栏高度与左侧按钮条一致）
        right_widget = QWidget()
        right_layout = QVBoxLayout()

        _lbl_rule_code = QLabel("规则代码：")
        _lbl_rule_code.setToolTip(_rule_code_hint_full)
        right_layout.addWidget(_lbl_rule_code)
        self.rule_code_unsaved_hint = QLabel(
            "⚠ 规则代码已修改，尚未保存！请点击下方「保存当前规则代码」，"
            "修改才会写入磁盘并在下次选股/重启后生效。"
        )
        self.rule_code_unsaved_hint.setWordWrap(True)
        self.rule_code_unsaved_hint.setStyleSheet(_RULE_CODE_UNSAVED_BANNER_STYLE)
        self.rule_code_unsaved_hint.hide()
        right_layout.addWidget(self.rule_code_unsaved_hint)
        self.rule_code_editor = QPlainTextEdit()
        self.rule_code_editor.setPlaceholderText(
            "# 参数\nN = 3\nM = 30\n\n"
            "def select(stock_code, stock_name, sectors, daily_data, as_of_date, ctx):\n"
            "    # 在此实现完整逻辑（使用 daily_data / stock_code / as_of_date）\n"
            "    return False, {}\n"
        )
        font = QFont("Consolas", 10)
        self.rule_code_editor.setFont(font)
        self.rule_code_editor.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.rule_code_editor.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.rule_code_editor.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        right_layout.addWidget(self.rule_code_editor, 1)

        right_btn_bar = QWidget()
        right_btn_bar.setMinimumHeight(44)
        right_btns = QHBoxLayout(right_btn_bar)
        right_btns.setContentsMargins(0, 4, 0, 0)
        self.rule_save_current_btn = QPushButton("保存当前规则代码")
        right_btns.addStretch()
        right_btns.addWidget(self.rule_save_current_btn)
        right_layout.addWidget(right_btn_bar, 0)
        right_widget.setLayout(right_layout)
        splitter.addWidget(right_widget)

        # 左侧加宽；rules_group 总高不能过小，否则分割器内高度不足会把底部按钮行挤没
        splitter.setSizes([500, 500])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setMinimumHeight(280)
        rules_layout.addWidget(splitter)
        rules_group.setLayout(rules_layout)
        rules_group.setMinimumHeight(380)
        outer_layout.addWidget(rules_group, 1)

        bottom_inner = QWidget()
        bottom_layout = QVBoxLayout(bottom_inner)
        
        as_of_layout = QHBoxLayout()
        _lbl_start = QLabel("选股基准日（起）：")
        _lbl_start.setToolTip("单日回测请把右侧「止」改成与「起」同一天；若止晚于起，会按区间内每个交易日各算一遍。")
        as_of_layout.addWidget(_lbl_start)
        self.as_of_date_edit = QDateEdit()
        self.as_of_date_edit.setCalendarPopup(True)
        self.as_of_date_edit.setDisplayFormat("yyyy-MM-dd")
        if self._initial_as_of is not None:
            d0 = self._initial_as_of
            self.as_of_date_edit.setDate(QDate(d0.year, d0.month, d0.day))
        else:
            self.as_of_date_edit.setDate(QDate.currentDate())
        as_of_layout.addWidget(self.as_of_date_edit)
        _lbl_end = QLabel(
            "止（与起相同=单日；起止不同=区间内逐日选股，日志里会列出每个选股日）"
        )
        _lbl_end.setToolTip("常见误操作：只把「起」改成历史日，「止」仍为今天 → 会连今天一起算，最近N根K线会含今日。")
        as_of_layout.addWidget(_lbl_end)
        self.as_of_date_end_edit = QDateEdit()
        self.as_of_date_end_edit.setCalendarPopup(True)
        self.as_of_date_end_edit.setDisplayFormat("yyyy-MM-dd")
        if self._initial_as_of is not None:
            self.as_of_date_end_edit.setDate(QDate(d0.year, d0.month, d0.day))
        else:
            self.as_of_date_end_edit.setDate(QDate.currentDate())
        as_of_layout.addWidget(self.as_of_date_end_edit)
        self.as_of_today_btn = QPushButton("今天")
        self.as_of_today_btn.setToolTip("将选股基准日的起、止都设为今天")
        self.as_of_today_btn.clicked.connect(self._set_as_of_dates_to_today)
        as_of_layout.addWidget(self.as_of_today_btn)
        as_of_layout.addStretch()
        bottom_layout.addLayout(as_of_layout)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        
        self.start_button = QPushButton("开始筛选")
        self.start_button.clicked.connect(self.start_filter)
        button_layout.addWidget(self.start_button)
        
        self.stop_button = QPushButton("停止筛选")
        self.stop_button.clicked.connect(self.stop_filter)
        self.stop_button.setEnabled(False)  # 初始状态禁用
        button_layout.addWidget(self.stop_button)
        
        button_layout.addStretch()
        bottom_layout.addLayout(button_layout)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        bottom_layout.addWidget(self.progress_bar)
        
        # 状态标签（需要在load_sectors_to_list之前创建，因为该方法会使用它）
        self.status_label = QLabel("")
        bottom_layout.addWidget(self.status_label)

        # 调试/日志输出框：用于显示筛选过程与“不通过原因”等多行信息
        self.debug_output = QTextEdit()
        self.debug_output.setReadOnly(True)
        self.debug_output.setPlaceholderText("运行日志将在这里滚动显示…")
        self.debug_output.setMinimumHeight(64)
        self.debug_output.setMaximumHeight(120)
        self.debug_output.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        bottom_layout.addWidget(self.debug_output)
        
        # 结果表格区域（按“规则”动态生成标签页）
        from PyQt5.QtWidgets import QTabWidget
        self.result_tabs = QTabWidget()
        # 压低最小高度，把垂直空间让给上方规则区（含按钮行）；表格在 Tab 内仍可滚动
        self.result_tabs.setMinimumHeight(100)
        self.result_tabs.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        bottom_layout.addWidget(self.result_tabs)

        outer_layout.addWidget(bottom_inner, 0)

        # 每次开始筛选前会清空并重建
        self._result_tables_by_rule = {}
        self._result_rows_by_rule = {}
        
        self.setLayout(outer_layout)

        # 加载板块列表（在 status_label 等控件创建之后）
        self.load_sectors_to_list()

        # 初始化规则面板
        self._refresh_rule_list()
        self.rule_list.currentRowChanged.connect(self._on_rule_selected)
        # 右键菜单改名/删除/复制
        self.rule_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.rule_list.customContextMenuRequested.connect(self._on_rule_list_context_menu)
        # 点击前缀图标切换启用
        self.rule_list.viewport().installEventFilter(self)
        self.rule_add_btn.clicked.connect(self._add_rule)
        self.rule_del_btn.clicked.connect(self._delete_rule)
        self.rule_rename_btn.clicked.connect(self._rename_rule)
        self.rule_dup_btn.clicked.connect(self._duplicate_rule)
        self.rule_import_btn.clicked.connect(self._import_rule)
        self.rule_export_btn.clicked.connect(self._export_rule)
        self.rule_save_current_btn.clicked.connect(self._save_current_rule_code)
        self.rule_code_editor.textChanged.connect(self._on_rule_code_changed)
        
        # 已移除筛选条件初始化

    def _refresh_rule_list(self):
        """刷新左侧规则列表显示"""
        # 注意：不要在 setCurrentRow 时屏蔽信号，否则右侧编辑器不会加载代码，只会看到占位文本。
        self.rule_list.blockSignals(True)
        try:
            current_id = ""
            try:
                current_item = self.rule_list.currentItem()
                current_id = str(current_item.data(Qt.UserRole) or "") if current_item is not None else ""
            except Exception:
                current_id = ""

            self.rule_list.clear()
            for r in (self.rules or []):
                name = str(r.get("name") or "未命名")
                enabled = bool(r.get("enabled", True))
                prefix = "✅ " if enabled else "⏸ "
                item = QListWidgetItem(prefix + name)
                item.setData(Qt.UserRole, str(r.get("id") or ""))
                self.rule_list.addItem(item)
        finally:
            self.rule_list.blockSignals(False)

        # 尽量恢复到刷新前的选中项，否则默认选中第一条
        if self.rule_list.count() == 0:
            self._current_rule_id = None
            try:
                self.rule_code_editor.blockSignals(True)
                self.rule_code_editor.setPlainText("")
                self.rule_code_editor.blockSignals(False)
                self._capture_rule_code_snapshot("")
            except Exception:
                pass
            return

        target_row = -1
        if current_id:
            for i in range(self.rule_list.count()):
                it = self.rule_list.item(i)
                if str(it.data(Qt.UserRole) or "") == current_id:
                    target_row = i
                    break
        if target_row < 0:
            target_row = 0

        self.rule_list.setCurrentRow(target_row)
        # 保险：有些情况下 setCurrentRow 不触发时，手动加载
        self._on_rule_selected(self.rule_list.currentRow())

    def _get_rule_by_id(self, rid: str) -> Optional[Dict[str, object]]:
        for r in (self.rules or []):
            if str(r.get("id") or "") == str(rid):
                return r
        return None

    def _get_selected_rule_id(self) -> str:
        """从规则列表当前选中项获取 rule_id（比 _current_rule_id 更可靠）。"""
        try:
            item = self.rule_list.currentItem() if hasattr(self, "rule_list") else None
            rid = str(item.data(Qt.UserRole) or "") if item is not None else ""
            return rid
        except Exception:
            return str(self._current_rule_id or "")

    def _make_unique_rule_name(self, desired_name: str, exclude_rule_id: Optional[str] = None) -> str:
        """生成不重复的规则名；重复时自动追加 (2)/(3)/...（用于新增/复制/导入）。"""
        return make_unique_rule_name(desired_name, self.rules or [], exclude_rule_id)

    def _is_rule_name_taken(self, name: str, exclude_rule_id: Optional[str] = None) -> bool:
        """系统内是否已有同名规则（重命名时用于拒绝冲突）。"""
        return rule_name_is_taken(name, self.rules or [], exclude_rule_id)

    def eventFilter(self, obj, event):
        # 股票搜索框回车：只勾选当前筛选可见行，并吞掉回车，避免触发对话框默认按钮（如「开始筛选」）或板块区「全选」导致列表被整表刷新且丢失过滤
        if obj is self.stock_search_input and event.type() == QEvent.KeyPress:
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                self._apply_stock_search_enter()
                event.accept()
                return True
        # 仅处理规则列表的“点击前缀图标切换启用”
        try:
            if obj is self.rule_list.viewport() and event.type() == QEvent.MouseButtonPress:
                pos = event.pos()
                item = self.rule_list.itemAt(pos)
                if item is not None:
                    rect = self.rule_list.visualItemRect(item)
                    # 点击区域在最左侧一定宽度内，认为是点了“启用图标”
                    if (pos.x() - rect.x()) <= 26:
                        rid = str(item.data(Qt.UserRole) or "")
                        r = self._get_rule_by_id(rid)
                        if r is not None:
                            r["enabled"] = not bool(r.get("enabled", True))
                            save_single_sector_rule(r)
                            self._refresh_rule_list()
                        return True
        except Exception:
            pass
        return super().eventFilter(obj, event)

    def _set_rule_code_dirty(self, dirty: bool) -> None:
        self._rule_code_dirty = bool(dirty)
        if dirty:
            self.rule_code_unsaved_hint.setText(
                "⚠ 规则代码已修改，尚未保存！请点击下方「保存当前规则代码」，"
                "修改才会写入磁盘并在下次选股/重启后生效。"
            )
            self.rule_code_unsaved_hint.setStyleSheet(_RULE_CODE_UNSAVED_BANNER_STYLE)
            self.rule_code_unsaved_hint.show()
            self.rule_save_current_btn.setText("保存当前规则代码（有未保存修改）")
            self.rule_save_current_btn.setStyleSheet(_RULE_SAVE_BTN_DIRTY_STYLE)
        else:
            self.rule_code_unsaved_hint.hide()
            self.rule_save_current_btn.setText("保存当前规则代码")
            self.rule_save_current_btn.setStyleSheet("")

    def _capture_rule_code_snapshot(self, code: str) -> None:
        self._rule_code_saved_snapshot = str(code or "")
        self._set_rule_code_dirty(False)

    def _flash_rule_code_saved_hint(self) -> None:
        self.rule_code_unsaved_hint.setText("✓ 规则代码已保存，修改已生效。")
        self.rule_code_unsaved_hint.setStyleSheet(_RULE_CODE_SAVED_BANNER_STYLE)
        self.rule_code_unsaved_hint.show()
        QTimer.singleShot(2500, self.rule_code_unsaved_hint.hide)

    def _revert_current_rule_code_to_snapshot(self) -> None:
        rid = self._current_rule_id
        if not rid:
            return
        r = self._get_rule_by_id(rid)
        if not r:
            return
        r["code"] = self._rule_code_saved_snapshot
        try:
            self.rule_code_editor.blockSignals(True)
            self.rule_code_editor.setPlainText(self._rule_code_saved_snapshot)
        finally:
            self.rule_code_editor.blockSignals(False)
        self._set_rule_code_dirty(False)

    def _prompt_unsaved_rule_code(self, action: str) -> str:
        """返回 save / discard / cancel"""
        mb = QMessageBox(self)
        mb.setIcon(QMessageBox.Warning)
        mb.setWindowTitle("规则代码未保存")
        mb.setText("当前规则代码有未保存的修改。")
        mb.setInformativeText(
            "修改仅在点击「保存当前规则代码」后才会写入磁盘并生效。\n"
            f"是否在{action}前保存？"
        )
        save_btn = mb.addButton("保存", QMessageBox.AcceptRole)
        discard_btn = mb.addButton("不保存", QMessageBox.DestructiveRole)
        cancel_btn = mb.addButton("取消", QMessageBox.RejectRole)
        mb.exec_()
        clicked = mb.clickedButton()
        if clicked == save_btn:
            return "save"
        if clicked == discard_btn:
            return "discard"
        return "cancel"

    def _handle_unsaved_rule_code_before_action(self, action: str) -> bool:
        """未保存时询问用户；返回 True 表示可以继续原操作。"""
        if not self._rule_code_dirty:
            return True
        choice = self._prompt_unsaved_rule_code(action)
        if choice == "cancel":
            return False
        if choice == "save":
            if not self._save_current_rule_code(show_success_message=False):
                return False
        elif choice == "discard":
            self._revert_current_rule_code_to_snapshot()
        return True

    def _load_rule_code_at_row(self, row: int) -> None:
        if row < 0 or row >= self.rule_list.count():
            self._current_rule_id = None
            return
        item = self.rule_list.item(row)
        rid = item.data(Qt.UserRole) if item else None
        rid = str(rid or "")
        r = self._get_rule_by_id(rid)
        self._current_rule_id = rid if r else None
        if not r:
            return
        code = str(r.get("code") or "")
        try:
            self.rule_code_editor.blockSignals(True)
            self.rule_code_editor.setPlainText(code)
        finally:
            self.rule_code_editor.blockSignals(False)
        self._capture_rule_code_snapshot(code)

    def _on_rule_selected(self, row: int):
        """切换选中规则时，加载到右侧编辑器"""
        if self._rule_switch_guard:
            return
        if row < 0 or row >= self.rule_list.count():
            self._current_rule_id = None
            self._rule_list_prev_row = row
            return

        prev_row = self._rule_list_prev_row
        if prev_row >= 0 and row != prev_row and self._rule_code_dirty:
            self._rule_switch_guard = True
            try:
                self.rule_list.blockSignals(True)
                self.rule_list.setCurrentRow(prev_row)
            finally:
                self.rule_list.blockSignals(False)
                self._rule_switch_guard = False

            if not self._handle_unsaved_rule_code_before_action("切换规则"):
                return

            self._rule_switch_guard = True
            try:
                self.rule_list.blockSignals(True)
                self.rule_list.setCurrentRow(row)
            finally:
                self.rule_list.blockSignals(False)
                self._rule_switch_guard = False

        self._rule_list_prev_row = row
        self._load_rule_code_at_row(row)

    def _on_rule_list_context_menu(self, pos):
        item = self.rule_list.itemAt(pos)
        if item is None:
            return
        rid = str(item.data(Qt.UserRole) or "")
        r = self._get_rule_by_id(rid)
        if r is None:
            return
        menu = QMenu(self)
        act_rename = menu.addAction("重命名…")
        act_dup = menu.addAction("复制")
        act_del = menu.addAction("删除")
        chosen = menu.exec_(self.rule_list.viewport().mapToGlobal(pos))
        if chosen == act_rename:
            self._current_rule_id = rid
            self._rename_rule()
        elif chosen == act_dup:
            self._current_rule_id = rid
            self._duplicate_rule()
        elif chosen == act_del:
            self._current_rule_id = rid
            self._delete_rule()

    def _on_rule_code_changed(self):
        rid = self._current_rule_id
        if not rid:
            return
        r = self._get_rule_by_id(rid)
        if not r:
            return
        text = self.rule_code_editor.toPlainText()
        r["code"] = text
        self._set_rule_code_dirty(text != self._rule_code_saved_snapshot)

    def _rename_rule(self):
        rid = str(self._get_selected_rule_id() or self._current_rule_id or "")
        if not rid:
            QMessageBox.information(self, "提示", "请先在左侧选择一条规则")
            return
        r = self._get_rule_by_id(rid)
        if not r:
            QMessageBox.information(self, "提示", "未找到当前规则")
            return
        new_name, ok = QInputDialog.getText(
            self, "重命名规则", "规则名称：", text=str(r.get("name") or "")
        )
        if not ok:
            return
        norm_name = _normalize_rule_name(new_name)
        if self._is_rule_name_taken(norm_name, exclude_rule_id=rid):
            QMessageBox.warning(
                self,
                "重命名失败",
                f"已存在同名规则「{norm_name}」，请使用其他名称。\n"
                "（导入规则时若同名会自动改名；手动重命名不允许与现有规则重名。）",
            )
            return
        old_name = str(r.get("name") or "")
        r["name"] = norm_name
        save_single_sector_rule(r)
        self._refresh_rule_list()
        if old_name != norm_name:
            try:
                self.debug_output.append(f"规则已重命名：{old_name} → {norm_name}")
            except Exception:
                pass

    def _add_rule(self):
        new_name = self._make_unique_rule_name("新规则")
        new_rule = {
            "id": str(uuid.uuid4()),
            "name": new_name,
            "enabled": True,
            "code": (
                "# 新规则\nN = 3\nM = 30\n\n"
                "def select(stock_code, stock_name, sectors, daily_data, as_of_date, ctx):\n"
                "    \"\"\"在此编写完整选股逻辑（勿调用规则代码外的引擎函数）。\"\"\"\n"
                "    return False, {}\n"
            ),
        }
        self.rules.append(new_rule)
        save_single_sector_rule(new_rule)
        self._refresh_rule_list()
        # 选中最后一条
        self.rule_list.setCurrentRow(self.rule_list.count() - 1)

    def _delete_rule(self):
        rid = self._current_rule_id
        if not rid:
            return
        r = self._get_rule_by_id(rid)
        if not r:
            return
        if QMessageBox.question(self, "确认删除", f"确定删除规则「{r.get('name','未命名')}」？") != QMessageBox.Yes:
            return
        # 先从磁盘删除该规则文件（含旧文件名残留）
        removed = delete_sector_rule_files(rid)
        self.rules = [x for x in self.rules if str(x.get("id") or "") != rid]
        self._current_rule_id = None
        self._refresh_rule_list()
        if removed > 0:
            try:
                self.debug_output.append(f"已删除规则文件 {removed} 个：{r.get('name','未命名')}")
            except Exception:
                pass

    def _duplicate_rule(self):
        rid = self._current_rule_id
        if not rid:
            return
        r = self._get_rule_by_id(rid)
        if not r:
            return
        new_name = self._make_unique_rule_name(str(r.get("name") or "未命名") + " - 副本")
        copy_rule = {
            "id": str(uuid.uuid4()),
            "name": new_name,
            "enabled": bool(r.get("enabled", True)),
            "code": str(r.get("code") or ""),
        }
        self.rules.append(copy_rule)
        save_single_sector_rule(copy_rule)
        self._refresh_rule_list()
        self.rule_list.setCurrentRow(self.rule_list.count() - 1)

    def _save_rules_to_disk(self):
        # 保留：批量保存（目前 UI 不暴露该入口，主要用于批量修复/迁移时）
        save_sector_rules(self.rules or [])
        QMessageBox.information(self, "已保存", f"规则已保存到目录：{_RULES_DIR}")

    def _save_current_rule_code(self, *, show_success_message: bool = True) -> bool:
        """保存当前选中的规则（主要是右侧代码编辑框内容）"""
        rid = str(self._get_selected_rule_id() or "")
        if not rid:
            if show_success_message:
                QMessageBox.information(self, "提示", "请先在左侧选择一条规则")
            return False
        r = self._get_rule_by_id(rid)
        if not r:
            if show_success_message:
                QMessageBox.information(self, "提示", "未找到当前规则")
            return False
        # 代码内容已在 _on_rule_code_changed 中实时写回；这里直接落盘
        save_single_sector_rule(r)
        self._capture_rule_code_snapshot(str(r.get("code") or ""))
        self._flash_rule_code_saved_hint()
        if show_success_message:
            QMessageBox.information(
                self,
                "已保存",
                f"已保存：{r.get('name', '规则')}\n"
                f"目录：{_RULES_DIR}\n\n"
                "修改已写入磁盘，选股与重启后将使用最新规则代码。",
            )
        return True

    def _export_rule(self):
        """导出当前选中的规则为一个 JSON 文件（便于分享/备份）"""
        rid = str(self._get_selected_rule_id() or "")
        if not rid:
            QMessageBox.information(self, "提示", "请先在左侧选择一条规则")
            return
        r = self._get_rule_by_id(rid)
        if not r:
            QMessageBox.information(self, "提示", "未找到当前规则")
            return
        name = _sanitize_rule_filename(str(r.get("name") or "规则"))
        short_id = rid.split("-")[0] if "-" in rid else rid[:8]
        default_fn = f"{name}__{short_id}.json"
        default_path = os.path.join(_RULES_DIR, default_fn)
        file_path, _ = QFileDialog.getSaveFileName(self, "导出规则", default_path, "规则文件 (*.json)")
        if not file_path:
            return
        payload = {
            "id": str(r.get("id") or ""),
            "name": str(r.get("name") or ""),
            "enabled": bool(r.get("enabled", True)),
            "code": str(r.get("code") or ""),
        }
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "导出成功", f"已导出到：\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出失败：{e}")

    def _import_rule(self):
        """从 JSON 文件导入一条规则（导入为新规则，自动去重名称/ID）"""
        file_path, _ = QFileDialog.getOpenFileName(self, "导入规则", _RULES_DIR, "规则文件 (*.json)")
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"读取规则文件失败：{e}")
            return
        if not isinstance(data, dict):
            QMessageBox.warning(self, "导入失败", "规则文件格式不正确（应为 JSON 对象）")
            return

        # 规范化字段
        name = str(data.get("name") or "导入规则").strip() or "导入规则"
        name = self._make_unique_rule_name(name)
        code = str(data.get("code") or "").strip()
        if not code:
            QMessageBox.warning(self, "导入失败", "规则代码为空（缺少 code 字段）")
            return
        code = ensure_rule_code_has_param_preamble(code)
        enabled = bool(data.get("enabled", True))

        # ID：如果文件里带 id 且不冲突可以复用；否则生成新 id
        imported_id = str(data.get("id") or "").strip()
        existing_ids = {str(r.get("id") or "") for r in (self.rules or [])}
        rid = imported_id if imported_id and imported_id not in existing_ids else str(uuid.uuid4())

        orig_name = _normalize_rule_name(str(data.get("name") or "导入规则"))
        new_rule = {"id": rid, "name": name, "enabled": enabled, "code": code}
        self.rules.append(new_rule)
        save_single_sector_rule(new_rule)
        self._refresh_rule_list()
        # 选中新导入的规则
        for i in range(self.rule_list.count()):
            it = self.rule_list.item(i)
            if str(it.data(Qt.UserRole) or "") == rid:
                self.rule_list.setCurrentRow(i)
                break
        rename_note = ""
        if orig_name != name:
            rename_note = f"\n（与现有规则重名，已自动改名为「{name}」）"
        QMessageBox.information(
            self, "导入成功", f"已导入规则：{name}{rename_note}\n你可以在右侧修改代码后保存。"
        )
    
    def load_sectors_to_list(self):
        """加载 QMT 板块到列表（申万行业 / 概念等）"""
        try:
            self.all_sectors_list = load_all_sectors()
            if not self.all_sectors_list:
                from utils.qmt_execution_config import get_qmt_mode

                mode = get_qmt_mode()
                if mode in ("builtin", "standalone"):
                    tip = (
                        "未读到板块列表（builtin 模式应使用大 QMT 同步的 "
                        "data/qmt_sector_index.json）。\n"
                        "请确认大 QMT 已运行「蚂蚁量化规则」，并等待板块同步完成"
                        "（启动约数十秒，或次日 0 点定时同步）。"
                    )
                else:
                    tip = (
                        "未从 QMT 获取到板块列表。\n"
                        "请确认 MiniQMT/行情服务已启动，并在 QMT 下载中心勾选「全部板块」。"
                    )
                QMessageBox.warning(self, "板块数据为空", tip)
            
            # 清空网格布局
            while self.sector_grid_layout.count():
                item = self.sector_grid_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            
            self.sector_checkboxes.clear()
            
            # 每行显示的列数（可以根据需要调整）
            columns_per_row = 5
            current_row = 0
            col = 0
            
            # 添加所有 QMT 板块
            for sector in self.all_sectors_list:
                checkbox = QCheckBox(sector)
                checkbox.stateChanged.connect(self.update_sector_count)
                # 启动时默认全选，避免每次都需要手动点“全选”
                checkbox.blockSignals(True)
                checkbox.setChecked(True)
                checkbox.blockSignals(False)
                self.sector_grid_layout.addWidget(checkbox, current_row, col)
                self.sector_checkboxes[sector] = checkbox
                col += 1
                if col >= columns_per_row:
                    col = 0
                    current_row += 1
            
            self.update_sector_count()
            # 更新板块选择标题，显示数量
            self.sector_group.setTitle(f"板块选择（可多选） {len(self.all_sectors_list)} 个")
            
        except Exception as e:
            logger.error(f"加载板块列表失败: {str(e)}", exc_info=True)
            QMessageBox.warning(self, "错误", f"加载板块列表失败: {str(e)}")
            self.status_label.setText("板块列表加载失败")
    
    def filter_sectors(self, text: str):
        """根据搜索文本过滤综合板块（通过重建布局实现）"""
        if not text:
            text = ""
        search_text = text.strip().lower()
        
        logger.info(f"搜索综合板块: 输入文本='{text}', 处理后='{search_text}', 复选框总数={len(self.sector_checkboxes)}")
        
        # 先隐藏所有widget
        for checkbox in self.sector_checkboxes.values():
            checkbox.hide()
        
        # 彻底清空当前布局
        while self.sector_grid_layout.count():
            item = self.sector_grid_layout.takeAt(0)
            if item.widget():
                self.sector_grid_layout.removeWidget(item.widget())
        
        # 确保所有widget都已从布局中移除
        for checkbox in self.sector_checkboxes.values():
            try:
                self.sector_grid_layout.removeWidget(checkbox)
            except:
                pass
        
        # 每行显示的列数
        columns_per_row = 5
        current_row = 0
        col = 0
        
        matched_sectors_result: Optional[List[str]] = None
        if not search_text:
            # 如果搜索文本为空，显示所有综合板块
            for sector in self.all_sectors_list:
                if sector in self.sector_checkboxes:
                    self.sector_grid_layout.addWidget(self.sector_checkboxes[sector], current_row, col)
                    self.sector_checkboxes[sector].show()
                    col += 1
                    if col >= columns_per_row:
                        col = 0
                        current_row += 1
        else:
            # 有搜索文本，只显示匹配的综合板块
            matched_sectors = []
            for sector in self.all_sectors_list:
                if search_text in sector.lower():
                    matched_sectors.append(sector)
            
            matched_sectors_result = matched_sectors
            logger.info(f"搜索 '{text}' 找到 {len(matched_sectors)} 个匹配的综合板块")
            
            # 只添加匹配的复选框
            for sector in matched_sectors:
                if sector in self.sector_checkboxes:
                    self.sector_grid_layout.addWidget(self.sector_checkboxes[sector], current_row, col)
                    self.sector_checkboxes[sector].show()
                    col += 1
                    if col >= columns_per_row:
                        col = 0
                        current_row += 1
        
        # 有搜索时：取消「当前结果集以外」板块的勾选，避免先全选再搜索后隐藏项仍参与并集导致股票列表像全市场
        if search_text and matched_sectors_result is not None:
            matched_set = set(matched_sectors_result)
            for sector, checkbox in self.sector_checkboxes.items():
                if sector not in matched_set:
                    checkbox.blockSignals(True)
                    checkbox.setChecked(False)
                    checkbox.blockSignals(False)
            self.update_sector_count()
        
        # 更新布局
        self.sector_grid_layout.invalidate()
        self.sector_grid_layout.activate()
        self.sector_container.updateGeometry()
        self.sector_container.adjustSize()
        
        # 强制更新
        QApplication.processEvents()
        
        logger.info(f"布局重建完成，可见复选框数: {sum(1 for cb in self.sector_checkboxes.values() if cb.isVisible())}")
    
    def select_all_sectors(self):
        """全选所有板块（只选择当前可见的复选框）"""
        # 临时阻止信号，提高性能
        checked_count = 0
        for checkbox in self.sector_checkboxes.values():
            if checkbox.isVisible():
                checkbox.blockSignals(True)
                checkbox.setChecked(True)
                checkbox.blockSignals(False)
                checked_count += 1
        # 更新板块计数（不立即更新股票列表，避免卡顿）
        selected_sectors = self.get_selected_sectors()
        selected_count = len(selected_sectors)
        mode_text = "交集" if self.intersection_mode_radio.isChecked() else "并集"
        
        # 显示选中的板块名称
        if selected_count == 0:
            self.sector_count_label.setText(f"已选择: 0 个板块")
        else:
            # 如果板块数量较少（<=5个），显示所有板块名称
            if selected_count <= 5:
                sectors_text = "、".join(selected_sectors)
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
            else:
                # 如果板块数量较多，只显示前3个，后面用省略号
                sectors_text = "、".join(selected_sectors[:3]) + f" 等{selected_count}个"
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
        
        # 立即显示加载状态，然后延迟更新股票列表
        self.stock_list_group.setTitle("选中板块对应的股票列表 (准备加载...)")
        QApplication.processEvents()
        # 延迟更新股票列表，避免全选时卡顿
        QTimer.singleShot(200, self.update_stock_list)
        logger.info(f"全选了 {checked_count} 个可见的板块")
    
    def deselect_all_sectors(self):
        """取消全选所有板块（只取消当前可见的复选框）"""
        # 临时阻止信号，提高性能
        unchecked_count = 0
        for checkbox in self.sector_checkboxes.values():
            if checkbox.isVisible():
                checkbox.blockSignals(True)
                checkbox.setChecked(False)
                checkbox.blockSignals(False)
                unchecked_count += 1
        # 更新板块计数（不立即更新股票列表，避免卡顿）
        selected_sectors = self.get_selected_sectors()
        selected_count = len(selected_sectors)
        mode_text = "交集" if self.intersection_mode_radio.isChecked() else "并集"
        
        # 显示选中的板块名称
        if selected_count == 0:
            self.sector_count_label.setText(f"已选择: 0 个板块")
        else:
            # 如果板块数量较少（<=5个），显示所有板块名称
            if selected_count <= 5:
                sectors_text = "、".join(selected_sectors)
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
            else:
                # 如果板块数量较多，只显示前3个，后面用省略号
                sectors_text = "、".join(selected_sectors[:3]) + f" 等{selected_count}个"
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
        
        # 立即更新股票列表（取消全选时应该很快）
        self.update_stock_list()
        logger.info(f"取消全选了 {unchecked_count} 个可见的板块")
    
    def on_sector_mode_changed(self):
        """当筛选模式改变时，重新更新股票列表和板块计数标签"""
        # 立即更新板块计数标签（显示当前模式）
        self.update_sector_count()
        # 延迟更新股票列表，避免卡顿
        QTimer.singleShot(300, self.update_stock_list)
    
    def update_sector_count(self):
        """更新已选择板块数量，并更新股票列表"""
        selected_sectors = self.get_selected_sectors()
        selected_count = len(selected_sectors)
        mode_text = "交集" if self.intersection_mode_radio.isChecked() else "并集"
        
        # 显示选中的板块名称
        if selected_count == 0:
            self.sector_count_label.setText(f"已选择: 0 个板块")
        else:
            # 如果板块数量较少（<=5个），显示所有板块名称
            if selected_count <= 5:
                sectors_text = "、".join(selected_sectors)
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
            else:
                # 如果板块数量较多，只显示前3个，后面用省略号
                sectors_text = "、".join(selected_sectors[:3]) + f" 等{selected_count}个"
                self.sector_count_label.setText(f"已选择: {selected_count} 个板块（{mode_text}模式）: {sectors_text}")
        
        # 延迟更新股票列表，避免卡顿（单个复选框变化时也延迟更新）
        QTimer.singleShot(300, self.update_stock_list)
    
    def get_selected_sectors(self) -> List[str]:
        """获取选中的板块列表"""
        selected = []
        for sector_name, checkbox in self.sector_checkboxes.items():
            if checkbox.isChecked():
                selected.append(sector_name)
        return selected
    
    def update_stock_list(self):
        """更新选中板块对应的股票列表（QMT 板块成分股）"""
        try:
            selected_sectors = self.get_selected_sectors()

            self.stock_list_group.setTitle("选中板块对应的股票列表 (加载中...)")
            QApplication.processEvents()

            self.stock_list_table.setRowCount(0)
            QApplication.processEvents()

            if not selected_sectors:
                self._stock_list_base = []
                self.stock_checkboxes.clear()
                self.stock_list_group.setTitle("选中板块对应的股票列表 (0 只)")
                self.update_stock_selected_count()
                self._schedule_auto_run_after_stock_list(ready=False, reason="未选中任何板块")
                return

            mode = "intersection" if self.intersection_mode_radio.isChecked() else "union"
            stock_dict = self._sector_store.stocks_for_sectors(selected_sectors, mode=mode)
            if stock_dict is None:
                self._stock_list_base = []
                self.stock_checkboxes.clear()
                self.stock_list_group.setTitle("选中板块对应的股票列表 (0 只)")
                self.update_stock_selected_count()
                self._schedule_auto_run_after_stock_list(ready=False, reason="QMT 板块数据不可用")
                return

            QApplication.processEvents()

            self._stock_sectors_map = {
                code: info.get("all_sectors") or info.get("matched_sectors") or []
                for code, info in stock_dict.items()
            }

            stock_list = []
            for code, info in stock_dict.items():
                all_sectors = info.get("all_sectors") or []
                sectors_str = ';'.join(all_sectors) if all_sectors else ''
                stock_list.append((code, str(info.get("name") or "未知"), sectors_str))

            stock_list.sort(key=lambda x: x[0])
            self._stock_list_base = stock_list
            self._populate_stock_table_from_base(preserve_unchecked=False)
            self.filter_stock_list(self.stock_search_input.text())

            logger.debug(
                "更新股票列表完成，底池 %d 只，开关过滤后 %d 只",
                len(self._stock_list_base),
                self.stock_list_table.rowCount(),
            )
            self._schedule_auto_run_after_stock_list(ready=bool(self.stock_list_table.rowCount()))

        except Exception as e:
            logger.error(f"更新股票列表失败: {str(e)}", exc_info=True)
            self._schedule_auto_run_after_stock_list(ready=False, reason=f"更新股票列表失败: {e}")
    
    def _apply_stock_search_enter(self) -> None:
        """回车：先按当前关键字过滤，再将勾选限定为当前可见行（典型场景：输入代码后只保留该只）。"""
        text = self.stock_search_input.text()
        self.filter_stock_list(text)
        if not text.strip():
            return
        for row in range(self.stock_list_table.rowCount()):
            checkbox = self.stock_list_table.cellWidget(row, 0)
            if checkbox is None or not isinstance(checkbox, QCheckBox):
                continue
            visible = not self.stock_list_table.isRowHidden(row)
            checkbox.blockSignals(True)
            checkbox.setChecked(visible)
            checkbox.blockSignals(False)
        self.update_stock_selected_count()

    @staticmethod
    def _is_star_or_bj_code(stock_code: str) -> bool:
        code = normalize_stock_code(stock_code)
        if code.startswith(("688", "689")):
            return True
        if code.startswith(("8", "4", "920")):
            return True
        return False

    def _stock_matches_exclude_filters(self, stock_code: str, stock_name: str = "") -> bool:
        """当前排除开关下，该股票是否应从备选池剔除。"""
        if getattr(self, "exclude_st_checkbox", None) is not None:
            if self.exclude_st_checkbox.isChecked() and is_st_stock(stock_name):
                return True
        if getattr(self, "exclude_star_bj_checkbox", None) is not None:
            if self.exclude_star_bj_checkbox.isChecked() and self._is_star_or_bj_code(stock_code):
                return True
        return False

    def _filter_stock_list_by_exclude(
        self, stock_list: List[Tuple[str, str, str]]
    ) -> List[Tuple[str, str, str]]:
        return [
            (code, name, sectors)
            for code, name, sectors in stock_list
            if not self._stock_matches_exclude_filters(code, name)
        ]

    def _collect_unchecked_stock_codes(self) -> set:
        unchecked = set()
        for row in range(self.stock_list_table.rowCount()):
            checkbox = self.stock_list_table.cellWidget(row, 0)
            code_item = self.stock_list_table.item(row, 1)
            if not code_item or not isinstance(checkbox, QCheckBox):
                continue
            if not checkbox.isChecked():
                unchecked.add(code_item.text().strip().zfill(6))
        return unchecked

    def _update_stock_list_group_title(self, shown_count: int) -> None:
        base_count = len(getattr(self, "_stock_list_base", []) or [])
        excluded = max(0, base_count - shown_count)
        if excluded > 0:
            self.stock_list_group.setTitle(
                f"选中板块对应的股票列表 ({shown_count} 只，已排除 {excluded})"
            )
        else:
            self.stock_list_group.setTitle(f"选中板块对应的股票列表 ({shown_count} 只)")

    def _populate_stock_table_from_base(
        self,
        *,
        preserve_unchecked: bool = True,
        force_checked_codes: Optional[set] = None,
    ) -> None:
        """按当前排除开关从底池生成备选表；开关关闭时被剔除标的会回到池中。"""
        unchecked = self._collect_unchecked_stock_codes() if preserve_unchecked else set()
        if force_checked_codes:
            unchecked -= {str(c).strip().zfill(6) for c in force_checked_codes}

        filtered = self._filter_stock_list_by_exclude(list(self._stock_list_base or []))
        self.stock_checkboxes.clear()
        self.stock_list_table.setUpdatesEnabled(False)
        table_batch_size = 500
        total_stocks = len(filtered)
        self.stock_list_table.setRowCount(total_stocks)

        for i, (code, name, sectors_str) in enumerate(filtered):
            code6 = str(code).strip().zfill(6)
            checkbox = QCheckBox()
            checkbox.setChecked(code6 not in unchecked)
            checkbox.stateChanged.connect(self._on_stock_checkbox_changed)
            self.stock_list_table.setCellWidget(i, 0, checkbox)
            self.stock_checkboxes[code6] = checkbox
            self.stock_list_table.setItem(i, 1, QTableWidgetItem(code6))
            self.stock_list_table.setItem(i, 2, QTableWidgetItem(name))
            self.stock_list_table.setItem(i, 3, QTableWidgetItem(sectors_str))
            if (i + 1) % table_batch_size == 0:
                self.stock_list_group.setTitle(
                    f"选中板块对应的股票列表 (加载中... {i+1}/{total_stocks})"
                )
                QApplication.processEvents()

        self.stock_list_table.setUpdatesEnabled(True)
        self._update_stock_list_group_title(total_stocks)
        self.update_stock_selected_count()

    def _load_stock_exclude_prefs(self) -> None:
        prefs = {"exclude_st": False, "exclude_star_bj": False}
        try:
            if os.path.isfile(_UI_PREFS_PATH):
                with open(_UI_PREFS_PATH, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict):
                    prefs["exclude_st"] = bool(raw.get("exclude_st", False))
                    prefs["exclude_star_bj"] = bool(raw.get("exclude_star_bj", False))
        except Exception:
            logger.debug("加载选股 UI 偏好失败", exc_info=True)
        for cb, key in (
            (self.exclude_st_checkbox, "exclude_st"),
            (self.exclude_star_bj_checkbox, "exclude_star_bj"),
        ):
            cb.blockSignals(True)
            cb.setChecked(bool(prefs[key]))
            cb.blockSignals(False)

    def _save_stock_exclude_prefs(self) -> None:
        try:
            os.makedirs(os.path.dirname(_UI_PREFS_PATH), exist_ok=True)
            payload = {
                "exclude_st": bool(self.exclude_st_checkbox.isChecked()),
                "exclude_star_bj": bool(self.exclude_star_bj_checkbox.isChecked()),
            }
            with open(_UI_PREFS_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            logger.debug("保存选股 UI 偏好失败", exc_info=True)

    def _on_stock_exclude_filter_changed(self, *_args) -> None:
        self._save_stock_exclude_prefs()
        self._populate_stock_table_from_base(preserve_unchecked=True)
        self.filter_stock_list(self.stock_search_input.text())

    def _on_stock_checkbox_changed(self, *_args) -> None:
        self.update_stock_selected_count()
    
    def filter_stock_list(self, text: str):
        """根据搜索文本过滤股票列表"""
        search_text = text.strip().lower()
        
        for row in range(self.stock_list_table.rowCount()):
            code_item = self.stock_list_table.item(row, 1)  # 股票代码列
            name_item = self.stock_list_table.item(row, 2)  # 股票名称列
            sectors_item = self.stock_list_table.item(row, 3)  # 所属板块列
            
            if code_item and name_item:
                code = code_item.text().lower()
                name = name_item.text().lower()
                sectors = sectors_item.text().lower() if sectors_item else ''
                
                if not search_text or search_text in code or search_text in name or search_text in sectors:
                    self.stock_list_table.setRowHidden(row, False)
                else:
                    self.stock_list_table.setRowHidden(row, True)
    
    def select_all_stocks(self):
        """全选股票列表中的所有股票（只选择当前可见的）"""
        for row in range(self.stock_list_table.rowCount()):
            if not self.stock_list_table.isRowHidden(row):
                checkbox = self.stock_list_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox):
                    checkbox.blockSignals(True)
                    checkbox.setChecked(True)
                    checkbox.blockSignals(False)
        self.update_stock_selected_count()
    
    def deselect_all_stocks(self):
        """取消全选股票列表中的所有股票（只取消当前可见的）"""
        for row in range(self.stock_list_table.rowCount()):
            if not self.stock_list_table.isRowHidden(row):
                checkbox = self.stock_list_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox):
                    checkbox.blockSignals(True)
                    checkbox.setChecked(False)
                    checkbox.blockSignals(False)
        self.update_stock_selected_count()
    
    def load_stocks_from_file(self):
        """从文件中选择股票并添加到股票列表"""
        try:
            # 打开文件选择对话框
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "选择股票文件",
                "",
                "所有文件 (*.*);;Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;文本文件 (*.txt)"
            )
            
            if not file_path:
                return
            
            # 读取文件内容
            stock_codes_from_file = set()
            
            # 判断文件类型
            file_ext = os.path.splitext(file_path)[1].lower()
            
            # 尝试按Excel格式读取
            if file_ext in ['.xlsx', '.xls']:
                df = None
                last_error = None
                
                # 方法1: 尝试使用pd.read_html（有些.xls文件实际上是HTML格式）
                # 尝试多种编码
                encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1']
                for encoding in encodings:
                    try:
                        tables = pd.read_html(file_path, encoding=encoding)
                        if tables and len(tables) > 0:
                            df = tables[0]
                            logger.info(f"使用read_html成功读取文件: {file_path} (编码: {encoding})")
                            break
                    except Exception as e:
                        last_error = e
                        logger.debug(f"read_html读取失败 (编码: {encoding}): {str(e)}")
                        continue
                
                # 方法2: 如果HTML读取失败，尝试Excel引擎
                if df is None or df.empty:
                    try:
                        if file_ext == '.xlsx':
                            # .xlsx文件优先使用openpyxl
                            try:
                                df = pd.read_excel(file_path, engine='openpyxl')
                            except ImportError:
                                QMessageBox.warning(self, "错误", "读取Excel文件需要安装openpyxl库。\n请运行: pip install openpyxl")
                                return
                            except Exception as e:
                                # 如果openpyxl失败，尝试使用xlrd
                                try:
                                    df = pd.read_excel(file_path, engine='xlrd')
                                except:
                                    raise e
                        else:
                            # .xls文件优先使用xlrd
                            try:
                                df = pd.read_excel(file_path, engine='xlrd')
                            except ImportError:
                                # 如果xlrd不可用，尝试openpyxl（某些.xls文件可能实际上是新格式）
                                try:
                                    df = pd.read_excel(file_path, engine='openpyxl')
                                except ImportError:
                                    QMessageBox.warning(self, "错误", "读取.xls文件需要安装xlrd或openpyxl库。\n请运行: pip install xlrd 或 pip install openpyxl")
                                    return
                                except Exception as e2:
                                    QMessageBox.warning(self, "错误", f"读取.xls文件失败: {str(e2)}")
                                    return
                            except Exception as e:
                                # 如果xlrd失败，尝试openpyxl
                                try:
                                    df = pd.read_excel(file_path, engine='openpyxl')
                                except:
                                    raise e
                        logger.info(f"使用read_excel成功读取文件: {file_path}")
                    except Exception as e:
                        last_error = e
                        logger.debug(f"read_excel读取失败: {str(e)}")
                
                # 方法3: 如果Excel读取失败，尝试作为CSV文件读取（有些.xls文件实际上是CSV格式）
                if df is None or df.empty:
                    logger.info(f"Excel读取失败，尝试作为CSV文件读取: {file_path}")
                    csv_encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1']
                    for encoding in csv_encodings:
                        try:
                            df = pd.read_csv(file_path, encoding=encoding)
                            if df is not None and not df.empty:
                                logger.info(f"使用read_csv成功读取文件: {file_path} (编码: {encoding})")
                                break
                        except Exception as e:
                            logger.debug(f"read_csv读取失败 (编码: {encoding}): {str(e)}")
                            continue
                
                if df is None or df.empty:
                    error_msg = f"无法读取文件: {file_path}\n已尝试Excel、HTML和CSV格式，均失败"
                    if last_error:
                        error_msg += f"\n最后错误: {str(last_error)}"
                    QMessageBox.warning(self, "错误", error_msg)
                    return
                
                # 尝试从文件中提取股票代码
                # 优先查找包含"代码"或"code"的列
                code_col = None
                for col in df.columns:
                    col_str = str(col).strip()
                    if '代码' in col_str or 'code' in col_str.lower():
                        code_col = col
                        break
                
                # 如果没有找到代码列，尝试使用第二列（第一列可能是序号）
                if code_col is None:
                    if len(df.columns) >= 2:
                        code_col = df.columns[1]  # 第二列
                    elif len(df.columns) >= 1:
                        code_col = df.columns[0]  # 第一列
                    else:
                        QMessageBox.warning(self, "错误", "文件中没有找到股票代码列")
                        return
                
                # 从代码列提取股票代码
                for code in df[code_col].dropna():
                    code_clean = _normalize_import_stock_code(code)
                    if code_clean:
                        stock_codes_from_file.add(code_clean)
            else:
                # 尝试按CSV格式读取
                try:
                    df = pd.read_csv(file_path, encoding='utf-8-sig')
                    # 尝试从第一列获取股票代码
                    if len(df.columns) > 0:
                        first_col = df.columns[0]
                        for code in df[first_col].dropna():
                            code_clean = _normalize_import_stock_code(code)
                            if code_clean:
                                stock_codes_from_file.add(code_clean)
                except:
                    # 如果不是CSV格式，按文本文件读取
                    try:
                        with open(file_path, 'r', encoding='utf-8-sig') as f:
                            for line in f:
                                line = line.strip()
                                if not line:
                                    continue
                                # 尝试提取股票代码（可能是单独一行，或者用逗号/制表符分隔）
                                parts = line.replace(',', ' ').replace('\t', ' ').split()
                                for part in parts:
                                    code_clean = _normalize_import_stock_code(part.strip())
                                    if code_clean:
                                        stock_codes_from_file.add(code_clean)
                                        break  # 每行只取第一个有效的股票代码
                    except Exception as e:
                        QMessageBox.warning(self, "错误", f"读取文件失败: {str(e)}")
                        return
            
            if not stock_codes_from_file:
                QMessageBox.warning(self, "提示", "文件中未找到有效的股票代码")
                return
            
            universe = self._sector_store.get_universe()
            stocks_to_add = []
            not_found_codes = []

            for code in stock_codes_from_file:
                code = code.zfill(6)
                if code.startswith('1') or code.startswith('5'):
                    continue
                if code in universe:
                    stock_name = self._sector_store.stock_name(code)
                    stocks_to_add.append((code, stock_name))
                else:
                    not_found_codes.append(code)

            if not stocks_to_add:
                msg = "文件中未找到有效的股票代码"
                if not_found_codes:
                    msg += f"\n\n以下股票代码不在沪深A股列表中：\n{', '.join(not_found_codes[:20])}"
                    if len(not_found_codes) > 20:
                        msg += f"\n... 还有 {len(not_found_codes) - 20} 只"
                QMessageBox.warning(self, "提示", msg)
                return
            
            self._sector_store.ensure_inverted_index()

            existing_codes = {
                str(code).strip().zfill(6)
                for code, _name, _sectors in (self._stock_list_base or [])
            }
            force_checked = set()
            added_count = 0
            for code, name in stocks_to_add:
                code6 = str(code).strip().zfill(6)
                force_checked.add(code6)
                if code6 not in self._stock_sectors_map:
                    self._stock_sectors_map[code6] = self._sector_store.sectors_for_stock(code6)
                sectors_list = self._stock_sectors_map.get(code6, [])
                sectors_str = ";".join(sectors_list) if sectors_list else ""
                if code6 in existing_codes:
                    self._stock_list_base = [
                        (code6, name, sectors_str) if str(c).strip().zfill(6) == code6 else (c, n, s)
                        for c, n, s in self._stock_list_base
                    ]
                else:
                    self._stock_list_base.append((code6, name, sectors_str))
                    existing_codes.add(code6)
                    added_count += 1

            self._stock_list_base.sort(key=lambda x: x[0])
            self._populate_stock_table_from_base(
                preserve_unchecked=True,
                force_checked_codes=force_checked,
            )
            self.filter_stock_list(self.stock_search_input.text())

            selected_count = sum(
                1
                for code in force_checked
                if code in self.stock_checkboxes and self.stock_checkboxes[code].isChecked()
            )
            skipped_by_filter = len(force_checked) - selected_count
            
            # 显示结果
            msg_parts = []
            if added_count > 0:
                msg_parts.append(f"已添加 {added_count} 只新股票到备选底池")
            if selected_count > 0:
                msg_parts.append(f"当前备选池已选中 {selected_count} 只")
            if skipped_by_filter > 0:
                msg_parts.append(
                    f"另有 {skipped_by_filter} 只因排除开关未进入当前备选池（关闭开关后会自动出现）"
                )
            if not_found_codes:
                msg_parts.append(f"\n\n以下股票代码不在沪深A股列表中：\n{', '.join(not_found_codes[:20])}")
                if len(not_found_codes) > 20:
                    msg_parts.append(f"\n... 还有 {len(not_found_codes) - 20} 只")
            
            if msg_parts:
                QMessageBox.information(self, "从文件选择完成", "\n".join(msg_parts))
            else:
                QMessageBox.information(self, "从文件选择完成", f"已成功处理 {len(force_checked)} 只股票")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"从文件选择股票失败: {str(e)}")
            logger.error(f"从文件选择股票失败: {str(e)}", exc_info=True)
    
    def get_selected_stocks(self) -> List[str]:
        """获取当前备选池中勾选的股票代码列表"""
        selected = []
        for row in range(self.stock_list_table.rowCount()):
            if not self.stock_list_table.isRowHidden(row):
                checkbox = self.stock_list_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                    code_item = self.stock_list_table.item(row, 1)
                    if code_item:
                        selected.append(code_item.text())
        return selected
    
    def update_stock_selected_count(self):
        """更新已选择股票数量"""
        selected_count = len(self.get_selected_stocks())
        self.stock_selected_count_label.setText(f"已选择: {selected_count} 只股票")
    
    def _get_trading_dates(self, count: int, as_of_date: Optional[date] = None) -> List[date]:
        """获取最近 N 个交易日（委托 utils.trading_day）"""
        return get_trading_dates(count, as_of_date)
    
    def _load_stocks_by_sectors(self, sector_names: List[str]) -> List[Tuple[str, str, str]]:
        """按 QMT 板块筛选股票，返回 (代码, 名称, 所属板块) 列表。"""
        try:
            stock_dict = self._sector_store.stocks_for_sectors(sector_names, mode="union")
            stock_list: List[Tuple[str, str, str]] = []
            for code, info in sorted(stock_dict.items()):
                name = str(info.get("name") or "未知")
                sectors = info.get("all_sectors") or []
                for sector in sectors:
                    stock_list.append((code, name, sector))
            logger.info("QMT 板块筛选: %d 条记录", len(stock_list))
            return stock_list
        except Exception as e:
            logger.error(f"加载股票信息失败: {str(e)}", exc_info=True)
            QMessageBox.warning(self, "错误", f"加载股票信息失败: {str(e)}")
            return []
    
    def _set_as_of_dates_to_today(self):
        """将选股基准日起、止都设为今天。"""
        today = QDate.currentDate()
        self.as_of_date_edit.setDate(today)
        self.as_of_date_end_edit.setDate(today)

    def start_filter(self):
        """开始筛选"""
        try:
            # 如果已有筛选线程在运行，禁止重复启动，避免旧线程仍在运行时被销毁
            if self.filter_thread and self.filter_thread.isRunning():
                if self._auto_run:
                    logger.warning("自动运行跳过：当前筛选尚未结束")
                else:
                    QMessageBox.warning(self, "提示", "当前筛选尚未结束，请先点击“停止筛选”或等待完成。")
                return

            # 获取选中的股票代码列表（从股票列表表格中）
            selected_stock_codes = self.get_selected_stocks()
            
            if not selected_stock_codes:
                if self._auto_run:
                    logger.error("自动运行失败：未选中任何股票")
                    self.status_label.setText("自动运行失败：未选中任何股票")
                    QTimer.singleShot(50, self.accept)
                else:
                    QMessageBox.warning(self, "错误", "请至少从股票列表中选择一只股票")
                return
            
            logger.info(f"选中的股票 ({len(selected_stock_codes)} 只): {selected_stock_codes[:10]}{'...' if len(selected_stock_codes) > 10 else ''}")
            
            # 从股票列表表格中构建stock_list
            stock_list = []
            selected_stock_codes_set = set(selected_stock_codes)
            
            for row in range(self.stock_list_table.rowCount()):
                code_item = self.stock_list_table.item(row, 1)  # 股票代码列
                name_item = self.stock_list_table.item(row, 2)  # 股票名称列
                
                if code_item and name_item:
                    code = code_item.text()
                    if code in selected_stock_codes_set:
                        name = name_item.text()
                        # 从保存的板块映射中获取板块信息
                        sectors = self._stock_sectors_map.get(code, [])
                        # 如果一只股票匹配多个板块，为每个板块创建一条记录
                        if sectors:
                            for sector in sectors:
                                if sector.strip():
                                    stock_list.append((code, name, sector.strip()))
                        else:
                            # 如果股票没有板块信息（比如从文件添加的），仍然添加一条记录，板块信息为空
                            stock_list.append((code, name, ''))
            
            if not stock_list:
                if self._auto_run:
                    logger.error("自动运行失败：未找到选中股票的有效记录")
                    self.status_label.setText("自动运行失败：未找到选中股票的有效记录")
                    QTimer.singleShot(50, self.accept)
                else:
                    QMessageBox.warning(self, "警告", f"未找到选中的股票")
                return
            
            # 统计实际要筛选的信息
            unique_stock_count = len(selected_stock_codes)
            total_records_count = len(stock_list)
            
            self.status_label.setText(f"找到 {unique_stock_count} 只选中的股票（共 {total_records_count} 条记录），开始筛选...")
            
            # 内置模式窗口参数由各规则代码顶部常量提供；线程入口仅使用模块默认初值（规则编译时再按代码覆盖）
            n = SECTOR_RULE_DEFAULT_N
            m_mode1 = SECTOR_RULE_DEFAULT_M
            n_mode2 = n
            m_mode2 = m_mode1
            n_mode3 = SECTOR_RULE_DEFAULT_N_MODE3
            m_mode3 = SECTOR_RULE_DEFAULT_M_MODE3
            l_mode3 = SECTOR_RULE_DEFAULT_L_MODE3

            enabled_rules = [r for r in (self.rules or []) if bool(r.get("enabled", True))]
            if not enabled_rules:
                if self._auto_run:
                    logger.error("自动运行失败：未启用任何规则")
                    self.status_label.setText("自动运行失败：未启用任何规则")
                    QTimer.singleShot(50, self.accept)
                else:
                    QMessageBox.warning(self, "错误", "请至少启用一条选股规则")
                return
            if self._rule_code_dirty:
                mb = QMessageBox(self)
                mb.setIcon(QMessageBox.Warning)
                mb.setWindowTitle("规则代码未保存")
                mb.setText("当前规则代码有未保存的修改。")
                mb.setInformativeText(
                    "请先点击「保存当前规则代码」，修改才会写入磁盘并生效。\n"
                    "您也可以直接保存并继续开始筛选。"
                )
                save_btn = mb.addButton("保存并继续", QMessageBox.AcceptRole)
                cancel_btn = mb.addButton("取消", QMessageBox.RejectRole)
                mb.exec_()
                if mb.clickedButton() == cancel_btn:
                    return
                if mb.clickedButton() == save_btn and not self._save_current_rule_code(
                    show_success_message=False
                ):
                    return
            # 每次执行前保存一次规则，避免忘记保存
            save_sector_rules(self.rules or [])
            rid = str(self._current_rule_id or "")
            if rid:
                r = self._get_rule_by_id(rid)
                if r is not None:
                    self._capture_rule_code_snapshot(str(r.get("code") or ""))
            
            # 清空结果，并复位进度/滚动（再次开始时勿沿用上次位置）
            self._result_tables_by_rule = {}
            self._result_rows_by_rule = {}
            self._rule_tab_index = {}
            while self.result_tabs.count() > 0:
                w = self.result_tabs.widget(0)
                self.result_tabs.removeTab(0)
                if w is not None:
                    w.deleteLater()
            self._reset_filter_run_ui()

            qd = self.as_of_date_edit.date()
            picked = date(qd.year(), qd.month(), qd.day())
            qd_end = self.as_of_date_end_edit.date()
            picked_end = date(qd_end.year(), qd_end.month(), qd_end.day())
            # 起止相同：单日（今天且未指定区间终点外逻辑 → None 表示走 15:00 日历）；起止不同：区间内每个交易日批量选股
            if picked == picked_end:
                as_of_end_for_thread = None
                as_of_for_thread = None if picked == date.today() else picked
            else:
                as_of_for_thread = picked
                as_of_end_for_thread = picked_end
            
            # 区间选股：工作量随「交易日数×股票数」增长，提前提示避免误以为卡死
            if as_of_end_for_thread is not None:
                lo, hi = (picked, picked_end) if picked <= picked_end else (picked_end, picked)
                day_list = get_trading_dates_in_range_sorted(lo, hi)
                n_td = len(day_list)
                est_steps = unique_stock_count * max(n_td, 1)
                nm = len(enabled_rules)
                if n_td > 0 and (est_steps > 500 or n_td > 10):
                    if self._auto_run:
                        logger.info(
                            f"自动运行：区间工作量较大但已按参数自动继续（交易日={n_td}, 股票={unique_stock_count}, 估算步数={est_steps}, 规则数={nm}）"
                        )
                    else:
                        reply = QMessageBox.question(
                            self,
                            "区间选股工作量",
                            f"当前区间内约有 {n_td} 个交易日，将对 {unique_stock_count} 只股票逐日筛选，"
                            f"合计约 {est_steps} 步进度（每步还会跑已勾选的 {nm} 个模式），耗时会明显长于单日。\n\n"
                            f"进度若长时间停在某只股票，多为正在下载该股历史日线，请稍候。\n\n"
                            f"若只需某一天的结果，请把「止」设为与「起」同一天。\n\n"
                            f"是否继续？",
                            QMessageBox.Yes | QMessageBox.No,
                            QMessageBox.Yes,
                        )
                        if reply != QMessageBox.Yes:
                            return
            
            self.filter_thread = SectorStockFilterThread(
                stock_list,
                n, m_mode1,
                n_mode2, m_mode2,
                n_mode3, m_mode3, l_mode3,
                self.rules,
                as_of_for_thread,
                as_of_end_for_thread,
                self,
            )
            self.filter_thread.progress_updated.connect(self.on_progress_updated)
            self.filter_thread.stock_found.connect(self.on_stock_found)
            self.filter_thread.finished.connect(self.on_finished)
            self.filter_thread.error_occurred.connect(self.on_error)
            self.filter_thread.debug_info.connect(self.on_debug_info)
            
            # 更新状态
            self.status_label.setText("正在筛选股票...")
            
            # 禁用开始按钮，启用停止按钮
            self.start_button.setEnabled(False)
            self.stop_button.setEnabled(True)
            
            # 启动线程
            self.filter_thread.start()
            
        except Exception as e:
            logger.error(f"启动筛选失败: {str(e)}", exc_info=True)
            QMessageBox.warning(self, "错误", f"启动筛选失败: {str(e)}")
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
    
    def stop_filter(self):
        """停止筛选"""
        try:
            if self.filter_thread and self.filter_thread.isRunning():
                self.filter_thread.stop()
                self.status_label.setText("正在停止筛选...")
                logger.info("用户点击停止筛选")
                # 阻塞等待线程真正结束，避免 QThread 在仍运行时被销毁
                try:
                    self.filter_thread.wait()
                except Exception as e_wait:
                    logger.error(f"等待筛选线程结束时出错: {str(e_wait)}", exc_info=True)
        except Exception as e:
            logger.error(f"停止筛选失败: {str(e)}", exc_info=True)
            QMessageBox.warning(self, "错误", f"停止筛选失败: {str(e)}")

    def _reset_filter_run_ui(self) -> None:
        """再次开始筛选时：进度条与各滚动条归零，日志清空。"""
        try:
            if hasattr(self, "progress_bar") and self.progress_bar is not None:
                self.progress_bar.setMinimum(0)
                self.progress_bar.setMaximum(100)
                self.progress_bar.setValue(0)
        except Exception:
            pass
        try:
            if hasattr(self, "debug_output") and self.debug_output is not None:
                self.debug_output.clear()
                sb = self.debug_output.verticalScrollBar()
                if sb is not None:
                    sb.setValue(0)
        except Exception:
            pass
        for attr in ("_main_scroll", "sector_scroll"):
            try:
                area = getattr(self, attr, None)
                if area is not None:
                    vsb = area.verticalScrollBar()
                    if vsb is not None:
                        vsb.setValue(0)
            except Exception:
                pass
        try:
            if hasattr(self, "stock_list_table") and self.stock_list_table is not None:
                vsb = self.stock_list_table.verticalScrollBar()
                if vsb is not None:
                    vsb.setValue(0)
        except Exception:
            pass

    def on_progress_updated(self, current: int, total: int, stock_code: str):
        """更新进度"""
        progress = int((current / total) * 100) if total > 0 else 0
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"正在筛选: {stock_code} ({current}/{total})")
    
    def _ensure_rule_tab(self, rule_id: str, rule_name: str, extra_keys: List[str]):
        """确保某条规则的结果 Tab 存在，并按 extra_keys 动态扩展列"""
        rule_id = str(rule_id or "")
        if not rule_id:
            return
        if rule_id not in self._result_tables_by_rule:
            w = QWidget()
            v = QVBoxLayout()
            table = QTableWidget()
            base_cols = ["股票代码", "股票名称", "所属板块", "选股日"]
            cols = base_cols + [k for k in (extra_keys or []) if k and k not in base_cols]
            table.setColumnCount(len(cols))
            table.setHorizontalHeaderLabels(cols)
            header = table.horizontalHeader()
            if len(cols) >= 1:
                header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
            if len(cols) >= 2:
                header.setSectionResizeMode(1, QHeaderView.Stretch)
            if len(cols) >= 3:
                header.setSectionResizeMode(2, QHeaderView.Stretch)
            header.setSectionResizeMode(len(cols) - 1, QHeaderView.ResizeToContents)
            v.addWidget(table)
            btn = QPushButton(f"保存「{rule_name}」结果")
            btn.clicked.connect(lambda _=False, rid=rule_id: self.save_rule_results(rid))
            v.addWidget(btn)
            w.setLayout(v)
            tab_idx = self.result_tabs.addTab(w, f"{rule_name} (0)")
            self._rule_tab_index[rule_id] = tab_idx
            self._result_tables_by_rule[rule_id] = table
            self._result_rows_by_rule[rule_id] = []
        else:
            table = self._result_tables_by_rule[rule_id]
            # 需要新增列时扩展（只增不减）
            existing = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
            to_add = [k for k in (extra_keys or []) if k and k not in existing]
            if to_add:
                new_cols = existing + to_add
                table.setColumnCount(len(new_cols))
                table.setHorizontalHeaderLabels(new_cols)

    def on_stock_found(self, rec: object):
        """找到符合条件的股票（规则驱动；rec 为 dict）"""
        if not isinstance(rec, dict):
            return
        rule_id = str(rec.get("rule_id") or "")
        rule_name = str(rec.get("rule_name") or "未命名规则")
        stock_code = str(rec.get("stock_code") or "")
        stock_name = str(rec.get("stock_name") or "")
        sectors = str(rec.get("sectors") or "")
        as_of = str(rec.get("as_of") or "")
        extra = rec.get("extra") if isinstance(rec.get("extra"), dict) else {}
        is_first_board = self._is_first_board_rule_name(rule_name)
        if is_first_board and isinstance(extra, dict):
            extra = dict(extra)
            extra.pop("今日连板数", None)
        tab_extra_keys: List[str] = []
        seen_k = set()
        for k in self.RESULT_RULE_CONFIG_COLUMNS:
            if k in extra and k not in seen_k:
                tab_extra_keys.append(k)
                seen_k.add(k)
        for k in self.RESULT_CONDITION_OUTCOME_COLUMNS:
            if k in extra and k not in seen_k:
                tab_extra_keys.append(k)
                seen_k.add(k)
        for k in extra.keys():
            if k and not str(k).startswith("_") and k not in seen_k:
                tab_extra_keys.append(k)
                seen_k.add(k)
        for k in self.RESULT_CONCEPT_INFLOW_COLUMNS:
            if k not in seen_k:
                tab_extra_keys.append(k)
                seen_k.add(k)

        self._ensure_rule_tab(rule_id, rule_name, tab_extra_keys)
        table = self._result_tables_by_rule.get(rule_id)
        rows = self._result_rows_by_rule.get(rule_id, [])
        if table is None:
            return

        merged_extra = self._enrich_stock_result_extra(stock_code, as_of, extra)
        if is_first_board:
            merged_extra.pop("今日连板数", None)

        # 去重/合并：同一规则、同一股票、同一选股日视为同一条记录，板块合并
        existing_idx = None
        for i, r in enumerate(rows):
            if r.get("code") == stock_code and r.get("as_of") == as_of:
                existing_idx = i
                break

        if existing_idx is None:
            row_idx = table.rowCount()
            table.insertRow(row_idx)
            base = {"code": stock_code, "name": stock_name, "sectors": sectors, "as_of": as_of}
            base.update(merged_extra)
            rows.append(base)

            col_map = {}
            for ci in range(table.columnCount()):
                col_map[table.horizontalHeaderItem(ci).text()] = ci
            table.setItem(row_idx, col_map.get("股票代码", 0), QTableWidgetItem(stock_code))
            table.setItem(row_idx, col_map.get("股票名称", 1), QTableWidgetItem(stock_name))
            table.setItem(row_idx, col_map.get("所属板块", 2), QTableWidgetItem(sectors))
            table.setItem(row_idx, col_map.get("选股日", 3), QTableWidgetItem(as_of))
            for k, v in merged_extra.items():
                if k in col_map:
                    table.setItem(row_idx, col_map[k], QTableWidgetItem(str(v) if v is not None else ""))
            table.scrollToItem(table.item(row_idx, 0))
        else:
            row_idx = existing_idx
            # 合并板块
            old = str(rows[existing_idx].get("sectors") or "")
            merged_sec = sorted(set([x for x in old.split(";") if x] + [x for x in sectors.split(";") if x]))
            rows[existing_idx]["sectors"] = ";".join(merged_sec)
            # 更新 extra（后来的覆盖空值）
            if isinstance(extra, dict):
                for k, v in extra.items():
                    if k not in rows[existing_idx] or rows[existing_idx].get(k) in ("", None):
                        rows[existing_idx][k] = v
            row_extra = {k: v for k, v in rows[existing_idx].items() if k not in ("code", "name", "sectors", "as_of")}
            rows[existing_idx].update(self._enrich_stock_result_extra(stock_code, as_of, row_extra))
            # 更新表格显示
            for ci in range(table.columnCount()):
                h = table.horizontalHeaderItem(ci).text()
                if h == "所属板块":
                    it = table.item(row_idx, ci)
                    if it:
                        it.setText(";".join(merged_sec))
                elif h in rows[existing_idx] and h not in ("code", "name", "sectors", "as_of"):
                    it = table.item(row_idx, ci)
                    val = rows[existing_idx][h]
                    if it is None:
                        table.setItem(row_idx, ci, QTableWidgetItem(str(val) if val is not None else ""))
                    else:
                        it.setText(str(val) if val is not None else "")

        self._result_rows_by_rule[rule_id] = rows
        # 更新 Tab 标题（用 rule_id 映射，避免规则名含特殊字符时匹配失败）
        ti = self._rule_tab_index.get(rule_id)
        if ti is not None and 0 <= ti < self.result_tabs.count():
            self.result_tabs.setTabText(ti, f"{rule_name} ({len(rows)})")
        self.status_label.setText(f"找到：{rule_name} - {stock_code} {stock_name}（{as_of}）")
    
    def on_finished(self, rule_counts):
        """筛选完成（rule_counts: dict {rule_id: count}）"""
        if not isinstance(rule_counts, dict):
            rule_counts = {}
        total_count = sum(int(v or 0) for v in rule_counts.values()) if rule_counts else 0
        
        # 检查是否是被停止的
        was_stopped = self.filter_thread and not self.filter_thread.is_running
        
        if was_stopped:
            info_text = f"筛选已停止，已找到 {total_count} 只符合条件的股票"
        else:
            rule_info = []
            # 展示前若干条非零规则统计
            shown = 0
            for r in (self.rules or []):
                rid = str(r.get("id") or "")
                c = int(rule_counts.get(rid, 0) or 0)
                if c > 0:
                    rule_info.append(f"{r.get('name','未命名')}:{c}")
                    shown += 1
                if shown >= 6:
                    break
            info_text = f"筛选完成，共找到 {total_count} 只符合条件的股票"
            if rule_info:
                info_text += f"（{', '.join(rule_info)}）"
            info_text += "。"
        
        self.status_label.setText(info_text)
        try:
            if hasattr(self, "progress_bar") and self.progress_bar is not None:
                mx = max(1, int(self.progress_bar.maximum() or 1))
                self.progress_bar.setValue(mx)
        except Exception:
            pass
        
        # 不再自动保存，用户可以通过各模式的保存按钮手动保存
        if self._auto_run and not self._auto_run_finished:
            self._auto_run_finished = True
            saved_n = 0
            failed_rules: List[str] = []
            for r in (self.rules or []):
                if not bool(r.get("enabled", True)):
                    continue
                rid = str(r.get("id") or "")
                rname = str(r.get("name") or "规则")
                if not rid:
                    continue
                try:
                    fp = self._default_rule_result_path(rname)
                    if self._save_rule_results_to_path(rid, fp):
                        saved_n += 1
                except Exception as e:
                    logger.error(f"自动保存规则结果失败: {rname}: {e}", exc_info=True)
                    failed_rules.append(rname)
            logger.info(f"自动运行完成：已自动保存 {saved_n} 个规则结果文件")
            if failed_rules:
                logger.warning(f"自动保存失败规则: {', '.join(failed_rules)}")
            # 自动模式下无需人工干预：直接退出
            QTimer.singleShot(50, self.accept)
            return
        
        # 重新启用开始按钮，禁用停止按钮
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
    
    def on_error(self, error_msg: str):
        """处理错误"""
        if self._auto_run:
            logger.error(f"自动运行错误: {error_msg}")
            self.status_label.setText(f"自动运行错误: {error_msg}")
            QTimer.singleShot(50, self.accept)
        else:
            QMessageBox.warning(self, "错误", error_msg)
        # 重新启用开始按钮，禁用停止按钮
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
    
    def on_debug_info(self, info: str):
        """处理调试信息"""
        # 状态栏只显示最后一行摘要；详细内容追加到可滚动日志框
        try:
            last_line = (info or "").strip().splitlines()[-1] if (info or "").strip() else ""
        except Exception:
            last_line = str(info)
        self.status_label.setText(last_line)
        if hasattr(self, "debug_output") and self.debug_output is not None:
            self.debug_output.append(info)

    def _schedule_auto_run_after_stock_list(self, *, ready: bool, reason: str = "") -> None:
        """auto-run 须等股票列表加载并默认勾选后再开始筛选。"""
        if not self._auto_run_pending:
            return
        if ready:
            self._auto_run_pending = False
            QTimer.singleShot(100, self._start_filter_auto)
            return
        self._auto_run_pending = False
        if self._auto_run_started:
            return
        self._auto_run_started = True
        msg = reason or "未选中任何股票"
        logger.error(f"自动运行失败：{msg}")
        self.status_label.setText(f"自动运行失败：{msg}")
        QTimer.singleShot(50, self.accept)

    def _start_filter_auto(self):
        """自动运行模式：启动后自动执行一次开始筛选。"""
        if self._auto_run_started:
            return
        selected = self.get_selected_stocks()
        if not selected:
            title = self.stock_list_group.title()
            still_loading = (
                self.stock_list_table.rowCount() == 0
                or "加载中" in title
                or "筛选中" in title
                or "准备加载" in title
            )
            if still_loading and self._auto_run_retry_count < 120:
                self._auto_run_retry_count += 1
                QTimer.singleShot(500, self._start_filter_auto)
                return
            self._auto_run_started = True
            logger.error("自动运行失败：未选中任何股票")
            self.status_label.setText("自动运行失败：未选中任何股票")
            QTimer.singleShot(50, self.accept)
            return
        self._auto_run_started = True
        try:
            self.start_filter()
        except Exception as e:
            logger.error(f"自动运行启动失败: {e}", exc_info=True)
            self.status_label.setText(f"自动运行失败: {e}")
            QTimer.singleShot(50, self.accept)

    def _default_rule_result_path(self, rule_name: str) -> str:
        """按当前起止选股日生成该规则默认导出路径。"""
        history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
        os.makedirs(history_dir, exist_ok=True)

        def _fmt_ymd(d: date) -> str:
            return d.strftime("%Y-%m-%d")

        try:
            qd_start = self.as_of_date_edit.date()
            qd_end = self.as_of_date_end_edit.date()
            d_start = date(qd_start.year(), qd_start.month(), qd_start.day())
            d_end = date(qd_end.year(), qd_end.month(), qd_end.day())
            lo, hi = (d_start, d_end) if d_start <= d_end else (d_end, d_start)
            suffix = _fmt_ymd(lo) if lo == hi else f"{_fmt_ymd(lo)}_{_fmt_ymd(hi)}"
        except Exception:
            suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_rule_name = _sanitize_rule_filename(str(rule_name or "规则"))
        return os.path.join(history_dir, f"选股结果_{safe_rule_name}_{suffix}.xls")

    def _save_rule_results_to_path(self, rule_id: str, file_path: str) -> bool:
        """保存指定规则结果到给定路径（无交互对话框）。"""
        rid = str(rule_id or "")
        rows = self._result_rows_by_rule.get(rid, [])
        if not rows:
            return False
        rows = reorder_selection_rows_for_export(rows, max_per_tag=2)
        r = self._get_rule_by_id(rid)
        rule_name = str(r.get("name") if r else "规则")
        is_first_board = self._is_first_board_rule_name(rule_name)
        if not file_path.lower().endswith(".xls"):
            file_path += ".xls"

        base_cols = ["股票代码", "股票名称", "所属板块", "选股日"]
        extra_cols: List[str] = []
        for rr in rows:
            for k in rr.keys():
                if k in ("code", "name", "sectors", "as_of"):
                    continue
                if is_first_board and k == "今日连板数":
                    continue
                if k not in extra_cols:
                    extra_cols.append(k)
        extra_cols = self._order_result_extra_columns(extra_cols)
        cols = base_cols + extra_cols

        out_rows = []
        for rr in rows:
            out = {
                "股票代码": rr.get("code", ""),
                "股票名称": rr.get("name", ""),
                "所属板块": rr.get("sectors", ""),
                "选股日": rr.get("as_of", ""),
            }
            for k in extra_cols:
                v = rr.get(k, "")
                out[k] = "" if v is None else v
            out_rows.append(out)
        df = pd.DataFrame(out_rows, columns=cols)
        if "股票代码" in df.columns:
            df["股票代码"] = df["股票代码"].astype(str).str.strip().str.zfill(6)
        save_xls_with_text_code(file_path, df)
        return True
    
    def save_rule_results(self, rule_id: str):
        """保存指定规则的结果（每条规则独立导出为 XLS）"""
        rid = str(rule_id or "")
        rows = self._result_rows_by_rule.get(rid, [])
        if not rows:
            QMessageBox.information(self, "提示", "该规则没有结果可保存")
            return
        rows = reorder_selection_rows_for_export(rows, max_per_tag=2)
        r = self._get_rule_by_id(rid)
        rule_name = str(r.get("name") if r else "规则")
        is_first_board = self._is_first_board_rule_name(rule_name)

        history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
        os.makedirs(history_dir, exist_ok=True)

        def _fmt_ymd(d: date) -> str:
            return d.strftime("%Y-%m-%d")

        # 默认按界面上的“起/止选股日”生成文件名后缀（含年，便于归档）：
        # - 单日：选股结果_模式一_2026-04-27.xls
        # - 区间：选股结果_模式一_2026-04-01_2026-04-27.xls
        try:
            qd_start = self.as_of_date_edit.date()
            qd_end = self.as_of_date_end_edit.date()
            d_start = date(qd_start.year(), qd_start.month(), qd_start.day())
            d_end = date(qd_end.year(), qd_end.month(), qd_end.day())
            lo, hi = (d_start, d_end) if d_start <= d_end else (d_end, d_start)
            if lo == hi:
                suffix = _fmt_ymd(lo)
            else:
                suffix = f"{_fmt_ymd(lo)}_{_fmt_ymd(hi)}"
        except Exception:
            # 兜底：异常时退回时间戳，避免阻塞保存
            suffix = datetime.now().strftime('%Y%m%d_%H%M%S')

        safe_rule_name = _sanitize_rule_filename(rule_name)
        filename = f"选股结果_{safe_rule_name}_{suffix}.xls"
        default_path = os.path.join(history_dir, filename)
        file_path, _ = QFileDialog.getSaveFileName(self, "保存结果", default_path, "Excel 97-2003 (*.xls)")
        if not file_path:
            return
        if not file_path.lower().endswith(".xls"):
            file_path += ".xls"

        base_cols = ["股票代码", "股票名称", "所属板块", "选股日"]
        extra_cols: List[str] = []
        for rr in rows:
            for k in rr.keys():
                if k in ("code", "name", "sectors", "as_of"):
                    continue
                if is_first_board and k == "今日连板数":
                    continue
                if k not in extra_cols:
                    extra_cols.append(k)
        extra_cols = self._order_result_extra_columns(extra_cols)
        cols = base_cols + extra_cols

        out_rows = []
        for rr in rows:
            out = {
                "股票代码": rr.get("code", ""),
                "股票名称": rr.get("name", ""),
                "所属板块": rr.get("sectors", ""),
                "选股日": rr.get("as_of", ""),
            }
            for k in extra_cols:
                v = rr.get(k, "")
                out[k] = "" if v is None else v
            out_rows.append(out)
        df = pd.DataFrame(out_rows, columns=cols)
        if "股票代码" in df.columns:
            df["股票代码"] = df["股票代码"].astype(str).str.strip().str.zfill(6)
        try:
            save_xls_with_text_code(file_path, df)
        except Exception as e:
            QMessageBox.warning(self, "保存失败", f"写入 .xls 失败：{e}\n请先安装 xlwt：pip install xlwt")
            return
        QMessageBox.information(self, "成功", f"结果已保存到:\n{file_path}")

    def save_mode_results(self, mode: int):
        """旧版模式导出入口（已升级为规则驱动）"""
        QMessageBox.information(self, "提示", "当前选股系统已改为“规则驱动”，请在结果标签页点击对应规则的“保存结果”按钮导出。")
    
    def export_stock_list_to_file(self):
        """导出当前股票列表中选中的股票到Excel文件"""
        try:
            row_count = self.stock_list_table.rowCount()
            if row_count == 0:
                QMessageBox.information(self, "提示", "当前没有可导出的股票")
                return
            
            # 创建history_data目录（如果不存在）
            history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
            os.makedirs(history_dir, exist_ok=True)
            
            # 生成文件名（Excel格式）
            filename = f"板块筛选_股票列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(history_dir, filename)
            
            # 准备数据：只导出选中的股票，包含"股票代码、股票名称、所属板块"三列（不含"是否选中"列）
            export_rows = []
            for row in range(row_count):
                # 跳过被隐藏的行（搜索过滤后）
                if self.stock_list_table.isRowHidden(row):
                    continue
                
                checkbox = self.stock_list_table.cellWidget(row, 0)
                code_item = self.stock_list_table.item(row, 1)
                name_item = self.stock_list_table.item(row, 2)
                sectors_item = self.stock_list_table.item(row, 3)
                
                if not code_item or not name_item:
                    continue
                
                # 只导出选中的股票
                is_selected = False
                if checkbox and isinstance(checkbox, QCheckBox):
                    is_selected = checkbox.isChecked()
                
                if not is_selected:
                    continue  # 跳过未选中的股票
                
                export_rows.append({
                    "code": code_item.text().strip(),
                    "name": name_item.text().strip(),
                    "sector": sectors_item.text().strip() if sectors_item else ''
                })
            
            if not export_rows:
                QMessageBox.information(self, "提示", "当前没有选中的股票可导出")
                return
            
            # 创建DataFrame（列名使用'code'、'name'和'sector'，以便save_excel_with_text_code函数识别）
            df = pd.DataFrame(export_rows)
            
            # 使用save_excel_with_text_code函数保存Excel（确保股票代码是文本格式）
            save_excel_with_text_code(file_path, df)
            
            # 重新读取并更新表头为中文
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active
                # 更新表头为中文
                if ws.cell(row=1, column=1).value == 'code':
                    ws.cell(row=1, column=1).value = '股票代码'
                if ws.cell(row=1, column=2).value == 'name':
                    ws.cell(row=1, column=2).value = '股票名称'
                if ws.cell(row=1, column=3).value == 'sector':
                    ws.cell(row=1, column=3).value = '所属板块'
                wb.save(file_path)
                wb.close()
            except Exception as e:
                logger.warning(f"更新Excel表头为中文时出错: {str(e)}，但文件已保存")
            
            logger.info(f"股票列表已导出到: {file_path}，共 {len(export_rows)} 条记录")
            self.status_label.setText(f"股票列表已导出到: {file_path}，共 {len(export_rows)} 条记录")
            QMessageBox.information(self, "导出成功", f"股票列表已导出到:\n{file_path}\n共 {len(export_rows)} 条记录")
        
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")
            logger.error(f"导出股票列表时出错: {str(e)}", exc_info=True)
    
    def auto_save_results(self):
        """自动保存结果（合并各规则 Tab 的结果）"""
        try:
            # 创建history_data目录（如果不存在）
            history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
            os.makedirs(history_dir, exist_ok=True)
            
            # 生成文件名
            filename = f"板块筛选_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
            file_path = os.path.join(history_dir, filename)
            
            all_results = []
            rule_counts: Dict[str, int] = {}
            for r in (self.rules or []):
                rid = str(r.get("id") or "")
                rname = str(r.get("name") or "规则")
                rows = self._result_rows_by_rule.get(rid, [])
                if not rows:
                    continue
                rows = reorder_selection_rows_for_export(rows, max_per_tag=2)
                rule_counts[rname] = len(rows)
                for rr in rows:
                    stock_copy = {
                        "股票代码": rr.get("code", ""),
                        "股票名称": rr.get("name", ""),
                        "所属板块": rr.get("sectors", ""),
                        "选股日": rr.get("as_of", ""),
                        "规则": rname,
                    }
                    for k, v in rr.items():
                        if k in ("code", "name", "sectors", "as_of"):
                            continue
                        if k not in stock_copy:
                            stock_copy[k] = v
                    all_results.append(stock_copy)
            
            if all_results:
                df = pd.DataFrame(all_results)
                if "股票代码" in df.columns:
                    df["股票代码"] = df["股票代码"].astype(str).str.strip().str.zfill(6)
            else:
                df = pd.DataFrame(columns=["股票代码", "股票名称", "所属板块", "选股日", "规则"])
            try:
                save_xls_with_text_code(file_path, df)
            except Exception as e:
                raise RuntimeError(f"写入 .xls 失败：{e}（请先安装 xlwt：pip install xlwt）")
            
            total_count = len(all_results)
            detail = ", ".join(f"{k}:{v}" for k, v in rule_counts.items()) if rule_counts else "无"
            logger.info(f"结果已保存到: {file_path}，共 {total_count} 条（{detail}）")
            
            # 不再自动关闭窗口，只显示保存成功信息
            if all_results:
                self.status_label.setText(f"结果已保存到: {file_path}")
            else:
                self.status_label.setText(f"未找到符合条件的股票，已保存空结果到: {file_path}")
            
            # 重新启用开始按钮
            self.start_button.setEnabled(True)
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")
            logger.error(f"保存结果时出错: {str(e)}", exc_info=True)
            # 重新启用开始按钮
            self.start_button.setEnabled(True)
    
    def start_countdown(self):
        """启动30秒倒计时"""
        # 如果已有定时器，先停止
        if self.countdown_timer:
            self.countdown_timer.stop()
            self.countdown_timer = None
        
        self.countdown_seconds = 30
        self.countdown_timer = QTimer()
        self.countdown_timer.timeout.connect(self.update_countdown)
        self.countdown_timer.start(1000)  # 每秒更新一次
        self.update_countdown()  # 立即更新一次
    
    def update_countdown(self):
        """更新倒计时"""
        if self.countdown_seconds > 0:
            # 更新状态标签显示倒计时
            current_text = self.status_label.text()
            # 如果状态标签包含倒计时信息，更新它；否则追加
            if "秒后自动关闭" in current_text:
                # 替换倒计时数字
                import re
                new_text = re.sub(r'\d+秒后自动关闭', f'{self.countdown_seconds}秒后自动关闭', current_text)
                self.status_label.setText(new_text)
            else:
                self.status_label.setText(f"{current_text}，{self.countdown_seconds}秒后自动关闭...")
            
            self.countdown_seconds -= 1
        else:
            # 倒计时结束，关闭程序
            if self.countdown_timer:
                self.countdown_timer.stop()
                self.countdown_timer = None
            logger.info("倒计时结束，自动关闭窗口")
            self.accept()  # 使用accept()而不是close()，确保对话框正确关闭
    
    def closeEvent(self, event):
        """关闭事件"""
        if self._rule_code_dirty:
            if not self._handle_unsaved_rule_code_before_action("关闭窗口"):
                event.ignore()
                return
        if self.filter_thread and self.filter_thread.isRunning():
            # 如果正在筛选，停止线程
            self.filter_thread.stop()
            self.filter_thread.wait()
        
        if self.countdown_timer:
            self.countdown_timer.stop()

        try:
            self._save_stock_exclude_prefs()
        except Exception:
            pass
        
        event.accept()


def main():
    parser = argparse.ArgumentParser(description="蚂蚁量化选股系统")
    parser.add_argument(
        "--as-of",
        metavar="YYYY-MM-DD",
        default=None,
        help="预设选股基准日（回测）；界面仍可修改",
    )
    parser.add_argument(
        "--auto-run",
        action="store_true",
        help="自动运行：启动后自动开始筛选，完成后自动保存结果并退出",
    )
    args, _unknown = parser.parse_known_args()
    initial_as_of = None
    if args.as_of:
        try:
            initial_as_of = datetime.strptime(args.as_of.strip(), "%Y-%m-%d").date()
        except ValueError:
            print(f"无效的 --as-of 日期: {args.as_of!r}，请使用 YYYY-MM-DD", file=sys.stderr)
            sys.exit(2)

    # 优先复用启动早期创建的 QApplication / 单例服务器 / splash（覆盖重依赖加载阶段）
    app = _EARLY_APP or QApplication.instance() or QApplication(sys.argv)

    server = _EARLY_SINGLETON_SERVER
    if server is None:
        # 与备份目录（原版 AntStockFilterSingleton）并存，便于 A/B 对比
        server_name = _SINGLETON_SERVER_NAME

        # 先尝试作为“第二实例”：连到已有的本地服务器，若成功则发送激活请求并退出
        socket = QLocalSocket()
        socket.connectToServer(server_name)
        if socket.waitForConnected(200):
            try:
                socket.write(b"activate")
                socket.flush()
                socket.waitForBytesWritten(200)
            except Exception:
                pass
            socket.disconnectFromServer()
            # 已有实例在运行，本实例直接退出
            return
        socket.abort()

        # 没有已运行实例：创建本地服务器，供后续实例发送“activate”指令
        server = QLocalServer()
        # 防止残留的同名服务器阻止绑定
        try:
            QLocalServer.removeServer(server_name)
        except Exception:
            pass
        server.listen(server_name)

    splash = _EARLY_SPLASH
    if splash is None:
        try:
            splash = QLabel("正在加载选股系统")
            splash.setWindowTitle("蚂蚁量化选股系统")
            splash.setAlignment(Qt.AlignCenter)
            splash.setFixedSize(360, 80)
            splash.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.SplashScreen)
            splash.setStyleSheet(
                "background:#FFF8E1; color:#333; font-size:12pt; border:1px solid #FFCC80;"
            )
            splash.show()
            app.processEvents()
        except Exception:
            splash = None
    else:
        try:
            app.processEvents()
        except Exception:
            pass

    dialog = SectorStockFilterDialog(initial_as_of=initial_as_of, auto_run=bool(args.auto_run))

    if splash is not None:
        try:
            splash.close()
        except Exception:
            pass
        splash = None

    def handle_new_connection():
        client = server.nextPendingConnection()
        if not client:
            return
        try:
            if client.waitForReadyRead(200):
                _ = client.readAll()  # 当前只关心有无消息，不解析内容
                # 收到激活请求：把窗口前置显示
                dialog.showNormal()
                dialog.raise_()
                dialog.activateWindow()
        except Exception:
            pass
        finally:
            client.disconnectFromServer()

    server.newConnection.connect(handle_new_connection)

    dialog.showMaximized()  # 最大化显示，但保留标题栏和最大化/最小化按钮

    sys.exit(app.exec_())


if __name__ == '__main__':
    # 如果今天不是交易日，则跳过
    # if not is_tradeday():
    #     print("今天不是交易日，跳过")
    #     sys.exit(0)
    main()


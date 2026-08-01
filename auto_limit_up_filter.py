#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
自动涨停无新高股票筛选程序
程序启动后自动开始运行，搜索完成后自动保存，保存后30秒倒计时关闭
"""

import sys
import os
import pandas as pd
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Tuple
import time
import logging
import warnings
import json
import csv
from utils.trading_day import is_tradeday

# 抑制Qt和log4cplus的警告/错误信息（必须在导入任何Qt或xtquant模块之前设置）
# 设置环境变量抑制Qt Windows版本警告
os.environ.setdefault('QT_LOGGING_RULES', '*.debug=false')

# 抑制Python警告
warnings.filterwarnings("ignore")

# 重定向log4cplus错误（来自第三方库xtquant/QMT）
# 保存原始的stderr
_original_stderr = sys.stderr

class FilteredStderr:
    """过滤stderr，隐藏log4cplus错误和Qt警告"""
    def __init__(self, original_stderr):
        self.original_stderr = original_stderr
    
    def write(self, text):
        # 过滤log4cplus错误信息
        if 'log4cplus' in text.lower() or 'adsyncnamespace' in text.lower():
            return
        # 过滤Qt Windows版本警告
        if 'Qt: Untested Windows version' in text:
            return
        # 其他信息正常输出
        self.original_stderr.write(text)
    
    def flush(self):
        self.original_stderr.flush()
    
    def __getattr__(self, name):
        return getattr(self.original_stderr, name)

# 应用stderr过滤器（必须在导入任何可能产生警告的模块之前）
sys.stderr = FilteredStderr(_original_stderr)

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QTableWidget, QTableWidgetItem, QLabel, QMessageBox,
                             QProgressBar, QHeaderView, QApplication, QSplitter,
                             QWidget, QFrame)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont

# 导入 matplotlib 用于绑定图表
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates
import numpy as np

try:
    import xtquant.xtdata as xtdata
    from utils.trading_day import is_tradeday
    from utils.stock_info_manager import get_stock_name
except ImportError as e:
    print(f"导入模块失败: {e}")

# 配置日志（必须在导入其他模块之前配置，以便在导入失败时可以使用logger）
_script_dir = os.path.dirname(os.path.abspath(__file__))
_logs_dir = os.path.join(_script_dir, 'logs')
os.makedirs(_logs_dir, exist_ok=True)
_log_file_path = os.path.join(_logs_dir, 'auto_limit_up_filter.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(_log_file_path, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 导入主力净流入添加工具
try:
    from auto_add_inflow import InflowAdder
    logger.info("成功导入主力净流入工具")
except ImportError as e:
    logger.warning(f"无法导入主力净流入工具: {e}")
    InflowAdder = None
except Exception as e:
    logger.error(f"导入主力净流入工具时出错: {e}", exc_info=True)
    InflowAdder = None


def save_excel_with_text_code(excel_file_path: str, df: pd.DataFrame):
    """保存Excel文件，确保股票代码列是文本格式
    
    Args:
        excel_file_path: Excel文件路径
        df: 要保存的DataFrame
    """
    try:
        # 确保code列是字符串类型，并补零
        if 'code' in df.columns:
            def clean_code(code):
                code_str = str(code).strip()
                if '.' in code_str:
                    code_str = code_str.split('.')[0]
                # 确保是6位数字（补零）
                if code_str.isdigit():
                    return code_str.zfill(6)
                return code_str
            df['code'] = df['code'].apply(clean_code)
            # 确保列类型是字符串
            df['code'] = df['code'].astype(str)
        
        # 使用ExcelWriter和openpyxl引擎，在写入前就设置格式
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # 找到code列的索引
            code_col_idx = None
            if 'code' in df.columns:
                for idx, col_name in enumerate(df.columns, start=1):
                    if col_name == 'code':
                        code_col_idx = idx
                        break
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    
                    # 如果是code列，设置为文本格式
                    if code_col_idx and c_idx == code_col_idx and r_idx > 1:  # 跳过表头
                        cell.number_format = '@'  # @表示文本格式
                        # 确保值是字符串，并补零
                        if value is not None:
                            value_str = str(value).strip()
                            if '.' in value_str:
                                value_str = value_str.split('.')[0]
                            if value_str.isdigit():
                                value_str = value_str.zfill(6)
                            cell.value = value_str
                        else:
                            cell.value = ''
                    else:
                        cell.value = value
            
            wb.save(excel_file_path)
            wb.close()
        except Exception as e:
            # 如果openpyxl方法失败，回退到原来的方法
            logger.warning(f"使用openpyxl直接写入失败: {str(e)}，尝试回退方法")
            # 保存Excel
            df.to_excel(excel_file_path, index=False, engine='openpyxl')
            
            # 使用openpyxl设置code列为文本格式
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path)
                ws = wb.active
                
                # 找到code列的索引
                if 'code' in df.columns:
                    for col_idx, header in enumerate(df.columns, start=1):
                        if header == 'code':
                            # 将该列的所有单元格设置为文本格式
                            for row_idx in range(2, len(df) + 2):  # 从第2行开始（第1行是标题）
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.number_format = '@'  # @表示文本格式
                                # 确保值是字符串，并补零
                                if cell.value is not None:
                                    value_str = str(cell.value).strip()
                                    if '.' in value_str:
                                        value_str = value_str.split('.')[0]
                                    if value_str.isdigit():
                                        value_str = value_str.zfill(6)
                                    cell.value = value_str
                            break
                
                wb.save(excel_file_path)
                wb.close()
            except Exception as e2:
                logger.warning(f"设置Excel文本格式时出错: {str(e2)}，但文件已保存")
            
    except Exception as e:
        raise Exception(f"保存Excel文件失败: {str(e)}")


def add_stock_info_columns(file_path: str) -> bool:
    """为CSV文件添加概念、行业和板块列
    
    Args:
        file_path: CSV文件路径
        
    Returns:
        bool: 是否成功
    """
    try:
        # 加载股票信息JSON文件
        current_dir = os.path.dirname(__file__)
        json_path = os.path.join(current_dir, 'data', 'all_a_stock_info.json')
        if not os.path.exists(json_path):
            logger.warning(f"股票信息文件不存在: {json_path}")
            return False
        
        logger.info(f"正在加载股票信息文件: {json_path}")
        with open(json_path, 'r', encoding='utf-8') as f:
            stock_info_dict = json.load(f)
        logger.info(f"成功加载 {len(stock_info_dict)} 只股票的信息")
        
        # 检测文件编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']
        encoding = None
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    f.readline()
                encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if encoding is None:
            logger.error("无法检测文件编码")
            return False
        
        logger.info(f"检测到文件编码: {encoding}")
        
        # 读取CSV文件
        rows = []
        fieldnames = None
        
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames) if reader.fieldnames else []
            
            for row in reader:
                rows.append(row)
        
        if not rows:
            logger.warning("文件中没有数据行")
            return False
        
        logger.info(f"读取到 {len(rows)} 行数据")
        
        # 添加新列（如果不存在）
        new_columns = ['概念', '行业', '板块']
        for col in new_columns:
            if col not in fieldnames:
                fieldnames.append(col)
                logger.info(f"添加新列: {col}")
        
        def normalize_code_for_lookup(raw):
            """统一为6位字符串，便于与 all_a_stock_info.json 的键匹配（JSON 键可能是 000001 或 1）"""
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                return ''
            s = str(raw).strip()
            if '.' in s and s.split('.')[0].isdigit():
                s = s.split('.')[0]  # 去掉 .0
            if s.replace('.', '').isdigit():
                s = s.split('.')[0].zfill(6)
            return s
        
        # 填充数据
        filled_count = 0
        empty_count = 0
        
        for row in rows:
            code_raw = row.get('code', '')
            code = normalize_code_for_lookup(code_raw)
            
            if code:
                # 从JSON中获取股票信息（键可能是 000001 或 1，先试6位再试数字形式）
                stock_info = stock_info_dict.get(code, {})
                if not stock_info and code.lstrip('0'):
                    stock_info = stock_info_dict.get(code.lstrip('0') or '0', {})
                if not stock_info and code.isdigit():
                    stock_info = stock_info_dict.get(str(int(code)), {})
                
                # 概念：列表转字符串（用分号分隔）
                concepts = stock_info.get('concepts', [])
                if concepts and isinstance(concepts, list):
                    # 确保所有元素都是字符串
                    row['概念'] = ';'.join(str(c) for c in concepts if c)
                else:
                    row['概念'] = ''
                
                # 行业：直接使用字符串
                industry = stock_info.get('industry', '')
                row['行业'] = str(industry) if industry else ''
                
                # 板块：列表转字符串（用分号分隔）
                plates = stock_info.get('plates', [])
                if plates and isinstance(plates, list):
                    # 确保所有元素都是字符串
                    row['板块'] = ';'.join(str(p) for p in plates if p)
                else:
                    row['板块'] = ''
                
                if concepts or stock_info.get('industry') or plates:
                    filled_count += 1
                else:
                    empty_count += 1
            else:
                row['概念'] = ''
                row['行业'] = ''
                row['板块'] = ''
                empty_count += 1
        
        logger.info(f"填充完成：成功 {filled_count} 条，未找到 {empty_count} 条")
        
        # 保存回原文件
        logger.info(f"正在保存文件...")
        with open(file_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        logger.info(f"完成！文件已保存: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"添加股票信息列时出错: {str(e)}", exc_info=True)
        return False


def _normalize_date_str(d) -> str:
    """将日期转为 YYYY-MM-DD 字符串"""
    if d is None or (isinstance(d, float) and pd.isna(d)):
        return ""
    s = str(d).strip()
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    try:
        return pd.Timestamp(d).strftime('%Y-%m-%d')
    except Exception:
        return s[:10] if len(s) >= 10 else s


# 在统计和选择“当日最多涨停概念”时，需要剔除的概念/板块/行业名称
EXCLUDED_TAGS = {
    '一级',
    '二级',
    '三级',
    '融资融券',
    '最近多板',
    '央国企改革',
    '标准普尔',
    '江苏板块',
    '西部大开发',
    '沪股通',
    '湖南板块',
    '北京板块',
    '东方财富热股',
    '上海板块',
    '广东板块',
    '深股通',
    '小盘股',
    '一带一路',
    '机构重仓',
    '上证',
    '反内卷概念',
    '微盘股',
    '新材料',
    '稀缺资源',
    '专精特新',
    '富时罗素',
    '公共事业',
    '福建板块',
    '海南板块',
    '化债(AMC)概念',
    '浙江板块',
    '长江三角',
    '并购重组概念',
    '股权转让',
    '转债标的',
    '低价股',
    '深成',
}


def _load_limit_up_codes_for_date(history_dir: str, date_str: str) -> List[str]:
    """加载某日涨停板数据，返回当日涨停股票代码列表（6位字符串）。"""
    if not date_str:
        return []
    date_norm = _normalize_date_str(date_str)
    if not date_norm:
        return []
    filename = f"涨停板数据_{date_norm}.csv"
    filepath = os.path.join(history_dir, filename)
    if not os.path.exists(filepath):
        logger.debug(f"未找到涨停板数据文件: {filepath}")
        return []
    codes = []
    for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                next(reader)  # 跳过表头
                for row in reader:
                    if len(row) < 1:
                        continue
                    code = str(row[0]).strip()
                    code = ''.join(c for c in code if c.isdigit())
                    if len(code) == 6:
                        codes.append(code.zfill(6))
            break
        except (UnicodeDecodeError, UnicodeError, Exception):
            continue
    return codes


def _get_stock_tags(stock_info_dict: Dict, code: str) -> List[str]:
    """获取股票所属的概念/行业/板块列表（去重、非空）。"""
    if not code:
        return []
    code = str(code).strip()
    if code.isdigit():
        code = code.zfill(6)
    info = stock_info_dict.get(code) or stock_info_dict.get(code.lstrip('0') or '0') or (stock_info_dict.get(str(int(code))) if code.isdigit() else None)
    if not info:
        return []
    tags = []
    industry = info.get('industry', '')
    if industry and str(industry).strip():
        tags.append(str(industry).strip())
    for c in (info.get('concepts') or []):
        if c and str(c).strip():
            tags.append(str(c).strip())
    for p in (info.get('plates') or []):
        if p and str(p).strip():
            tags.append(str(p).strip())
    # 去重并剔除不需要参与统计的概念/板块/行业
    unique_tags = list(dict.fromkeys(tags))
    return [t for t in unique_tags if t not in EXCLUDED_TAGS]


def _build_date_concept_stats(
    history_dir: str,
    stock_info_dict: Dict,
    dates_list: List[str],
) -> Dict[str, Dict[str, Tuple[int, int]]]:
    """
    按日期预计算：当日各概念/板块/行业的涨停数和排名。
    返回 { 日期_str: { 概念名: (涨停数, 排名) } }，每个日期只读一次涨停板数据。
    """
    unique_dates = []
    seen = set()
    for date_val in dates_list:
        date_norm = _normalize_date_str(date_val)
        if date_norm and date_norm not in seen:
            seen.add(date_norm)
            unique_dates.append(date_norm)
    result = {}
    for date_norm in unique_dates:
        codes = _load_limit_up_codes_for_date(history_dir, date_norm)
        concept_count: Dict[str, int] = {}
        for code in codes:
            for tag in _get_stock_tags(stock_info_dict, code):
                concept_count[tag] = concept_count.get(tag, 0) + 1
        sorted_items = sorted(concept_count.items(), key=lambda x: (-x[1], x[0]))
        rank_map = {}
        for rank, (concept, count) in enumerate(sorted_items, start=1):
            rank_map[concept] = (count, rank)
        result[date_norm] = rank_map
    return result


def add_concept_rank_columns(df: pd.DataFrame, history_dir: str) -> pd.DataFrame:
    """
    为 DataFrame 增加三列（基于涨停板数据 + all_a_stock_info）：
    - 当日最多涨停概念：该股所属概念中当日涨停数最多的那个名称
    - 该概念当日涨停数
    - 该概念当日排名
    每个涨停日只读取一次涨停板数据并计算一次排名。
    """
    if df is None or df.empty:
        return df
    if 'limit_date' not in df.columns or 'code' not in df.columns:
        return df
    json_path = os.path.join(os.path.dirname(__file__), 'data', 'all_a_stock_info.json')
    if not os.path.exists(json_path):
        logger.warning("all_a_stock_info.json 不存在，跳过概念排名列")
        df['当日最多涨停概念'] = ''
        df['该概念当日涨停数'] = ''
        df['该概念当日排名'] = ''
        return df
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            stock_info_dict = json.load(f)
    except Exception as e:
        logger.warning(f"加载 all_a_stock_info 失败: {e}，跳过概念排名列")
        df['当日最多涨停概念'] = ''
        df['该概念当日涨停数'] = ''
        df['该概念当日排名'] = ''
        return df
    dates_used = df['limit_date'].dropna().unique().tolist()
    date_to_stats = _build_date_concept_stats(history_dir, stock_info_dict, dates_used)
    col_name = '当日最多涨停概念'
    col_count = '该概念当日涨停数'
    col_rank = '该概念当日排名'
    df[col_name] = ''
    df[col_count] = ''
    df[col_rank] = ''
    for idx, row in df.iterrows():
        code = row.get('code', '')
        code = str(code).strip()
        if code.isdigit():
            code = code.zfill(6)
        dt = row.get('limit_date', '')
        date_key = _normalize_date_str(dt)
        stats = date_to_stats.get(date_key, {})
        tags = _get_stock_tags(stock_info_dict, code)
        if not tags or not stats:
            continue
        best_name, best_count, best_rank = '', 0, 0
        for tag in tags:
            tup = stats.get(tag)
            if tup and tup[0] > best_count:
                best_count, best_rank = tup[0], tup[1]
                best_name = tag
        if best_name:
            df.at[idx, col_name] = best_name
            df.at[idx, col_count] = int(best_count)
            df.at[idx, col_rank] = int(best_rank)
    return df


# ==================== 赚钱指数相关类 ====================

class ProfitIndexCalculatorThread(QThread):
    """赚钱指数计算线程"""
    
    progress_updated = pyqtSignal(int, int, str)  # 当前进度, 总数, 当前股票
    calculation_finished = pyqtSignal(dict)  # 计算结果
    error_occurred = pyqtSignal(str)  # 错误信息
    
    def __init__(self, days: int = 30, parent=None):
        super().__init__(parent)
        self.days = days  # 要显示的天数
        self.extra_days = 10  # 额外获取的天数（用于计算10日均线）
        self.is_running = True
    
    def stop(self):
        self.is_running = False
    
    def _get_trading_dates(self) -> List[str]:
        """获取最近N+extra+1个交易日（需要额外天数来计算均线）
        
        使用 xtdata.get_trading_dates() 获取交易日历，比 akshare 更可靠
        """
        from datetime import datetime, timedelta
        
        current_date = datetime.now()
        
        # 如果当前时间 < 15:00，从昨天开始算
        if current_date.hour < 15:
            current_date = current_date - timedelta(days=1)
        
        # 需要的交易日数量
        # +1 是因为需要前一天来计算涨跌
        total_days_needed = self.days + self.extra_days + 1
        
        try:
            # 使用 xtdata 获取交易日历（更可靠）
            # 往前推足够多的自然日以确保获取到足够的交易日
            start_date = (current_date - timedelta(days=total_days_needed * 2)).strftime('%Y%m%d')
            end_date = current_date.strftime('%Y%m%d')
            
            # 获取沪市交易日历
            trading_dates_ts = xtdata.get_trading_dates('SH', start_time=start_date, end_time=end_date)
            
            if trading_dates_ts and len(trading_dates_ts) > 0:
                # 转换时间戳为日期字符串
                trading_dates = []
                for ts in trading_dates_ts:
                    if isinstance(ts, (int, float)):
                        # 时间戳格式（毫秒）
                        dt = datetime.fromtimestamp(ts / 1000)
                        trading_dates.append(dt.strftime('%Y%m%d'))
                    elif isinstance(ts, str):
                        trading_dates.append(ts.replace('-', ''))
                    else:
                        trading_dates.append(str(ts))
                
                # 按日期升序排列
                trading_dates.sort()
                
                # 取最后 total_days_needed 个
                if len(trading_dates) >= total_days_needed:
                    trading_dates = trading_dates[-total_days_needed:]
                
                logger.info(f"使用 xtdata 获取交易日历成功，共 {len(trading_dates)} 个交易日")
                return trading_dates
        except Exception as e:
            logger.warning(f"xtdata 获取交易日历失败: {e}，回退到 is_tradeday 方式")
        
        # 回退方案：使用 is_tradeday 逐日判断
        trading_dates = []
        check_date = current_date
        while len(trading_dates) < total_days_needed:
            if is_tradeday(check_date):
                trading_dates.append(check_date.strftime('%Y%m%d'))
            check_date = check_date - timedelta(days=1)
            # 安全限制：最多往前查150天
            if (current_date - check_date).days > 150:
                break
        
        # 按日期升序排列
        trading_dates.reverse()
        return trading_dates
    
    def run(self):
        """运行计算"""
        try:
            # 获取所有A股代码
            all_stocks = xtdata.get_stock_list_in_sector('沪深A股')
            if not all_stocks:
                # 如果获取失败，尝试其他方式
                all_stocks = []
                for sector in ['上证A股', '深证A股']:
                    stocks = xtdata.get_stock_list_in_sector(sector)
                    if stocks:
                        all_stocks.extend(stocks)
                all_stocks = list(set(all_stocks))
            
            logger.info(f"赚钱指数计算：共 {len(all_stocks)} 只股票")
            
            # 获取交易日列表
            trading_dates = self._get_trading_dates()
            if len(trading_dates) < 2:
                self.error_occurred.emit("交易日数据不足")
                return
            
            logger.info(f"赚钱指数计算：交易日范围 {trading_dates[0]} ~ {trading_dates[-1]}")
            
            # 初始化统计数据
            # date -> {'up': 上涨家数, 'down': 下跌家数, 'flat': 平盘家数}
            daily_stats = {d: {'up': 0, 'down': 0, 'flat': 0} for d in trading_dates[1:]}
            
            total = len(all_stocks)
            start_date = trading_dates[0]
            end_date = trading_dates[-1]
            
            # 不使用 download_history_data2 批量下载（会卡住）
            # 改为在遍历每只股票时逐只下载，和涨停筛选一致
            logger.info(f"开始计算赚钱指数（共 {total} 只股票）...")
            
            # 遍历所有股票
            for idx, stock_code in enumerate(all_stocks):
                if not self.is_running:
                    break
                
                # 更新进度
                if idx % 100 == 0:
                    self.progress_updated.emit(idx + 1, total, stock_code)
                
                try:
                    # 先下载该股票的日K数据（确保数据最新）
                    try:
                        xtdata.download_history_data(stock_code, '1d', start_date, end_date)
                    except:
                        pass  # 下载失败时使用本地缓存
                    
                    # 获取K线数据
                    df = xtdata.get_market_data_ex([], [stock_code], period='1d',
                                                   start_time=start_date,
                                                   end_time=end_date,
                                                   count=-1)
                    
                    if df is None or stock_code not in df or df[stock_code].empty:
                        continue
                    
                    stock_df = df[stock_code]
                    
                    # 确保有收盘价列
                    if 'close' not in stock_df.columns:
                        continue
                    
                    # 将索引转换为日期字符串格式，方便匹配
                    # xtdata 返回的索引可能是时间戳(毫秒)或日期字符串
                    index_to_date = {}
                    for idx in stock_df.index:
                        try:
                            if isinstance(idx, (int, float)):
                                # 时间戳格式（毫秒）
                                dt = datetime.fromtimestamp(idx / 1000)
                                date_str = dt.strftime('%Y%m%d')
                            elif isinstance(idx, str):
                                date_str = idx.replace('-', '')[:8]
                            else:
                                date_str = str(idx)[:8]
                            index_to_date[date_str] = idx
                        except:
                            continue
                    
                    # 遍历每个交易日，计算涨跌
                    prev_close = None
                    for trade_date in trading_dates:
                        if trade_date in index_to_date:
                            try:
                                original_idx = index_to_date[trade_date]
                                close = stock_df.loc[original_idx, 'close']
                                
                                if prev_close is not None and trade_date in daily_stats:
                                    if close > prev_close:
                                        daily_stats[trade_date]['up'] += 1
                                    elif close < prev_close:
                                        daily_stats[trade_date]['down'] += 1
                                    else:
                                        daily_stats[trade_date]['flat'] += 1
                                
                                prev_close = close
                            except:
                                pass
                        else:
                            prev_close = None  # 该股票当天无数据，重置
                    
                except Exception as e:
                    # 单只股票出错不影响整体
                    continue
            
            # 打印每天的统计数据，用于调试
            logger.info("=" * 50)
            logger.info("每日涨跌统计：")
            for trade_date in sorted(daily_stats.keys()):
                stats = daily_stats[trade_date]
                total_stocks = stats['up'] + stats['down'] + stats['flat']
                logger.info(f"  {trade_date}: 上涨={stats['up']}, 下跌={stats['down']}, 平盘={stats['flat']}, 总计={total_stocks}")
            logger.info("=" * 50)
            
            # 计算所有天数的赚钱指数（包括额外的天数用于计算均线）
            all_dates = []
            all_profit_index = []
            all_up_count = []
            all_down_count = []
            all_flat_count = []
            
            for trade_date in trading_dates[1:]:  # 跳过第一天（没有前一天对比）
                stats = daily_stats[trade_date]
                up = stats['up']
                down = stats['down']
                flat = stats['flat']
                total = up + down + flat
                
                # 方案A：涨跌强度指数 = (上涨 - 下跌) / 总数 * 100
                # 范围：-100 ~ +100，0是平衡点
                if total > 0:
                    profit_idx = (up - down) / total * 100
                else:
                    profit_idx = 0
                
                all_dates.append(trade_date)
                all_profit_index.append(profit_idx)
                all_up_count.append(up)
                all_down_count.append(down)
                all_flat_count.append(flat)
            
            # 计算5日均线和10日均线（基于完整数据）
            all_ma5 = []
            all_ma10 = []
            for i in range(len(all_profit_index)):
                # 5日均线
                if i < 4:
                    all_ma5.append(None)
                else:
                    ma5 = sum(all_profit_index[i-4:i+1]) / 5
                    all_ma5.append(ma5)
                
                # 10日均线
                if i < 9:
                    all_ma10.append(None)
                else:
                    ma10 = sum(all_profit_index[i-9:i+1]) / 10
                    all_ma10.append(ma10)
            
            # 只返回最后 days 天的数据（用于显示）
            display_count = min(self.days, len(all_dates))
            result = {
                'dates': all_dates[-display_count:],
                'profit_index': all_profit_index[-display_count:],
                'up_count': all_up_count[-display_count:],
                'down_count': all_down_count[-display_count:],
                'flat_count': all_flat_count[-display_count:],
                'ma5': all_ma5[-display_count:],
                'ma10': all_ma10[-display_count:]
            }
            
            logger.info(f"赚钱指数计算完成，共 {len(result['dates'])} 个交易日（含均线数据）")
            self.calculation_finished.emit(result)
            
        except Exception as e:
            logger.error(f"赚钱指数计算出错: {str(e)}", exc_info=True)
            self.error_occurred.emit(f"计算出错: {str(e)}")


class ProfitIndexChartWidget(QWidget):
    """赚钱指数图表组件"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 创建 matplotlib 图表
        self.figure = Figure(figsize=(10, 6), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        
        # 状态标签
        self.status_label = QLabel("正在计算赚钱指数...")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False
    
    def update_chart(self, data: dict):
        """更新图表"""
        self.figure.clear()
        
        if not data or not data.get('dates'):
            self.status_label.setText("无数据")
            return
        
        dates = data['dates']
        profit_index = data['profit_index']
        up_count = data['up_count']
        down_count = data['down_count']
        flat_count = data['flat_count']
        ma5 = data['ma5']
        ma10 = data.get('ma10', [None] * len(dates))  # 兼容旧数据
        
        # 转换日期格式用于显示
        date_labels = [f"{d[4:6]}/{d[6:8]}" for d in dates]  # MM/DD 格式
        x = np.arange(len(dates))
        
        # 创建两个子图：上面是赚钱指数，下面是涨跌家数堆叠柱状图
        gs = self.figure.add_gridspec(2, 1, height_ratios=[2, 1], hspace=0.1)
        
        # 上图：赚钱指数折线图
        ax1 = self.figure.add_subplot(gs[0])
        
        # 绘制赚钱指数折线
        ax1.plot(x, profit_index, 'o-', color='#333333', linewidth=1.5, markersize=4, label='涨跌强度')
        
        # 用颜色填充区域（>0红色区域，<0绿色区域）
        for i in range(len(x)):
            if profit_index[i] >= 0:
                ax1.fill_between([x[i]-0.3, x[i]+0.3], [0, 0], [profit_index[i], profit_index[i]], 
                               color='#FFCCCC', alpha=0.7)
            else:
                ax1.fill_between([x[i]-0.3, x[i]+0.3], [profit_index[i], profit_index[i]], [0, 0], 
                               color='#CCFFCC', alpha=0.7)
        
        # 绘制5日均线
        ma5_valid = [(i, v) for i, v in enumerate(ma5) if v is not None]
        if ma5_valid:
            ma5_x, ma5_y = zip(*ma5_valid)
            ax1.plot(ma5_x, ma5_y, '--', color='#FF8800', linewidth=2, label='MA5')
        
        # 绘制10日均线
        ma10_valid = [(i, v) for i, v in enumerate(ma10) if v is not None]
        if ma10_valid:
            ma10_x, ma10_y = zip(*ma10_valid)
            ax1.plot(ma10_x, ma10_y, '--', color='#0088FF', linewidth=2, label='MA10')
        
        # 绘制0基准线
        ax1.axhline(y=0, color='#666666', linestyle='--', linewidth=1.5, label='平衡线(0)')
        
        # 设置标题和标签
        ax1.set_title('A股涨跌强度指数（最近30个交易日）', fontsize=14, fontweight='bold')
        ax1.set_ylabel('涨跌强度', fontsize=10)
        ax1.legend(loc='upper left', fontsize=8)
        ax1.grid(True, alpha=0.3)
        
        # 隐藏x轴标签（下图会显示）
        ax1.set_xticks(x)
        ax1.set_xticklabels([])
        
        # 设置y轴范围，确保0在中间附近，留出一些空间
        y_abs_max = max(abs(min(profit_index)), abs(max(profit_index)), 10) * 1.2
        ax1.set_ylim(-y_abs_max, y_abs_max)
        
        # 在数据点上显示数值
        for i, v in enumerate(profit_index):
            color = '#FF4444' if v >= 0 else '#00AA00'
            offset_y = 8 if v >= 0 else -12
            ax1.annotate(f'{v:.0f}', (x[i], v), textcoords="offset points", 
                        xytext=(0, offset_y), ha='center', fontsize=7, color=color)
        
        # 下图：涨跌家数堆叠柱状图
        ax2 = self.figure.add_subplot(gs[1])
        
        bar_width = 0.6
        ax2.bar(x, up_count, bar_width, label='上涨', color='#FF4444', alpha=0.8)
        ax2.bar(x, flat_count, bar_width, bottom=up_count, label='平盘', color='#888888', alpha=0.8)
        ax2.bar(x, down_count, bar_width, bottom=[u+f for u,f in zip(up_count, flat_count)], 
               label='下跌', color='#00AA00', alpha=0.8)
        
        ax2.set_ylabel('家数', fontsize=10)
        ax2.set_xticks(x)
        ax2.set_xticklabels(date_labels, rotation=45, ha='right', fontsize=8)
        ax2.legend(loc='upper left', fontsize=8, ncol=3)
        ax2.grid(True, alpha=0.3, axis='y')
        
        # 调整布局
        self.figure.tight_layout()
        self.canvas.draw()
        
        # 更新状态
        latest_idx = profit_index[-1]
        if latest_idx > 20:
            status = "赚钱很容易 📈📈"
        elif latest_idx > 0:
            status = "赚钱容易 📈"
        elif latest_idx > -20:
            status = "赚钱困难 📉"
        else:
            status = "赚钱很困难 📉📉"
        self.status_label.setText(f"最新涨跌强度: {latest_idx:.1f} ({status})")


class LimitUpNoNewHighFilterThread(QThread):
    """涨停无新高筛选线程"""
    
    # 信号：更新进度
    progress_updated = pyqtSignal(int, int, str)  # (当前数量, 总数, 当前股票代码)
    # 信号：找到符合条件的股票
    stock_found = pyqtSignal(str, str, str, str)  # (股票代码, 股票名称, 涨停日期, 涨停价)
    # 信号：筛选完成
    finished = pyqtSignal(int)  # (找到的股票数量)
    # 信号：错误信息
    error_occurred = pyqtSignal(str)  # (错误信息)
    # 信号：调试信息（用于显示统计）
    debug_info = pyqtSignal(str)  # (调试信息)
    
    def __init__(self, stock_list: List[Tuple[str, str]], trading_dates: List[date], parent=None):
        super().__init__(parent)
        self.stock_list = stock_list
        self.trading_dates = trading_dates
        self.is_running = True
        
    def stop(self):
        """停止筛选"""
        self.is_running = False
    
    def _get_limit_ratio(self, stock_code: str, as_of_date=None) -> float:
        """获取股票的涨停幅度"""
        try:
            from utils.limit_ratio import get_limit_ratio

            stock_name = get_stock_name(stock_code) or ""
            return get_limit_ratio(stock_code, stock_name, as_of_date)
        except Exception:
            return 0.10
    
    def _get_full_stock_code(self, stock_code: str) -> str:
        """获取完整的股票代码（带市场后缀）"""
        if '.' in stock_code:
            return stock_code
        
        if stock_code.startswith(('0', '1', '3')):
            return f"{stock_code}.SZ"
        elif stock_code.startswith('6'):
            return f"{stock_code}.SH"
        elif stock_code.startswith(('8', '4', '920')):
            return f"{stock_code}.BJ"
        else:
            return stock_code
    
    def _calculate_technical_indicators(self, daily_data: pd.DataFrame) -> pd.DataFrame:
        """计算技术指标：布林线上轨
        
        注意：计算布林线需要至少20个交易日的数据。
        在 _get_daily_data 中，我们获取了60天的数据以确保有足够的历史数据来计算布林线。
        """
        try:
            # 确保数据按日期排序
            daily_data = daily_data.sort_values('date').copy()
            
            # 计算布林线
            # 中轨：MA20（20日移动平均）
            # 使用 min_periods=1 确保即使数据不足20天也能计算（但精度会降低）
            daily_data['MA20'] = daily_data['close'].rolling(window=20, min_periods=1).mean()
            # 标准差
            daily_data['STD20'] = daily_data['close'].rolling(window=20, min_periods=1).std()
            # 上轨 = 中轨 + 2 * 标准差
            daily_data['BOLL_UPPER'] = daily_data['MA20'] + 2.0 * daily_data['STD20']
            
            return daily_data
            
        except Exception as e:
            logger.error(f"计算技术指标失败: {str(e)}", exc_info=True)
            return daily_data
    
    def _get_daily_data(self, stock_code: str, include_prev_day: bool = True) -> Optional[pd.DataFrame]:
        """获取日线数据
        
        Args:
            stock_code: 股票代码
            include_prev_day: 是否包含前一个交易日（用于计算涨停价）
                注意：当 include_prev_day=True 时，返回所有获取的数据（约60天），
                用于计算布林线等技术指标；当 include_prev_day=False 时，只返回最近10个交易日的数据。
        """
        try:
            full_stock_code = self._get_full_stock_code(stock_code)
            
            # 计算日期范围（往前推更多天数以确保获取到足够的交易日和前一个交易日，以及计算布林线需要的数据）
            # 布林线需要至少20个交易日的数据，所以需要获取足够的历史数据
            end_date = self.trading_dates[-1] if self.trading_dates else date.today()
            start_date = end_date - timedelta(days=60)  # 往前推60天，确保包含前一个交易日和足够的数据计算布林线（约40-45个交易日）
            
            startdate = start_date.strftime("%Y%m%d") + "000000"
            enddate = end_date.strftime("%Y%m%d") + "235959"
            
            # 下载历史数据
            try:
                xtdata.download_history_data(full_stock_code, '1d', startdate, enddate)
                time.sleep(0.05)  # 短暂等待
            except Exception as e:
                logger.error(f"[{stock_code}] 下载历史数据失败: {str(e)}", exc_info=True)
                return None
            
            # 获取历史行情数据
            try:
                df = xtdata.get_market_data_ex([], [full_stock_code], period='1d',
                                             start_time=startdate,
                                             end_time=enddate,
                                             count=-1)
            except Exception as e:
                logger.error(f"[{stock_code}] 获取市场数据失败: {str(e)}", exc_info=True)
                return None
            
            if df is None:
                return None
            
            if full_stock_code not in df:
                return None
            
            if len(df[full_stock_code]) == 0:
                return None
            
            # 转换为DataFrame
            stock_data = df[full_stock_code]
            
            daily_data = pd.DataFrame({
                'time': stock_data['time'],
                'open': stock_data['open'],
                'high': stock_data['high'],
                'low': stock_data['low'],
                'close': stock_data['close'],
                'volume': stock_data['volume']
            })
            
            # 设置time为索引并转换为日期
            daily_data.set_index('time', inplace=True)
            
            # 转换时间戳为日期（考虑时区：先转换为UTC，再转换为东八区）
            time_index = pd.to_datetime(daily_data.index, unit='ms')
            # 对于DatetimeIndex，直接使用时区方法
            if time_index.tz is None:
                # 如果没有时区信息，先设置为UTC，再转换为东八区
                time_index = time_index.tz_localize('UTC').tz_convert('Asia/Shanghai')
            else:
                # 如果已有时区信息，直接转换为东八区
                time_index = time_index.tz_convert('Asia/Shanghai')
            # 取东八区的日期：将DatetimeIndex转换为Series，使用原始索引，然后取日期
            date_series = pd.Series(time_index, index=daily_data.index)
            daily_data['date'] = date_series.dt.date
            
            # 筛选交易日数据
            daily_data = daily_data[daily_data['date'].apply(is_tradeday)]
            
            # 按日期排序
            daily_data = daily_data.sort_values('date')
            
            # 计算技术指标（布林线）
            daily_data = self._calculate_technical_indicators(daily_data)
            
            # 如果只需要最近10个交易日的数据，进行筛选
            if not include_prev_day and len(daily_data) > 0:
                # 只保留在trading_dates范围内的数据
                daily_data = daily_data[daily_data['date'].isin(self.trading_dates)]
            
            return daily_data
            
        except Exception as e:
            logger.error(f"[{stock_code}] 获取日线数据异常: {str(e)}", exc_info=True)
            return None
    
    def _check_stock(self, stock_code: str, stock_name: str) -> Optional[Tuple[str, str]]:
        """检查股票是否符合条件"""
        try:
            # 获取日线数据（包含前一个交易日）
            # 注意：include_prev_day=True 时，会返回所有获取的数据（约60天），
            # 这样才有足够的历史数据来计算布林线（需要至少20个交易日）
            daily_data = self._get_daily_data(stock_code, include_prev_day=True)
            if daily_data is None or daily_data.empty:
                return None
            
            # 获取涨停幅度
            # limit_ratio 在循环内按 trade_date 取值（ST 主板 2026-07-06 前 5%、之后 10%）
            
            # 筛选出最近10个交易日的数据
            trading_dates_set = set(self.trading_dates)
            recent_data = daily_data[daily_data['date'].isin(trading_dates_set)]
            
            if recent_data.empty:
                return None
            
            # 获取最后一个交易日（用于排除）
            last_trading_date = recent_data['date'].max()
            
            # 按日期从早到晚排序，确保从早往后查找
            recent_data_sorted = recent_data.sort_values('date')
            
            # 遍历最近10个交易日，查找第一个涨停（排除最后一个交易日）
            for idx, row in recent_data_sorted.iterrows():
                trade_date = row['date']
                
                limit_ratio = self._get_limit_ratio(stock_code, trade_date)
                
                # 排除最后一个交易日，因为无法判断后续是否有更高的价格
                if trade_date >= last_trading_date:
                    continue
                
                close_price = row['close']
                high_price = row['high']
                
                # 从全部数据中找到前一个交易日
                prev_trading_days = daily_data[daily_data['date'] < trade_date]
                if prev_trading_days.empty:
                    continue  # 没有前一日数据，跳过
                
                prev_row = prev_trading_days.iloc[-1]  # 获取最后一个（最近的）前一个交易日
                prev_close = prev_row['close']
                prev_date = prev_row['date']
                
                # 计算涨停价（四舍五入到分）
                limit_up_price = round(prev_close * (1 + limit_ratio), 2)
                
                # 判断是否涨停
                # 方法1：收盘价达到涨停价（允许0.02的误差）
                price_diff = abs(close_price - limit_up_price)
                # 方法2：收盘价相对于前一日收盘价的涨幅（用于处理数据偏差）
                price_increase_ratio = (close_price - prev_close) / prev_close if prev_close > 0 else 0
                expected_increase_ratio = limit_ratio
                
                # 判断条件：
                # 1. 收盘价接近涨停价（差值<0.02）
                # 2. 或者收盘价相对前一日涨幅接近涨停幅度（在涨停幅度的99%以上）
                is_limit_up = (price_diff < 0.02) or (price_increase_ratio >= expected_increase_ratio * 0.99)
                
                if is_limit_up:
                    # 找到第一个涨停，立即检查条件
                    
                    # 1. 检查涨停板价格是否 >= 布林线上轨
                    # 获取涨停板当天的布林线上轨
                    limit_day_data = daily_data[daily_data['date'] == trade_date]
                    if limit_day_data.empty:
                        continue
                    
                    boll_upper = limit_day_data['BOLL_UPPER'].iloc[0]
                    # 如果涨停板价格 < 布林线上轨，不满足条件，直接返回None
                    # 允许0.01的误差，因为价格可能有微小的浮动
                    if limit_up_price < boll_upper - 0.01:
                        return None
                    
                    # 2. 检查涨停板之前的交易日（在10个交易日内）的最高价是否都低于涨停板价
                    before_limit_data = recent_data_sorted[recent_data_sorted['date'] < trade_date]
                    if not before_limit_data.empty:
                        # 检查涨停板之前的所有交易日的最高价
                        max_high_before = before_limit_data['high'].max()
                        # 如果涨停板之前的最高价 >= 涨停板价，不满足条件，直接返回None
                        # 允许0.01的误差，因为价格可能有微小的浮动
                        if max_high_before >= limit_up_price - 0.01:
                            return None
                    
                    # 3. 检查涨停板之后的交易日（在10个交易日内）的最高价不能超过涨停板价格的一定比例
                    # 比例是涨停板的50%：10%涨停不能高出5%，20%涨停不能高出10%，5%涨停不能高出2.5%
                    after_limit_data = daily_data[daily_data['date'] > trade_date]
                    
                    # 如果涨停后没有数据，跳过（不应该出现，因为已经排除了最后一个交易日）
                    if after_limit_data.empty:
                        continue
                    
                    # 计算允许的最大价格：涨停板价格 * (1 + 涨停板涨幅 * 0.5)
                    # 例如：10%涨停，允许最高价 = 涨停价 * 1.05（即不能高出5%）
                    max_allowed_price = limit_up_price * (1 + limit_ratio * 0.5)
                    # 检查涨停板之后的所有交易日的最高价
                    max_high_after = after_limit_data['high'].max()
                    # 如果涨停板之后的最高价 > 允许的最大价格，不满足条件，直接返回None
                    # 允许0.01的误差，因为价格可能有微小的浮动
                    if max_high_after > max_allowed_price + 0.01:
                        return None
                    
                    # 如果所有条件都满足，符合条件，返回结果（不再继续查找）
                    return (trade_date.strftime('%Y-%m-%d'), f"{limit_up_price:.2f}")
            
            return None
            
        except Exception as e:
            logger.error(f"[{stock_code}] 检查过程中出错: {str(e)}", exc_info=True)
            return None
    
    def run(self):
        """运行筛选"""
        found_count = 0
        limit_up_count = 0  # 统计有涨停的股票数量
        
        try:
            total = len(self.stock_list)
            for idx, (stock_code, stock_name) in enumerate(self.stock_list):
                if not self.is_running:
                    break
                
                # 更新进度
                self.progress_updated.emit(idx + 1, total, stock_code)
                
                # 检查股票
                result = self._check_stock(stock_code, stock_name)
                if result:
                    limit_date, limit_price = result
                    self.stock_found.emit(stock_code, stock_name, limit_date, limit_price)
                    found_count += 1
                    # 每找到10只股票，发送一次调试信息
                    if found_count % 10 == 0:
                        self.debug_info.emit(f"已找到 {found_count} 只符合条件的股票")
                
                # 每100只股票发送一次统计信息
                if (idx + 1) % 100 == 0:
                    self.debug_info.emit(f"已筛选 {idx + 1}/{total} 只股票，找到 {found_count} 只符合条件的股票")
                
                # 短暂延迟，避免请求过快
                time.sleep(0.1)
            
            self.finished.emit(found_count)
            
        except Exception as e:
            self.error_occurred.emit(f"筛选过程出错: {str(e)}")
            self.finished.emit(found_count)


class AutoLimitUpNoNewHighDialog(QDialog):
    """自动涨停无新高股票筛选对话框"""
    
    def __init__(self, parent=None, trading_days_count: int = 10):
        super().__init__(parent)
        self.trading_days_count = max(1, min(int(trading_days_count), 60))  # 1~60 天
        self.setWindowTitle("蚂蚁量化 - 涨停无新高筛选")
        self.setMinimumSize(900, 620)
        self.resize(1000, 700)
        
        # 去掉右上角的问号（帮助按钮）
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        self.filter_thread = None
        self.profit_index_thread = None  # 兼容旧变量（本版本不再启动赚钱指数）
        self.result_stocks = []  # 存储筛选结果
        self.countdown_timer = None  # 倒计时定时器
        self.countdown_seconds = 10  # 倒计时秒数（完成后10秒自动退出）
        
        self.setup_ui()
        
        # 启动后直接筛选
        QTimer.singleShot(300, self.start_filter)
    
    def setup_ui(self):
        """设置UI"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(5)
        
        filter_frame = QFrame()
        filter_frame.setFrameStyle(QFrame.StyledPanel)
        filter_layout = QVBoxLayout(filter_frame)
        filter_layout.setContentsMargins(5, 5, 5, 5)
        
        # 说明标签
        info_label = QLabel(f"涨停无新高筛选条件（最近{self.trading_days_count}个交易日）：\n"
                           f"1. 涨停板价格 >= 布林线上轨  2. 涨停前无新高  3. 涨停后未大幅超越")
        info_label.setWordWrap(True)
        font = QFont()
        font.setPointSize(9)
        info_label.setFont(font)
        filter_layout.addWidget(info_label)
        
        # 按钮和进度区域
        control_layout = QHBoxLayout()
        
        self.close_button = QPushButton("关闭")
        self.close_button.clicked.connect(self.close)
        control_layout.addWidget(self.close_button)
        
        control_layout.addStretch()
        
        # 状态标签
        self.status_label = QLabel("正在启动...")
        control_layout.addWidget(self.status_label)
        
        control_layout.addStretch()
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedWidth(200)
        control_layout.addWidget(self.progress_bar)
        
        filter_layout.addLayout(control_layout)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["股票代码", "股票名称", "涨停日期", "涨停价"])
        
        # 设置表格属性
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        
        filter_layout.addWidget(self.result_table)
        
        main_layout.addWidget(filter_frame)
        self.setLayout(main_layout)
    
    def start_profit_index_calculation(self):
        """启动赚钱指数计算"""
        try:
            self.profit_index_thread = ProfitIndexCalculatorThread(days=30)
            self.profit_index_thread.progress_updated.connect(self.on_profit_index_progress)
            self.profit_index_thread.calculation_finished.connect(self.on_profit_index_finished)
            self.profit_index_thread.error_occurred.connect(self.on_profit_index_error)
            self.profit_index_thread.start()
            logger.info("赚钱指数计算线程已启动")
        except Exception as e:
            logger.error(f"启动赚钱指数计算失败: {str(e)}", exc_info=True)
    
    def on_profit_index_progress(self, current, total, stock_code):
        """赚钱指数计算进度更新"""
        self.profit_index_chart.status_label.setText(f"计算赚钱指数... {current}/{total} ({stock_code})")
    
    def on_profit_index_finished(self, data):
        """赚钱指数计算完成"""
        try:
            self.profit_index_chart.update_chart(data)
            logger.info("赚钱指数图表更新完成")
        except Exception as e:
            logger.error(f"更新赚钱指数图表失败: {str(e)}", exc_info=True)
    
    def on_profit_index_error(self, error_msg):
        """赚钱指数计算出错"""
        self.profit_index_chart.status_label.setText(f"计算出错: {error_msg}")
        logger.error(f"赚钱指数计算出错: {error_msg}")
    
    def _get_trading_dates(self) -> List[date]:
        """获取最近 N 个交易日（N = self.trading_days_count）

        规则：
        - 如果今天是交易日且当前时间 ≥ 09:30，则把今天算作最近的一个交易日；
        - 否则，从昨天开始往前找 N 个交易日。
        """
        try:
            current_time = datetime.now()
            current_date = current_time.date()
            current_hour = current_time.hour
            current_minute = current_time.minute
            
            # 交易日且已过 9:30 才包含当天
            include_today = is_tradeday(current_date) and (
                (current_hour > 9) or (current_hour == 9 and current_minute >= 30)
            )
            
            trading_dates = []
            search_date = current_date if include_today else current_date - timedelta(days=1)
            found_days = 0
            n = getattr(self, 'trading_days_count', 10)
            # 节后第一天等场景：假期+周末会导致自然日内不足 n 个交易日，故放宽到 60 天
            max_search_days = 60
            
            while found_days < n and max_search_days > 0:
                if is_tradeday(search_date):
                    trading_dates.append(search_date)
                    found_days += 1
                search_date -= timedelta(days=1)
                max_search_days -= 1
            
            # 按时间顺序排列（从早到晚）
            trading_dates.sort()
            
            return trading_dates
            
        except Exception as e:
            return []
    
    def _load_stock_list(self) -> List[Tuple[str, str]]:
        """从CSV文件加载股票列表"""
        csv_path = os.path.join(os.path.dirname(__file__), 'data', 'all_a_stocks.csv')
        if not os.path.exists(csv_path):
            QMessageBox.warning(self, "错误", f"找不到股票列表文件: {csv_path}")
            return []
        
        # 尝试多种编码方式
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
        df = None
        last_error = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                break  # 成功读取，跳出循环
            except (UnicodeDecodeError, UnicodeError) as e:
                last_error = e
                continue  # 尝试下一个编码
            except Exception as e:
                last_error = e
                continue
        
        if df is None:
            QMessageBox.warning(self, "错误", f"加载股票列表失败: 无法使用常见编码（utf-8, gbk, gb2312, gb18030）读取文件\n最后错误: {str(last_error)}")
            return []
        
        try:
            stock_list = []
            
            for _, row in df.iterrows():
                stock_code = str(row['证券代码']).zfill(6)  # 确保是6位数字
                # 只保留0、3、6开头的股票代码
                if not stock_code.startswith(('0', '3', '4','6' ,'8' ,'9')):
                    continue
                stock_name = str(row['证券简称'])
                stock_list.append((stock_code, stock_name))
            
            return stock_list
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"解析股票列表失败: {str(e)}")
            return []
    
    def start_filter(self):
        """开始筛选"""
        try:
            # 加载股票列表
            stock_list = self._load_stock_list()
            if not stock_list:
                return
            
            # 获取交易日列表（节后第一天等会放宽搜索范围，尽量凑满指定天数）
            trading_dates = self._get_trading_dates()
            n = getattr(self, 'trading_days_count', 10)
            if len(trading_dates) < 2:
                QMessageBox.warning(self, "警告", f"无法获取足够的交易日（只找到{len(trading_dates)}个）")
                return
            if len(trading_dates) < n:
                self.status_label.setText(f"提示：仅找到最近{len(trading_dates)}个交易日，将按此范围筛选（目标{n}个）")
            
            # 清空结果
            # 注意：不在主线程批量下载数据，避免UI卡死
            # 筛选线程中的 _get_daily_data() 会逐只股票下载数据
            self.result_stocks = []
            self.result_table.setRowCount(0)
            
            # 创建筛选线程
            self.filter_thread = LimitUpNoNewHighFilterThread(stock_list, trading_dates, self)
            self.filter_thread.progress_updated.connect(self.on_progress_updated)
            self.filter_thread.stock_found.connect(self.on_stock_found)
            self.filter_thread.finished.connect(self.on_finished)
            self.filter_thread.error_occurred.connect(self.on_error)
            self.filter_thread.debug_info.connect(self.on_debug_info)
            
            # 更新状态
            self.status_label.setText("正在筛选股票...")
            
            # 启动线程
            self.filter_thread.start()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"启动筛选失败: {str(e)}")
    
    def on_progress_updated(self, current: int, total: int, stock_code: str):
        """更新进度"""
        progress = int((current / total) * 100) if total > 0 else 0
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"正在筛选: {stock_code} ({current}/{total})")
    
    def on_stock_found(self, stock_code: str, stock_name: str, limit_date: str, limit_price: str):
        """找到符合条件的股票"""
        # 添加到结果列表
        self.result_stocks.append({
            'code': stock_code,
            'name': stock_name,
            'limit_date': limit_date,
            'limit_price': limit_price
        })
        
        # 添加到表格
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        self.result_table.setItem(row, 0, QTableWidgetItem(stock_code))
        self.result_table.setItem(row, 1, QTableWidgetItem(stock_name))
        self.result_table.setItem(row, 2, QTableWidgetItem(limit_date))
        self.result_table.setItem(row, 3, QTableWidgetItem(limit_price))
        
        # 滚动到最新添加的行，确保用户能看到
        self.result_table.scrollToItem(self.result_table.item(row, 0))
        
        # 更新状态标签显示最新找到的股票
        self.status_label.setText(f"找到符合条件的股票: {stock_code} {stock_name} (涨停日期: {limit_date}, 涨停价: {limit_price})")
    
    def on_finished(self, count: int):
        """筛选完成"""
        self.status_label.setText(f"筛选完成，找到 {count} 只符合条件的股票，正在保存...")
        
        # 自动保存
        self.auto_save_results()
    
    def on_error(self, error_msg: str):
        """处理错误"""
        QMessageBox.warning(self, "错误", error_msg)
        # 即使出错也尝试保存已有结果
        if self.result_stocks:
            self.auto_save_results()
    
    def on_debug_info(self, info: str):
        """处理调试信息"""
        # 在状态标签中显示调试信息
        self.status_label.setText(info)
    
    def auto_save_results(self):
        """自动保存结果（只保存Excel格式）"""
        try:
            # 创建history_data目录（如果不存在）
            history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
            os.makedirs(history_dir, exist_ok=True)
            
            # 生成文件名（只生成Excel文件名）
            n = getattr(self, 'trading_days_count', 10)
            base_filename = f"{n}日内涨停_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            excel_file_path = os.path.join(history_dir, f"{base_filename}.xlsx")
            
            # 准备数据
            if self.result_stocks:
                df = pd.DataFrame(self.result_stocks)
                # 确保股票代码格式为6位数字（去掉可能的后缀）
                if 'code' in df.columns:
                    # 使用更兼容的方法处理股票代码格式
                    def clean_code(code):
                        code_str = str(code).strip()
                        # 去掉市场后缀（如 .SZ, .SH, .BJ）
                        if '.' in code_str:
                            code_str = code_str.split('.')[0]
                        # 确保是6位数字
                        return code_str.zfill(6) if code_str.isdigit() else code_str
                    df['code'] = df['code'].apply(clean_code)
                # 添加三列：当日最多涨停概念、该概念当日涨停数、该概念当日排名（每个涨停日只算一次排名）
                try:
                    self.status_label.setText(f"正在计算当日概念涨停排名...")
                    self.status_label.repaint()
                    QApplication.processEvents()
                    df = add_concept_rank_columns(df, history_dir)
                except Exception as e:
                    logger.warning(f"添加概念排名列时出错: {e}", exc_info=True)
            else:
                # 即使没有结果也创建一个空DataFrame
                df = pd.DataFrame(columns=['code', 'name', 'limit_date', 'limit_price'])
            
            # 保存为Excel
            try:
                save_excel_with_text_code(excel_file_path, df)
                logger.info(f"已保存 {len(self.result_stocks)} 条结果到Excel: {excel_file_path}")
                if self.result_stocks:
                    self.status_label.setText(f"结果已保存到Excel: {base_filename}，正在添加主力净流入...")
                else:
                    self.status_label.setText(f"未找到符合条件的股票，已保存空结果到Excel: {base_filename}，正在添加主力净流入...")
            except ImportError:
                # 如果没有安装openpyxl，提示错误
                logger.error("未安装openpyxl，无法保存Excel文件")
                QMessageBox.warning(self, "错误", "未安装openpyxl，无法保存Excel文件。请安装: pip install openpyxl")
                self.status_label.setText(f"保存失败：未安装openpyxl")
                self.start_countdown()
                return
            except Exception as e:
                logger.error(f"保存Excel文件时出错: {str(e)}", exc_info=True)
                QMessageBox.warning(self, "错误", f"保存Excel文件失败: {str(e)}")
                self.status_label.setText(f"保存失败: {str(e)}")
                self.start_countdown()
                return
            
            # 更新UI
            self.status_label.repaint()
            
            # 自动添加主力净流入列
            if InflowAdder is not None:
                try:
                    self.status_label.setText(f"正在为文件添加主力净流入列...")
                    self.status_label.repaint()
                    
                    # 先读取Excel文件为DataFrame，添加主力净流入列，再保存回Excel
                    logger.info(f"开始为文件添加主力净流入列: {excel_file_path}")
                    try:
                        df_temp = pd.read_excel(excel_file_path, engine='openpyxl')
                        # 临时保存为CSV用于InflowAdder处理
                        import tempfile
                        temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8-sig')
                        temp_csv_path = temp_csv.name
                        temp_csv.close()
                        df_temp.to_csv(temp_csv_path, index=False, encoding='utf-8-sig')
                        
                        adder = InflowAdder()
                        success = adder.add_inflow_column(temp_csv_path)
                        
                        if success:
                            # 读取更新后的CSV并保存回Excel
                            df_updated = pd.read_csv(temp_csv_path, encoding='utf-8-sig')
                            save_excel_with_text_code(excel_file_path, df_updated)
                            logger.info(f"已更新Excel文件（添加主力净流入列）: {excel_file_path}")
                            self.status_label.setText(f"结果已保存到Excel: {base_filename}（已添加主力净流入列），正在添加概念、行业和板块列...")
                        else:
                            self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列失败），正在添加概念、行业和板块列...")
                            logger.warning(f"为文件添加主力净流入列失败: {excel_file_path}")
                        
                        # 删除临时CSV文件
                        try:
                            os.unlink(temp_csv_path)
                        except:
                            pass
                    except Exception as e:
                        logger.error(f"添加主力净流入列时出错: {str(e)}", exc_info=True)
                        self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列时出错），正在添加概念、行业和板块列...")
                except Exception as e:
                    logger.error(f"添加主力净流入列时出错: {str(e)}", exc_info=True)
                    self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列时出错），正在添加概念、行业和板块列...")
            else:
                logger.warning("主力净流入工具未启用（InflowAdder 为 None）")
                self.status_label.setText(f"结果已保存到Excel: {base_filename}（未启用主力净流入功能），正在添加概念、行业和板块列...")
            
            # 自动添加概念、行业和板块列
            try:
                self.status_label.setText(f"正在为文件添加概念、行业和板块列...")
                self.status_label.repaint()
                QApplication.processEvents()
                
                # 先读取Excel文件，添加列，再保存回Excel
                try:
                    df_temp = pd.read_excel(excel_file_path, engine='openpyxl')
                    # 临时保存为CSV用于add_stock_info_columns处理
                    import tempfile
                    temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8-sig')
                    temp_csv_path = temp_csv.name
                    temp_csv.close()
                    df_temp.to_csv(temp_csv_path, index=False, encoding='utf-8-sig')
                    
                    success = add_stock_info_columns(temp_csv_path)
                    
                    if success:
                        # 读取更新后的CSV并保存回Excel
                        df_updated = pd.read_csv(temp_csv_path, encoding='utf-8-sig')
                        save_excel_with_text_code(excel_file_path, df_updated)
                        logger.info(f"已更新Excel文件（添加概念、行业和板块列）: {excel_file_path}")
                        self.status_label.setText(f"✅ 筛选完成！结果已保存: {base_filename}")
                    else:
                        self.status_label.setText(f"✅ 筛选完成！结果已保存: {base_filename}（部分列添加失败）")
                        logger.warning(f"为文件添加概念、行业和板块列失败: {excel_file_path}")
                    
                    # 删除临时CSV文件
                    try:
                        os.unlink(temp_csv_path)
                    except:
                        pass
                except Exception as e:
                    logger.error(f"添加概念、行业和板块列时出错: {str(e)}", exc_info=True)
                    self.status_label.setText(f"✅ 筛选完成！结果已保存: {base_filename}")
            except Exception as e:
                logger.error(f"添加概念、行业和板块列时出错: {str(e)}", exc_info=True)
                self.status_label.setText(f"✅ 筛选完成！结果已保存: {base_filename}")
            
            # 启动倒计时
            self.start_countdown()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")
            logger.error(f"保存结果时出错: {str(e)}", exc_info=True)
            # 即使保存失败也启动倒计时
            self.start_countdown()
    
    def start_countdown(self):
        """保存结束后开始10秒倒计时自动退出。"""
        if self.countdown_timer:
            self.countdown_timer.stop()
            self.countdown_timer.deleteLater()
            self.countdown_timer = None
        self.countdown_seconds = 10
        self.status_label.setText(f"处理完成，{self.countdown_seconds} 秒后自动退出...")
        self.countdown_timer = QTimer(self)
        self.countdown_timer.timeout.connect(self.update_countdown)
        self.countdown_timer.start(1000)
    
    def update_countdown(self):
        """更新倒计时并在归零时退出程序。"""
        self.countdown_seconds -= 1
        if self.countdown_seconds <= 0:
            if self.countdown_timer:
                self.countdown_timer.stop()
            self.close()
            return
        self.status_label.setText(f"处理完成，{self.countdown_seconds} 秒后自动退出...")
    
    def closeEvent(self, event):
        """关闭事件"""
        # 停止赚钱指数计算线程
        if self.profit_index_thread and self.profit_index_thread.isRunning():
            self.profit_index_thread.stop()
            self.profit_index_thread.wait()
        
        # 停止筛选线程
        if self.filter_thread and self.filter_thread.isRunning():
            self.filter_thread.stop()
            self.filter_thread.wait()
        
        if self.countdown_timer:
            self.countdown_timer.stop()
        
        event.accept()


def main():
    app = QApplication(sys.argv)
    
    # 命令行参数：python auto_limit_up_filter.py [天数]，如 python auto_limit_up_filter.py 5 表示最近5个交易日；不带参数默认10天
    trading_days_count = 10
    if len(sys.argv) >= 2 and str(sys.argv[1]).strip().isdigit():
        trading_days_count = int(sys.argv[1].strip())
        trading_days_count = max(1, min(trading_days_count, 60))
    
    dialog = AutoLimitUpNoNewHighDialog(trading_days_count=trading_days_count)
    dialog.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    #如果今天不是交易日，则跳过
    if not is_tradeday():
        print("今天不是交易日，跳过")
        sys.exit(0)
    main()


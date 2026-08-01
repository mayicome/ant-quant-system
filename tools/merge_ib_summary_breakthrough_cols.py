#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
临时：将 history_data 目录下「各日选股收益汇总_智能买入」类 xlsx 的每一行，
按「选股日（或买入日）+ 股票代码」到同目录「突破详情_买入信号.xlsx」中查找三列：
  突破量/均量(倍)、委卖/委买比、突破量/前tick被吃卖档合计(倍)（旧表头「突破量/前卖一量(倍)」仍兼容）
追加在汇总表各列后面，写出新文件（默认文件名加后缀 _带突破指标）。

依赖：pandas、openpyxl

用法:
  python tools/merge_ib_summary_breakthrough_cols.py
  python tools/merge_ib_summary_breakthrough_cols.py --data-dir history_data
  python tools/merge_ib_summary_breakthrough_cols.py --data-dir D:\\你的目录
  python tools/merge_ib_summary_breakthrough_cols.py --breakthrough 突破详情_买入信号.xlsx
  python tools/merge_ib_summary_breakthrough_cols.py --list-xlsx
  python tools/merge_ib_summary_breakthrough_cols.py --inplace

说明：默认目录为「项目根/history_data」（与 intelligentbuy 导出一致）。
      若文件在别处请用 --data-dir 指定绝对路径。
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parent.parent

DATE_COL_CANDS = ("选股日", "选择日期", "选股日期")
# 汇总表常只有「买入日」；突破表里两列都有，合并时两键都建索引
BUY_DAY_COL_CANDS = ("买入日", "buy_trade_day", "交易日期")
CODE_COL_CANDS = ("股票代码", "代码", "证券代码")
SIGNAL_COL_CANDS = ("买入信号", "买入")
TIME_COL_CANDS = ("突破时刻", "突破时间")
TARGET_COLS_OUT = ("突破量/均量(倍)", "委卖/委买比", "突破量/前tick被吃卖档合计(倍)")
# 突破详情表内第三列：新表头或旧表头
BT_BREAK_MUL_COL_CANDS = ("突破量/前tick被吃卖档合计(倍)", "突破量/前卖一量(倍)")
# 未指定 --pattern 时依次尝试的汇总表 glob
SUMMARY_GLOB_FALLBACKS = (
    "*选股收益汇总*智能买入*.xlsx",
    "*各日选股收益汇总*.xlsx",
    "*收益汇总*智能买入*.xlsx",
)


def _default_data_dir() -> Path:
    """与 intelligentbuy 导出一致：项目根下 history_data。"""
    return (ROOT / "history_data").resolve()


def _find_breakthrough_file(data_dir: Path, name: str) -> Optional[Path]:
    """精确名 -> 目录内「突破详情*买入信号*.xlsx」-> 「突破详情*.xlsx」。"""
    exact = (data_dir / name).resolve()
    if exact.is_file():
        return exact
    cands = sorted(data_dir.glob("突破详情*买入信号*.xlsx"))
    if cands:
        return cands[0]
    cands2 = sorted(data_dir.glob("突破详情*.xlsx"))
    return cands2[0] if cands2 else None


def _list_xlsx(data_dir: Path, limit: int = 80) -> List[str]:
    xs = sorted(data_dir.glob("*.xlsx"))
    return [p.name for p in xs[:limit]]


def _find_col(df, cands: Tuple[str, ...]) -> Optional[str]:
    cols = list(df.columns)
    col_set = {str(c).strip(): c for c in cols}
    for c in cands:
        if c in col_set:
            return col_set[c]
    for raw in cols:
        s = str(raw).strip().replace(" ", "")
        for c in cands:
            if s == c.replace(" ", ""):
                return raw
    return None


def _norm_code(v) -> str:
    import pandas as pd

    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    # Excel 数字 28 / 28.0 → 统一 6 位代码
    if "." in s:
        try:
            n = int(float(s))
            return str(n).zfill(6)
        except (ValueError, TypeError, OverflowError):
            pass
    s = s.replace(".", "").split(".")[0]
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 6:
        return digits[-6:].zfill(6) if len(digits) > 6 else digits.zfill(6)
    if digits:
        try:
            return str(int(digits)).zfill(6)
        except ValueError:
            return digits.zfill(6)[:6]
    try:
        n = int(float(s))
        return str(n).zfill(6)
    except (ValueError, TypeError, OverflowError):
        return (s[:6].zfill(6) if s[:6].isdigit() else s[:6])[:6]


def _norm_date(v) -> str:
    import pandas as pd

    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    ts = pd.to_datetime(v, errors="coerce")
    if not pd.isna(ts):
        return ts.strftime("%Y-%m-%d")
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    s = s.replace("/", "-")[:10]
    return s


def _cell_blank(v) -> bool:
    import pandas as pd

    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    s = str(v).strip().lower()
    return s in ("", "nan", "none", "nat")


def _read_excel_all_sheets(path: Path) -> Tuple[Dict[str, Any], List[str]]:
    """整表按字符串读入，避免股票代码被读成 float 丢前导零、键对不上。"""
    import pandas as pd

    xl = pd.ExcelFile(path)
    out = {
        name: pd.read_excel(path, sheet_name=name, dtype=str, keep_default_na=False)
        for name in xl.sheet_names
    }
    return out, list(xl.sheet_names)


def _xlsx_set_text_columns(path: Path, sheet_to_headers: Dict[str, List[str]]) -> None:
    """将指定表头列设为 Excel 文本格式 @，并把股票代码规范为 6 位字符串（避免显示成 28）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(path)
    for sheet_name, headers in sheet_to_headers.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        titles = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            titles.append(str(v).strip() if v is not None else "")
        for h in headers:
            if not h:
                continue
            try:
                ci = titles.index(h) + 1
            except ValueError:
                continue
            is_code_col = str(h).strip() in {str(x).strip() for x in CODE_COL_CANDS}
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                if is_code_col:
                    cell.value = _norm_code(cell.value)
                cell.number_format = "@"
    wb.save(path)


def _build_lookup(bt_df) -> Dict[Tuple[str, str], Tuple[str, str, str]]:
    """键：(日期, 六位代码)。日期同时建「选股日」「买入日」两套键（与汇总表对齐）。"""
    import pandas as pd

    bt_df = bt_df.rename(columns={c: str(c).strip() for c in bt_df.columns})
    miss = [c for c in TARGET_COLS_OUT[:2] if c not in bt_df.columns]
    c2_src = _find_col(bt_df, BT_BREAK_MUL_COL_CANDS)
    if miss:
        raise SystemExit(
            f"突破详情表缺少列 {miss}。当前列: {list(bt_df.columns)}"
        )
    if not c2_src:
        raise SystemExit(
            f"突破详情表缺少列 {BT_BREAK_MUL_COL_CANDS} 之一。当前列: {list(bt_df.columns)}"
        )

    sel_c = _find_col(bt_df, DATE_COL_CANDS)
    buy_c = _find_col(bt_df, BUY_DAY_COL_CANDS)
    code_c = _find_col(bt_df, CODE_COL_CANDS)
    if not code_c:
        raise SystemExit(
            f"突破详情表缺少代码列 {CODE_COL_CANDS}。\n当前列: {list(bt_df.columns)}"
        )
    if not sel_c and not buy_c:
        raise SystemExit(
            f"突破详情表需要「选股日」或「买入日」之一。\n当前列: {list(bt_df.columns)}"
        )

    sig_c = _find_col(bt_df, SIGNAL_COL_CANDS)
    time_c = _find_col(bt_df, TIME_COL_CANDS)

    work = bt_df.copy()
    work["_k_code"] = work[code_c].map(_norm_code).astype(str)
    work = work[work["_k_code"].str.len() == 6]

    if sig_c:
        s = work[sig_c].astype(str).str.strip()

        def _prio(x: str) -> int:
            if x in ("是", "True", "1", "YES", "yes", "Y"):
                return 0
            return 1

        work["_prio"] = s.map(_prio)
    else:
        work["_prio"] = 1

    if time_c:
        work["_t"] = work[time_c].astype(str)
    else:
        work["_t"] = ""

    work = work.sort_values(by=["_prio", "_t"], ascending=[True, True])
    c0, c1 = TARGET_COLS_OUT[0], TARGET_COLS_OUT[1]
    out: Dict[Tuple[str, str], Tuple[str, str, str]] = {}
    for _, r in work.iterrows():
        code = str(r["_k_code"])
        v1 = "" if _cell_blank(r[c0]) else str(r[c0]).strip()
        v2 = "" if _cell_blank(r[c1]) else str(r[c1]).strip()
        v3 = "" if _cell_blank(r[c2_src]) else str(r[c2_src]).strip()
        d_sel = _norm_date(r[sel_c]) if sel_c else ""
        d_buy = _norm_date(r[buy_c]) if buy_c else ""
        keys = set()
        if d_sel:
            keys.add((d_sel, code))
        if d_buy:
            keys.add((d_buy, code))
        for k in keys:
            if not k[0] or len(k[1]) != 6:
                continue
            if k not in out:
                out[k] = (v1, v2, v3)
    return out


def _process_one_summary(
    path_in: Path,
    lookup: Dict[Tuple[str, str], Tuple[str, str, str]],
    path_out: Path,
    sheet: str | int = 0,
) -> None:
    import pandas as pd

    all_sheets, sheet_names_list = _read_excel_all_sheets(path_in)
    if not sheet_names_list:
        raise SystemExit(f"{path_in.name}: 工作簿无工作表")
    if isinstance(sheet, int):
        if sheet < 0 or sheet >= len(sheet_names_list):
            raise SystemExit(
                f"{path_in.name}: 工作表序号 {sheet} 无效，共 {len(sheet_names_list)} 个表: {sheet_names_list}"
            )
        sheet_name = sheet_names_list[sheet]
    else:
        sheet_name = sheet
        if sheet_name not in sheet_names_list:
            raise SystemExit(
                f"{path_in.name}: 无工作表 {sheet_name!r}，可选: {sheet_names_list}"
            )

    for name in sheet_names_list:
        sdf = all_sheets[name]
        all_sheets[name] = sdf.rename(columns={c: str(c).strip() for c in sdf.columns})

    df = all_sheets[sheet_name].copy()

    sel_c = _find_col(df, DATE_COL_CANDS)
    buy_c = _find_col(df, BUY_DAY_COL_CANDS)
    code_c = _find_col(df, CODE_COL_CANDS)
    if not code_c or (not sel_c and not buy_c):
        raise SystemExit(
            f"{path_in.name}: 汇总表需要「股票代码/代码」及「选股日」或「买入日」之一。\n"
            f"当前列: {list(df.columns)}"
        )

    for c in list(TARGET_COLS_OUT):
        if c in df.columns:
            df = df.drop(columns=[c])
    _old_mul = "突破量/前卖一量(倍)"
    if _old_mul in df.columns and _old_mul != TARGET_COLS_OUT[2]:
        df = df.drop(columns=[_old_mul])

    def _row_merge_date(r: pd.Series) -> str:
        if sel_c:
            u = _norm_date(r[sel_c])
            if u:
                return u
        if buy_c:
            return _norm_date(r[buy_c])
        return ""

    dates = df.apply(_row_merge_date, axis=1)
    codes = df[code_c].map(_norm_code)
    v1, v2, v3 = [], [], []
    for d, co in zip(dates, codes):
        t = lookup.get((str(d), str(co)), ("", "", ""))
        v1.append(t[0])
        v2.append(t[1])
        v3.append(t[2])
    df[TARGET_COLS_OUT[0]] = v1
    df[TARGET_COLS_OUT[1]] = v2
    df[TARGET_COLS_OUT[2]] = v3
    df[code_c] = codes.astype(str)
    all_sheets[sheet_name] = df

    fmt_map: Dict[str, List[str]] = {}
    for name, sdf in all_sheets.items():
        cc = _find_col(sdf, CODE_COL_CANDS)
        if cc:
            fmt_map[name] = [cc]

    path_out.parent.mkdir(parents=True, exist_ok=True)
    same = path_out.resolve() == path_in.resolve()

    def _write(to: Path) -> None:
        with pd.ExcelWriter(to, engine="openpyxl") as writer:
            for name in sheet_names_list:
                all_sheets[name].to_excel(writer, sheet_name=name, index=False)
        if fmt_map:
            _xlsx_set_text_columns(to, fmt_map)

    if same:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path_out.parent))
        os.close(fd)
        tmp_path = Path(tmp)
        try:
            _write(tmp_path)
            os.replace(tmp_path, path_out)
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise
    else:
        _write(path_out)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="汇总 xlsx 按选股日+代码合并突破详情三列（临时工具）"
    )
    ap.add_argument(
        "--data-dir",
        default="",
        help="含汇总与突破详情 xlsx 的目录。默认：项目根/history_data",
    )
    ap.add_argument(
        "--breakthrough",
        default="突破详情_买入信号.xlsx",
        help="突破详情文件名（位于 data-dir）",
    )
    ap.add_argument(
        "--pattern",
        default="",
        help="汇总表 glob；留空则依次尝试多种文件名（见脚本内 SUMMARY_GLOB_FALLBACKS）",
    )
    ap.add_argument(
        "--suffix",
        default="_带突破指标",
        help="输出文件名在 .xlsx 前插入的后缀；空则加默认后缀",
    )
    ap.add_argument(
        "--inplace",
        action="store_true",
        help="直接覆盖原汇总文件（建议先备份）",
    )
    ap.add_argument("--sheet", default="0", help="工作表名或 0 起序号，默认 0")
    ap.add_argument(
        "--list-xlsx",
        action="store_true",
        help="仅列出目录内 .xlsx 文件名后退出（用于核对路径）",
    )
    args = ap.parse_args()

    data_dir = (
        Path(args.data_dir).expanduser().resolve()
        if (args.data_dir or "").strip()
        else _default_data_dir()
    )
    if not data_dir.is_dir():
        print(f"目录不存在: {data_dir}", file=sys.stderr)
        return 1

    if args.list_xlsx:
        names = _list_xlsx(data_dir, limit=200)
        print(f"目录: {data_dir}")
        print(f"共 {len(names)} 个 .xlsx（最多列 200 个）：")
        for n in names:
            print(f"  {n}")
        return 0

    bt_path = _find_breakthrough_file(data_dir, args.breakthrough)
    if bt_path is None or not bt_path.is_file():
        print(f"未找到突破详情文件（试过精确名「{args.breakthrough}」及 突破详情*.xlsx）。", file=sys.stderr)
        print(f"目录: {data_dir}", file=sys.stderr)
        xs = _list_xlsx(data_dir, limit=40)
        if xs:
            print("当前目录下部分 .xlsx：", file=sys.stderr)
            for n in xs:
                print(f"  {n}", file=sys.stderr)
        else:
            print("该目录下没有任何 .xlsx。", file=sys.stderr)
        return 1

    try:
        import pandas as pd
    except ImportError:
        print("需要安装: pip install pandas openpyxl", file=sys.stderr)
        return 1

    sheets_bt, sn_bt = _read_excel_all_sheets(bt_path)
    if not sn_bt:
        print(f"突破详情工作簿无表: {bt_path}", file=sys.stderr)
        return 1
    lookup: Dict[Tuple[str, str], Tuple[str, str, str]] = {}
    for sn in sn_bt:
        part = _build_lookup(sheets_bt[sn])
        for k, v in part.items():
            if k not in lookup:
                lookup[k] = v
    print(f"使用目录: {data_dir}")
    print(f"突破详情: {bt_path.name}，索引键数: {len(lookup)}")

    try:
        sheet_arg: str | int = int(args.sheet)
    except ValueError:
        sheet_arg = str(args.sheet)

    pats = [args.pattern.strip()] if (args.pattern or "").strip() else list(SUMMARY_GLOB_FALLBACKS)
    hits: List[Path] = []
    seen: set = set()
    for pat in pats:
        for p in sorted(data_dir.glob(pat)):
            if not p.is_file():
                continue
            if p.name.startswith("~$"):
                continue
            if p.resolve() == bt_path.resolve():
                continue
            if "_带突破指标" in p.stem and not args.inplace:
                continue
            rp = p.resolve()
            if rp in seen:
                continue
            seen.add(rp)
            hits.append(p)

    if not hits:
        print(f"未在 {data_dir} 下找到汇总 xlsx（已尝试 glob: {pats}）。", file=sys.stderr)
        xs = _list_xlsx(data_dir, limit=40)
        if xs:
            print("当前目录下部分 .xlsx：", file=sys.stderr)
            for n in xs:
                print(f"  {n}", file=sys.stderr)
        return 1

    suf = (args.suffix or "_带突破指标").strip()
    for path_in in hits:
        if args.inplace:
            path_out = path_in
        else:
            path_out = path_in.parent / f"{path_in.stem}{suf}{path_in.suffix}"
        _process_one_summary(path_in, lookup, path_out, sheet=sheet_arg)
        print(f"已写出: {path_out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

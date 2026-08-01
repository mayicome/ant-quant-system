#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""批量读取历史涨停 CSV，生成对应日期的封单结构 Excel。"""

from __future__ import annotations

import argparse
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

# 从 tools/ 子目录运行时，确保能 import 项目根目录下的模块
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import xtquant.xtdata as xtdata
except Exception:
    xtdata = None

from limit_up_structure_analysis_gui import (
    _apply_cross_sectional_score_adjustment,
    _calc_weighted_total_score,
    _calc_close_order_amount_yi,
    _calc_day_turnover_yuan,
    _calc_float_market_cap_yuan,
    _calc_order_trend_and_score,
    _calc_order_volatility_coeff,
    _classify_order_stability,
    _confidence_tag,
    _extract_tick_preview_rows,
    _label_order_rating,
    _load_or_fetch_tick_df,
    _load_limitup_codes,
    _score_close_order_amount,
    _score_rush_intensity,
    _score_seal_hardness,
)


def repo_root() -> str:
    return _ROOT


def history_dir() -> str:
    return os.path.join(repo_root(), "history_data")


def _st_name(name: str) -> bool:
    n = (name or "").strip()
    trans = str.maketrans({"Ｓ": "S", "Ｔ": "T", "＊": "*", "－": "-", "％": "%"})
    n2 = n.translate(trans).upper()
    return ("ST" in n2) or ("*ST" in n2) or n2.startswith("ST")


def list_limitup_csv_files() -> List[Tuple[str, str]]:
    """返回 [(date_yyyymmdd, csv_path)]，按日期升序。"""
    hd = history_dir()
    cands: List[Tuple[str, str]] = []
    if not os.path.isdir(hd):
        return cands
    for name in os.listdir(hd):
        if not (name.startswith("涨停板数据_") and name.lower().endswith(".csv")):
            continue
        m = re.search(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})", name)
        if not m:
            continue
        d = f"{m.group(1)}{m.group(2)}{m.group(3)}"
        cands.append((d, os.path.join(hd, name)))
    cands.sort(key=lambda x: x[0])
    return cands


def out_xlsx_path(date_yyyymmdd: str) -> str:
    return os.path.join(history_dir(), f"封单结构_{date_yyyymmdd}.xlsx")


def _excel_stock_cell_str_for_export(v: str) -> str:
    """6 位数字代码前加零宽空格，减轻 Excel/WPS 智能格式提醒。"""
    s = str(v or "").strip()
    return ("\u200b" + s) if (len(s) == 6 and s.isdigit()) else s


def _save_df_with_code_text_format(df: pd.DataFrame, out_path: str) -> None:
    """保存 Excel，并对代码列应用文本格式 + 忽略数字文本告警。"""
    x = df.copy()
    if "code" in x.columns:
        x["code"] = x["code"].map(_excel_stock_cell_str_for_export)
    x.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    code_col_1based: Optional[int] = None
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=c).value or "").strip()
        if header in {"code", "股票代码"}:
            code_col_1based = c
            break

    if code_col_1based is not None:
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=code_col_1based)
            cell.number_format = "@"
            qp = getattr(cell, "quote_prefix", None)
            if qp is not None:
                try:
                    cell.quote_prefix = True
                except Exception:
                    pass
        try:
            from openpyxl.worksheet.cell_range import CellRange
            from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors

            letter = get_column_letter(code_col_1based)
            rng = f"{letter}1:{letter}{ws.max_row}"
            ws.ignored_errors = IgnoredErrors(
                ignoredError=(IgnoredError(sqref=CellRange(rng), numberStoredAsText=True),)
            )
        except Exception:
            pass

    wb.save(out_path)


def build_one_day(date_yyyymmdd: str, csv_path: str) -> Tuple[pd.DataFrame, int]:
    """
    返回 (结果表, tick 有数据的股票数)。
    tick 为空时各项得分为 0，封单评级会全部落在「🔴 虚封高危」（与主界面一致），
    并非 CSV 与输出文件名日期不一致。
    """
    if xtdata is None:
        raise RuntimeError("xtdata 不可用，请先确保 QMT/xtquant 环境正常。")

    try:
        xtdata.enable_hello = False
    except Exception:
        pass

    code_to_name: Dict[str, str] = _load_limitup_codes(csv_path)
    rows: List[dict] = []
    tick_ok_count = 0

    for code, name in sorted(code_to_name.items(), key=lambda x: x[0]):
        if _st_name(name):
            continue
        try:
            tick_df = _load_or_fetch_tick_df(code, date_yyyymmdd, save_cache=True)
        except Exception:
            tick_df = None

        if tick_df is not None and not getattr(tick_df, "empty", True):
            tick_ok_count += 1

        # 保持与主界面同口径：无 tick 则留空并继续
        _headers, tick_rows = _extract_tick_preview_rows(tick_df, limit_n=5)
        _ = tick_rows  # 仅触发字段兼容逻辑，不写入结果

        close_amt_yi = _calc_close_order_amount_yi(tick_df)
        day_turnover_yuan = _calc_day_turnover_yuan(tick_df)
        float_cap_yuan = _calc_float_market_cap_yuan(code, tick_df)

        seal_ratio = None
        seal_hardness = ""
        if close_amt_yi is not None and day_turnover_yuan is not None and day_turnover_yuan > 0:
            seal_ratio = (close_amt_yi * 1e8) / day_turnover_yuan
            seal_hardness = f"{seal_ratio * 100.0:.2f}%"

        rush_ratio = None
        rush_intensity = ""
        if close_amt_yi is not None and float_cap_yuan is not None and float_cap_yuan > 0:
            rush_ratio = (close_amt_yi * 1e8) / float_cap_yuan
            rush_intensity = f"{rush_ratio * 100.0:.2f}%"

        vol_coeff = _calc_order_volatility_coeff(tick_df)
        stability_label, stability_score = _classify_order_stability(vol_coeff) if vol_coeff is not None else ("", None)
        trend_label, trend_score = _calc_order_trend_and_score(tick_df)

        score_hard = _score_seal_hardness(seal_ratio)
        score_rush = _score_rush_intensity(rush_ratio)
        score_amt = _score_close_order_amount(close_amt_yi)
        score_stability = int(stability_score or 0)
        score_trend = int(trend_score or 0)
        total_score = _calc_weighted_total_score(
            score_hard,
            score_rush,
            score_amt,
            score_stability,
            score_trend,
        )
        rows.append(
            {
                "code": code,
                "name": name or "",
                "close_order_amount_yi": "" if close_amt_yi is None else f"{close_amt_yi:.4f}",
                "seal_hardness": seal_hardness,
                "rush_intensity": rush_intensity,
                "order_stability": stability_label,
                "order_trend": trend_label,
                "order_rating": "",
                "confidence_tag": "",
                "_close_amt_yi": close_amt_yi,
                "_seal_ratio": seal_ratio,
                "_rush_ratio": rush_ratio,
                "_stability_score": score_stability,
                "_trend_score": score_trend,
                "_base_total_score": total_score,
                "_total_score": total_score,
            }
        )

    _apply_cross_sectional_score_adjustment(rows)
    for r in rows:
        r.pop("_close_amt_yi", None)
        r.pop("_seal_ratio", None)
        r.pop("_rush_ratio", None)
        r.pop("_stability_score", None)
        r.pop("_trend_score", None)
        r.pop("_base_total_score", None)
        r.pop("_total_score", None)
    return pd.DataFrame(rows), tick_ok_count


def main() -> int:
    parser = argparse.ArgumentParser(description="批量生成历史封单结构文件")
    parser.add_argument("--start", help="起始日期 YYYYMMDD（可选）")
    parser.add_argument("--end", help="结束日期 YYYYMMDD（可选）")
    parser.add_argument("--overwrite", action="store_true", help="覆盖已存在的封单结构文件")
    args = parser.parse_args()

    files = list_limitup_csv_files()
    if not files:
        print("未找到历史涨停 CSV：history_data/涨停板数据_*.csv")
        return 1

    start = args.start.strip() if args.start else ""
    end = args.end.strip() if args.end else ""
    if start:
        files = [x for x in files if x[0] >= start]
    if end:
        files = [x for x in files if x[0] <= end]
    if not files:
        print("筛选后无可处理日期。")
        return 1

    ok = 0
    skip = 0
    fail = 0
    for d, csv_path in files:
        out_path = out_xlsx_path(d)
        if (not args.overwrite) and os.path.exists(out_path):
            print(f"[跳过] {d} -> 已存在 {os.path.basename(out_path)}")
            skip += 1
            continue
        try:
            df, tick_n = build_one_day(d, csv_path)
            _save_df_with_code_text_format(df, out_path)
            n = len(df)
            line = f"[完成] {d} -> {os.path.basename(out_path)} rows={n} tick有数据={tick_n}/{n}"
            if tick_n == 0 and n > 0:
                line += (
                    "  [警告] 未拉到任何 tick，评级将全为「虚封高危」口径；"
                    "请确认 QMT 已连网/已下载该日 tick，而非 CSV 与文件名日期不一致。"
                )
            print(line)
            ok += 1
        except Exception as e:
            print(f"[失败] {d} -> {e}")
            fail += 1

    print(f"\n汇总: 完成={ok} 跳过={skip} 失败={fail}")
    return 0 if fail == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())

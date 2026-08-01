# -*- coding: utf-8 -*-
"""
从「主力净流入统计_YYYYMMDD.xlsx」提取近 N 日累计占比前5，生成策略生成器可选股文件。

默认：选股日 ≥ 2026-07-16，N ∈ {2,3,4,5}。
输出：每个 N 一份 xlsx/csv，另有「全部」合并表（含列「近N日」）。
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
ARCH = ROOT / "history_data" / "存档"
HIST = ROOT / "history_data"
OUT_DIR = ROOT / "history_data" / "主力净流入回测"
DATE_RE = re.compile(r"主力净流入统计_(\d{8})\.xlsx$", re.I)
TYPE_RE = re.compile(r"^近(\d+)日累计占比前5$")
ZWSP = "\u200b"


def _norm_code(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip().replace(ZWSP, "").replace("\xa0", "")
    if "." in s:
        s = s.split(".", 1)[0]
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits:
        return digits.zfill(6)[-6:]
    return s


def _parse_date(path: Path) -> str | None:
    m = DATE_RE.search(path.name)
    if not m:
        return None
    s = m.group(1)
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def _collect_files(dirs: list[Path], since: str) -> list[Path]:
    since_ymd = since.replace("-", "")
    found: dict[str, Path] = {}
    for d in dirs:
        if not d.is_dir():
            continue
        for fp in d.glob("主力净流入统计_*.xlsx"):
            m = DATE_RE.search(fp.name)
            if not m:
                continue
            ymd = m.group(1)
            if ymd < since_ymd:
                continue
            # 现行目录优先于存档
            if ymd not in found or d == HIST:
                found[ymd] = fp
    return [found[k] for k in sorted(found)]


def _write_excel(df: pd.DataFrame, path: Path, sheet: str = "选股") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        from openpyxl.styles.numbers import FORMAT_TEXT

        header = {cell.value: cell.column for cell in ws[1]}
        for key in ("股票代码", "代码"):
            if key not in header:
                continue
            col_idx = header[key]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = FORMAT_TEXT
                if cell.value is not None:
                    cell.value = _norm_code(cell.value)


def _prep_block(raw: pd.DataFrame, sel: str, typ: str, n_tag) -> pd.DataFrame | None:
    sub = raw[raw["统计类型"].astype(str).str.strip() == typ].copy()
    if sub.empty:
        return None
    sub.insert(0, "选股日", sel)
    if n_tag is not None:
        sub.insert(1, "近N日", n_tag)
    sub["股票代码"] = sub["股票代码"].map(_norm_code)
    if "股票名称" not in sub.columns:
        sub["股票名称"] = ""
    else:
        sub["股票名称"] = sub["股票名称"].astype(str).where(sub["股票名称"].notna(), "")
    sub["代码"] = sub["股票代码"]
    insert_at = 2 if n_tag is not None else 1
    sub.insert(insert_at, "名次", range(1, len(sub) + 1))
    return sub


def _finalize_cols(out: pd.DataFrame, with_n: bool) -> pd.DataFrame:
    front = ["选股日", "近N日", "名次", "股票代码", "股票名称", "统计类型", "累计占流通%"]
    if not with_n:
        front = ["选股日", "名次", "股票代码", "股票名称", "统计类型", "累计占流通%"]
    rest = [c for c in out.columns if c not in front and c != "代码"]
    cols = front + rest + (["代码"] if "代码" in out.columns else [])
    return out.reindex(columns=[c for c in cols if c in out.columns])


def build(since: str, ns: list[int], out_dir: Path) -> Path:
    files = _collect_files([HIST, ARCH], since)
    if not files:
        raise SystemExit(f"未找到 ≥ {since} 的 主力净流入统计_*.xlsx")

    want_types = {f"近{n}日累计占比前5": n for n in ns}
    frames: list[pd.DataFrame] = []
    for fp in files:
        sel = _parse_date(fp)
        if not sel:
            continue
        raw = pd.read_excel(fp)
        if "统计类型" not in raw.columns:
            print(f"跳过无「统计类型」: {fp.name}")
            continue
        for typ, n in want_types.items():
            sub = _prep_block(raw, sel, typ, n)
            if sub is None:
                continue
            frames.append(sub)
            print(f"{sel} N={n}: {len(sub)} 条 ← {fp.name}")

    if not frames:
        raise SystemExit("没有提取到近N日累计占比前5 行")

    out = _finalize_cols(pd.concat(frames, ignore_index=True), with_n=True)
    dates = sorted(out["选股日"].astype(str).unique())
    d0, d1 = dates[0].replace("-", ""), dates[-1].replace("-", "")
    out_dir.mkdir(parents=True, exist_ok=True)

    all_path = out_dir / f"选股_近N日累计占比前5_N{''.join(map(str, ns))}_{d0}_{d1}.xlsx"
    _write_excel(out, all_path, "全部")
    out.to_csv(all_path.with_suffix(".csv"), index=False, encoding="utf-8-sig")

    for n in ns:
        part = out[out["近N日"] == n].copy()
        if part.empty:
            continue
        p = out_dir / f"选股_近{n}日累计占比前5_{d0}_{d1}.xlsx"
        _write_excel(part, p, "选股")
        part.to_csv(p.with_suffix(".csv"), index=False, encoding="utf-8-sig")
        print(f"N={n}: {part['选股日'].nunique()} 日 × {len(part)} 行 → {p.name}")

    print(f"\n合并: {len(dates)} 日, {len(out)} 行 → {all_path}")
    return all_path


def build_today_top5(since: str, out_dir: Path) -> Path:
    """提取「今日占比前5」→ 选股文件。"""
    files = _collect_files([HIST, ARCH], since)
    if not files:
        raise SystemExit(f"未找到 ≥ {since} 的 主力净流入统计_*.xlsx")

    typ = "今日占比前5"
    frames: list[pd.DataFrame] = []
    for fp in files:
        sel = _parse_date(fp)
        if not sel:
            continue
        raw = pd.read_excel(fp)
        if "统计类型" not in raw.columns:
            print(f"跳过无「统计类型」: {fp.name}")
            continue
        sub = _prep_block(raw, sel, typ, None)
        if sub is None:
            print(f"{sel}: 无「{typ}」← {fp.name}")
            continue
        frames.append(sub)
        print(f"{sel} 今日占比前5: {len(sub)} 条 ← {fp.name}")

    if not frames:
        raise SystemExit("没有提取到今日占比前5 行")

    out = _finalize_cols(pd.concat(frames, ignore_index=True), with_n=False)
    dates = sorted(out["选股日"].astype(str).unique())
    d0, d1 = dates[0].replace("-", ""), dates[-1].replace("-", "")
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"选股_今日占比前5_{d0}_{d1}.xlsx"
    _write_excel(out, path, "选股")
    out.to_csv(path.with_suffix(".csv"), index=False, encoding="utf-8-sig")
    print(f"\n今日占比前5: {len(dates)} 日, {len(out)} 行 → {path}")
    return path


def main():
    ap = argparse.ArgumentParser(description="主力净流入统计 → 选股文件")
    ap.add_argument("--since", default="2026-07-16")
    ap.add_argument("--n", default="2,3,4,5", help="逗号分隔；空则不生成近N日")
    ap.add_argument("--today", action="store_true", help="生成今日占比前5")
    ap.add_argument("--today-only", action="store_true", help="只生成今日占比前5")
    ap.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = ap.parse_args()
    if args.today_only or args.today:
        build_today_top5(args.since, args.out_dir)
    if not args.today_only:
        ns = [int(x.strip()) for x in str(args.n).split(",") if x.strip()]
        if ns:
            build(args.since, ns, args.out_dir)


if __name__ == "__main__":
    main()

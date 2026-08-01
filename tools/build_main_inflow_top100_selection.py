# -*- coding: utf-8 -*-
"""
将 history_data/个股主力净流入 下每日 CSV 整理为策略生成器可批量回测的选股文件。

规则：
- 按「净流入占流通%」降序取每日前 TOP_N（默认 100）
- 旧版双行表头 CSV 归一到与近期一致的列名
- 保留归一后全部数据列，并增加「选股日」「股票代码」「股票名称」
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SRC = ROOT / "history_data" / "个股主力净流入"
DATE_RE = re.compile(r"个股主力净流入_(\d{8})\.csv$", re.I)
RANK_COL = "净流入占流通%"

# 与 2026-07-17 起导出格式一致
CANONICAL_COLS = [
    "序号",
    "代码",
    "名称",
    "最新价",
    "今日涨跌幅",
    "今日主力净流入-净额",
    "流通市值",
    "净流入占流通%",
    "今日主力净流入-净占比",
    "今日超大单净流入-净额",
    "今日超大单净流入-净占比",
    "今日大单净流入-净额",
    "今日大单净流入-净占比",
    "今日中单净流入-净额",
    "今日中单净流入-净占比",
    "今日小单净流入-净额",
    "今日小单净流入-净占比",
]

# 东财旧版「相关 + 换行表头」解析后的错位列名 → 标准列
OLD_FORMAT_RENAME = {
    "今日\n涨跌幅": "今日涨跌幅",
    "今日主力净流入": "今日主力净流入-净额",
    "今日超大单净流入": "今日主力净流入-净占比",
    "今日大单净流入": "今日超大单净流入-净额",
    "今日中单净流入": "今日超大单净流入-净占比",
    "今日小单净流入": "今日大单净流入-净额",
    "净额": "今日大单净流入-净占比",
    "净占比": "今日中单净流入-净额",
    "净额.1": "今日中单净流入-净占比",
    "净占比.1": "今日小单净流入-净额",
    "净额.2": "今日小单净流入-净占比",
}


def _parse_date_from_name(path: Path) -> str | None:
    m = DATE_RE.search(path.name)
    if not m:
        return None
    s = m.group(1)
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def _norm_code(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if "." in s:
        s = s.split(".", 1)[0]
    if s.isdigit():
        return s.zfill(6)
    return s


def _normalize_frame(df: pd.DataFrame) -> pd.DataFrame:
    """统一旧/新两种导出表头。"""
    work = df.copy()
    # 去掉列名首尾空白
    work.columns = [str(c).strip() for c in work.columns]

    if "相关" in work.columns or "今日\n涨跌幅" in work.columns or "净额.2" in work.columns:
        work = work.rename(columns=OLD_FORMAT_RENAME)
        if "相关" in work.columns:
            work = work.drop(columns=["相关"])

    # 已是新格式则原样；缺列补空，多列丢弃无关
    for c in CANONICAL_COLS:
        if c not in work.columns:
            work[c] = pd.NA
    return work[CANONICAL_COLS].copy()


def build(src_dir: Path, top_n: int, out_path: Path | None) -> Path:
    files = sorted(src_dir.glob("个股主力净流入_*.csv"))
    if not files:
        raise SystemExit(f"未找到 CSV：{src_dir}")

    frames: list[pd.DataFrame] = []
    for fp in files:
        sel = _parse_date_from_name(fp)
        if not sel:
            print(f"跳过无法解析日期的文件: {fp.name}", file=sys.stderr)
            continue
        raw = pd.read_csv(fp, encoding="utf-8-sig")
        df = _normalize_frame(raw)
        if RANK_COL not in df.columns:
            raise SystemExit(f"{fp.name} 缺少列「{RANK_COL}」")

        work = df.copy()
        work[RANK_COL] = pd.to_numeric(work[RANK_COL], errors="coerce")
        work = work.dropna(subset=[RANK_COL])
        work = work.sort_values(RANK_COL, ascending=False, kind="mergesort")
        work = work.head(int(top_n)).copy()

        work.insert(0, "选股日", sel)
        work.insert(1, "股票代码", work["代码"].map(_norm_code))
        work.insert(2, "股票名称", work["名称"].astype(str).where(work["名称"].notna(), ""))
        work["代码"] = work["股票代码"]
        frames.append(work)
        print(f"{sel}: {len(work)} 条 (源 {len(raw)})")

    if not frames:
        raise SystemExit("没有可合并的数据")

    out = pd.concat(frames, ignore_index=True)
    cols = ["选股日", "股票代码", "股票名称"] + CANONICAL_COLS
    out = out.reindex(columns=cols)

    dates = sorted(out["选股日"].astype(str).unique())
    if out_path is None:
        out_path = src_dir / (
            f"选股_主力净流入占流通前{top_n}_{dates[0].replace('-', '')}_{dates[-1].replace('-', '')}.xlsx"
        )
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="选股")
        ws = writer.sheets["选股"]
        from openpyxl.styles.numbers import FORMAT_TEXT

        header = {cell.value: cell.column for cell in ws[1]}
        for key in ("股票代码", "代码"):
            if key not in header:
                continue
            col_idx = header[key]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = FORMAT_TEXT
                if cell.value is not None:
                    cell.value = _norm_code(cell.value)

    csv_path = out_path.with_suffix(".csv")
    out.to_csv(csv_path, index=False, encoding="utf-8-sig")

    print(
        f"完成: {len(dates)} 个选股日, {len(out)} 行 →\n  {out_path}\n  {csv_path}"
    )
    return out_path


def main():
    ap = argparse.ArgumentParser(description="主力净流入占流通 TopN → 策略生成器选股文件")
    ap.add_argument("--src", type=Path, default=DEFAULT_SRC, help="每日 CSV 目录")
    ap.add_argument("--top", type=int, default=100, help="每日取前 N 名")
    ap.add_argument("--out", type=Path, default=None, help="输出 xlsx 路径（可选）")
    args = ap.parse_args()
    build(args.src, args.top, args.out)


if __name__ == "__main__":
    main()

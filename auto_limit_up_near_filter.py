#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
自动接近涨停股票筛选程序
程序启动后自动开始运行，搜索完成后自动保存，保存后30秒倒计时关闭
"""

import sys
import os
import pandas as pd
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Tuple
import time
import logging
import warnings
import json
import csv
from utils.trading_day import is_tradeday

# 抑制Qt和log4cplus的警告/错误信息（必须在导入任何Qt或xtquant模块之前设置）
# 设置环境变量抑制Qt Windows版本警告
os.environ.setdefault('QT_LOGGING_RULES', '*.debug=false')

# 抑制Python警告
warnings.filterwarnings("ignore")

# 重定向log4cplus错误（来自第三方库xtquant/QMT）
# 保存原始的stderr
_original_stderr = sys.stderr

class FilteredStderr:
    """过滤stderr，隐藏log4cplus错误和Qt警告"""
    def __init__(self, original_stderr):
        self.original_stderr = original_stderr
    
    def write(self, text):
        # 过滤log4cplus错误信息
        if 'log4cplus' in text.lower() or 'adsyncnamespace' in text.lower():
            return
        # 过滤Qt Windows版本警告
        if 'Qt: Untested Windows version' in text:
            return
        # 其他信息正常输出
        self.original_stderr.write(text)
    
    def flush(self):
        self.original_stderr.flush()
    
    def __getattr__(self, name):
        return getattr(self.original_stderr, name)

# 应用stderr过滤器（必须在导入任何可能产生警告的模块之前）
sys.stderr = FilteredStderr(_original_stderr)

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QTableWidget, QTableWidgetItem, QLabel, QMessageBox,
                             QProgressBar, QHeaderView, QApplication)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont

try:
    import xtquant.xtdata as xtdata
    from utils.trading_day import is_tradeday
    from utils.stock_info_manager import get_stock_name
except ImportError as e:
    print(f"导入模块失败: {e}")

# 配置日志
_LOGS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(_LOGS_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_LOGS_DIR, 'auto_limit_up_near_filter.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 导入主力净流入添加工具
try:
    from auto_add_inflow import InflowAdder
    logger.info("成功导入主力净流入工具")
except ImportError as e:
    logger.warning(f"无法导入主力净流入工具: {e}")
    InflowAdder = None
except Exception as e:
    logger.error(f"导入主力净流入工具时出错: {e}", exc_info=True)
    InflowAdder = None


def save_excel_with_text_code(excel_file_path: str, df: pd.DataFrame):
    """保存Excel文件，确保股票代码列是文本格式
    
    Args:
        excel_file_path: Excel文件路径
        df: 要保存的DataFrame
    """
    try:
        # 确保code列是字符串类型，并补零
        if 'code' in df.columns:
            def clean_code(code):
                code_str = str(code).strip()
                if '.' in code_str:
                    code_str = code_str.split('.')[0]
                # 确保是6位数字（补零）
                if code_str.isdigit():
                    return code_str.zfill(6)
                return code_str
            df['code'] = df['code'].apply(clean_code)
            # 确保列类型是字符串
            df['code'] = df['code'].astype(str)
        
        # 使用ExcelWriter和openpyxl引擎，在写入前就设置格式
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # 找到code列的索引
            code_col_idx = None
            if 'code' in df.columns:
                for idx, col_name in enumerate(df.columns, start=1):
                    if col_name == 'code':
                        code_col_idx = idx
                        break
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    
                    # 如果是code列，设置为文本格式
                    if code_col_idx and c_idx == code_col_idx and r_idx > 1:  # 跳过表头
                        cell.number_format = '@'  # @表示文本格式
                        # 确保值是字符串，并补零
                        if value is not None:
                            value_str = str(value).strip()
                            if '.' in value_str:
                                value_str = value_str.split('.')[0]
                            if value_str.isdigit():
                                value_str = value_str.zfill(6)
                            cell.value = value_str
                        else:
                            cell.value = ''
                    else:
                        cell.value = value
            
            wb.save(excel_file_path)
            wb.close()
        except Exception as e:
            # 如果openpyxl方法失败，回退到原来的方法
            logger.warning(f"使用openpyxl直接写入失败: {str(e)}，尝试回退方法")
            # 保存Excel
            df.to_excel(excel_file_path, index=False, engine='openpyxl')
            
            # 使用openpyxl设置code列为文本格式
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path)
                ws = wb.active
                
                # 找到code列的索引
                if 'code' in df.columns:
                    for col_idx, header in enumerate(df.columns, start=1):
                        if header == 'code':
                            # 将该列的所有单元格设置为文本格式
                            for row_idx in range(2, len(df) + 2):  # 从第2行开始（第1行是标题）
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.number_format = '@'  # @表示文本格式
                                # 确保值是字符串，并补零
                                if cell.value is not None:
                                    value_str = str(cell.value).strip()
                                    if '.' in value_str:
                                        value_str = value_str.split('.')[0]
                                    if value_str.isdigit():
                                        value_str = value_str.zfill(6)
                                    cell.value = value_str
                            break
                
                wb.save(excel_file_path)
                wb.close()
            except Exception as e2:
                logger.warning(f"设置Excel文本格式时出错: {str(e2)}，但文件已保存")
            
    except Exception as e:
        raise Exception(f"保存Excel文件失败: {str(e)}")


def add_stock_info_columns(file_path: str) -> bool:
    """为CSV文件添加概念、行业和板块列
    
    Args:
        file_path: CSV文件路径
        
    Returns:
        bool: 是否成功
    """
    try:
        # 加载股票信息JSON文件
        current_dir = os.path.dirname(__file__)
        json_path = os.path.join(current_dir, 'data', 'all_a_stock_info.json')
        if not os.path.exists(json_path):
            logger.warning(f"股票信息文件不存在: {json_path}")
            return False
        
        logger.info(f"正在加载股票信息文件: {json_path}")
        with open(json_path, 'r', encoding='utf-8') as f:
            stock_info_dict = json.load(f)
        logger.info(f"成功加载 {len(stock_info_dict)} 只股票的信息")
        
        # 检测文件编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']
        encoding = None
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    f.readline()
                encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if encoding is None:
            logger.error("无法检测文件编码")
            return False
        
        logger.info(f"检测到文件编码: {encoding}")
        
        # 读取CSV文件
        rows = []
        fieldnames = None
        
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames) if reader.fieldnames else []
            
            for row in reader:
                rows.append(row)
        
        if not rows:
            logger.warning("文件中没有数据行")
            return False
        
        logger.info(f"读取到 {len(rows)} 行数据")
        
        # 添加新列（如果不存在）
        new_columns = ['概念', '行业', '板块']
        for col in new_columns:
            if col not in fieldnames:
                fieldnames.append(col)
                logger.info(f"添加新列: {col}")
        
        def normalize_code_for_lookup(raw):
            """统一为6位字符串，便于与 all_a_stock_info.json 的键匹配（JSON 键可能是 000001 或 1）"""
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                return ''
            s = str(raw).strip()
            if '.' in s and s.split('.')[0].isdigit():
                s = s.split('.')[0]  # 去掉 .0
            if s.replace('.', '').isdigit():
                s = s.split('.')[0].zfill(6)
            return s
        
        # 填充数据
        filled_count = 0
        empty_count = 0
        
        for row in rows:
            code_raw = row.get('code', '')
            code = normalize_code_for_lookup(code_raw)
            
            if code:
                # 从JSON中获取股票信息（键可能是 000001 或 1，先试6位再试数字形式）
                stock_info = stock_info_dict.get(code, {})
                if not stock_info and code.lstrip('0'):
                    stock_info = stock_info_dict.get(code.lstrip('0') or '0', {})
                if not stock_info and code.isdigit():
                    stock_info = stock_info_dict.get(str(int(code)), {})
                
                # 概念：列表转字符串（用分号分隔）
                concepts = stock_info.get('concepts', [])
                if concepts and isinstance(concepts, list):
                    # 确保所有元素都是字符串
                    row['概念'] = ';'.join(str(c) for c in concepts if c)
                else:
                    row['概念'] = ''
                
                # 行业：直接使用字符串
                industry = stock_info.get('industry', '')
                row['行业'] = str(industry) if industry else ''
                
                # 板块：列表转字符串（用分号分隔）
                plates = stock_info.get('plates', [])
                if plates and isinstance(plates, list):
                    # 确保所有元素都是字符串
                    row['板块'] = ';'.join(str(p) for p in plates if p)
                else:
                    row['板块'] = ''
                
                if concepts or stock_info.get('industry') or plates:
                    filled_count += 1
                else:
                    empty_count += 1
            else:
                row['概念'] = ''
                row['行业'] = ''
                row['板块'] = ''
                empty_count += 1
        
        logger.info(f"填充完成：成功 {filled_count} 条，未找到 {empty_count} 条")
        
        # 保存回原文件
        logger.info(f"正在保存文件...")
        with open(file_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        logger.info(f"完成！文件已保存: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"添加股票信息列时出错: {str(e)}", exc_info=True)
        return False


class LimitUpNearFilterThread(QThread):
    """接近涨停筛选线程"""
    
    # 信号：更新进度
    progress_updated = pyqtSignal(int, int, str)  # (当前数量, 总数, 当前股票代码)
    # 信号：找到符合条件的股票
    stock_found = pyqtSignal(str, str, str, str)  # (股票代码, 股票名称, 接近涨停日期, 收盘价)
    # 信号：筛选完成
    finished = pyqtSignal(int)  # (找到的股票数量)
    # 信号：错误信息
    error_occurred = pyqtSignal(str)  # (错误信息)
    # 信号：调试信息（用于显示统计）
    debug_info = pyqtSignal(str)  # (调试信息)
    
    def __init__(self, stock_list: List[Tuple[str, str]], trading_dates: List[date], parent=None):
        super().__init__(parent)
        self.stock_list = stock_list
        self.trading_dates = trading_dates
        self.is_running = True
        
    def stop(self):
        """停止筛选"""
        self.is_running = False
        
    def _get_limit_ratio(self, stock_code: str, as_of_date=None) -> float:
        """获取股票的涨停幅度"""
        try:
            from utils.limit_ratio import get_limit_ratio

            stock_name = get_stock_name(stock_code) or ""
            return get_limit_ratio(stock_code, stock_name, as_of_date)
        except Exception:
            return 0.10
    
    def _get_full_stock_code(self, stock_code: str) -> str:
        """获取完整的股票代码（带市场后缀）"""
        if '.' in stock_code:
            return stock_code
        
        if stock_code.startswith(('0', '1', '3')):
            return f"{stock_code}.SZ"
        elif stock_code.startswith('6'):
            return f"{stock_code}.SH"
        elif stock_code.startswith(('8', '4', '920')):
            return f"{stock_code}.BJ"
        else:
            return stock_code
    
    def _calculate_technical_indicators(self, daily_data: pd.DataFrame) -> pd.DataFrame:
        """计算技术指标：布林线上轨
        
        注意：计算布林线需要至少20个交易日的数据。
        在 _get_daily_data 中，我们获取了60天的数据以确保有足够的历史数据来计算布林线。
        """
        try:
            # 确保数据按日期排序
            daily_data = daily_data.sort_values('date').copy()
            
            # 计算布林线
            # 中轨：MA20（20日移动平均）
            # 使用 min_periods=1 确保即使数据不足20天也能计算（但精度会降低）
            daily_data['MA20'] = daily_data['close'].rolling(window=20, min_periods=1).mean()
            # 标准差
            daily_data['STD20'] = daily_data['close'].rolling(window=20, min_periods=1).std()
            # 上轨 = 中轨 + 2 * 标准差
            daily_data['BOLL_UPPER'] = daily_data['MA20'] + 2.0 * daily_data['STD20']
            
            return daily_data
            
        except Exception as e:
            logger.error(f"计算技术指标失败: {str(e)}", exc_info=True)
            return daily_data
    
    def _get_daily_data(self, stock_code: str, include_prev_day: bool = True) -> Optional[pd.DataFrame]:
        """获取日线数据
        
        Args:
            stock_code: 股票代码
            include_prev_day: 是否包含前一个交易日（用于计算涨停价）
                注意：当 include_prev_day=True 时，返回所有获取的数据（约60天），
                用于计算布林线等技术指标；当 include_prev_day=False 时，只返回最近10个交易日的数据。
        """
        try:
            full_stock_code = self._get_full_stock_code(stock_code)
            
            # 计算日期范围（往前推更多天数以确保获取到足够的交易日和前一个交易日，以及计算布林线需要的数据）
            # 布林线需要至少20个交易日的数据，所以需要获取足够的历史数据
            end_date = self.trading_dates[-1] if self.trading_dates else date.today()
            start_date = end_date - timedelta(days=60)  # 往前推60天，确保包含前一个交易日和足够的数据计算布林线（约40-45个交易日）
            
            startdate = start_date.strftime("%Y%m%d") + "000000"
            enddate = end_date.strftime("%Y%m%d") + "235959"
            
            # 下载历史数据
            try:
                xtdata.download_history_data(full_stock_code, '1d', startdate, enddate)
                time.sleep(0.05)  # 短暂等待
            except Exception as e:
                logger.error(f"[{stock_code}] 下载历史数据失败: {str(e)}", exc_info=True)
                return None
            
            # 获取历史行情数据
            try:
                df = xtdata.get_market_data_ex([], [full_stock_code], period='1d',
                                             start_time=startdate,
                                             end_time=enddate,
                                             count=-1)
            except Exception as e:
                logger.error(f"[{stock_code}] 获取市场数据失败: {str(e)}", exc_info=True)
                return None
            
            if df is None:
                return None
            
            if full_stock_code not in df:
                return None
            
            if len(df[full_stock_code]) == 0:
                return None
            
            # 转换为DataFrame
            stock_data = df[full_stock_code]
            
            daily_data = pd.DataFrame({
                'time': stock_data['time'],
                'open': stock_data['open'],
                'high': stock_data['high'],
                'low': stock_data['low'],
                'close': stock_data['close'],
                'volume': stock_data['volume']
            })
            
            # 设置time为索引并转换为日期
            daily_data.set_index('time', inplace=True)
            
            # 转换时间戳为日期（考虑时区：先转换为UTC，再转换为东八区）
            time_index = pd.to_datetime(daily_data.index, unit='ms')
            # 对于DatetimeIndex，直接使用时区方法
            if time_index.tz is None:
                # 如果没有时区信息，先设置为UTC，再转换为东八区
                time_index = time_index.tz_localize('UTC').tz_convert('Asia/Shanghai')
            else:
                # 如果已有时区信息，直接转换为东八区
                time_index = time_index.tz_convert('Asia/Shanghai')
            # 取东八区的日期：将DatetimeIndex转换为Series，使用原始索引，然后取日期
            date_series = pd.Series(time_index, index=daily_data.index)
            daily_data['date'] = date_series.dt.date
            
            # 筛选交易日数据
            daily_data = daily_data[daily_data['date'].apply(is_tradeday)]
            
            # 按日期排序
            daily_data = daily_data.sort_values('date')
            
            # 计算技术指标（布林线）
            daily_data = self._calculate_technical_indicators(daily_data)
            
            # 如果只需要最近10个交易日的数据，进行筛选
            if not include_prev_day and len(daily_data) > 0:
                # 只保留在trading_dates范围内的数据
                daily_data = daily_data[daily_data['date'].isin(self.trading_dates)]
            
            return daily_data
            
        except Exception as e:
            logger.error(f"[{stock_code}] 获取日线数据异常: {str(e)}", exc_info=True)
            return None
    
    def _check_stock(self, stock_code: str, stock_name: str) -> Optional[Tuple[str, str]]:
        """检查股票是否符合条件"""
        try:
            # 获取日线数据（包含前一个交易日）
            # 注意：include_prev_day=True 时，会返回所有获取的数据（约60天），
            # 这样才有足够的历史数据来计算布林线（需要至少20个交易日）
            daily_data = self._get_daily_data(stock_code, include_prev_day=True)
            if daily_data is None or daily_data.empty:
                return None
            
            # limit_ratio 在循环内按 trade_date 取值
            
            # 筛选出最近10个交易日的数据
            trading_dates_set = set(self.trading_dates)
            recent_data = daily_data[daily_data['date'].isin(trading_dates_set)]
            
            if recent_data.empty:
                return None
            
            # 检查最近10个交易日中是否出现过涨停板，如果有则直接排除
            recent_data_sorted = recent_data.sort_values('date')
            for idx, row in recent_data_sorted.iterrows():
                trade_date = row['date']
                limit_ratio = self._get_limit_ratio(stock_code, trade_date)
                close_price = row['close']
                
                # 从全部数据中找到前一个交易日
                prev_trading_days = daily_data[daily_data['date'] < trade_date]
                if prev_trading_days.empty:
                    continue  # 没有前一日数据，跳过
                
                prev_row = prev_trading_days.iloc[-1]  # 获取最后一个（最近的）前一个交易日
                prev_close = prev_row['close']
                
                # 计算涨停价
                limit_up_price = prev_close * (1 + limit_ratio)
                
                # 判断是否涨停：收盘价 >= 涨停价（允许0.01的误差）
                if close_price >= limit_up_price - 0.01:
                    # 出现过涨停板，直接排除
                    return None
            
            # 获取最后一个交易日（用于排除）
            last_trading_date = recent_data['date'].max()
            
            # 按日期从早到晚排序，确保从早往后查找
            recent_data_sorted = recent_data.sort_values('date')
            
            # 遍历最近10个交易日，查找第一个接近涨停（排除最后一个交易日）
            for idx, row in recent_data_sorted.iterrows():
                trade_date = row['date']
                
                # 排除最后一个交易日，因为无法判断后续是否有更高的价格
                if trade_date >= last_trading_date:
                    continue
                
                limit_ratio = self._get_limit_ratio(stock_code, trade_date)
                
                close_price = row['close']
                high_price = row['high']
                
                # 从全部数据中找到前一个交易日
                prev_trading_days = daily_data[daily_data['date'] < trade_date]
                if prev_trading_days.empty:
                    continue  # 没有前一日数据，跳过
                
                prev_row = prev_trading_days.iloc[-1]  # 获取最后一个（最近的）前一个交易日
                prev_close = prev_row['close']
                prev_date = prev_row['date']
                
                # 计算接近涨停价（75%涨停幅度）
                # 例如：10%涨停的股票，接近涨停价 = 前收盘 * (1 + 0.10 * 0.75) = 前收盘 * 1.075
                near_limit_price = prev_close * (1 + limit_ratio * 0.75)
                
                # 判断是否接近涨停：收盘价 >= 接近涨停价
                # 允许0.01的误差，因为价格可能有微小的浮动
                if close_price < near_limit_price - 0.01:
                    continue  # 不满足接近涨停条件，继续查找
                
                # 找到第一个接近涨停，立即检查其他条件
                
                # 1. 检查收盘价是否 >= 布林线上轨
                # 获取当天的布林线上轨
                day_data = daily_data[daily_data['date'] == trade_date]
                if day_data.empty:
                    continue
                
                boll_upper = day_data['BOLL_UPPER'].iloc[0]
                # 如果收盘价 < 布林线上轨，不满足条件，直接返回None
                # 允许0.01的误差，因为价格可能有微小的浮动
                if close_price < boll_upper - 0.01:
                    return None
                
                # 2. 检查该日的最高价是否满足条件
                # 2.1 检查该日之前的交易日（在10个交易日内）的最高价是否低于该日最高价
                before_day_data = recent_data_sorted[recent_data_sorted['date'] < trade_date]
                has_before_data = not before_day_data.empty
                if has_before_data:
                    max_high_before = before_day_data['high'].max()
                    # 如果该日的最高价 < 之前的最高价，不满足条件
                    # 允许0.01的误差，因为价格可能有微小的浮动
                    if high_price < max_high_before - 0.01:
                        return None
                
                # 2.2 检查该日之后的交易日最高价不能超过当日最高价的(1+涨停板涨幅×40%)
                after_day_data = daily_data[daily_data['date'] > trade_date]
                if after_day_data.empty:
                    continue  # 如果该日后没有数据，跳过（不应该出现，因为已经排除了最后一个交易日）
                
                max_high_after = after_day_data['high'].max()
                
                # 计算允许的最大价格：当日最高价 × (1 + 涨停板涨幅 × 40%)
                # 例如：10%涨停的股票，允许最高价 = 当日最高价 × 1.04（不能高出4%）
                #      20%涨停的股票，允许最高价 = 当日最高价 × 1.08（不能高出8%）
                allowed_max_price = high_price * (1 + limit_ratio * 0.4)
                
                # 如果涨停后的最高价超过允许的最大价格，不满足条件
                # 允许0.01的误差，因为价格可能有微小的浮动
                if max_high_after > allowed_max_price + 0.01:
                    return None
                
                # 如果所有条件都满足，符合条件，返回结果（不再继续查找）
                return (trade_date.strftime('%Y-%m-%d'), f"{close_price:.2f}")
            
            return None
            
        except Exception as e:
            logger.error(f"[{stock_code}] 检查过程中出错: {str(e)}", exc_info=True)
            return None
    
    def run(self):
        """运行筛选"""
        found_count = 0
        
        try:
            total = len(self.stock_list)
            for idx, (stock_code, stock_name) in enumerate(self.stock_list):
                if not self.is_running:
                    break
                
                # 更新进度
                self.progress_updated.emit(idx + 1, total, stock_code)
                
                # 检查股票
                result = self._check_stock(stock_code, stock_name)
                if result:
                    near_limit_date, close_price = result
                    self.stock_found.emit(stock_code, stock_name, near_limit_date, close_price)
                    found_count += 1
                    # 每找到10只股票，发送一次调试信息
                    if found_count % 10 == 0:
                        self.debug_info.emit(f"已找到 {found_count} 只符合条件的股票")
                
                # 每100只股票发送一次统计信息
                if (idx + 1) % 100 == 0:
                    self.debug_info.emit(f"已筛选 {idx + 1}/{total} 只股票，找到 {found_count} 只符合条件的股票")
                
                # 短暂延迟，避免请求过快
                time.sleep(0.1)
            
            self.finished.emit(found_count)
            
        except Exception as e:
            self.error_occurred.emit(f"筛选过程出错: {str(e)}")
            self.finished.emit(found_count)


class AutoLimitUpNearDialog(QDialog):
    """自动接近涨停股票筛选对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("自动接近涨停股票筛选")
        self.setMinimumSize(800, 600)
        self.resize(1000, 700)
        
        # 去掉右上角的问号（帮助按钮）
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        self.filter_thread = None
        self.result_stocks = []  # 存储筛选结果
        self.countdown_timer = None  # 倒计时定时器
        self.countdown_seconds = 30  # 倒计时秒数
        
        self.setup_ui()
        
        # 自动开始筛选
        QTimer.singleShot(500, self.start_filter)  # 延迟500ms后自动开始，确保UI已完全加载
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout()
        
        # 说明标签
        info_label = QLabel("筛选条件：最近10个交易日内有接近涨停板（收盘价达到涨停比率的75%，例如10%涨停的股票收盘价达到7.5%以上），该日必须满足：\n1. 收盘价 >= 布林线上轨\n2. 该日之前的交易日最高价低于该日最高价\n3. 该日之后的交易日最高价不能超过该日最高价的(1+涨停板涨幅×40%)\n4. 最近10个交易日内未出现过涨停板")
        info_label.setWordWrap(True)
        font = QFont()
        font.setPointSize(10)
        info_label.setFont(font)
        layout.addWidget(info_label)
        
        # 按钮区域（只保留关闭按钮）
        button_layout = QHBoxLayout()
        
        self.close_button = QPushButton("关闭")
        self.close_button.clicked.connect(self.close)
        button_layout.addWidget(self.close_button)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # 状态标签
        self.status_label = QLabel("正在启动...")
        layout.addWidget(self.status_label)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["股票代码", "股票名称", "接近涨停日期", "收盘价"])
        
        # 设置表格属性
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        
        layout.addWidget(self.result_table)
        
        self.setLayout(layout)
    
    def _get_trading_dates(self) -> List[date]:
        """获取最近10个交易日"""
        try:
            current_time = datetime.now()
            current_date = current_time.date()
            current_hour = current_time.hour
            
            # 交易日的下午3点以后才包含当天（您通常在15:00后运行，此时会包含今天）
            include_today = is_tradeday(current_date) and (current_hour >= 15)
            
            trading_dates = []
            search_date = current_date if include_today else current_date - timedelta(days=1)
            found_days = 0
            # 节后第一天等场景：假期+周末会导致30自然日内不足10个交易日，故放宽到60天
            max_search_days = 60  # 最多往前找60天，确保长假后也能凑满10个交易日
            
            while found_days < 10 and max_search_days > 0:
                if is_tradeday(search_date):
                    trading_dates.append(search_date)
                    found_days += 1
                search_date -= timedelta(days=1)
                max_search_days -= 1
            
            # 按时间顺序排列（从早到晚）
            trading_dates.sort()
            
            return trading_dates
            
        except Exception as e:
            return []
    
    def _load_stock_list(self) -> List[Tuple[str, str]]:
        """从CSV文件加载股票列表"""
        csv_path = os.path.join(os.path.dirname(__file__), 'data', 'all_a_stocks.csv')
        if not os.path.exists(csv_path):
            QMessageBox.warning(self, "错误", f"找不到股票列表文件: {csv_path}")
            return []
        
        # 尝试多种编码方式
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
        df = None
        last_error = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                break  # 成功读取，跳出循环
            except (UnicodeDecodeError, UnicodeError) as e:
                last_error = e
                continue  # 尝试下一个编码
            except Exception as e:
                last_error = e
                continue
        
        if df is None:
            QMessageBox.warning(self, "错误", f"加载股票列表失败: 无法使用常见编码（utf-8, gbk, gb2312, gb18030）读取文件\n最后错误: {str(last_error)}")
            return []
        
        try:
            stock_list = []
            
            for _, row in df.iterrows():
                stock_code = str(row['证券代码']).zfill(6)  # 确保是6位数字
                # 只保留0、3、6开头的股票代码
                if not stock_code.startswith(('0', '3', '4','6' ,'8' ,'9')):
                    continue
                stock_name = str(row['证券简称'])
                stock_list.append((stock_code, stock_name))
            
            return stock_list
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"解析股票列表失败: {str(e)}")
            return []
    
    def start_filter(self):
        """开始筛选"""
        try:
            # 加载股票列表
            stock_list = self._load_stock_list()
            if not stock_list:
                return
            
            # 获取交易日列表（节后第一天等会放宽搜索范围，尽量凑满10个交易日）
            trading_dates = self._get_trading_dates()
            if len(trading_dates) < 2:
                QMessageBox.warning(self, "警告", f"无法获取足够的交易日（只找到{len(trading_dates)}个）")
                return
            if len(trading_dates) < 10:
                self.status_label.setText(f"提示：仅找到最近{len(trading_dates)}个交易日，将按此范围筛选")
            
            # 清空结果
            self.result_stocks = []
            self.result_table.setRowCount(0)
            
            # 创建筛选线程
            self.filter_thread = LimitUpNearFilterThread(stock_list, trading_dates, self)
            self.filter_thread.progress_updated.connect(self.on_progress_updated)
            self.filter_thread.stock_found.connect(self.on_stock_found)
            self.filter_thread.finished.connect(self.on_finished)
            self.filter_thread.error_occurred.connect(self.on_error)
            self.filter_thread.debug_info.connect(self.on_debug_info)
            
            # 更新状态
            self.status_label.setText("正在筛选股票...")
            
            # 启动线程
            self.filter_thread.start()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"启动筛选失败: {str(e)}")
    
    def on_progress_updated(self, current: int, total: int, stock_code: str):
        """更新进度"""
        progress = int((current / total) * 100) if total > 0 else 0
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"正在筛选: {stock_code} ({current}/{total})")
    
    def on_stock_found(self, stock_code: str, stock_name: str, near_limit_date: str, close_price: str):
        """找到符合条件的股票"""
        # 添加到结果列表
        self.result_stocks.append({
            'code': stock_code,
            'name': stock_name,
            'near_limit_date': near_limit_date,
            'close_price': close_price
        })
        
        # 添加到表格
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        self.result_table.setItem(row, 0, QTableWidgetItem(stock_code))
        self.result_table.setItem(row, 1, QTableWidgetItem(stock_name))
        self.result_table.setItem(row, 2, QTableWidgetItem(near_limit_date))
        self.result_table.setItem(row, 3, QTableWidgetItem(close_price))
        
        # 滚动到最新添加的行，确保用户能看到
        self.result_table.scrollToItem(self.result_table.item(row, 0))
        
        # 更新状态标签显示最新找到的股票
        self.status_label.setText(f"找到符合条件的股票: {stock_code} {stock_name} (接近涨停日期: {near_limit_date}, 收盘价: {close_price})")
    
    def on_finished(self, count: int):
        """筛选完成"""
        self.status_label.setText(f"筛选完成，找到 {count} 只符合条件的股票，正在保存...")
        
        # 自动保存
        self.auto_save_results()
    
    def on_error(self, error_msg: str):
        """处理错误"""
        QMessageBox.warning(self, "错误", error_msg)
        # 即使出错也尝试保存已有结果
        if self.result_stocks:
            self.auto_save_results()
    
    def on_debug_info(self, info: str):
        """处理调试信息"""
        # 在状态标签中显示调试信息
        self.status_label.setText(info)
    
    def auto_save_results(self):
        """自动保存结果（只保存Excel格式）"""
        try:
            # 创建history_data目录（如果不存在）
            history_dir = os.path.join(os.path.dirname(__file__), 'history_data')
            os.makedirs(history_dir, exist_ok=True)
            
            # 生成文件名（只生成Excel文件名）
            base_filename = f"10日内接近涨停_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            excel_file_path = os.path.join(history_dir, f"{base_filename}.xlsx")
            
            # 准备数据
            if self.result_stocks:
                df = pd.DataFrame(self.result_stocks)
                # 确保股票代码格式为6位数字（去掉可能的后缀）
                if 'code' in df.columns:
                    # 使用更兼容的方法处理股票代码格式
                    def clean_code(code):
                        code_str = str(code).strip()
                        # 去掉市场后缀（如 .SZ, .SH, .BJ）
                        if '.' in code_str:
                            code_str = code_str.split('.')[0]
                        # 确保是6位数字
                        return code_str.zfill(6) if code_str.isdigit() else code_str
                    df['code'] = df['code'].apply(clean_code)
            else:
                # 即使没有结果也创建一个空DataFrame
                df = pd.DataFrame(columns=['code', 'name', 'near_limit_date', 'close_price'])
            
            # 保存为Excel
            try:
                save_excel_with_text_code(excel_file_path, df)
                logger.info(f"已保存 {len(self.result_stocks)} 条结果到Excel: {excel_file_path}")
                if self.result_stocks:
                    self.status_label.setText(f"结果已保存到Excel: {base_filename}，正在添加主力净流入...")
                else:
                    self.status_label.setText(f"未找到符合条件的股票，已保存空结果到Excel: {base_filename}，正在添加主力净流入...")
            except ImportError:
                # 如果没有安装openpyxl，提示错误
                logger.error("未安装openpyxl，无法保存Excel文件")
                QMessageBox.warning(self, "错误", "未安装openpyxl，无法保存Excel文件。请安装: pip install openpyxl")
                self.status_label.setText(f"保存失败：未安装openpyxl")
                self.start_countdown()
                return
            except Exception as e:
                logger.error(f"保存Excel文件时出错: {str(e)}", exc_info=True)
                QMessageBox.warning(self, "错误", f"保存Excel文件失败: {str(e)}")
                self.status_label.setText(f"保存失败: {str(e)}")
                self.start_countdown()
                return
            
            # 更新UI
            self.status_label.repaint()
            QApplication.processEvents()  # 确保UI更新
            
            # 自动添加主力净流入列
            if InflowAdder is not None:
                try:
                    self.status_label.setText(f"正在为文件添加主力净流入列...")
                    self.status_label.repaint()
                    QApplication.processEvents()
                    
                    # 先读取Excel文件为DataFrame，添加主力净流入列，再保存回Excel
                    logger.info(f"开始为文件添加主力净流入列: {excel_file_path}")
                    try:
                        df_temp = pd.read_excel(excel_file_path, engine='openpyxl')
                        # 临时保存为CSV用于InflowAdder处理
                        import tempfile
                        temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8-sig')
                        temp_csv_path = temp_csv.name
                        temp_csv.close()
                        df_temp.to_csv(temp_csv_path, index=False, encoding='utf-8-sig')
                        
                        adder = InflowAdder()
                        success = adder.add_inflow_column(temp_csv_path)
                        
                        if success:
                            # 读取更新后的CSV并保存回Excel
                            df_updated = pd.read_csv(temp_csv_path, encoding='utf-8-sig')
                            save_excel_with_text_code(excel_file_path, df_updated)
                            logger.info(f"已更新Excel文件（添加主力净流入列）: {excel_file_path}")
                            self.status_label.setText(f"结果已保存到Excel: {base_filename}（已添加主力净流入列），正在添加概念、行业和板块列...")
                        else:
                            self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列失败），正在添加概念、行业和板块列...")
                            logger.warning(f"为文件添加主力净流入列失败: {excel_file_path}")
                        
                        # 删除临时CSV文件
                        try:
                            os.unlink(temp_csv_path)
                        except:
                            pass
                    except Exception as e:
                        logger.error(f"添加主力净流入列时出错: {str(e)}", exc_info=True)
                        self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列时出错），正在添加概念、行业和板块列...")
                except Exception as e:
                    logger.error(f"添加主力净流入列时出错: {str(e)}", exc_info=True)
                    self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加主力净流入列时出错），正在添加概念、行业和板块列...")
            else:
                logger.warning("主力净流入工具未启用（InflowAdder 为 None）")
                self.status_label.setText(f"结果已保存到Excel: {base_filename}（未启用主力净流入功能），正在添加概念、行业和板块列...")
            
            # 自动添加概念、行业和板块列
            try:
                self.status_label.setText(f"正在为文件添加概念、行业和板块列...")
                self.status_label.repaint()
                QApplication.processEvents()
                
                # 先读取Excel文件，添加列，再保存回Excel
                try:
                    df_temp = pd.read_excel(excel_file_path, engine='openpyxl')
                    # 临时保存为CSV用于add_stock_info_columns处理
                    import tempfile
                    temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8-sig')
                    temp_csv_path = temp_csv.name
                    temp_csv.close()
                    df_temp.to_csv(temp_csv_path, index=False, encoding='utf-8-sig')
                    
                    success = add_stock_info_columns(temp_csv_path)
                    
                    if success:
                        # 读取更新后的CSV并保存回Excel
                        df_updated = pd.read_csv(temp_csv_path, encoding='utf-8-sig')
                        save_excel_with_text_code(excel_file_path, df_updated)
                        logger.info(f"已更新Excel文件（添加概念、行业和板块列）: {excel_file_path}")
                        self.status_label.setText(f"结果已保存到Excel: {base_filename}（已添加所有列），{self.countdown_seconds}秒后自动关闭...")
                    else:
                        self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加概念、行业和板块列失败），{self.countdown_seconds}秒后自动关闭...")
                        logger.warning(f"为文件添加概念、行业和板块列失败: {excel_file_path}")
                    
                    # 删除临时CSV文件
                    try:
                        os.unlink(temp_csv_path)
                    except:
                        pass
                except Exception as e:
                    logger.error(f"添加概念、行业和板块列时出错: {str(e)}", exc_info=True)
                    self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加概念、行业和板块列时出错），{self.countdown_seconds}秒后自动关闭...")
            except Exception as e:
                logger.error(f"添加概念、行业和板块列时出错: {str(e)}", exc_info=True)
                self.status_label.setText(f"结果已保存到Excel: {base_filename}（添加概念、行业和板块列时出错），{self.countdown_seconds}秒后自动关闭...")
            
            # 启动倒计时
            self.start_countdown()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")
            logger.error(f"保存结果时出错: {str(e)}", exc_info=True)
            # 即使保存失败也启动倒计时
            self.start_countdown()
    
    def start_countdown(self):
        """启动30秒倒计时"""
        self.countdown_seconds = 30
        self.countdown_timer = QTimer()
        self.countdown_timer.timeout.connect(self.update_countdown)
        self.countdown_timer.start(1000)  # 每秒更新一次
        self.update_countdown()  # 立即更新一次
    
    def update_countdown(self):
        """更新倒计时"""
        if self.countdown_seconds > 0:
            # 更新状态标签显示倒计时
            current_text = self.status_label.text()
            # 如果状态标签包含倒计时信息，更新它；否则追加
            if "秒后自动关闭" in current_text:
                # 替换倒计时数字
                import re
                new_text = re.sub(r'\d+秒后自动关闭', f'{self.countdown_seconds}秒后自动关闭', current_text)
                self.status_label.setText(new_text)
            else:
                self.status_label.setText(f"{current_text}，{self.countdown_seconds}秒后自动关闭...")
            
            self.countdown_seconds -= 1
        else:
            # 倒计时结束，关闭程序
            if self.countdown_timer:
                self.countdown_timer.stop()
            self.close()
    
    def closeEvent(self, event):
        """关闭事件"""
        if self.filter_thread and self.filter_thread.isRunning():
            # 如果正在筛选，停止线程
            self.filter_thread.stop()
            self.filter_thread.wait()
        
        if self.countdown_timer:
            self.countdown_timer.stop()
        
        event.accept()


def main():
    app = QApplication(sys.argv)
    
    dialog = AutoLimitUpNearDialog()
    dialog.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    #如果今天不是交易日，则跳过
    if not is_tradeday():
        print("今天不是交易日，跳过")
        sys.exit(0)
    main()

